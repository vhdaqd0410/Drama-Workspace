# -*- coding: utf-8 -*-
"""融合功能回归测试：交付联动、全局待办、交付日期同步。"""
import os
import sys
import pytest

# ============================================================
# 1. 交付日期联动：置为已交付时自动补记 delivered_date
# ============================================================

class TestDeliveredDateLinkage:
    def test_delivered_auto_sets_date(self, tmp_db):
        tmp_db.upsert_project("联动A", "", "")
        tmp_db.update_project_status("联动A", delivery_status="delivered",
                                     last_delivered_at="2026-08-20 10:00:00")
        p = tmp_db.get_project("联动A")
        assert p["delivered_date"] == "2026-08-20"

    def test_delivered_uses_today_if_no_timestamp(self, tmp_db):
        tmp_db.upsert_project("联动B", "", "")
        tmp_db.update_project_status("联动B", delivery_status="delivered")
        p = tmp_db.get_project("联动B")
        # 无时间戳时用当天日期（YYYY-MM-DD 格式）
        assert p["delivered_date"] and len(p["delivered_date"]) == 10

    def test_keeps_existing_delivered_date(self, tmp_db):
        tmp_db.upsert_project("联动C", "", "")
        tmp_db.update_project_status("联动C", delivered_date="2026-08-01")
        tmp_db.update_project_status("联动C", delivery_status="delivered",
                                     last_delivered_at="2026-08-20 10:00:00")
        p = tmp_db.get_project("联动C")
        assert p["delivered_date"] == "2026-08-01"  # 已有日期不被覆盖

    def test_not_delivered_no_auto(self, tmp_db):
        tmp_db.upsert_project("联动D", "", "")
        tmp_db.update_project_status("联动D", delivery_status="pending")
        p = tmp_db.get_project("联动D")
        assert not (p.get("delivered_date") or "")


# ============================================================
# 2. 全局待办（get_all_todos）
# ============================================================

class TestGlobalTodos:
    def test_get_all_todos(self, tmp_db):
        tmp_db.upsert_project("待办项目A", "", "")
        tmp_db.upsert_project("待办项目B", "", "")
        tmp_db.add_project_todo("待办项目A", "完成第一集", priority=1)
        tmp_db.add_project_todo("待办项目A", "完成第二集", priority=0)
        tmp_db.add_project_todo("待办项目B", "校色", priority=0)

        todos = tmp_db.get_all_todos()
        assert len(todos) == 3
        # 未完成优先，priority 高的在前
        assert todos[0]["text"] == "完成第一集"

    def test_include_done_flag(self, tmp_db):
        tmp_db.upsert_project("待办项目C", "", "")
        t = tmp_db.add_project_todo("待办项目C", "已完成的待办", priority=0)
        tmp_db.update_project_todo(t, done=1)
        assert len(tmp_db.get_all_todos(include_done=False)) == 0
        assert len(tmp_db.get_all_todos(include_done=True)) == 1

    def test_keyword_filter(self, tmp_db):
        tmp_db.upsert_project("待办项目D", "", "")
        tmp_db.add_project_todo("待办项目D", "配音检查", priority=0)
        tmp_db.add_project_todo("待办项目D", "字幕导出", priority=0)
        assert len(tmp_db.get_all_todos(keyword="配音")) == 1
        assert len(tmp_db.get_all_todos(keyword="不存在的内容")) == 0


# ============================================================
# 3. 交付日期同步（sync_delivery_dates_from_target 解析）
# ============================================================

class TestSyncDeliveryDates:
    def test_parse_short_and_long(self, tmp_db, tmp_path):
        import openpyxl
        p = str(tmp_path / "target.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:A2")
        ws["A1"] = "同步项目A"
        ws["D1"] = "8.15上午10点交"
        ws.merge_cells("A3:A3")
        ws["A3"] = "同步项目B"
        ws["D3"] = "2026-8-20"
        wb.save(p)

        tmp_db.upsert_project("同步项目A", "", "")
        tmp_db.upsert_project("同步项目B", "", "")
        from features import sync_delivery_dates_from_target
        updated = sync_delivery_dates_from_target(p, tmp_db)
        assert len(updated) == 2
        dates = {u["project"]: u["date"] for u in updated}
        assert dates["同步项目A"] == "2026-08-15"
        assert dates["同步项目B"] == "2026-08-20"
        # 验证已写入
        assert tmp_db.get_project("同步项目A")["delivered_date"] == "2026-08-15"


# ============================================================
# 4. 分集工作量同步（sync_episode_plan_from_target）
# ============================================================

class TestEpisodePlanSync:
    def test_sync_rebuilds_plan(self, tmp_db, tmp_path):
        import openpyxl
        p = str(tmp_path / "master.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["项目X", None, None, None])
        ws.append([None, None, "张三：1-3", None])
        ws.append([None, None, "李四：4-6", None])
        ws.append(["项目Y", None, None, None])
        ws.append([None, None, "张三：1-2，5", None])
        wb.save(p)

        tmp_db.upsert_project("项目X", "", "")
        tmp_db.upsert_project("项目Y", "", "")
        from features import sync_episode_plan_from_target
        res = sync_episode_plan_from_target(p, tmp_db, clear_stale=False)
        assert res["total_projects"] == 2
        assert res["total_episodes"] == 9   # 项目X 6集 + 项目Y 3集
        px = tmp_db.get_episode_plan("项目X")
        assert px["1"] == "张三" and px["6"] == "李四"
        py = tmp_db.get_episode_plan("项目Y")
        assert py["5"] == "张三"

    def test_clear_stale(self, tmp_db, tmp_path):
        import openpyxl
        p = str(tmp_path / "m.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["项目A", None, None, None])
        ws.append([None, None, "张三：1-2", None])
        wb.save(p)
        tmp_db.upsert_project("项目A", "", "")
        tmp_db.upsert_project("项目B", "", "")   # 不在目标文件
        tmp_db.set_episode_plan("项目B", {"1": "假人", "2": "假人"})
        from features import sync_episode_plan_from_target
        res = sync_episode_plan_from_target(p, tmp_db, clear_stale=True)
        assert "项目B" in res["cleared"]
        assert tmp_db.get_episode_plan("项目B") == {}

    def test_sync_populates_independent_workload(self, tmp_db, tmp_path):
        # 同一项目内分集范围重叠（张靖杰21-25 与 张淯升16-28）：
        # editor_workload 必须独立计数，不被 episode_plan 覆盖丢失。
        import openpyxl
        p = str(tmp_path / "ov.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["项目O", None, None, None])
        ws.append([None, None, "张靖杰：4-7,21-25", None])
        ws.append([None, None, "张淯升：16-28", None])
        wb.save(p)
        tmp_db.upsert_project("项目O", "", "")
        from features import sync_episode_plan_from_target
        res = sync_episode_plan_from_target(p, tmp_db, clear_stale=False)
        wl = tmp_db.get_editor_workload("项目O")
        # 独立计数：张靖杰 4-7(4) + 21-25(5) = 9；张淯升 16-28(13)
        assert wl["张靖杰"] == 9
        assert wl["张淯升"] == 13
        # episode_plan 因重叠，张靖杰只剩4集(21-25被张淯升覆盖)
        plan = tmp_db.get_episode_plan("项目O")
        zhang = sum(1 for v in plan.values() if v == "张靖杰")
        assert zhang == 4


# ============================================================
# 5. 统一剪辑师工作量口径（aggregate_editor_workload）
# ============================================================

class TestAggregateEditorWorkload:
    def test_aggregate_all(self, tmp_db):
        tmp_db.upsert_project("P1", "", "")
        tmp_db.upsert_project("P2", "", "")
        tmp_db.set_episode_plan("P1", {"1": "张三", "2": "张三", "3": "李四"})
        tmp_db.set_episode_plan("P2", {"1": "李四"})
        from features import aggregate_editor_workload
        wl = aggregate_editor_workload(tmp_db)
        by_name = {e["name"]: e["assigned"] for e in wl}
        assert by_name["张三"] == 2 and by_name["李四"] == 2
        assert len(wl) == 2

    def test_aggregate_by_month(self, tmp_db):
        tmp_db.upsert_project("P3", "", "")
        tmp_db.upsert_project("P4", "", "")
        tmp_db.update_project_status("P3", project_month="2026-08")
        tmp_db.update_project_status("P4", project_month="2026-07")
        tmp_db.set_episode_plan("P3", {"1": "张三"})
        tmp_db.set_episode_plan("P4", {"2": "李四"})
        from features import aggregate_editor_workload
        wl = aggregate_editor_workload(tmp_db, month="2026-08")
        by_name = {e["name"]: e["assigned"] for e in wl}
        assert by_name.get("张三") == 1
        assert "李四" not in by_name

    def test_aggregate_uses_episode_plan_as_final(self, tmp_db):
        # 分集详情(episode_plan)是最终数据源：即使存在 editor_workload，也以 episode_plan 为准
        # （用户在分集管理手动校准过，episode_plan 即最终口径）
        tmp_db.upsert_project("P5", "", "")
        tmp_db.set_episode_plan("P5", {"4": "张靖杰", "5": "张靖杰", "6": "张靖杰",
                                        "7": "张靖杰", "21": "张淯升", "22": "张淯升"})
        # editor_workload 是旧目标文件推导的旧值，应被忽略
        tmp_db.set_editor_workload("P5", {"张靖杰": 9, "张淯升": 13})
        from features import aggregate_editor_workload
        wl = aggregate_editor_workload(tmp_db)
        by_name = {e["name"]: e["assigned"] for e in wl}
        # 以当前 episode_plan 为准：张靖杰=4（重叠已由用户在UI校准解决）
        assert by_name["张靖杰"] == 4
        assert by_name["张淯升"] == 2

    def test_aggregate_from_episode_plan(self, tmp_db):
        tmp_db.upsert_project("P6", "", "")
        tmp_db.set_episode_plan("P6", {"1": "张三", "2": "张三"})
        from features import aggregate_editor_workload
        wl = aggregate_editor_workload(tmp_db)
        by_name = {e["name"]: e["assigned"] for e in wl}
        assert by_name["张三"] == 2


# ============================================================
# 6. 以分集数据为准同步工作量（sync_workload_from_episode_plan）
# ============================================================

class TestSyncWorkloadFromEpisodePlan:
    def test_rebuilds_workload_from_episode_plan(self, tmp_db):
        # 分集数据是最终口径：sync 应从 episode_plan 重建 editor_workload，不读Excel
        tmp_db.upsert_project("PA", "", "")
        tmp_db.set_episode_plan("PA", {"1": "张三", "2": "张三", "3": "李四"})
        # 先放一个错误旧值
        tmp_db.set_editor_workload("PA", {"张三": 99})
        from features import sync_workload_from_episode_plan
        res = sync_workload_from_episode_plan(tmp_db)
        assert res["total_projects"] == 1
        assert res["total_episodes"] == 3
        wl = tmp_db.get_editor_workload("PA")
        assert wl == {"张三": 2, "李四": 1}

    def test_leaves_episode_plan_untouched(self, tmp_db):
        tmp_db.upsert_project("PB", "", "")
        tmp_db.set_episode_plan("PB", {"1": "王五", "2": "王五"})
        from features import sync_workload_from_episode_plan
        sync_workload_from_episode_plan(tmp_db)
        # episode_plan 不被覆盖
        assert tmp_db.get_episode_plan("PB") == {"1": "王五", "2": "王五"}
        assert tmp_db.get_editor_workload("PB") == {"王五": 2}


# ============================================================
# 7. 数据洞察 KPI 口径（compute_insights_summary）
# ============================================================

class TestInsightsSummary:
    def _mk(self, tmp_db, name, month, status, total_eps=70):
        tmp_db.upsert_project(name, "", "")
        tmp_db.update_project_status(name, project_month=month, custom_status=status)
        try:
            tmp_db.set_episodes(name, total_eps, total_eps)
        except Exception:
            pass
        return name

    def test_month_completed_uses_custom_status(self, tmp_db):
        # 关键口径：本月已完成 = 本月项目里 custom_status=='已完成'（而非 delivered_date）
        self._mk(tmp_db, "A", "2026-08", "已完成")
        self._mk(tmp_db, "B", "2026-08", "剪辑中")
        self._mk(tmp_db, "C", "2026-07", "已完成")  # 其他月不算
        from features import compute_insights_summary
        s = compute_insights_summary(tmp_db, month="2026-08")
        assert s["monthProjectCount"] == 2
        assert s["monthCompleted"] == 1
        assert s["inProgress"] == 1

    def test_active_project_filter(self, tmp_db):
        # 空壳项目（无状态/无集数/未交付）不计入
        tmp_db.upsert_project("空壳", "", "")   # 无任何制作痕迹
        self._mk(tmp_db, "实项目", "2026-08", "剪辑中")
        from features import compute_insights_summary
        s = compute_insights_summary(tmp_db, month="2026-08")
        assert s["projectCount"] == 1
        assert "空壳" not in s["statusMap"]
        assert s["monthProjectCount"] == 1

    def test_editor_episodes_from_episode_plan(self, tmp_db):
        self._mk(tmp_db, "P", "2026-08", "剪辑中")
        tmp_db.set_episode_plan("P", {"1": "张三", "2": "张三", "3": "李四"})
        from features import compute_insights_summary
        s = compute_insights_summary(tmp_db, month="2026-08")
        assert s["editorEpisodes"] == {"张三": 2, "李四": 1}


# ============================================================
# 8. 统一分集解析器（fenji_parser）—— 消除多份重复漂移
# ============================================================

class TestFenjiParser:
    def test_parse_assign_line_fullwidth_separators(self):
        from fenji_parser import parse_assign_line
        plan = {}
        # 全角分号、加号、顿号、中横线都要解析
        parse_assign_line(plan, "张三：1-3，44-45")
        parse_assign_line(plan, "李四：5；6-7")
        parse_assign_line(plan, "王五：8 9")  # 空格分隔
        assert plan == {"1": "张三", "2": "张三", "3": "张三", "44": "张三", "45": "张三",
                        "5": "李四", "6": "李四", "7": "李四",
                        "8": "王五", "9": "王五"}

    def test_parse_assign_line_descending_range(self):
        from fenji_parser import parse_assign_line
        plan = {}
        parse_assign_line(plan, "张三：10-7")  # 降序范围
        assert plan == {"7": "张三", "8": "张三", "9": "张三", "10": "张三"}

    def test_expand_ranges(self):
        from fenji_parser import expand_ranges
        assert expand_ranges("4-7,21-25") == {4, 5, 6, 7, 21, 22, 23, 24, 25}
        assert expand_ranges("3") == {3}

    def test_all_three_call_sites_delegate_to_same_impl(self):
        # 三处解析入口最终都落到同一实现，行为一致
        from fenji_parser import parse_assign_line
        from features import _parse_assign_range_str
        from enhanced_routes import _parse_assign_line as er_parse
        p1, p2, p3 = {}, {}, {}
        line = "张三：1-3；44-45"  # 含全角分号
        parse_assign_line(p1, line)
        _parse_assign_range_str(p2, line)
        er_parse(p3, line)
        assert p1 == p2 == p3


# ============================================================
# 9. 交付日历增强（compute_delivery_stats）—— 延迟预警 + 按时交付率
# ============================================================

class TestDeliveryStats:
    def _mk(self, tmp_db, name, month, status, delivered="", due=""):
        tmp_db.upsert_project(name, "", "")
        kw = {"project_month": month, "custom_status": status}
        if delivered: kw["delivered_date"] = delivered
        if due: kw["due_date"] = due
        tmp_db.update_project_status(name, **kw)

    def test_ontime_rate_and_late(self, tmp_db):
        self._mk(tmp_db, "A", "2026-08", "已完成", delivered="2026-08-10", due="2026-08-15")  # 按时
        self._mk(tmp_db, "B", "2026-08", "已完成", delivered="2026-08-20", due="2026-08-15")  # 迟交
        from features import compute_delivery_stats
        s = compute_delivery_stats(tmp_db, month="2026-08")
        assert s["delivered_count"] == 2
        assert s["ontime_count"] == 1
        assert s["late_count"] == 1
        assert s["on_time_rate"] == 50.0

    def test_overdue_undelivered(self, tmp_db):
        # due_date 已过（相对今天）但未交付 → 预警
        self._mk(tmp_db, "C", "2026-08", "剪辑中", delivered="", due="2020-01-01")
        from features import compute_delivery_stats
        s = compute_delivery_stats(tmp_db, month="2026-08")
        assert s["undelivered_count"] == 1
        assert s["overdue_count"] == 1
        assert s["overdue"][0]["name"] == "C"


# ============================================================
# 10. 提成/绩效全链路（commission_service.compute_commission_breakdown）
# ============================================================

class TestCommissionBreakdown:
    def _mk_editor(self, name, eps):
        return {"name": name, "assigned": eps, "projects": 1}

    def test_一卡_超额(self):
        from commission_service import compute_commission_breakdown
        rows, summary = compute_commission_breakdown(
            [self._mk_editor("陈春阳", 90), self._mk_editor("程梦", 80)])
        by = {r["name"]: r for r in rows}
        # 陈春阳 一卡(基准70)，超额20集×20=400
        assert by["陈春阳"]["role"] == "一卡剪辑"
        assert by["陈春阳"]["commission"] == (90 - 70) * 20
        assert by["陈春阳"]["is_complete"] is True

    def test_二卡_缺集扣款(self):
        from commission_service import compute_commission_breakdown
        rows, summary = compute_commission_breakdown(
            [self._mk_editor("王田田", 100)])  # 二卡基准120，缺20×50
        by = {r["name"]: r for r in rows}
        assert by["王田田"]["role"] == "二卡剪辑"
        assert by["王田田"]["is_complete"] is False
        assert by["王田田"]["commission"] == -(120 - 100) * 50

    def test_组长_组奖(self):
        from commission_service import compute_commission_breakdown
        rows, summary = compute_commission_breakdown(
            [{"name": "张大强", "assigned": 8, "projects": 4}])
        by = {r["name"]: r for r in rows}
        # 组长：8×20 + 4×100 = 160+400=560
        assert by["张大强"]["role"] == "剪辑组长"
        assert by["张大强"]["commission"] == 8 * 20 + 4 * 100
        assert summary["total_people"] == 1

    def test_summary_totals(self):
        from commission_service import compute_commission_breakdown
        rows, summary = compute_commission_breakdown([
            self._mk_editor("陈春阳", 90),
            self._mk_editor("王田田", 100),
        ])
        # 陈春阳 +400，王田田 -1000，合计 -600
        assert summary["total_commission"] == (90 - 70) * 20 - (120 - 100) * 50
        assert summary["met_quota"] == 1
        assert summary["unmet_quota"] == 1


# ============================================================
# 11. 个人工作量卡片（compute_person_cards）—— 年度逐月趋势
# ============================================================

class TestPersonCards:
    def test_annual_trend(self, tmp_db):
        tmp_db.upsert_project("P1", "", "")
        tmp_db.upsert_project("P2", "", "")
        tmp_db.update_project_status("P1", project_month="2026-07")
        tmp_db.update_project_status("P2", project_month="2026-08")
        tmp_db.set_episode_plan("P1", {"1": "张三", "2": "张三"})
        tmp_db.set_episode_plan("P2", {"1": "张三", "3": "李四"})
        from commission_service import compute_person_cards
        data = compute_person_cards(tmp_db, year="2026")
        by = {c["name"]: c for c in data["cards"]}
        assert by["张三"]["annual_total"] == 3  # 7月2集 + 8月1集
        assert by["张三"]["month_count"] == 2
        assert by["张三"]["best_month"] == "2026-07"
        assert by["李四"]["annual_total"] == 1
        assert data["year"] == "2026"

    def test_filters_by_year(self, tmp_db):
        tmp_db.upsert_project("P1", "", "")
        tmp_db.update_project_status("P1", project_month="2025-12")
        tmp_db.set_episode_plan("P1", {"1": "张三"})
        from commission_service import compute_person_cards
        data = compute_person_cards(tmp_db, year="2026")
        assert data["cards"] == []  # 2025年的不算进2026


# ============================================================
# 12. 审核流超时提醒（compute_audit_alerts）—— 责任到人 + 超时预警
# ============================================================

class TestAuditAlerts:
    def test_alert_stale_review(self, tmp_db):
        from datetime import datetime, timedelta
        import features
        # 卡在质检中10天的项目
        tmp_db.upsert_project("慢项目", "", "")
        old = (datetime.now() - timedelta(days=10)).strftime("%Y-%m-%d %H:%M:%S")
        tmp_db.update_project_status("慢项目", custom_status="质检中", status_changed_at=old)
        tmp_db.set_episode_plan("慢项目", {"1": "张三"})
        # 刚变状态的项目不预警
        tmp_db.upsert_project("新项目", "", "")
        tmp_db.update_project_status("新项目", custom_status="质检中")
        # 已完成不算
        tmp_db.upsert_project("完成", "", "")
        tmp_db.update_project_status("完成", custom_status="已完成")
        data = features.compute_audit_alerts(tmp_db, stale_days=3)
        names = {a["name"]: a for a in data["alerts"]}
        assert "慢项目" in names
        assert names["慢项目"]["days_stuck"] >= 9.9
        # 责任到人：从分集回退到张三
        assert names["慢项目"]["owner"] == "张三"
        assert "新项目" not in names
        assert "完成" not in names

    def test_owner_explicit(self, tmp_db):
        from datetime import datetime, timedelta
        import features
        tmp_db.upsert_project("P", "", "")
        old = (datetime.now() - timedelta(days=6)).strftime("%Y-%m-%d %H:%M:%S")
        tmp_db.update_project_status("P", custom_status="修改中", status_changed_at=old, owner="李四")
        data = features.compute_audit_alerts(tmp_db, stale_days=3)
        assert any(a["name"] == "P" and a["owner"] == "李四" for a in data["alerts"])


