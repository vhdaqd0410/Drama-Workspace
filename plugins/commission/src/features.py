# -*- coding: utf-8 -*-
"""AI后期剪辑提成工具 - 扩展功能模块 v2.0"""
import os, re, json, shutil, datetime, html as html_mod, random
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 共享工具函数
from utils import parse_episode_ranges, parse_overtime_episodes, merge_episode_ranges


# ===================== 工具函数 =====================

def _find_chinese_font():
    """自动查找系统中可用的中文字体"""
    candidates = [
        r'C:\Windows\Fonts\simsun.ttc',
        r'C:\Windows\Fonts\msyh.ttc',
        r'C:\Windows\Fonts\msyhbd.ttc',
        r'C:\Windows\Fonts\simkai.ttf',
    ]
    for f in candidates:
        if os.path.exists(f):
            return f
    return None


def _escape_html(text):
    """HTML转义"""
    return html_mod.escape(str(text))


# ===================== 1. 生成下月模板 =====================

def generate_next_month_template(template_path, output_dir=None):
    if output_dir is None:
        output_dir = os.path.dirname(template_path)
    CN_MONTHS = ['', '一月','二月','三月','四月','五月','六月',
                 '七月','八月','九月','十月','十一月','十二月']
    wb = load_workbook(template_path)
    ws = wb.active

    title_cell = None
    for col in [2, 1]:
        val = ws.cell(1, col).value
        if val and re.search(r'\d+年\d+月', str(val)):
            title_cell = (1, col)
            break
    if not title_cell:
        return None, "无法识别模板中的年月标题"

    old_title = str(ws.cell(*title_cell).value)
    m = re.search(r'(\d+)年(\d+)月', old_title)
    if not m:
        return None, "无法解析月份"

    year, month = int(m.group(1)), int(m.group(2))
    month += 1
    if month > 12:
        month = 1; year += 1

    new_title = re.sub(r'\d+年\d+月', f'{year}年{month:02d}月', old_title)
    ws.cell(*title_cell, new_title)

    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= 4:
            ws.unmerge_cells(str(mc))
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=20):
        for cell in row:
            try: cell.value = None
            except AttributeError: pass

    cn = CN_MONTHS[month]
    output = os.path.join(output_dir, f'AI后期剪辑提成一组模板_{year}年{month:02d}月.xlsx')
    wb.save(output); wb.close()
    return output, f"下月模板已生成：{year}年{month:02d}月（{cn}）"


# ===================== 2. 导出PDF =====================

def export_to_pdf(excel_path, pdf_path=None):
    if pdf_path is None:
        pdf_path = excel_path.replace('.xlsx', '.pdf')

    font_path = _find_chinese_font()
    if not font_path:
        raise RuntimeError("未找到可用中文字体，PDF导出失败")

    from fpdf import FPDF
    wb = load_workbook(excel_path)
    ws = wb.active

    pdf = FPDF(orientation='L', unit='mm', format='A3')
    pdf.add_font('CJK', '', font_path, uni=True)
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    col_widths = [8, 12, 12, 5, 8, 10, 18, 10, 10, 15, 8, 8, 8, 8, 8, 8, 15, 8, 6, 12]
    headers = [str(ws.cell(3, c).value)[:6] if ws.cell(3, c).value else '' for c in range(1, 21)]

    def draw_header(y):
        for i, hdr in enumerate(headers):
            x = 5 + sum(col_widths[:i])
            pdf.set_xy(x, y)
            pdf.set_font('CJK', '', 5)
            pdf.cell(col_widths[i], 6, hdr, border=1, align='C')

    draw_header(10)
    row_y = 16
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=20):
        if row_y > 280:
            pdf.add_page(); row_y = 10
            draw_header(row_y); row_y += 6
        for ci, cell in enumerate(row[:20]):
            v = str(cell.value)[:12] if cell.value is not None else ''
            x = 5 + sum(col_widths[:ci])
            pdf.set_xy(x, row_y)
            pdf.set_font('CJK', '', 5)
            pdf.cell(col_widths[ci], 6, v, border=1, align='C')
        row_y += 6

    pdf.output(pdf_path); wb.close()
    return pdf_path


# ===================== 3. 个人绩效卡片 =====================

def generate_person_cards(records, commission_data, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    persons = {}
    for r in records:
        nm = r['身份证姓名']
        if nm not in persons:
            persons[nm] = {'projects': [], 'total_eps': 0}
        persons[nm]['projects'].append(r)
        persons[nm]['total_eps'] += r['单项目数/集数']

    cards = []
    for nm in sorted(persons.keys()):
        cd = commission_data.get(nm, {})
        proj_rows = ''
        for p in persons[nm]['projects']:
            pid = _escape_html(str(p['项目ID']))
            pname = _escape_html(str(p['AI项目名称'])[:30])
            d = p['结束日期'].strftime('%m/%d') if p['结束日期'] else ''
            ot = p.get('超时集数', 0)
            eps_display = p['单项目数/集数']
            ot_mark = ''
            if ot > 0:
                ot_mark = f' <span style="background:#fef3c7;color:#d97706;font-size:10px;padding:1px 4px;border-radius:3px">含{ot}集超4分</span>'
            proj_rows += f'<tr><td>{pid}</td><td>{pname}</td><td>{eps_display}集{ot_mark}</td><td>{d}</td></tr>'

        sc = '#27ae60' if cd.get('is_complete') == '是' else '#e74c3c'
        st = '达标' if cd.get('is_complete') == '是' else '未达标'
        q = cd.get('quota', 0)
        qt = f'基准{q}集' if q > 0 else '无基准'
        tc = cd.get('total_commission', 0)
        cc = '#27ae60' if tc >= 0 else '#e74c3c'

        card = f'''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{_escape_html(nm)} - 绩效卡片</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft YaHei','PingFang SC',sans-serif;background:#f0f4f8;padding:16px}}
.card{{max-width:600px;margin:0 auto;background:#fff;border-radius:16px;box-shadow:0 4px 24px rgba(0,0,0,.1);overflow:hidden}}
.hd{{background:linear-gradient(135deg,#1e3a5f,#3b82f6);color:#fff;padding:24px 20px;text-align:center}}
.hd h1{{font-size:clamp(18px,5vw,22px)}}.hd .role{{font-size:clamp(11px,3vw,13px);opacity:.85;margin-top:2px}}
.stats{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;padding:16px 20px}}
.stat{{text-align:center;padding:12px 6px;background:#f8fafc;border-radius:10px}}
.stat .v{{font-size:clamp(20px,6vw,28px);font-weight:700}}
.stat .l{{font-size:clamp(10px,2.5vw,11px);color:#64748b;margin-top:2px}}
.status{{display:inline-block;padding:3px 12px;border-radius:12px;font-size:clamp(11px,3vw,13px);font-weight:600;color:#fff;background:{sc}}}
.ps{{padding:0 20px 16px}}.ps h3{{font-size:clamp(12px,3.5vw,14px);color:#334155;margin-bottom:8px}}
.table-wrap{{overflow-x:auto;-webkit-overflow-scrolling:touch}}
.ps table{{width:100%;border-collapse:collapse;font-size:clamp(10px,2.5vw,12px);min-width:350px}}
.ps th{{background:#e2e8f0;padding:6px 8px;text-align:left;white-space:nowrap}}
.ps td{{padding:5px 8px;border-bottom:1px solid #e2e8f0}}
.ft{{background:#f8fafc;padding:14px 20px;text-align:center;font-size:clamp(10px,2.5vw,12px);color:#94a3b8}}
@media(max-width:480px){{
  body{{padding:8px}}
  .card{{border-radius:12px}}
  .stats{{grid-template-columns:repeat(3,1fr);gap:6px;padding:12px 12px}}
  .stat{{padding:10px 4px}}
  .hd{{padding:18px 14px}}
  .ps{{padding:0 12px 12px}}
}}
</style></head><body><div class="card">
<div class="hd"><h1>{_escape_html(nm)}</h1><div class="role">{cd.get("role","")} · {qt}</div></div>
<div class="stats">
<div class="stat"><div class="v">{persons[nm]["total_eps"]}</div><div class="l">完成集数</div></div>
<div class="stat"><div class="v"><span class="status">{st}</span></div><div class="l">绩效状态</div></div>
<div class="stat"><div class="v" style="color:{cc}">{tc:,}</div><div class="l">提成(元)</div></div>
</div>
<div class="ps"><h3>本月项目明细 ({len(persons[nm]["projects"])}个)</h3>
<div class="table-wrap"><table><tr><th>项目ID</th><th>名称</th><th>集数</th><th>交付</th></tr>{proj_rows}</table></div></div>
<div class="ft">AI后期剪辑提成 · {datetime.date.today().strftime("%Y年%m月")}</div>
</div></body></html>'''
        cards.append((nm, card))

    paths = []
    for nm, html in cards:
        fpath = os.path.join(output_dir, f'{nm}_绩效卡片.html')
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(html)
        paths.append(fpath)

    idx_html = '''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>全员绩效卡片</title>
<style>*{margin:0;padding:0;box-sizing:border-box}body{font-family:'Microsoft YaHei','PingFang SC',sans-serif;background:#f0f4f8;padding:20px}h1{text-align:center;color:#1e3a5f;margin-bottom:20px;font-size:clamp(18px,5vw,28px)}.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:16px;max-width:1400px;margin:0 auto}@media(max-width:600px){body{padding:10px}.grid{grid-template-columns:1fr;gap:10px}}</style>
</head><body><h1>AI后期剪辑 · 全员绩效卡片</h1><div class="grid">'''
    for nm, _ in cards:
        idx_html += f'<iframe src="{_escape_html(nm)}_绩效卡片.html" style="width:100%;height:340px;border:none;border-radius:12px;background:#fff;box-shadow:0 2px 12px rgba(0,0,0,.06)" loading="lazy"></iframe>\n'
    idx_html += '</div></body></html>'
    idx_path = os.path.join(output_dir, 'index.html')
    with open(idx_path, 'w', encoding='utf-8') as f:
        f.write(idx_html)
    paths.insert(0, idx_path)
    return paths


# ===================== 4. 多月份对比 =====================

def compare_months(file1_path, file2_path):
    wb1 = load_workbook(file1_path, data_only=True); wb2 = load_workbook(file2_path, data_only=True)
    ws1, ws2 = wb1.active, wb2.active

    m1 = re.search(r'(\d+)年(\d+)月', str(ws1.cell(1, 2).value or ws1.cell(1, 1).value or ''))
    m2 = re.search(r'(\d+)年(\d+)月', str(ws2.cell(1, 2).value or ws2.cell(1, 1).value or ''))
    label1 = f"{m1.group(1)}年{m1.group(2)}月" if m1 else "文件1"
    label2 = f"{m2.group(1)}年{m2.group(2)}月" if m2 else "文件2"

    def parse_sheet(ws):
        data = {}
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=20):
            name = str(row[1].value).strip() if row[1].value else ''
            if not name: continue
            eps = int(row[10].value) if row[10].value else 0
            comm = int(row[17].value) if row[17].value else 0
            if name not in data: data[name] = {'eps': 0, 'comm': 0}
            data[name]['eps'] += eps
            data[name]['comm'] += comm
        return data

    d1, d2 = parse_sheet(ws1), parse_sheet(ws2)
    all_names = sorted(set(list(d1.keys()) + list(d2.keys())))
    diff_rows = []
    for nm in all_names:
        e1, e2 = d1.get(nm, {}).get('eps', 0), d2.get(nm, {}).get('eps', 0)
        c1, c2 = d1.get(nm, {}).get('comm', 0), d2.get(nm, {}).get('comm', 0)
        de, dc = e2 - e1, c2 - c1
        if de != 0 or dc != 0:
            diff_rows.append((nm, e1, e2, de, c1, c2, dc))
    wb1.close(); wb2.close()
    return label1, label2, diff_rows


# ===================== 5. 备份 =====================

def backup_output(excel_path, html_path, backup_dir):
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    backed = []
    for fp in [excel_path, html_path]:
        if fp and os.path.exists(fp):
            name, ext = os.path.splitext(os.path.basename(fp))
            dest = os.path.join(backup_dir, f'{name}_{ts}{ext}')
            shutil.copy2(fp, dest)
            backed.append(dest)
    return backed


# ===================== 6. 数据预览 =====================

def data_preview(records, commission_data):
    """返回数据预览信息：每人集数/项目数/绩效/提成"""
    persons = {}
    for r in records:
        nm = r['身份证姓名']
        if nm not in persons:
            persons[nm] = {'eps': 0, 'projects': set()}
        persons[nm]['eps'] += r['单项目数/集数']
        if r['项目ID']:
            persons[nm]['projects'].add(r['项目ID'])

    preview = []
    for nm in sorted(persons.keys()):
        cd = commission_data.get(nm, {})
        preview.append({
            'name': nm,
            'role': cd.get('role', ''),
            'episodes': persons[nm]['eps'],
            'projects': len(persons[nm]['projects']),
            'quota': cd.get('quota', 0),
            'status': cd.get('is_complete', ''),
            'commission': cd.get('total_commission', 0),
        })
    return preview


# ===================== 7. 组内排名 =====================

def generate_ranking_html(commission_data, output_path=None):
    """生成组内排名HTML"""
    persons = []
    for nm, cd in commission_data.items():
        if not cd: continue
        persons.append({
            'name': nm, 'role': cd.get('role', ''),
            'eps': cd.get('total_episodes', 0),
            'comm': cd.get('total_commission', 0),
            'status': cd.get('is_complete', ''),
        })

    by_eps = sorted(persons, key=lambda x: x['eps'], reverse=True)
    by_comm = sorted(persons, key=lambda x: x['comm'], reverse=True)

    def rank_rows(data, key):
        rows = ''
        for i, p in enumerate(data):
            medal = ['🥇','🥈','🥉'][i] if i < 3 else str(i+1)
            v = p[key]
            color = '#27ae60' if v >= 0 else '#e74c3c'
            sign = '集' if key == 'eps' else '元'
            rows += f'<tr><td>{medal}</td><td><b>{_escape_html(p["name"])}</b></td><td>{p["role"]}</td><td style="color:{color};font-weight:600">{v:,}{sign}</td><td>{"✅" if p["status"]=="是" else "❌"}</td></tr>'
        return rows

    html = f'''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><title>剪辑一组 · 组内排名</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft YaHei',sans-serif;background:#f0f4f8;padding:20px;color:#2c3e50}}
h1{{text-align:center;color:#1e3a5f;margin:20px 0}}
.ct{{max-width:1100px;margin:0 auto}}
.row{{display:grid;grid-template-columns:1fr 1fr;gap:20px}}
.card{{background:#fff;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.06);overflow:hidden}}
.card h2{{background:linear-gradient(135deg,#1e3a5f,#3b82f6);color:#fff;padding:14px 20px;font-size:16px}}
table{{width:100%;border-collapse:collapse}}
th{{background:#e2e8f0;padding:8px 12px;font-size:12px;text-align:left}}
td{{padding:8px 12px;font-size:13px;border-bottom:1px solid #e2e8f0}}
tr:hover{{background:#f8fafc}}
</style></head><body>
<h1>剪辑一组 · 月度排名</h1>
<div class="ct"><div class="row">
<div class="card"><h2>📊 集数排行</h2><table><tr><th>排名</th><th>姓名</th><th>角色</th><th>集数</th><th>绩效</th></tr>{rank_rows(by_eps, 'eps')}</table></div>
<div class="card"><h2>💰 提成排行</h2><table><tr><th>排名</th><th>姓名</th><th>角色</th><th>提成</th><th>绩效</th></tr>{rank_rows(by_comm, 'comm')}</table></div>
</div></div></body></html>'''

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)
    return html


# ===================== 8. 项目管理 =====================

def generate_project_management_html(records, output_path=None):
    """生成项目管理视图HTML"""
    projects = {}
    for r in records:
        pid = r['项目ID']
        if not pid: continue
        if pid not in projects:
            projects[pid] = {
                'name': r['AI项目名称'][:60],
                'type': r['项目类型'],
                'date': r['结束日期'],
                'people': {},
                'total_eps': 0,
            }
        nm = r['身份证姓名']
        if nm not in projects[pid]['people']:
            projects[pid]['people'][nm] = 0
        projects[pid]['people'][nm] += r['单项目数/集数']
        projects[pid]['total_eps'] += r['单项目数/集数']

    sorted_projs = sorted(projects.items(), key=lambda x: x[1]['date'] or datetime.date(2000,1,1))

    rows = ''
    for i, (pid, info) in enumerate(sorted_projs):
        d = info['date'].strftime('%m/%d') if info['date'] else '-'
        people_str = ', '.join(f'{n}({e}集)' for n, e in info['people'].items())
        rows += f'''<tr>
<td>{i+1}</td><td><b>{pid}</b></td><td>{_escape_html(info["name"])}</td>
<td>{info["total_eps"]}</td><td>{len(info["people"])}人</td><td>{d}</td>
<td style="font-size:11px">{_escape_html(people_str)}</td></tr>\n'''

    html = f'''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><title>项目管理 · 项目清单</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft YaHei',sans-serif;background:#f0f4f8;padding:20px;color:#2c3e50}}
h1{{text-align:center;color:#1e3a5f;margin:20px 0}}
.ct{{max-width:1200px;margin:0 auto;background:#fff;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.06);overflow:hidden}}
table{{width:100%;border-collapse:collapse}}
th{{background:linear-gradient(135deg,#1e3a5f,#3b82f6);color:#fff;padding:10px 12px;font-size:12px;text-align:left}}
td{{padding:8px 12px;font-size:13px;border-bottom:1px solid #e2e8f0}}
tr:hover{{background:#f8fafc}}
.summary{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;padding:20px;max-width:1200px;margin:0 auto 20px}}
.s{{background:#fff;border-radius:10px;padding:16px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.05)}}
.s .v{{font-size:28px;font-weight:700;color:#1e3a5f}}
.s .l{{font-size:12px;color:#64748b}}
</style></head><body>
<h1>项目管理 · 项目清单</h1>
<div class="summary">
<div class="s"><div class="v">{len(projects)}</div><div class="l">项目总数</div></div>
<div class="s"><div class="v">{sum(p["total_eps"] for p in projects.values())}</div><div class="l">总集数</div></div>
<div class="s"><div class="v">{len(set(n for p in projects.values() for n in p["people"]))}</div><div class="l">参与人数</div></div>
<div class="s"><div class="v">{sum(len(p["people"]) for p in projects.values())}</div><div class="l">人次</div></div>
</div>
<div class="ct"><table>
<tr><th>#</th><th>项目ID</th><th>项目名称</th><th>集数</th><th>人数</th><th>交付</th><th>参与人员(集数)</th></tr>
{rows}</table></div></body></html>'''

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)
    return html


# ===================== 9. 智能分集 =====================

def smart_episode_assignment(total_eps, selected_people, role_map, leader_prefix=None, seed=None):
    """五段式分集：
    ①小组长前段 ②一卡区间 ③中段大块 ④一卡补尾 ⑤小组长尾段
    """
    if total_eps < 1:
        raise ValueError('总集数必须大于 0')
    selected_people = list(dict.fromkeys(selected_people))
    if not selected_people:
        raise ValueError('请至少选择一位剪辑人员')

    rng = random.Random(seed)

    # 角色分类
    leaders    = [p for p in selected_people if '小组长' in role_map.get(p, '')]
    card1      = [p for p in selected_people if '一卡' in role_map.get(p, '') and p not in leaders]
    card2_as   = [p for p in selected_people if '二卡' in role_map.get(p, '') or '助理' in role_map.get(p, '')]
    big_ld     = [p for p in selected_people if '组长' in role_map.get(p, '') and p not in leaders and p not in card2_as]
    result = {p: [] for p in selected_people}

    # ---- 固定段长度 ----
    leader_head = min(5, 3 + max(0, total_eps - 80) // 50)       # 3→4→5
    leader_tail = 3 if total_eps >= 50 else max(1, total_eps // 10)
    card1_tail_extra = max(1, leader_tail - 1)                    # 一卡尾前补丁 2集
    card1_zone  = leader_prefix if leader_prefix is not None else leader_head + max(8, total_eps // 6)
    card1_zone  = min(card1_zone, total_eps - leader_tail - card1_tail_extra - 1)

    cursor = 1

    def emit(person, n):
        nonlocal cursor
        n = min(n, total_eps - cursor + 1)
        if n > 0:
            result[person].extend(range(cursor, cursor + n))
            cursor += n
        return n

    # ===== ① 小组长前段 =====
    if leaders:
        hc = _weighted_counts(leader_head, leaders)
        for p in leaders:
            emit(p, hc[p])

    # ===== ② 一卡前段（一卡区间） =====
    if card1 and cursor <= card1_zone:
        space = card1_zone - cursor + 1
        total_card1 = len(card1)
        for i, p in enumerate(card1):
            cnt = space // total_card1 + (1 if i < space % total_card1 else 0)
            emit(p, cnt)

    # ===== ③ 中段大块（二卡/助理/组长 按人数均分） =====
    # 给一卡尾前补丁留空间
    mid_end = total_eps - leader_tail - card1_tail_extra
    mid_ppl = card2_as + big_ld
    rng.shuffle(mid_ppl)

    if mid_ppl and cursor <= mid_end:
        mid_space = mid_end - cursor + 1
        mq = _weighted_counts(mid_space, mid_ppl)
        for p in mid_ppl:
            emit(p, mq[p])
            if cursor > mid_end:
                break

    # ===== ④ 一卡补尾（组长尾前） =====
    if card1 and cursor < total_eps - leader_tail + 1:
        pre = total_eps - leader_tail - cursor + 1
        if pre > 0:
            ct = _weighted_counts(pre, card1)
            for p in card1:
                emit(p, ct[p])

    # ===== ⑤ 小组长尾段 =====
    if leaders and cursor <= total_eps:
        tail_eps = total_eps - cursor + 1
        tc = _weighted_counts(tail_eps, leaders)
        for p in leaders:
            emit(p, tc[p])

    # ---- 兜底 ----
    if cursor <= total_eps:
        for p in selected_people:
            emit(p, total_eps - cursor + 1)
            if cursor > total_eps:
                break

    all_others = card2_as + big_ld
    formatted = _format_result(result, leaders, card1, all_others, big_ld,
                               selected_people, leader_head,
                               max(0, card1_zone - leader_head),
                               max(0, total_eps - card1_zone + 1), card1_zone)
    formatted['targets'] = {p: len(result.get(p, [])) for p in selected_people}
    formatted['leader_prefix'] = leader_head
    return formatted


def _weighted_counts(total, people):
    """按最大余数法均匀分配整数集数，确保总和恰好等于 total。"""
    if not people or total <= 0:
        return {person: 0 for person in people}
    n = len(people)
    # 均分 + 余数按剩余最大小数分配
    base = total // n
    raw_frac = {person: total / n - base for person in people}
    counts = {person: base for person in people}
    remainder = total - base * n
    for person in sorted(people, key=lambda p: raw_frac[p], reverse=True)[:remainder]:
        counts[person] += 1
    return counts


def validate_episode_assignments(lines, selected_people, total_eps):
    """校验人工编辑后的分集行，并返回标准化的分集结果。

    每行格式为 ``姓名：1-3,5``。要求所有集数恰好覆盖 1 到 total_eps 一次，
    防止确认后把漏集或重集写入项目表。
    """
    selected = set(selected_people)
    assignments = {}
    seen = {}
    for raw_line in lines.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = re.match(r'^([^：:]+)[：:]\s*(.+)$', line)
        if not match:
            raise ValueError(f'格式错误：{line}（请使用“姓名：1-3,5”）')
        name, episode_text = match.group(1).strip(), match.group(2).strip()
        if name not in selected:
            raise ValueError(f'“{name}”不在本次已选择人员中')
        if name in assignments:
            raise ValueError(f'“{name}”出现了多行，请合并为一行')
        episodes, _ = parse_episode_ranges(episode_text)
        if not episodes:
            raise ValueError(f'“{name}”没有有效集数')
        for ep in episodes:
            if ep < 1 or ep > total_eps:
                raise ValueError(f'第 {ep} 集超出项目范围 1-{total_eps}')
            if ep in seen:
                raise ValueError(f'第 {ep} 集重复分配给“{seen[ep]}”和“{name}”')
            seen[ep] = name
        assignments[name] = episodes

    missing = [str(ep) for ep in range(1, total_eps + 1) if ep not in seen]
    if missing:
        preview = '、'.join(missing[:12])
        suffix = '…' if len(missing) > 12 else ''
        raise ValueError(f'仍有未分配集数：{preview}{suffix}')

    return {
        'formatted': {
            name: '；'.join(f'{start}-{end}' if start != end else str(start)
                             for start, end in merge_episode_ranges(episodes))
            for name, episodes in assignments.items()
        },
        'summary': {name: len(episodes) for name, episodes in assignments.items()},
        'assignments': {name: merge_episode_ranges(episodes) for name, episodes in assignments.items()},
        'stats': {'总人数': len(assignments), '总集数': total_eps},
    }


def _format_result(result, leader, card1_only, card2_people, big_leader,
                   selected_people, n_first, n_mid, n_tail, card1_range=15):
    assignments, fmt, summary = {}, {}, {}
    for name in result:
        r = merge_episode_ranges(result[name])
        assignments[name] = r
        summary[name] = len(result[name])
        fmt[name] = '；'.join(f'{s}-{e}' if s != e else str(s) for s, e in r)

    stats = {
        '小组长人数': len(leader),
        '一卡人数': len(card1_only),
        '二卡助理人数': len(card2_people),
        '剪辑组长人数': len(big_leader),
        '总人数': len(selected_people),
        '前三集': n_first, f'中段(4-{card1_range})': n_mid,
        f'后段({card1_range + 1}+)': n_tail,
    }
    return {'assignments': assignments, 'summary': summary, 'formatted': fmt, 'stats': stats}


# ===================== 10. 个人月度趋势 =====================

def generate_person_trend_html(person_name, role_map, output_dir):
    """扫描output_dir下所有AI后期剪辑提成一组*xlsx，提取某人的月度集数和提成趋势"""
    pattern = re.compile(r'AI后期剪辑提成一组(.*?)\.xlsx')
    months = []
    for fname in sorted(os.listdir(output_dir)):
        m = pattern.match(fname)
        if not m or '_仪表盘' in fname: continue
        month_label = m.group(1)
        fpath = os.path.join(output_dir, fname)
        try:
            wb = load_workbook(fpath, data_only=True)
            ws = wb.active
            found_eps, found_comm = 0, 0
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=20):
                name_cell = str(row[1].value).strip() if row[1].value else ''
                if person_name not in name_cell: continue
                if row[10].value:
                    found_eps += int(row[10].value) if str(row[10].value).strip().lstrip('-').isdigit() else 0
                if row[17].value:
                    found_comm += int(row[17].value) if str(row[17].value).strip().lstrip('-').isdigit() else 0
            wb.close()
            months.append({'month': month_label, 'eps': found_eps, 'comm': found_comm})
        except Exception:
            continue

    if not months:
        return None

    # CSS 柱状图
    max_eps = max(m['eps'] for m in months) or 1
    max_comm = max(abs(m['comm']) for m in months) or 1
    eps_bars = ''
    comm_bars = ''
    quota = 70 if '一卡' in role_map.get(person_name, '') else (120 if '二卡' in role_map.get(person_name, '') or '助理' in role_map.get(person_name, '') else 0)

    for m in months:
        ep_pct = min(100, int(m['eps'] / max_eps * 100)) if max_eps else 0
        ec = '#27ae60' if m['eps'] >= quota else '#e74c3c'
        cp = abs(m['comm']) / max_comm * 100 if max_comm else 0
        cc = '#27ae60' if m['comm'] >= 0 else '#e74c3c'
        eps_bars += f'<div class="br"><span class="bl">{m["month"]}</span><div class="bt"><div class="bf" style="width:{ep_pct}%;background:{ec}"></div></div><span class="bv">{m["eps"]}集</span></div>'
        comm_bars += f'<div class="br"><span class="bl">{m["month"]}</span><div class="bt"><div class="bf" style="width:{min(cp,100):.0f}%;background:{cc}"></div></div><span class="bv" style="color:{cc}">{m["comm"]:,}元</span></div>'

    quota_info = f'基准{quota}集' if quota > 0 else '无基准'
    html = f'''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><title>{_escape_html(person_name)} · 月度趋势</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft YaHei',sans-serif;background:#f0f4f8;padding:20px;color:#2c3e50}}
h1{{text-align:center;color:#1e3a5f;margin:20px 0}}
.ct{{max-width:900px;margin:0 auto}}
.row{{display:grid;grid-template-columns:1fr 1fr;gap:24px}}
.card{{background:#fff;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.06);padding:20px}}
.card h2{{font-size:16px;margin-bottom:16px;padding-bottom:8px;border-bottom:2px solid #e2e8f0}}
.br{{display:flex;align-items:center;margin:6px 0;gap:8px}}
.bl{{width:50px;font-size:12px;text-align:right;flex-shrink:0}}
.bt{{flex:1;height:18px;background:#e2e8f0;border-radius:9px;overflow:hidden;position:relative}}
.bf{{height:100%;border-radius:9px;transition:width .3s}}
.bv{{width:70px;font-size:12px;font-weight:600;flex-shrink:0;text-align:right}}
.info{{text-align:center;color:#64748b;font-size:13px;margin:8px 0 20px}}
</style></head><body>
<h1>{_escape_html(person_name)} · 月度趋势</h1>
<div class="info">{role_map.get(person_name, '')} · {quota_info} · 共{len(months)}个月数据</div>
<div class="ct"><div class="row">
<div class="card"><h2>📊 集数变化</h2>{eps_bars}</div>
<div class="card"><h2>💰 提成变化</h2>{comm_bars}</div>
</div></div></body></html>'''
    trend_path = os.path.join(output_dir, f'{person_name}_月度趋势.html')
    with open(trend_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return trend_path


# ===================== 11. 数据校验 =====================

def validate_project_data(project_file, role_map):
    """扫描项目数据文件，返回校验问题列表"""
    import pandas as pd
    issues = []
    try:
        df = pd.read_excel(project_file, header=None)
    except Exception as e:
        return [('文件', '无法读取', str(e))]

    known_names = set(role_map.keys())
    proj_name = ''; proj_id = ''
    seen_eps = {}  # 项目ID -> 集数集合（检测重复分配）

    for idx in range(len(df)):
        col1 = str(df.iloc[idx, 0]) if pd.notna(df.iloc[idx, 0]) else ''
        col2 = str(df.iloc[idx, 1]) if pd.notna(df.iloc[idx, 1]) else ''
        col3 = str(df.iloc[idx, 2]) if pd.notna(df.iloc[idx, 2]) else ''
        col4 = str(df.iloc[idx, 3]) if pd.notna(df.iloc[idx, 3]) else ''

        if col1 not in ['', 'nan'] and '月份' not in col1:
            proj_name = col1[:60]
            m = re.search(r'\d{4}', col1)
            if m: 
                proj_id = m.group()
                if proj_id not in seen_eps:
                    seen_eps[proj_id] = set()

        if col3 not in ['', 'nan']:
            for match in re.finditer(r'([^\s：:]+)[：:]\s*(.+)', col3):
                name = match.group(1).strip()
                eps_part = match.group(2).strip()
                if name and name not in known_names:
                    issues.append((f'行{idx+1}', name, f'"{name}" 不在人员配置中'))
                # 检测集数范围
                nums = re.findall(r'\d+', eps_part)
                if nums:
                    for n in nums:
                        ep = int(n)
                        if proj_id:
                            if ep in seen_eps.get(proj_id, set()):
                                issues.append((f'行{idx+1}', name, f'集数{ep}在项目{proj_id}中已被分配过'))
                            seen_eps[proj_id].add(ep)

        # 检测日期格式
        if col4 not in ['', 'nan', '多版本已交付', '已分集', '已分发']:
            if not re.search(r'\d+\.\d+', col4):
                issues.append((f'行{idx+1}', '日期', f'日期格式异常: {col4[:20]}'))

    return issues if issues else []


# ===================== 12. 备份管理 =====================

def list_backups(backup_dir):
    """列出备份目录中所有备份"""
    if not os.path.exists(backup_dir):
        return []
    files = [f for f in os.listdir(backup_dir) if f.endswith('.xlsx') or f.endswith('.html')]
    files.sort(reverse=True)
    return [{'name': f, 'path': os.path.join(backup_dir, f),
             'size': os.path.getsize(os.path.join(backup_dir, f)),
             'mtime': datetime.datetime.fromtimestamp(os.path.getmtime(os.path.join(backup_dir, f)))}
            for f in files]

def cleanup_backups(backup_dir, keep=30):
    """清理旧备份，只保留最近keep个"""
    if not os.path.exists(backup_dir):
        return 0
    files = [f for f in os.listdir(backup_dir) if f.endswith('.xlsx') or f.endswith('.html')]
    files.sort()
    to_remove = max(0, len(files) - keep)
    for f in files[:to_remove]:
        os.remove(os.path.join(backup_dir, f))
    return to_remove


# ===================== 13. 高级查询与筛选 =====================

def advanced_filter(records, commission_data, filters):
    """
    高级筛选：按姓名、角色、集数范围、项目ID、绩效状态、提成范围组合筛选
    filters: {
        'name_keyword': str,      # 姓名模糊匹配
        'roles': list,            # 角色列表
        'min_eps': int,           # 最小编剧
        'max_eps': int,           # 最大集数
        'project_id': str,        # 项目ID精确匹配
        'status': str,            # '是'/'否'/'全部'
        'min_commission': int,    # 最低提成
        'max_commission': int,    # 最高提成
    }
    """
    # 先按人汇总
    persons = {}
    for r in records:
        nm = r['身份证姓名']
        if nm not in persons:
            persons[nm] = {'eps': 0, 'projects': set(), 'dates': [], 'records': []}
        persons[nm]['eps'] += r['单项目数/集数']
        persons[nm]['projects'].add(r['项目ID'])
        persons[nm]['dates'].append(r['结束日期'])
        persons[nm]['records'].append(r)

    result = []
    for nm, info in persons.items():
        cd = commission_data.get(nm, {})

        # 姓名模糊匹配
        if filters.get('name_keyword'):
            if filters['name_keyword'] not in nm:
                continue

        # 角色筛选
        if filters.get('roles'):
            role = cd.get('role', '')
            if role not in filters['roles']:
                continue

        # 集数范围
        min_eps = filters.get('min_eps')
        max_eps = filters.get('max_eps')
        if min_eps is not None and info['eps'] < min_eps:
            continue
        if max_eps is not None and info['eps'] > max_eps:
            continue

        # 项目ID筛选
        if filters.get('project_id'):
            if filters['project_id'] not in info['projects']:
                continue

        # 绩效状态
        if filters.get('status') and filters['status'] != '全部':
            if cd.get('is_complete', '') != filters['status']:
                continue

        # 提成范围
        min_c = filters.get('min_commission')
        max_c = filters.get('max_commission')
        comm = cd.get('total_commission', 0)
        if min_c is not None and comm < min_c:
            continue
        if max_c is not None and comm > max_c:
            continue

        result.append({
            'name': nm,
            'role': cd.get('role', ''),
            'episodes': info['eps'],
            'projects': len(info['projects']),
            'quota': cd.get('quota', 0),
            'status': cd.get('is_complete', ''),
            'commission': comm,
            'desc': cd.get('desc', ''),
        })

    return result


# ===================== 14. 数据修正工具 =====================

def correct_record(records, commission_data, correction):
    """
    修正单条记录或重新计算
    correction: {
        'action': 'update_eps' | 'update_name' | 'delete_record' | 'add_record',
        'record_index': int,        # 要修改的记录索引
        'new_eps': int,             # 新集数 (update_eps)
        'new_name': str,            # 新姓名 (update_name)
        'new_record': dict,         # 新增记录 (add_record)
    }
    返回: (new_records, new_commission_data)
    """
    action = correction.get('action', '')

    if action == 'update_eps':
        idx = correction['record_index']
        records[idx]['单项目数/集数'] = correction['new_eps']
        # 更新明细
        eps = correction['new_eps']
        records[idx]['完成明细'] = ','.join(str(i) for i in range(1, eps + 1))

    elif action == 'update_name':
        idx = correction['record_index']
        records[idx]['身份证姓名'] = correction['new_name']

    elif action == 'delete_record':
        idx = correction['record_index']
        records.pop(idx)

    elif action == 'add_record':
        records.append(correction['new_record'])

    # 重新计算提成 —— 此处需要外部调用compute_commission
    return records


# ===================== 15. 项目数据模板下载 =====================

def generate_project_template(output_path):
    """生成标准化的项目数据录入模板Excel"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '项目数据录入模板'

    # 表头样式
    header_font = Font(name='宋体', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    data_font = Font(name='宋体', size=11)
    data_align = Alignment(horizontal='center', vertical='center')
    example_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')

    # 标题行
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(1, 1, 'AI后期剪辑项目数据录入模板')
    title_cell.font = Font(name='宋体', size=16, bold=True, color='1E3A5F')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # 说明行
    ws.merge_cells('A2:E2')
    ws.cell(2, 1, '使用说明：按以下格式填写项目信息，每个项目占用连续多行（项目名×1 + 人员分配×N）').font = Font(name='宋体', size=9, color='64748B')
    ws.row_dimensions[2].height = 22

    # 列头
    headers = ['A: 项目名称/备注', 'B: 路径/目录', 'C: 人员分配', 'D: 交付日期', 'E: 状态']
    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border
    ws.row_dimensions[3].height = 28

    # 示例数据
    examples = [
        ['2200-客服千金追夫99次（海外）', r'O:\AI漫剧剪辑一组\客服千金追夫99次', None, '7.25下午18点交', None],
        [None, None, '张大强：1-3', None, None],
        [None, None, '任显翔：4-15', None, None],
        [None, None, '金文龙：18-40', None, None],
        [None, None, '陈冰洁：41-67', None, None],
        [None, None, '李钊琦：68-99', None, None],
        ['2199-绝世豪门（海外）', r'O:\AI漫剧剪辑一组\绝世豪门', None, '7.28下午18点交', None],
        [None, None, '陈春阳：1-3', None, None],
        [None, None, '程梦：4-15', None, None],
        [None, None, '刘梦真：16-40', None, None],
        [None, None, '杨倩：41-68', None, None],
    ]

    for i, row_data in enumerate(examples):
        row = 4 + i
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row, j, val)
            c.font = data_font
            c.alignment = data_align
            c.border = thin_border
            if i == 0:
                c.fill = example_fill
        ws.row_dimensions[row].height = 24

    # 合并示例项目单元格
    ws.merge_cells('A4:A9')
    ws.merge_cells('B4:B9')
    ws.merge_cells('D4:D9')
    ws.merge_cells('A10:A14')
    ws.merge_cells('B10:B14')
    ws.merge_cells('D10:D14')

    # 列宽
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12

    # 冻结表头
    ws.freeze_panes = 'A4'

    # 添加第二个sheet：录入规范
    ws2 = wb.create_sheet('录入规范')
    rules = [
        ['录入规范', '说明'],
        ['项目标题行', '放在 A 列，包含项目ID（4位数字）+ 项目名称。如：2200-客服千金追夫99次（海外）'],
        ['目录路径行', '放在 B 列，项目文件所在路径。如：O:\\AI漫剧剪辑一组\\项目名'],
        ['人员分配行', '放在 C 列，格式：姓名：集数范围。如：张大强：1-3 或 任显翔：4-10,12-15'],
        ['交付日期', '放在 D 列，格式：月.日+备注。如：7.25下午18点交'],
        ['状态', '放在 E 列（可选），填写：已分集 / 已交付 / 多版本已交付'],
        ['', ''],
        ['集数格式参考', ''],
        ['单集', '5'],
        ['连续范围', '1-10'],
        ['不连续范围', '1-5,8-12,15'],
        ['混合', '1-3,5,7-9'],
    ]
    for i, (k, v) in enumerate(rules):
        c1 = ws2.cell(i + 1, 1, k)
        c2 = ws2.cell(i + 1, 2, v)
        if i == 0:
            c1.font = Font(name='宋体', size=12, bold=True)
            c2.font = Font(name='宋体', size=12, bold=True)
        else:
            c1.font = Font(name='宋体', size=10, bold=True, color='1E3A5F')
            c2.font = Font(name='宋体', size=10)
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 65

    wb.save(output_path)
    return output_path


# ===================== 16. 配置快照与回滚 =====================

SNAPSHOT_DIR_NAME = 'config_snapshots'


def create_config_snapshot(config_path, snapshots_dir):
    """保存当前config.json快照"""
    if not os.path.exists(snapshots_dir):
        os.makedirs(snapshots_dir)
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    snap_name = f'config_{ts}.json'
    snap_path = os.path.join(snapshots_dir, snap_name)
    shutil.copy2(config_path, snap_path)
    return snap_path


def list_config_snapshots(snapshots_dir):
    """列出所有配置快照"""
    if not os.path.exists(snapshots_dir):
        return []
    files = [f for f in os.listdir(snapshots_dir) if f.endswith('.json')]
    files.sort(reverse=True)
    return [{'name': f, 'path': os.path.join(snapshots_dir, f),
             'mtime': datetime.datetime.fromtimestamp(
                 os.path.getmtime(os.path.join(snapshots_dir, f)))}
            for f in files]


def restore_config_snapshot(snapshot_path, config_path):
    """恢复配置快照"""
    shutil.copy2(snapshot_path, config_path)
    return True


# ===================== 17. Web仪表盘服务 =====================

import http.server
import socketserver
import urllib.parse


def start_web_server(serve_dir, port=8080):
    """启动本地Web服务器，提供仪表盘浏览"""
    os.chdir(serve_dir)

    class DashboardHandler(http.server.SimpleHTTPRequestHandler):
        def do_GET(self):
            # 解析路径
            parsed = urllib.parse.urlparse(self.path)
            path = parsed.path.lstrip('/')

            # 首页：列出所有HTML文件
            if path == '' or path == '/':
                self.send_response(200)
                self.send_header('Content-type', 'text/html; charset=utf-8')
                self.end_headers()

                html_files = []
                for f in sorted(os.listdir('.'), reverse=True):
                    if f.endswith('.html') or f.endswith('.htm'):
                        size = os.path.getsize(f)
                        mtime = datetime.datetime.fromtimestamp(os.path.getmtime(f))
                        html_files.append((f, size, mtime))

                # 分组
                dashboards = [(f, s, m) for f, s, m in html_files if '仪表盘' in f]
                rankings = [(f, s, m) for f, s, m in html_files if '排名' in f]
                projects = [(f, s, m) for f, s, m in html_files if '项目' in f]
                cards_dir = os.path.join(serve_dir, '个人绩效卡片')
                has_cards = os.path.isdir(cards_dir) and any(
                    f.endswith('.html') for f in os.listdir(cards_dir))

                def file_rows(files_list):
                    rows = ''
                    for fn, sz, mt in files_list:
                        rows += f'''<tr>
<td><a href="{_escape_html(fn)}">📄 {_escape_html(fn)}</a></td>
<td>{sz/1024:.1f} KB</td><td>{mt.strftime('%Y-%m-%d %H:%M')}</td>
</tr>'''
                    return rows

                cards_link = ''
                if has_cards:
                    cards_link = '''
<div class="card"><h2>🃏 个人绩效卡片</h2>
<p>共 {} 张</p>
<a class="btn" href="个人绩效卡片/index.html">📂 查看全员卡片</a>
</div>'''.format(len([f for f in os.listdir(cards_dir) if f.endswith('.html')]))

                html = f'''<!DOCTYPE html><html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI后期剪辑 · 数据看板</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft YaHei',sans-serif;background:#0f172a;color:#e2e8f0;min-height:100vh}}
.header{{background:linear-gradient(135deg,#1e3a5f,#3b82f6);padding:24px 32px;text-align:center}}
.header h1{{font-size:24px;color:#fff}}.header p{{font-size:13px;opacity:.8;margin-top:4px}}
.ct{{max-width:1200px;margin:0 auto;padding:20px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(360px,1fr));gap:16px}}
.card{{background:#1e293b;border-radius:12px;padding:20px;box-shadow:0 4px 16px rgba(0,0,0,.3)}}
.card h2{{font-size:16px;color:#93c5fd;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid #334155}}
.card p{{font-size:13px;color:#94a3b8;margin:8px 0}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
th{{text-align:left;padding:6px 8px;color:#94a3b8;border-bottom:1px solid #334155}}
td{{padding:6px 8px;border-bottom:1px solid #1e293b}}
a{{color:#60a5fa;text-decoration:none}}a:hover{{text-decoration:underline;color:#93c5fd}}
.btn{{display:inline-block;background:#3b82f6;color:#fff;padding:8px 16px;border-radius:8px;font-size:13px;margin-top:8px;text-decoration:none}}
.btn:hover{{background:#2563eb;text-decoration:none;color:#fff}}
.footer{{text-align:center;padding:20px;color:#475569;font-size:12px}}
@media(max-width:600px){{.grid{{grid-template-columns:1fr}}.header{{padding:16px}}}}
</style></head><body>
<div class="header"><h1>🎬 AI后期剪辑 · 数据看板</h1>
<p>{datetime.datetime.now().strftime('%Y年%m月%d日 %H:%M')} · 本地服务 :{port}</p></div>
<div class="ct"><div class="grid">
<div class="card"><h2>📊 月度仪表盘</h2><table>{file_rows(dashboards)}</table></div>
<div class="card"><h2>🏆 组内排名</h2><table>{file_rows(rankings)}</table></div>
<div class="card"><h2>🗂️ 项目管理</h2><table>{file_rows(projects)}</table></div>
{cards_link}
</div></div>
<div class="footer">AI后期剪辑提成工具 · 本地Web仪表盘 · 仅供内部使用</div>
</body></html>'''
                self.wfile.write(html.encode('utf-8'))
                return

            return super().do_GET()

        def log_message(self, fmt, *args):
            pass  # 静默日志

    try:
        server = socketserver.TCPServer(("", port), DashboardHandler)
        return server
    except OSError as e:
        if '10048' in str(e) or 'Address already in use' in str(e):
            # 端口被占用，尝试下一个
            return start_web_server(serve_dir, port + 1)
        raise


# ===================== 多格式导出 =====================

def export_to_csv(records, commission_data, output_dir):
    """导出人员提成 CSV 和项目明细 CSV。返回生成的文件路径列表。"""
    import csv
    os.makedirs(output_dir, exist_ok=True)
    created = []

    # 人员提成 CSV
    person_csv = os.path.join(output_dir, '人员提成.csv')
    with open(person_csv, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['姓名', '角色', '总集数', '基准', '绩效', '超额/组奖', '缺集扣除', '提成合计'])
        for name in commission_data:
            cd = commission_data.get(name, {})
            w.writerow([
                name, cd.get('role', ''), cd.get('total_episodes', 0),
                cd.get('quota', 0), cd.get('is_complete', ''),
                cd.get('overtime_bonus', 0) + cd.get('group_project_bonus', 0),
                cd.get('shortage_penalty', 0), cd.get('total_commission', 0),
            ])
    created.append(person_csv)

    # 项目明细 CSV
    proj_csv = os.path.join(output_dir, '项目明细.csv')
    with open(proj_csv, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['人员', '角色', '项目ID', '项目名称', '开始日期', '结束日期', '完成明细', '集数'])
        for r in records:
            w.writerow([
                r.get('身份证姓名', ''), r.get('角色', ''), r.get('项目ID', ''),
                r.get('AI项目名称', ''), r.get('开始日期', ''), r.get('结束日期', ''),
                r.get('完成明细', ''), r.get('单项目数/集数', 0),
            ])
    created.append(proj_csv)
    return created


def export_to_json(records, commission_data, output_dir):
    """导出结构化 JSON（含 records、commission、summary），美化格式。"""
    os.makedirs(output_dir, exist_ok=True)
    json_path = os.path.join(output_dir, '提成数据.json')
    summary = {
        'total_people': len(commission_data),
        'total_commission': sum(cd.get('total_commission', 0) for cd in commission_data.values()),
        'total_episodes': sum(r.get('单项目数/集数', 0) for r in records),
    }
    data = {
        'summary': summary,
        'records': records,
        'commission': commission_data,
    }
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    return json_path


# ===================== #1 历史数据累积与跨月趋势库 =====================

def archive_history(records, commission_data, history_dir, month_label):
    """将本次生成的数据归档到 history/ 目录。

    Args:
        records: 解析后的记录列表
        commission_data: 提成计算结果
        history_dir: 历史数据目录（默认 history/）
        month_label: 月份标签，如 '2026年08月'

    Returns:
        归档文件路径
    """
    os.makedirs(history_dir, exist_ok=True)
    summary = {
        'total_people': len(commission_data),
        'total_commission': sum(cd.get('total_commission', 0) for cd in commission_data.values()),
        'total_episodes': sum(r.get('单项目数/集数', 0) for r in records),
    }
    data = {
        'month': month_label,
        'summary': summary,
        'records': records,
        'commission': commission_data,
    }
    safe_month = re.sub(r'[\/:*?"<>|]', '_', str(month_label))
    json_path = os.path.join(history_dir, f'{safe_month}.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    return json_path


def load_history(history_dir):
    """读取 history/ 目录下所有月份的归档数据。

    Returns:
        list[dict]：每个元素含 month/summary/records/commission，按月份排序
    """
    if not os.path.isdir(history_dir):
        return []
    entries = []
    for fn in sorted(os.listdir(history_dir)):
        if not fn.endswith('.json'):
            continue
        try:
            with open(os.path.join(history_dir, fn), 'r', encoding='utf-8') as f:
                entries.append(json.load(f))
        except Exception:
            continue
    return entries


def generate_trend_report(history_dir, output_dir):
    """基于历史数据生成跨月趋势 HTML（收入/集数折线图 + 明细表）。

    Returns:
        html_path 或 None（无历史数据时）
    """
    entries = load_history(history_dir)
    if not entries:
        return None
    # 收集所有人员
    persons = set()
    for e in entries:
        for nm in e.get('commission', {}):
            persons.add(nm)
    persons = sorted(persons)
    months = [e.get('month', '?') for e in entries]

    # 每人各月提成/集数
    comm_matrix = {nm: [] for nm in persons}
    eps_matrix = {nm: [] for nm in persons}
    for e in entries:
        cd = e.get('commission', {})
        for nm in persons:
            comm_matrix[nm].append(cd.get(nm, {}).get('total_commission', 0))
            eps_matrix[nm].append(cd.get(nm, {}).get('total_episodes', 0))

    # 生成 SVG 折线图（人员提成趋势）
    def _polyline(values, color, max_v):
        if max_v <= 0:
            return ''
        w, h = 400, 120
        pts = []
        n = len(values)
        for i, v in enumerate(values):
            x = 10 + i * ((w - 20) / max(1, n - 1))
            y = h - 10 - (abs(v) / max_v) * (h - 20)
            pts.append(f'{x:.1f},{y:.1f}')
        return f'<polyline points="{" ".join(pts)}" fill="none" stroke="{color}" stroke-width="2"/>'

    def _build_chart(per_person_values, title, color, unit):
        rows = ''
        for nm in persons:
            vals = per_person_values[nm]
            max_v = max([abs(v) for v in vals] + [1])
            svg = (f'<svg viewBox="0 0 400 120" width="100%" height="120">'
                   f'<polyline points="" fill="none"/>'
                   f'{_polyline(vals, color, max_v)}'
                   f'</svg>')
            labels = ''.join(f'<span class="m">{months[i]}<b>{vals[i]}</b>{unit}</span>' for i in range(len(vals)))
            rows += f'<div class="tr"><div class="tl">{nm}</div><div class="tc">{svg}</div><div class="tv">{labels}</div></div>'
        return rows

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>跨月趋势总览</title>
<style>
body{{font-family:'Segoe UI','Microsoft YaHei',sans-serif;background:#f0f4f8;color:#2c3e50;margin:0}}
.hd{{background:linear-gradient(135deg,#1a5276,#2e86c1);color:#fff;padding:28px 40px;text-align:center}}
.hd h1{{font-size:26px;margin:0;letter-spacing:2px}}.hd p{{opacity:.85;margin-top:6px}}
.ct{{max-width:1100px;margin:0 auto;padding:24px 20px}}
.pn{{background:#fff;border-radius:12px;padding:20px 24px;box-shadow:0 2px 12px rgba(0,0,0,.06);margin-bottom:24px}}
.pn h2{{font-size:18px;color:#1a5276;margin-bottom:14px;border-bottom:2px solid #d6e4f0;padding-bottom:10px}}
.tr{{display:flex;align-items:center;gap:12px;padding:8px 0;border-bottom:1px solid #eef2f7}}
.tl{{min-width:80px;font-weight:600;font-size:14px}}
.tc{{flex:1}}.tv{{display:flex;gap:8px;font-size:11px;color:#7f8c8d;flex-wrap:wrap}}
.tv span{{background:#eaf2f8;padding:2px 8px;border-radius:8px;white-space:nowrap}}
.tv b{{color:#2980b9;margin:0 3px}}
.summary{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin-bottom:24px}}
.sd{{background:#fff;border-radius:10px;padding:16px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.05)}}
.sd .v{{font-size:28px;font-weight:700;color:#2980b9}}.sd .l{{font-size:12px;color:#7f8c8d;margin-top:4px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#d6e4f0;color:#1a5276;padding:8px;border-bottom:2px solid #b0c4de}}
td{{padding:6px 8px;border-bottom:1px solid #eef2f7;text-align:center}}
.ft{{text-align:center;padding:18px;color:#95a5a6;font-size:12px}}
</style></head><body>
<div class="hd"><h1>📈 跨月趋势总览</h1><p>{len(months)} 个月 · {len(persons)} 人 · 自动生成</p></div>
<div class="ct">
<div class="summary">
<div class="sd"><div class="v">{len(months)}</div><div class="l">累计月份</div></div>
<div class="sd"><div class="v">{len(persons)}</div><div class="l">参与人员</div></div>
<div class="sd"><div class="v">{sum(e.get("summary",{}).get("total_commission",0) for e in entries):,}</div><div class="l">累计提成</div></div>
<div class="sd"><div class="v">{sum(e.get("summary",{}).get("total_episodes",0) for e in entries):,}</div><div class="l">累计集数</div></div>
</div>
<div class="pn"><h2>💵 提成趋势（各月）</h2>{_build_chart(comm_matrix, "提成", "#2e86c1", "元")}</div>
<div class="pn"><h2>🎬 集数趋势（各月）</h2>{_build_chart(eps_matrix, "集数", "#27ae60", "集")}</div>
<div class="pn"><h2>📊 明细汇总表</h2><div style="overflow-x:auto"><table>
<thead><tr><th>人员</th>{''.join(f'<th>{m}</th>' for m in months)}</tr></thead>
<tbody>
{''.join(f'<tr><td><b>{nm}</b></td>' + ''.join(f'<td>{comm_matrix[nm][i]}</td>' for i in range(len(months))) + '</tr>' for nm in persons)}
</tbody></table></div></div>
</div>
<div class="ft">AI后期剪辑提成表生成工具 v8.0 · 跨月趋势总览</div>
</body></html>'''
    os.makedirs(output_dir, exist_ok=True)
    html_path = os.path.join(output_dir, '跨月趋势总览.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return html_path


# ===================== #3 年度汇总 =====================

def generate_annual_summary(history_dir, output_dir):
    """基于历史数据生成年度汇总对比 Excel（每人各月集数/提成矩阵）。

    Returns:
        excel_path 或 None（无历史数据时）
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    entries = load_history(history_dir)
    if not entries:
        return None
    months = [e.get('month', '?') for e in entries]
    persons = sorted(set(
        nm for e in entries for nm in e.get('commission', {})
    ))

    wb = Workbook()
    # Sheet1: 集数汇总
    ws1 = wb.active
    ws1.title = '集数汇总'
    title_font = Font(name='宋体', size=14, bold=True)
    hdr_font = Font(name='宋体', size=11, bold=True, color='1F4E79')
    data_font = Font(name='宋体', size=11)
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='D6E4F0')
    total_fill = PatternFill('solid', fgColor='F3F6FA')

    # 标题
    ws1.merge_cells(f'A1:{chr(64+2+len(months))}1')
    ws1.cell(1, 1, f'年度集数汇总（{months[0]} ~ {months[-1]}）').font = title_font
    ws1.cell(1, 1).alignment = center

    # 表头：人员 + 各月
    ws1.cell(2, 1, '人员').font = hdr_font
    for i, m in enumerate(months):
        c = ws1.cell(2, 2 + i, m); c.font = hdr_font; c.alignment = center
    # 每行数据
    r = 3
    for nm in persons:
        ws1.cell(r, 1, nm).font = data_font
        for i in range(len(months)):
            c = ws1.cell(r, 2 + i, 0); c.font = data_font; c.alignment = center
        r += 1
    for i, e in enumerate(entries):
        cd = e.get('commission', {})
        for pi, nm in enumerate(persons):
            ws1.cell(3 + pi, 2 + i, cd.get(nm, {}).get('total_episodes', 0))

    # Sheet2: 提成汇总
    ws2 = wb.create_sheet('提成汇总')
    ws2.merge_cells(f'A1:{chr(64+2+len(months))}1')
    ws2.cell(1, 1, f'年度提成汇总（{months[0]} ~ {months[-1]}）').font = title_font
    ws2.cell(1, 1).alignment = center
    ws2.cell(2, 1, '人员').font = hdr_font
    for i, m in enumerate(months):
        c = ws2.cell(2, 2 + i, m); c.font = hdr_font; c.alignment = center
    r = 3
    for nm in persons:
        ws2.cell(r, 1, nm).font = data_font
        r += 1
    for i, e in enumerate(entries):
        cd = e.get('commission', {})
        for pi, nm in enumerate(persons):
            ws2.cell(3 + pi, 2 + i, cd.get(nm, {}).get('total_commission', 0))

    # 应用边框/对齐
    for ws in [ws1, ws2]:
        for row in ws.iter_rows(min_row=2, max_row=len(persons) + 2, max_col=1 + len(months)):
            for cell in row:
                cell.border = bdr
                cell.alignment = center
        ws.column_dimensions['A'].width = 12
        for i in range(len(months)):
            ws.column_dimensions[chr(66 + i)].width = 16

    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, '年度汇总对比.xlsx')
    wb.save(excel_path)
    return excel_path


# ===================== #4 异常/风险预警 =====================

def scan_anomalies(records, commission_data, history_dir=None, monthly_goals=None):
    """扫描异常/风险，返回告警列表。

    检测项：
    - 未达标人员（is_complete == '否'）
    - 集数骤降（与上一月相比下降 >= 40%）
    - 超时占比过高（超时集数 / 总集数 >= 30%）
    - 目标严重滞后（实际集数 < 目标集数的 70%）

    Returns:
        list[dict]：每条 {type, level, person, msg}
    """
    warnings = []
    monthly_goals = monthly_goals or {}

    # 1) 未达标
    for nm, cd in commission_data.items():
        if cd.get('is_complete') == '否':
            warnings.append({
                'type': '未达标', 'level': '高', 'person': nm,
                'msg': f'未完成基准集数（{cd.get("total_episodes", 0)}/{cd.get("quota", 0)}）'
            })

    # 2) 目标滞后
    for nm, goal in monthly_goals.items():
        eps_goal = goal.get('episodes')
        if eps_goal:
            actual = commission_data.get(nm, {}).get('total_episodes', 0)
            if eps_goal > 0 and actual < eps_goal * 0.7:
                warnings.append({
                    'type': '目标滞后', 'level': '中', 'person': nm,
                    'msg': f'集数 {actual} < 目标 {eps_goal} 的70%'
                })

    # 3) 超时占比过高
    person_overtime = {}
    person_total = {}
    for r in records:
        nm = r.get('身份证姓名', '')
        person_total[nm] = person_total.get(nm, 0) + r.get('单项目数/集数', 0)
        person_overtime[nm] = person_overtime.get(nm, 0) + r.get('超时集数', 0)
    for nm in person_total:
        tot = person_total[nm]
        ot = person_overtime.get(nm, 0)
        if tot > 0 and ot / tot >= 0.3:
            warnings.append({
                'type': '超时偏高', 'level': '中', 'person': nm,
                'msg': f'超时 {ot} 集 / 总 {tot} 集（{int(ot/tot*100)}%）'
            })

    # 4) 集数骤降（与上一月对比）
    if history_dir:
        entries = load_history(history_dir)
        if len(entries) >= 2:
            prev = entries[-2].get('commission', {})
            for nm, cd in commission_data.items():
                prev_eps = prev.get(nm, {}).get('total_episodes', 0)
                cur_eps = cd.get('total_episodes', 0)
                if prev_eps > 0 and cur_eps < prev_eps * 0.6:
                    warnings.append({
                        'type': '集数骤降', 'level': '中', 'person': nm,
                        'msg': f'本月 {cur_eps} 集 < 上月 {prev_eps} 集的60%'
                    })

    # 按等级排序（高优先）
    order = {'高': 0, '中': 1, '低': 2}
    warnings.sort(key=lambda w: order.get(w.get('level', '低'), 9))
    return warnings


# ===================== #5 数据导入去重与清洗 =====================

def clean_import_data(df, id_col=None, name_col=None):
    """清洗导入数据：去重、去空行、标记异常。

    Args:
        df: pandas DataFrame 原始导入数据
        id_col: 项目ID列索引（0-based），None 则自动探测
        name_col: 人员名列索引，None 则自动探测

    Returns:
        (cleaned_df, report)：清洗后的 DataFrame 与报告列表
    """
    import pandas as pd
    report = []
    df = df.copy()

    # 自动探测列：找含"项目"且含数字ID的列、含人名的列
    if id_col is None or name_col is None:
        for ci in range(min(df.shape[1], 10)):
            sample = [str(x) for x in df.iloc[:30, ci].tolist() if pd.notna(x)]
            joined = ' '.join(sample)
            if name_col is None and any(len(s) in (2, 3) and s.isalpha() for s in sample if isinstance(s, str)):
                # 姓名列通常短文本
                name_col = ci
            if id_col is None and any(s.isdigit() and len(s) >= 3 for s in sample if isinstance(s, str)):
                id_col = ci

    # 去空行
    before = len(df)
    df = df.dropna(how='all')
    if len(df) < before:
        report.append(f'移除空行 {before - len(df)} 行')

    # 去重（按项目ID+人员，若可用）
    if id_col is not None and id_col < df.shape[1]:
        col_names = list(df.columns)
        # subset 需要列名；若列名是默认整数则直接用索引名
        def _col_ref(idx):
            c = col_names[idx]
            return c
        dup_key = [_col_ref(id_col)]
        if name_col is not None and name_col < df.shape[1]:
            dup_key.append(_col_ref(name_col))
        dup_before = len(df)
        df = df.drop_duplicates(subset=dup_key, keep='first')
        if len(df) < dup_before:
            report.append(f'按项目ID/人员去重，移除重复 {dup_before - len(df)} 行')

    # 标记空项目ID
    if id_col is not None and id_col < df.shape[1]:
        empty_ids = df[df.iloc[:, id_col].isna() | (df.iloc[:, id_col].astype(str).str.strip() == '')]
        if len(empty_ids):
            report.append(f'发现 {len(empty_ids)} 行缺少项目ID（已保留，请检查）')

    report.append(f'清洗后共 {len(df)} 行（原始 {before} 行）')
    return df, report


# ===================== #7 导出增强 =====================

def export_enhanced(records, commission_data, output_dir, split_by_role=False):
    """增强导出：人员提成总表 + 可选按角色拆分 + Excel 多工作表。

    Args:
        records: 项目记录
        commission_data: 提成结果
        output_dir: 输出目录
        split_by_role: 是否按角色拆分 CSV

    Returns:
        list[str]：生成的文件路径列表
    """
    import csv
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    os.makedirs(output_dir, exist_ok=True)
    created = []

    # ---- 1) 增强 CSV：自定义列 ----
    person_csv = os.path.join(output_dir, '人员提成_增强.csv')
    with open(person_csv, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['姓名', '角色', '总集数', '基准', '绩效', '超额/组奖', '缺集扣除', '提成合计', '是否达标'])
        for name in commission_data:
            cd = commission_data.get(name, {})
            extra = cd.get('overtime_bonus', 0) + cd.get('group_project_bonus', 0)
            w.writerow([
                name, cd.get('role', ''), cd.get('total_episodes', 0),
                cd.get('quota', 0), cd.get('is_complete', ''),
                extra, cd.get('shortage_penalty', 0), cd.get('total_commission', 0),
                '✅' if cd.get('is_complete') == '是' else '❌',
            ])
    created.append(person_csv)

    # ---- 2) 按角色拆分 CSV ----
    if split_by_role:
        roles = {}
        for name, cd in commission_data.items():
            r = cd.get('role', '未知')
            roles.setdefault(r, []).append((name, cd))
        for r, items in roles.items():
            safe = re.sub(r'[\/:*?"<>|]', '_', str(r))
            r_csv = os.path.join(output_dir, f'角色_{safe}.csv')
            with open(r_csv, 'w', newline='', encoding='utf-8-sig') as f:
                w = csv.writer(f)
                w.writerow(['姓名', '角色', '总集数', '基准', '提成合计'])
                for name, cd in items:
                    w.writerow([name, r, cd.get('total_episodes', 0),
                                cd.get('quota', 0), cd.get('total_commission', 0)])
            created.append(r_csv)

    # ---- 3) Excel 多工作表 ----
    xlsx = os.path.join(output_dir, '提成汇总_多表.xlsx')
    wb = Workbook()
    title_font = Font(name='宋体', size=13, bold=True)
    hdr_font = Font(name='宋体', size=11, bold=True, color='1F4E79')
    data_font = Font(name='宋体', size=11)
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Sheet1: 汇总
    ws = wb.active
    ws.title = '汇总'
    ws.merge_cells('A1:H1')
    ws.cell(1, 1, '人员提成汇总').font = title_font
    ws.cell(1, 1).alignment = center
    hdrs = ['姓名', '角色', '总集数', '基准', '绩效', '超额/组奖', '缺集扣除', '提成合计']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(2, ci, h); c.font = hdr_font; c.alignment = center
    r = 3
    for name, cd in commission_data.items():
        extra = cd.get('overtime_bonus', 0) + cd.get('group_project_bonus', 0)
        vals = [name, cd.get('role', ''), cd.get('total_episodes', 0), cd.get('quota', 0),
                cd.get('is_complete', ''), extra, cd.get('shortage_penalty', 0), cd.get('total_commission', 0)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v); c.font = data_font; c.alignment = center
        r += 1

    # Sheet2: 按角色分组
    ws2 = wb.create_sheet('按角色分组')
    ws2.merge_cells('A1:D1')
    ws2.cell(1, 1, '按角色分组').font = title_font
    ws2.cell(1, 1).alignment = center
    for ci, h in enumerate(['角色', '人数', '总集数', '提成合计'], 1):
        c = ws2.cell(2, ci, h); c.font = hdr_font; c.alignment = center
    role_agg = {}
    for name, cd in commission_data.items():
        r = cd.get('role', '未知')
        a = role_agg.setdefault(r, {'count': 0, 'eps': 0, 'comm': 0})
        a['count'] += 1
        a['eps'] += cd.get('total_episodes', 0)
        a['comm'] += cd.get('total_commission', 0)
    r = 3
    for role, a in role_agg.items():
        for ci, v in enumerate([role, a['count'], a['eps'], a['comm']], 1):
            c = ws2.cell(r, ci, v); c.font = data_font; c.alignment = center
        r += 1

    for ws_ in [ws, ws2]:
        for row in ws_.iter_rows(min_row=2, max_row=max(ws_.max_row, 3), max_col=8):
            for cell in row:
                cell.border = bdr
    wb.save(xlsx)
    created.append(xlsx)
    return created


# ===================== #9 数据备份自动策略 =====================

def auto_backup_config(config_path, history_dir, target_dir, keep=15):
    """备份 config.json 和 history/ 目录到指定位置，保留最近 keep 份。

    Returns:
        list[str]：本次备份生成的文件路径
    """
    import shutil as _sh
    os.makedirs(target_dir, exist_ok=True)
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    created = []

    # 备份 config.json
    if os.path.exists(config_path):
        dest = os.path.join(target_dir, f'config_{ts}.json')
        _sh.copy2(config_path, dest)
        created.append(dest)

    # 备份 history/ 目录（打包为 zip）
    if os.path.isdir(history_dir) and os.listdir(history_dir):
        import zipfile
        zip_path = os.path.join(target_dir, f'history_{ts}.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fn in os.listdir(history_dir):
                fp = os.path.join(history_dir, fn)
                if os.path.isfile(fp):
                    zf.write(fp, fn)
        created.append(zip_path)

    # 清理旧备份，保留最近 keep 份
    if keep > 0:
        backups = sorted([f for f in os.listdir(target_dir)
                          if f.startswith(('config_', 'history_')) and f.endswith(('.json', '.zip'))],
                         reverse=True)
        for old in backups[keep * 2:]:
            try:
                os.unlink(os.path.join(target_dir, old))
            except OSError:
                pass

    return created


def backup_due(policy, last_backup_time):
    """判断自动备份是否到期。

    Args:
        policy: 备份策略 dict {enabled, interval_days}
        last_backup_time: 上次备份时间（datetime 或 None）

    Returns:
        bool
    """
    if not policy or not policy.get('enabled'):
        return False
    interval = int(policy.get('interval_days', 7))
    if interval <= 0:
        return False
    if last_backup_time is None:
        return True
    return (datetime.datetime.now() - last_backup_time).days >= interval


# ===================== #11 移动端/网页只读版 =====================

def generate_readonly_dashboard(history_dir, serve_dir):
    """生成只读看板页（供主管在线查看，无需安装软件）。

    从 history/ 聚合各月数据，生成一个自包含 HTML 总览页。

    Returns:
        html_path 或 None
    """
    entries = load_history(history_dir)
    if not entries:
        return None
    months = [e.get('month', '?') for e in entries]
    persons = sorted(set(nm for e in entries for nm in e.get('commission', {})))

    # 总览卡片
    total_comm = sum(e.get('summary', {}).get('total_commission', 0) for e in entries)
    total_eps = sum(e.get('summary', {}).get('total_episodes', 0) for e in entries)

    # 最近一月明细表
    last = entries[-1].get('commission', {})
    detail_rows = ''
    for nm in persons:
        cd = last.get(nm, {})
        st = cd.get('is_complete', '')
        badge = '<span class="b ok">✓</span>' if st == '是' else ('<span class="b no">✗</span>' if st == '否' else '-')
        detail_rows += f'<tr><td>{nm}</td><td>{cd.get("role","")}</td><td>{cd.get("total_episodes",0)}</td><td>{cd.get("total_commission",0)}</td><td>{badge}</td></tr>'

    # 趋势数据（各月提成合计）
    monthly_comm = [e.get('summary', {}).get('total_commission', 0) for e in entries]
    max_c = max(monthly_comm + [1])
    bars = ''
    for i, m in enumerate(months):
        pct = int(monthly_comm[i] / max_c * 100)
        bars += f'<div class="bar-row"><span class="bl">{m}</span><div class="bt"><div class="bf" style="width:{pct}%;background:#3b82f6"></div></div><span class="bv">{monthly_comm[i]:,}</span></div>'

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>提成数据 · 只读看板</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft YaHei',sans-serif;background:#0f172a;color:#e2e8f0;min-height:100vh}}
.header{{background:linear-gradient(135deg,#1e3a5f,#3b82f6);padding:24px;text-align:center}}
.header h1{{font-size:24px;color:#fff;letter-spacing:2px}}.header p{{font-size:13px;opacity:.8;margin-top:4px}}
.ct{{max-width:1100px;margin:0 auto;padding:20px}}
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:14px;margin-bottom:20px}}
.sd{{background:#1e293b;border-radius:12px;padding:18px;text-align:center}}
.sd .v{{font-size:30px;font-weight:700;color:#60a5fa}}.sd .l{{font-size:12px;color:#94a3b8;margin-top:4px}}
.pn{{background:#1e293b;border-radius:12px;padding:20px;margin-bottom:20px}}
.pn h2{{font-size:16px;color:#93c5fd;margin-bottom:14px;border-bottom:1px solid #334155;padding-bottom:8px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{text-align:left;padding:8px;color:#94a3b8;border-bottom:1px solid #334155}}
td{{padding:8px;border-bottom:1px solid #1e293b}}
.b{{display:inline-block;padding:1px 8px;border-radius:8px;font-size:12px}}
.b.ok{{background:#065f46;color:#6ee7b7}}.b.no{{background:#7f1d1d;color:#fca5a5}}
.bar-row{{display:flex;align-items:center;gap:10px;margin:6px 0}}
.bl{{min-width:90px;font-size:12px;color:#94a3b8;text-align:right}}
.bt{{flex:1;height:18px;background:#334155;border-radius:4px}}
.bf{{height:100%;border-radius:4px}}
.bv{{min-width:70px;font-size:13px;font-weight:600;color:#60a5fa;text-align:left}}
.footer{{text-align:center;padding:18px;color:#475569;font-size:12px}}
@media(max-width:600px){{.header{{padding:16px}}.ct{{padding:12px}}.cards{{grid-template-columns:1fr 1fr}}}}
</style></head><body>
<div class="header"><h1>📊 提成数据 · 只读看板</h1>
<p>{len(months)} 个月 · {len(persons)} 人 · 只读模式</p></div>
<div class="ct">
<div class="cards">
<div class="sd"><div class="v">{len(months)}</div><div class="l">累计月份</div></div>
<div class="sd"><div class="v">{len(persons)}</div><div class="l">参与人员</div></div>
<div class="sd"><div class="v">{total_comm:,}</div><div class="l">累计提成（元）</div></div>
<div class="sd"><div class="v">{total_eps:,}</div><div class="l">累计集数</div></div>
</div>
<div class="pn"><h2>📈 月度提成趋势</h2>{bars}</div>
<div class="pn"><h2>👥 最新月份明细（{months[-1]}）</h2><div style="overflow-x:auto"><table>
<thead><tr><th>姓名</th><th>角色</th><th>集数</th><th>提成</th><th>绩效</th></tr></thead>
<tbody>{detail_rows}</tbody></table></div></div>
</div>
<div class="footer">AI后期剪辑提成工具 · 只读看板 · 数据自动汇总</div>
</body></html>'''
    os.makedirs(serve_dir, exist_ok=True)
    html_path = os.path.join(serve_dir, '只读看板.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return html_path
