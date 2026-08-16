# -*- coding: utf-8 -*-
"""
AI后期剪辑提成表生成工具 v4.0
==============================
方案：复制模板 → 解析数据 → 计算绩效/提成 → 填入完整提成表。
角色配置：编辑 config.json，自由选择一卡/二卡/组长。

用法：双击 run.bat 或 python generate_commission.py
"""

import pandas as pd
import re
import json
import datetime
import os
import sys
import shutil
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy

# 强制 stdout 使用 UTF-8，避免 Windows GBK 编码报错（仅在直接运行时）
if __name__ == '__main__' or (len(sys.argv) >= 2 and sys.stdout.encoding != 'utf-8'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

# ===================== 路径 =====================
try:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    SCRIPT_DIR = os.getcwd()
ROOT_DIR = os.path.dirname(SCRIPT_DIR)  # 项目根目录

# 支持命令行参数：python generate_commission.py [项目文件] [模板文件] [输出目录]
# 可选：--overtime-file 本次生成使用的超时标记 JSON 文件。
CLI_ARGS = sys.argv[1:]
OVERTIME_FILE = None
if '--overtime-file' in CLI_ARGS:
    overtime_index = CLI_ARGS.index('--overtime-file')
    try:
        OVERTIME_FILE = CLI_ARGS[overtime_index + 1]
    except IndexError:
        raise ValueError('--overtime-file 需要提供文件路径')
    del CLI_ARGS[overtime_index:overtime_index + 2]

if len(CLI_ARGS) >= 1:
    PROJECT_FILE = CLI_ARGS[0]
else:
    PROJECT_FILE = os.path.join(ROOT_DIR, '一组AI项目.xlsx')

if len(CLI_ARGS) >= 2:
    TEMPLATE_FILE = CLI_ARGS[1]
else:
    TEMPLATE_FILE = os.path.join(ROOT_DIR, 'AI后期剪辑提成一组模板.xlsx')
    if not os.path.exists(TEMPLATE_FILE):
        TEMPLATE_FILE = os.path.join(ROOT_DIR, 'AI后期剪辑提成一组最新.xlsx')

# 输出目录：命令行第3参数 > 默认脚本目录
if len(CLI_ARGS) >= 3:
    OUTPUT_DIR = CLI_ARGS[2]
else:
    OUTPUT_DIR = ROOT_DIR

CONFIG_FILE = os.path.join(ROOT_DIR, 'config.json')

# ===================== 从模板提取月份 =====================

def get_month_from_template(template_path):
    """从模板表头提取中文月份和完整日期字符串，如 ('七月', '2026年07月')"""
    CN_MONTHS = ['', '一月','二月','三月','四月','五月','六月',
                 '七月','八月','九月','十月','十一月','十二月']
    try:
        wb = load_workbook(template_path, read_only=True)
        ws = wb.active
        title = ws.cell(1, 2).value or ws.cell(1, 1).value or ''
        wb.close()
        m = re.search(r'(\d+)年(\d+)月', str(title))
        if m:
            year, month_num = int(m.group(1)), int(m.group(2))
            cn = CN_MONTHS[month_num] if 1 <= month_num <= 12 else '当月'
            return cn, f'{year}年{month_num:02d}月'
    except Exception:
        pass
    return '当月', '当月'


def get_month_from_data(df):
    """从项目数据文件提取真实月份。

    数据文件第一行通常是"八月份"这样的中文月份标题；
    若识别到，返回 (中文月份, 月份数字)，否则返回 (None, None)。
    """
    # 中文月份（第1个是空串占位，索引=月份数字）
    CN_MONTHS = ['', '一月','二月','三月','四月','五月','六月',
                 '七月','八月','九月','十月','十一月','十二月']
    # 带"份"的标题形式（数据文件里常见）
    CN_MONTHS_FEN = ['', '一月份','二月份','三月份','四月份','五月份','六月份',
                     '七月份','八月份','九月份','十月份','十一月份','十二月份']
    try:
        for i in range(min(3, len(df))):
            cell = clean(df.iloc[i, 0])
            if cell in CN_MONTHS_FEN:
                month_num = CN_MONTHS_FEN.index(cell)
                return CN_MONTHS[month_num], month_num
            if cell in CN_MONTHS:
                month_num = CN_MONTHS.index(cell)
                return CN_MONTHS[month_num], month_num
    except Exception:
        pass
    return None, None

OUTPUT_MONTH, TEMPLATE_DATE = get_month_from_template(TEMPLATE_FILE)
template_year_match = re.match(r'(\d{4})年', TEMPLATE_DATE)
TEMPLATE_YEAR = int(template_year_match.group(1)) if template_year_match else None
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'AI后期剪辑提成一组{OUTPUT_MONTH}.xlsx')

# ===================== 加载配置 =====================

def load_config():
    """加载并校验配置，且保证所有已配置人员都会进入输出顺序。"""
    from config_loader import load_config as load_validated_config
    return load_validated_config(CONFIG_FILE).to_dict()


def set_config(config):
    """注入配置并同步派生映射，供 CLI 与 GUI 复用。"""
    global cfg, ROLE_MAP, RULES, GROUPS, NAME_ORDER, ALL_NAMES
    cfg = config
    ROLE_MAP = cfg['人员角色']
    RULES = cfg['rules']
    GROUPS = cfg.get('小组', {})
    configured_order = cfg.get('人员排序', [])
    NAME_ORDER = list(dict.fromkeys(
        [name for name in configured_order if name in ROLE_MAP]
        + [name for name in ROLE_MAP if name not in configured_order]
    ))
    ALL_NAMES = list(ROLE_MAP.keys())


set_config(load_config())

def normalize_role(role_str):
    """容错：将用户输入的角色名规范化为标准名"""
    if not role_str:
        return '一卡剪辑'
    s = role_str.strip()
    # 先检查一卡/二卡/助理（避免"一卡组长"被误判为组长）
    if '一卡' in s:
        return '一卡剪辑'
    if '二卡' in s:
        return '二卡剪辑'
    if '助理' in s:
        return '剪辑助理'
    if '小组长' in s:
        return '一卡剪辑'   # 小组长和一卡计算方式一样
    if '组长' in s:
        return '剪辑组长'
    return '一卡剪辑'  # 默认

# 不再强制覆盖——GROUPS 中的组长仅表示谁是组长，不强制其角色
# （只有 config.json 人员角色中标记了"剪辑组长"的才是真正的组长）

# ===================== 文本清洗 =====================

def clean(text):
    if pd.isna(text) or text == '':
        return ''
    text = str(text).replace('\xa0', ' ').replace('\n', '').replace('\r', '')
    return re.sub(r'\s+', ' ', text).strip()


def extract_project_id(text):
    if not text:
        return ''
    text = clean(text)
    all_4 = re.findall(r'(?<!\d)\d{4}(?!\d)', text)
    if not all_4:
        longer = re.findall(r'(?<!\d)\d{4,}(?!\d)', text)
        return longer[0] if longer else ''
    if len(all_4) == 1:
        return all_4[0]
    good = []
    for m in all_4:
        idx = text.find(m)
        if idx > 0 and text[idx - 1].isalpha():
            continue
        if m.startswith('0'):
            continue
        good.append(m)
    if good:
        return max(good, key=int)
    non_zero = [m for m in all_4 if not m.startswith('0')]
    return non_zero[0] if non_zero else all_4[-1]


def clean_project_name(name):
    name = clean(name)
    if not name:
        return ''
    name = re.sub(r'^\d+-\d+-[（(]海外[）)]?\s*', '', name)
    name = re.sub(r'^\d+-[A-Za-z]\d+\s*', '', name)
    name = re.sub(r'^[A-Za-z]\d+-\d+\s*', '', name)
    name = re.sub(r'^\d+-\d+[_-]\d+[_-]\s*', '', name)
    name = re.sub(r'^\d+[A-Za-z]-\d+[_-]\d+[_-]\s*', '', name)
    name = re.sub(r'^[\dA-Za-z\-_/]+[_-]\s*', '', name)
    name = re.sub(r'^\d+-\d+-\s*', '', name)
    return name.strip()


def parse_episode_ranges(text):
    if not text:
        return [], 0
    text = clean(text)
    if not text:
        return [], 0
    # 多分隔符统一：逗号/分号/加号/空格 -> 逗号
    text = re.sub(r'[；;，,。+、\s]+', ',', text).strip(',')
    episodes = []
    for part in re.split(r',', text):
        part = part.strip()
        if not part:
            continue
        m = re.match(r'(\d+)\s*[-–—]\s*(\d+)', part)
        if m:
            s, e = int(m.group(1)), int(m.group(2))
            episodes.extend(range(min(s, e), max(s, e) + 1))
        else:
            m = re.match(r'(\d+)', part)
            if m:
                episodes.append(int(m.group(1)))
    result = sorted(set(episodes))
    return result, len(result)


def parse_overtime_episodes(f_col_text):
    """解析F列的超时集数标注，返回 set of int。支持：4,5,10-12 或 4 5 10 等多种格式"""
    if not f_col_text:
        return set()
    text = clean(str(f_col_text))
    if not text:
        return set()
    eps, _ = parse_episode_ranges(text)
    return set(eps)


def parse_person_assignment(text):
    if not text:
        return None
    text = clean(text)
    m = re.match(r'^([^\d:：]+)[：:]\s*(.+)', text)
    if m:
        eps, cnt = parse_episode_ranges(m.group(2).strip())
        if cnt > 0:
            return (m.group(1).strip(), ','.join(str(e) for e in eps), cnt)
    for known in sorted(ALL_NAMES, key=len, reverse=True):
        if text.startswith(known):
            rest = text[len(known):].strip()
            if rest and re.match(r'^[,\d]', rest):
                rest = re.sub(r'^[，,]\s*', '', rest)
                eps, cnt = parse_episode_ranges(rest)
                if cnt > 0:
                    return (known, ','.join(str(e) for e in eps), cnt)
            break
    m = re.match(r'^([^\d]+?)\s{2,}(\d.*)', text)
    if m:
        eps, cnt = parse_episode_ranges(m.group(2).strip())
        if cnt > 0:
            return (m.group(1).strip(), ','.join(str(e) for e in eps), cnt)
    m = re.match(r'^([^\d]+)\s+(\d[\d\s,\-–—;；+]+)', text)
    if m:
        eps, cnt = parse_episode_ranges(m.group(2).strip())
        if cnt > 0:
            return (m.group(1).strip(), ','.join(str(e) for e in eps), cnt)
    return None


def parse_delivery_date(date_str, default_year=None):
    if not date_str:
        return None
    date_str = clean(date_str)
    m = re.search(r'(\d+)\.(\d+)', date_str)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        year = default_year or datetime.date.today().year
        return datetime.date(year, month, day)
    return None


# ===================== 数据解析 =====================

def load_overtime_map(overtime_file):
    """加载本次生成指定的超时集数，拒绝格式不正确的输入。"""
    if not overtime_file:
        return {}
    with open(overtime_file, 'r', encoding='utf-8-sig') as f:
        raw_map = json.load(f)
    if not isinstance(raw_map, dict):
        raise ValueError('超时标记文件必须是 JSON 对象，格式为 {项目ID: [集数]}')
    result = {}
    for pid, episodes in raw_map.items():
        if not isinstance(episodes, list):
            raise ValueError(f'项目 {pid} 的超时集数必须是数组')
        result[str(pid)] = {int(ep) for ep in episodes}
    return result


def parse_projects(df, default_year=None, overtime_map=None):
    """解析项目表；项目标题不完整时跳过其分配行并输出明确警告。"""
    records = []
    proj_name, proj_id, end_date = '', '', None
    project_catalog = {}
    group_pids = {g: set() for g in GROUPS}
    overtime_map = overtime_map or {}

    for i in range(len(df)):
        c0 = clean(df.iloc[i, 0])
        c2 = clean(df.iloc[i, 2])
        c3 = clean(df.iloc[i, 3])

        # 跳过月份标题
        months = ['一月份','二月份','三月份','四月份','五月份','六月份',
                  '七月份','八月份','九月份','十月份','十一月份','十二月份']
        if c0 in months:
            continue

        if c0:
            # 新项目必须从空状态开始，不能沿用上一项目的 ID/日期。
            proj_name = clean_project_name(c0)
            proj_id = extract_project_id(c0)
            end_date = parse_delivery_date(c3, default_year)

        if c3:
            dt = parse_delivery_date(c3, default_year)
            if dt:
                end_date = dt

        if proj_id and proj_name and end_date:
            project_catalog[proj_id] = {
                '项目ID': proj_id,
                'AI项目名称': proj_name,
                '开始日期': end_date - datetime.timedelta(days=1),
                '结束日期': end_date,
            }

        # 解析当前项目的超时集数（从 overtime_map 取）
        overtime_set = overtime_map.get(proj_id, set())

        if c2:
            res = parse_person_assignment(c2)
            if res:
                name, detail, cnt = res
                if not proj_id or not proj_name or not end_date:
                    print(f'⚠️ 第{i + 1}行跳过：项目 ID、名称或交付日期不完整')
                elif name not in ROLE_MAP:
                    print(f'⚠️ 第{i + 1}行跳过：人员 "{name}" 未配置角色')
                elif cnt > 0:
                    start_date = end_date - datetime.timedelta(days=1)
                    # 该人员在该项目中分到的集数
                    person_eps = set()
                    for ep_str in detail.split(','):
                        if ep_str.strip().isdigit():
                            person_eps.add(int(ep_str.strip()))

                    # 该人员该项目的超时集数
                    person_overtime = person_eps & overtime_set
                    overtime_cnt = len(person_overtime)
                    # 有效集数 = 普通集数 + 超时集数（每个超时集算2集，即额外+1）
                    effective_cnt = cnt + overtime_cnt

                    records.append({
                        '身份证姓名': name,
                        '角色': normalize_role(ROLE_MAP[name]),
                        '项目ID': proj_id,
                        '项目类型': 'AI海外真人',
                        'AI项目名称': proj_name,
                        '开始日期': start_date,
                        '结束日期': end_date,
                        '完成明细': detail,
                        '单项目数/集数': effective_cnt,
                        '原始集数': cnt,
                        '超时集数': overtime_cnt,
                        '超时明细': sorted(person_overtime),
                        '参与剪辑': True,
                    })
                    # 统计组长所在组的项目
                    for gname, ginfo in GROUPS.items():
                        if name in ginfo['成员'] and proj_id:
                            group_pids[gname].add(proj_id)

    # 剪辑组长需要在明细中覆盖全组所有项目；未参与项目只展示项目信息。
    leader_names = [
        name for name in NAME_ORDER
        if normalize_role(ROLE_MAP.get(name)) == '剪辑组长'
    ]
    assigned_pairs = {
        (r['身份证姓名'], r['项目ID']) for r in records
    }
    for leader_name in leader_names:
        for pid, project in project_catalog.items():
            if (leader_name, pid) in assigned_pairs:
                continue
            records.append({
                '身份证姓名': leader_name,
                '角色': '剪辑组长',
                '项目ID': pid,
                '项目类型': 'AI海外真人',
                'AI项目名称': project['AI项目名称'],
                '开始日期': project['开始日期'],
                '结束日期': project['结束日期'],
                '完成明细': '',
                '单项目数/集数': 0,
                '原始集数': 0,
                '超时集数': 0,
                '超时明细': [],
                '参与剪辑': False,
            })

    return records, group_pids


# ===================== 自查 =====================

def self_check(records):
    id_names = defaultdict(set)
    for r in records:
        if r['项目ID']:
            id_names[r['项目ID']].add(r['AI项目名称'])
    issues = [(pid, names) for pid, names in id_names.items() if len(names) > 1]
    if issues:
        print("\n⚠️  自查警告：以下项目ID对应多个不同的项目名称：")
        for pid, names in issues:
            print(f"    ID={pid}: {names}")
    else:
        print("\n✅ 自查通过：相同项目ID的项目名称均一致。")
    return len(issues) == 0


# ===================== 提成计算 =====================

def compute_commission(records, group_pids):
    """
    提成计算：
    - 一卡：(总集数-70)×20，不足则-(70-集数)×50
    - 二卡/剪辑助理：(总集数-120)×20，不足则-(120-集数)×50
    - 组长：总集数×20 + 全组去重项目数×100
    """
    person_episodes = defaultdict(int)
    for r in records:
        person_episodes[r['身份证姓名']] += r['单项目数/集数']

    # 全组去重项目数（所有小组的项目ID取并集）
    global_pids = {r['项目ID'] for r in records if r.get('项目ID')}
    for pids in group_pids.values():
        global_pids |= pids
    total_unique_projects = len(global_pids)

    result = {}
    for name in NAME_ORDER:
        total = person_episodes.get(name, 0)
        role = normalize_role(ROLE_MAP.get(name, '一卡剪辑'))
        rule = RULES.get(role, RULES['一卡剪辑'])

        if role == '剪辑组长':
            group_bonus = total_unique_projects * rule.get('组内每部提成', 100)
            total_commission = total * rule['每集单价'] + group_bonus
            result[name] = {
                'total_episodes': total, 'role': role, 'rule': rule,
                'quota': 0, 'is_complete': '是',
                'overtime_bonus': 0, 'shortage_penalty': 0,
                'group_project_bonus': group_bonus,
                'project_count': total_unique_projects,
                'total_commission': total_commission,
                'desc': rule['提成构成描述'],
            }
        else:
            # 一卡/二卡/剪辑助理：一卡基准70，其余120
            quota = rule.get('基准集数', 120)
            if total >= quota:
                overtime = (total - quota) * rule['超额每集']
                result[name] = {
                    'total_episodes': total, 'role': role, 'rule': rule,
                    'quota': quota, 'is_complete': '是',
                    'overtime_bonus': overtime, 'shortage_penalty': 0,
                    'group_project_bonus': 0, 'project_count': 0,
                    'total_commission': overtime,
                    'desc': rule['提成构成描述'],
                }
            else:
                shortage = (quota - total) * rule['缺集每集扣']
                result[name] = {
                    'total_episodes': total, 'role': role, 'rule': rule,
                    'quota': quota, 'is_complete': '否',
                    'overtime_bonus': 0, 'shortage_penalty': shortage,
                    'group_project_bonus': 0, 'project_count': 0,
                    'total_commission': -shortage,
                    'desc': rule['提成构成描述'],
                }

    return result


# ===================== Excel生成 =====================

def generate_excel(records, commission_data, template_path, output_path):
    print(f"\n📝 正在生成Excel...")

    # 复制模板
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except PermissionError:
            base, ext = os.path.splitext(output_path)
            output_path = f"{base}_new{ext}"
    shutil.copy2(template_path, output_path)
    print(f"   复制模板: {template_path}")

    wb = load_workbook(output_path)
    ws = wb.active

    # 更新 Excel 主标题月份为数据文件真实月份（如 '后期剪辑部2026年07月提成表' → 八月）
    try:
        title_cell = ws.cell(1, 2)
        if title_cell.value:
            title_cell.value = re.sub(
                r'\d+年\d+月', TEMPLATE_DATE, str(title_cell.value))
    except Exception:
        pass

    data_font = copy(ws.cell(4, 2).font) if ws.cell(4, 2).font else Font(name='宋体', size=11)
    date_fmt = 'yyyy"年"m"月"d"日";@'
    thin = Side(style='thin')
    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 清除旧数据区合并
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= 4:
            ws.unmerge_cells(str(mc))

    data_start = 4
    old_max = ws.max_row
    data_count = len(records)

    # 先彻底清空模板旧数据：第4行到末尾，A-T列全部清空
    for rr in range(data_start, old_max + 1):
        for cc in range(1, 21):
            ws.cell(rr, cc).value = None

    if data_count > old_max - data_start + 1:
        ws.insert_rows(data_start, data_count - (old_max - data_start + 1) + 20)
    elif data_count < old_max - data_start + 1:
        ws.delete_rows(data_start + data_count, old_max - data_start - data_count + 1)

    # 排序
    sorted_records = []
    for nm in NAME_ORDER:
        prs = [r for r in records if r['身份证姓名'] == nm]
        prs.sort(key=lambda x: x['结束日期'])
        sorted_records.extend(prs)

    # 写入数据——每人独立序号
    person_seq = 0
    prev_name = ''
    global_idx = 0

    for r in sorted_records:
        global_idx += 1
        name = r['身份证姓名']
        if name != prev_name:
            person_seq = 0  # 换人，序号重置
        person_seq += 1
        ri = data_start + global_idx - 1
        is_first = (name != prev_name)

        # A: 组别（不写值，最后统一合并为剪辑一组）
        ws.cell(ri, 1).font = data_font
        ws.cell(ri, 1).alignment = center_align

        # B: 姓名
        if is_first:
            ws.cell(ri, 2, name)
        ws.cell(ri, 2).font = data_font
        ws.cell(ri, 2).alignment = center_align

        # C: 职位
        if is_first:
            ws.cell(ri, 3, r['角色'])
        ws.cell(ri, 3).font = data_font
        ws.cell(ri, 3).alignment = center_align

        # D: 序号（每人独立从1开始）
        ws.cell(ri, 4, person_seq)
        ws.cell(ri, 4).font = data_font
        ws.cell(ri, 4).alignment = center_align

        # E: 项目ID
        eid = int(r['项目ID']) if r['项目ID'].isdigit() else r['项目ID']
        ws.cell(ri, 5, eid)
        ws.cell(ri, 5).font = data_font
        ws.cell(ri, 5).alignment = center_align

        # F: 项目类型
        ws.cell(ri, 6, r['项目类型'])
        ws.cell(ri, 6).font = data_font
        ws.cell(ri, 6).alignment = center_align

        # G: AI项目名称
        ws.cell(ri, 7, r['AI项目名称'])
        ws.cell(ri, 7).font = data_font
        ws.cell(ri, 7).alignment = center_align

        # H: 开始日期
        c = ws.cell(ri, 8, r['开始日期'])
        c.font = data_font; c.alignment = center_align; c.number_format = date_fmt

        # I: 结束日期
        c = ws.cell(ri, 9, r['结束日期'])
        c.font = data_font; c.alignment = center_align; c.number_format = date_fmt

        # J: 完成明细
        ws.cell(ri, 10, r['完成明细'])
        ws.cell(ri, 10).font = data_font; ws.cell(ri, 10).alignment = center_align

        # K: 单项目数/集数 — 有超时集数时高亮
        episode_value = r['单项目数/集数'] if r.get('参与剪辑', True) else None
        c = ws.cell(ri, 11, episode_value)
        c.font = data_font; c.alignment = center_align
        if r.get('超时集数', 0) > 0:
            c.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            c.font = Font(name='宋体', size=11, bold=True)

        # L: 提成——组长填集数提成（集数×20），其他人留空
        if r['角色'] == '剪辑组长' and r.get('参与剪辑', True):
            ws.cell(ri, 12, r['单项目数/集数'] * 20)
        else:
            ws.cell(ri, 12).value = None
        ws.cell(ri, 12).font = data_font; ws.cell(ri, 12).alignment = center_align

        # M-T 清空（合并时统一填）
        for cc in range(13, 21):
            ws.cell(ri, cc).value = None
            ws.cell(ri, cc).font = data_font
            ws.cell(ri, cc).alignment = center_align

        for cc in range(1, 21):
            ws.cell(ri, cc).border = full_border

        prev_name = name

    last_row = data_start + len(sorted_records) - 1

    # 清除超出范围的数据并删除多余行（避免底部空白行）
    for rr in range(last_row + 1, ws.max_row + 1):
        for cc in range(1, 21):
            ws.cell(rr, cc).value = None
    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)

    # ---- 合并单元格 ----
    # A列：全部合并为"剪辑一组"
    ws.merge_cells(f'A{data_start}:A{last_row}')
    ws.cell(data_start, 1, '剪辑一组')
    ws.cell(data_start, 1).font = data_font
    ws.cell(data_start, 1).alignment = center_align
    ws.cell(data_start, 1).border = full_border

    # B/C列及M-T列：按人员汇总，多条记录时再合并
    cur_person = ''
    merge_start = data_start

    for idx, r in enumerate(sorted_records):
        if r['身份证姓名'] != cur_person:
            if cur_person:
                person_end = data_start + idx - 1
                _apply_person_merge(ws, merge_start, person_end, cur_person,
                                    sorted_records, commission_data,
                                    data_font, center_align, center_wrap, full_border)
            cur_person = r['身份证姓名']
            merge_start = data_start + idx

    # 最后一组
    if cur_person:
        _apply_person_merge(ws, merge_start, last_row, cur_person,
                            sorted_records, commission_data,
                            data_font, center_align, center_wrap, full_border)

    ws.freeze_panes = 'A4'
    _auto_fit_sheet(ws, data_start, last_row)

    # ---- 生成统计简报 (Sheet2) ----
    generate_summary_sheet(wb, sorted_records, commission_data)

    wb.save(output_path)
    print(f"✅ 已生成: {output_path}")
    print(f"   数据: 第{data_start}行 ~ 第{last_row}行 ({len(sorted_records)}条)")

    html_path = generate_html_dashboard(sorted_records, commission_data, output_path)

    # 自动打开文件
    try:
        os.startfile(output_path)
        print(f"📂 已自动打开 Excel")
    except Exception:
        pass
    if html_path:
        try:
            os.startfile(html_path)
            print(f"📂 已自动打开 统计仪表盘")
        except Exception:
            pass

    return output_path, html_path


def _auto_fit_sheet(ws, data_start, last_row):
    """自动调整列宽和行高，让表格更美观"""
    from openpyxl.utils import get_column_letter as _gcl

    # 每列最大字符宽度估算
    col_max_len = {}  # col_idx -> max chars
    for row in ws.iter_rows(min_row=3, max_row=last_row):
        for cell in row:
            col_idx = cell.column
            if cell.value:
                # 中文字符算2个宽度，英文算1个
                text = str(cell.value).replace('\n', ' ')
                w = sum(2 if '\u4e00' <= c <= '\u9fff' or '\u3000' <= c <= '\u303f' else 1 for c in text)
                col_max_len[col_idx] = max(col_max_len.get(col_idx, 0), w)

    # 设置列宽（带上下限）
    for col_idx, char_w in col_max_len.items():
        letter = _gcl(col_idx)
        # 加一点边距
        width = min(char_w + 4, 60) if col_idx == 7 else min(char_w + 3, 45)
        width = max(width, 6)
        ws.column_dimensions[letter].width = width

    # G列（项目名称）适当放宽
    ws.column_dimensions['G'].width = min(max(col_max_len.get(7, 20) + 4, 28), 55)

    # J列（完成明细）稍宽
    ws.column_dimensions['J'].width = min(max(col_max_len.get(10, 15) + 3, 18), 40)

    # Q列（提成构成）适当宽
    ws.column_dimensions['Q'].width = min(max(col_max_len.get(17, 20) + 4, 24), 45)

    # R列（提成合计，算式文本）适当宽
    ws.column_dimensions['R'].width = min(max(col_max_len.get(18, 20) + 4, 24), 40)

    # 数据行高：统一 20pt
    for r in range(data_start, last_row + 1):
        ws.row_dimensions[r].height = 20

    # 表头行高
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 33
    ws.row_dimensions[3].height = 48

    # F: 打印设置 —— 横向、适应页宽、缩放
    try:
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
    except Exception:
        pass


def _apply_person_merge(ws, start, end, name, sorted_records, comm_data,
                        data_font, center_align, center_wrap, full_border):
    """按人员合并B-C列和M-T列，并填入汇总数据"""
    if end > start:
        # B/C列合并（A列已在外部全局合并）
        for col_letter in ['B', 'C']:
            ws.merge_cells(f'{col_letter}{start}:{col_letter}{end}')

        # M-T列按人员合并
        for col_letter in ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
            ws.merge_cells(f'{col_letter}{start}:{col_letter}{end}')

    cd = comm_data.get(name, {})
    total_ep = cd.get('total_episodes', 0)
    overtime = cd.get('overtime_bonus', 0)
    shortage = cd.get('shortage_penalty', 0)
    group_bonus = cd.get('group_project_bonus', 0)
    is_complete = cd.get('is_complete', '')
    total_comm = cd.get('total_commission', 0)
    desc = cd.get('desc', '')
    role = cd.get('role', '')
    rule = cd.get('rule', {})
    project_count = cd.get('project_count', 0)

    # F: 按角色给 B/C 列（姓名/角色区）着色
    role_fill = {
        '剪辑组长': 'FEF3EB', '一卡剪辑': 'E6F7F0',
        '二卡剪辑': 'EAF2FF', '剪辑助理': 'F3F0FF', '小组长': 'FEF6EC',
    }
    try:
        _rf = role_fill.get(role, '')
        if _rf:
            for _col_letter in ['B', 'C']:
                for _rr in range(start, end + 1):
                    ws.cell(_rr, ord(_col_letter) - 64).fill = PatternFill('solid', fgColor=_rf)
    except Exception:
        pass

    # M: 总项目数/集数（组长填"项目数/集数"，其他人填集数）
    if role == '剪辑组长':
        m_value = f'{project_count}/{total_ep}'
    else:
        m_value = total_ep
    c = ws.cell(start, 13, m_value)
    c.font = data_font; c.alignment = center_align; c.border = full_border

    # N: 绩效是否完成——组长永远"是"
    c = ws.cell(start, 14, is_complete)
    c.font = data_font; c.alignment = center_align; c.border = full_border

    # O: 任务未完成扣除金额
    if shortage > 0:
        c = ws.cell(start, 15, shortage)
        c.font = data_font; c.alignment = center_align; c.border = full_border

    # P: 任务超额提成金额
    if role == '剪辑组长':
        # 组长P列填集数提成（集数×20）
        episode_money = total_ep * rule.get('每集单价', 20)
        c = ws.cell(start, 16, episode_money)
        c.font = data_font; c.alignment = center_align; c.border = full_border
    else:
        # 其他人P列填超额
        if overtime > 0:
            c = ws.cell(start, 16, overtime)
            c.font = data_font; c.alignment = center_align; c.border = full_border

    # Q: 提成构成（本月规则变动需标黄）
    c = ws.cell(start, 17, desc)
    c.font = Font(name='宋体', size=8, bold=False)
    c.alignment = center_wrap; c.border = full_border

    # R: 提成合计（写清计算过程：算式=结果）
    if role == '剪辑组长':
        eps_price = rule.get('每集单价', 20)
        proj_price = rule.get('组内每部提成', 100)
        r_value = f'{total_ep}×{eps_price}+{project_count}×{proj_price}={total_comm}'
    else:
        quota = rule.get('基准集数', 120)
        if total_ep >= quota:
            over_price = rule.get('超额每集', 20)
            r_value = f'({total_ep}-{quota})×{over_price}={total_comm}'
        else:
            short_price = rule.get('缺集每集扣', 50)
            r_value = f'-({quota}-{total_ep})×{short_price}={total_comm}'
    c = ws.cell(start, 18, r_value)
    c.font = data_font; c.alignment = center_align; c.border = full_border
    # F: 条件格式 —— 提成合计为正标绿、为负标红
    try:
        if total_comm > 0:
            c.fill = PatternFill('solid', fgColor='E6F7F0')
        elif total_comm < 0:
            c.fill = PatternFill('solid', fgColor='FEF0F0')
            c.font = Font(name='宋体', size=14, bold=False, color='CC0000')
    except Exception:
        pass

    # S: 奖/罚（不填，留空）

    # T: 备注
    # 收集该人员的超时集数信息
    overtime_notes = []
    for r in sorted_records:
        if r['身份证姓名'] == name and r.get('超时集数', 0) > 0:
            pid = r['项目ID']
            eps_str = ','.join(str(e) for e in sorted(r.get('超时明细', [])))
            overtime_notes.append(f'项目{pid}: 集{eps_str}超4分(+{r["超时集数"]}集)')

    remark_parts = []
    if group_bonus > 0:
        remark_parts.append(f"组内项目提成 +{group_bonus}")
    if overtime_notes:
        remark_parts.append('；'.join(overtime_notes))

    if remark_parts:
        c = ws.cell(start, 20, '\n'.join(remark_parts))
        c.font = Font(name='宋体', size=9, bold=False)
        c.alignment = center_wrap; c.border = full_border


# ===================== 统计简报 (Sheet2) =====================

def generate_summary_sheet(wb, records, commission_data):
    """在 Sheet2 生成可视化月度统计简报"""
    from openpyxl.utils import get_column_letter as _gcl

    ws2 = wb.create_sheet("月度统计简报")

    # 样式
    title_font = Font(name='宋体', size=16, bold=True, color='1F4E79')
    section_font = Font(name='宋体', size=12, bold=True, color='2E75B6')
    hdr_font = Font(name='宋体', size=11, bold=True)
    data_font = Font(name='宋体', size=11)
    note_font = Font(name='宋体', size=9, color='808080')
    green_font = Font(name='宋体', size=11, color='008000')
    red_font = Font(name='宋体', size=11, color='CC0000')
    blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    thin = Side(style='thin')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    ca = Alignment(horizontal='center', vertical='center')
    la = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 统计
    total_people = len(set(r['身份证姓名'] for r in records))
    seen_pids = {}
    for r in records:
        pid = r['项目ID']
        if pid:
            if pid not in seen_pids:
                seen_pids[pid] = {'name': r['AI项目名称'], 'date': r['结束日期'], 'people': set(), 'episodes': 0}
            seen_pids[pid]['people'].add(r['身份证姓名'])
            seen_pids[pid]['episodes'] += r['单项目数/集数']
    total_projects = len(seen_pids)
    total_episodes_all = sum(r['单项目数/集数'] for r in records)
    total_comm_all = sum(cd['total_commission'] for cd in commission_data.values())
    done = sum(1 for cd in commission_data.values() if cd['is_complete'] == '是')
    fail = sum(1 for cd in commission_data.values() if cd['is_complete'] == '否')
    card1 = sum(1 for cd in commission_data.values() if cd['role'] == '一卡剪辑')
    card2 = sum(1 for cd in commission_data.values() if cd['role'] == '二卡剪辑')

    row = 1; ws2.column_dimensions['A'].width = 8

    # ---------- 标题 ----------
    ws2.merge_cells('A1:H1')
    ws2.cell(1, 1, f'后期剪辑部 · {TEMPLATE_DATE} · 月度统计简报').font = title_font
    ws2.cell(1, 1).alignment = ca; ws2.row_dimensions[1].height = 36
    ws2.merge_cells('A2:H2')
    ws2.cell(2, 1, '剪辑一组').font = note_font
    ws2.cell(2, 1).alignment = ca
    row = 4

    # ---------- 一、月度总览 ----------
    ws2.merge_cells(f'A{row}:H{row}')
    ws2.cell(row, 1, '一、月度总览').font = section_font
    ws2.row_dimensions[row].height = 24; row += 1

    overview = [
        ('总人数', f'{total_people}人', '项目总数', f'{total_projects}部',
         '全组总集数', f'{total_episodes_all}集', '全组总提成', f'{total_comm_all:,}元'),
        ('绩效达标', f'{done}人', '未达标', f'{fail}人',
         '一卡剪辑', f'{card1}人', '二卡剪辑', f'{card2}人'),
    ]
    for items in overview:
        for ci, txt in enumerate(items):
            c = ws2.cell(row, ci + 1, txt)
            c.font = hdr_font if ci % 2 == 0 else data_font
            c.border = bdr; c.alignment = ca
            if ci % 2 == 0:
                c.fill = blue_fill
            elif ci == 3 and fail > 0 and '未达标' in items[2]:
                c.font = red_font
        row += 1
    row += 1

    # ---------- 二、项目清单 ----------
    ws2.merge_cells(f'A{row}:H{row}')
    ws2.cell(row, 1, '二、本月完成项目清单').font = section_font
    ws2.row_dimensions[row].height = 24; row += 1

    proj_hdrs = ['#', '项目ID', '项目名称', '类型', '集数', '参与人数', '交付日期']
    proj_widths = [5, 10, 45, 14, 8, 10, 14]
    for ci, (h, w) in enumerate(zip(proj_hdrs, proj_widths)):
        c = ws2.cell(row, ci + 1, h)
        c.font = hdr_font; c.fill = blue_fill; c.border = bdr; c.alignment = ca
        ws2.column_dimensions[_gcl(ci + 1)].width = w
    row += 1

    sorted_projs = sorted(seen_pids.items(), key=lambda x: (x[1]['date'] or datetime.date(2000,1,1)))
    for pi, (pid, info) in enumerate(sorted_projs):
        ws2.cell(row, 1, pi + 1).font = data_font
        ws2.cell(row, 1).border = bdr; ws2.cell(row, 1).alignment = ca
        ws2.cell(row, 2, pid).font = data_font
        ws2.cell(row, 2).border = bdr; ws2.cell(row, 2).alignment = ca
        ws2.cell(row, 3, info['name'][:60]).font = data_font
        ws2.cell(row, 3).border = bdr; ws2.cell(row, 3).alignment = la
        ws2.cell(row, 4, 'AI海外真人').font = data_font
        ws2.cell(row, 4).border = bdr; ws2.cell(row, 4).alignment = ca
        ws2.cell(row, 5, info['episodes']).font = data_font
        ws2.cell(row, 5).border = bdr; ws2.cell(row, 5).alignment = ca
        ws2.cell(row, 6, len(info['people'])).font = data_font
        ws2.cell(row, 6).border = bdr; ws2.cell(row, 6).alignment = ca
        d = info['date'].strftime('%Y-%m-%d') if info['date'] else ''
        ws2.cell(row, 7, d).font = data_font
        ws2.cell(row, 7).border = bdr; ws2.cell(row, 7).alignment = ca
        row += 1
    row += 1

    # ---------- 三、人员绩效明细 ----------
    ws2.merge_cells(f'A{row}:H{row}')
    ws2.cell(row, 1, '三、人员绩效明细').font = section_font
    ws2.row_dimensions[row].height = 24; row += 1

    person_hdrs = ['姓名', '角色', '总集数', '基准', '绩效', '超额/组奖', '缺集扣除', '提成合计']
    for ci, h in enumerate(person_hdrs):
        ws2.cell(row, ci + 1, h).font = hdr_font
        ws2.cell(row, ci + 1).fill = blue_fill; ws2.cell(row, ci + 1).border = bdr
        ws2.cell(row, ci + 1).alignment = ca
    row += 1

    for nm in NAME_ORDER:
        cd = commission_data.get(nm)
        if not cd: continue
        extra = cd['overtime_bonus'] + cd['group_project_bonus']
        vals = [nm, cd['role'], cd['total_episodes'],
                f'{cd["quota"]}集' if cd['quota'] > 0 else '无',
                cd['is_complete'], extra, cd['shortage_penalty'], cd['total_commission']]
        for ci, v in enumerate(vals):
            c = ws2.cell(row, ci + 1, v if (v or ci <= 1) else '')
            c.font = data_font; c.border = bdr; c.alignment = ca
            if ci == 4:
                c.font = green_font if v == '是' else (red_font if v == '否' else data_font)
            elif ci == 7 and isinstance(v, (int, float)):
                c.font = green_font if v > 0 else (red_font if v < 0 else data_font)
            elif ci == 6 and isinstance(v, (int, float)) and v > 0:
                c.font = red_font
        row += 1

    # 合计行
    row_vals = ['合计', '', total_episodes_all, '', '',
                sum(cd['overtime_bonus'] + cd['group_project_bonus'] for cd in commission_data.values()),
                sum(cd['shortage_penalty'] for cd in commission_data.values()),
                total_comm_all]
    for ci, v in enumerate(row_vals):
        c = ws2.cell(row, ci + 1, v if (v or ci <= 1) else '')
        c.font = hdr_font; c.fill = blue_fill; c.border = bdr; c.alignment = ca

    ws2.freeze_panes = 'A3'
    print(f"📊 统计简报已生成 (Sheet: 月度统计简报)")


# ===================== HTML 可视化仪表盘 =====================

def generate_html_dashboard(records, commission_data, excel_path):
    """纯CSS自包含HTML仪表盘——零外部依赖，离线可用"""
    from datetime import date as dt_date

    total_people = len(set(r['身份证姓名'] for r in records))
    total_comm_all = sum(cd['total_commission'] for cd in commission_data.values())
    total_ep_all = sum(r['单项目数/集数'] for r in records)
    done = sum(1 for cd in commission_data.values() if cd['is_complete'] == '是')
    fail = sum(1 for cd in commission_data.values() if cd['is_complete'] == '否')
    card1 = sum(1 for cd in commission_data.values() if cd['role'] == '一卡剪辑')
    card2 = sum(1 for cd in commission_data.values() if cd['role'] == '二卡剪辑')

    seen_pids = {}
    for r in records:
        pid = r['项目ID']
        if pid:
            if pid not in seen_pids:
                seen_pids[pid] = {'name': r['AI项目名称'], 'date': r['结束日期'], 'people': set(), 'episodes': 0}
            seen_pids[pid]['people'].add(r['身份证姓名'])
            seen_pids[pid]['episodes'] += r['单项目数/集数']
    total_projects = len(seen_pids)

    persons = []
    for nm in NAME_ORDER:
        cd = commission_data.get(nm)
        if not cd: continue
        persons.append({
            'name': nm, 'role': cd['role'], 'episodes': cd['total_episodes'],
            'quota': cd['quota'], 'status': cd['is_complete'],
            'extra': cd['overtime_bonus'] + cd['group_project_bonus'],
            'penalty': cd['shortage_penalty'], 'total': cd['total_commission']
        })

    # 项目清单
    proj_rows = ''
    sorted_projs = sorted(seen_pids.items(), key=lambda x: (x[1]['date'] or dt_date(2000,1,1)))
    for i, (pid, info) in enumerate(sorted_projs):
        proj_rows += f'<tr><td>{i+1}</td><td>{pid}</td><td class="tl">{info["name"][:60]}</td><td>{info["episodes"]}</td><td>{len(info["people"])}</td><td>{info["date"].strftime("%m/%d") if info["date"] else ""}</td></tr>\n'

    # CSS柱状图：集数完成
    max_ep = max((cd['total_episodes'] for cd in commission_data.values() if cd['total_episodes'] > 0), default=1)
    ep_bars = ''
    for p in persons:
        pct = min(100, int(p['episodes'] / max(max_ep, 1) * 100))
        color = '#27ae60' if (p['quota'] <= 0 or p['episodes'] >= p['quota']) else '#e74c3c'
        qpct = min(100, int(p['quota'] / max(max_ep, 1) * 100)) if p['quota'] > 0 else 0
        qmark = f'<div class="ql" style="left:{qpct}%" title="基准{p["quota"]}集"></div>' if p['quota'] > 0 else ''
        ep_bars += f'<div class="br"><span class="bl">{p["name"]}</span><div class="bt"><div class="bf" style="width:{pct}%;background:{color};"></div>{qmark}</div><span class="bv">{p["episodes"]}集</span></div>'

    # CSS柱状图：提成排行
    rank = sorted(persons, key=lambda x: x['total'], reverse=True)
    max_c = max((abs(p['total']) for p in rank), default=1) or 1
    rank_bars = ''
    for i, p in enumerate(rank):
        medal = ['🥇','🥈','🥉'][i] if i < 3 else str(i+1)
        pct = int(abs(p['total']) / max_c * 100)
        color = '#27ae60' if p['total'] > 0 else '#e74c3c'
        rank_bars += f'<div class="br"><span class="rn">{medal}</span><span class="bl">{p["name"]}</span><div class="bt"><div class="bf" style="width:{max(pct,2)}%;background:{color};"></div></div><span class="bv" style="color:{color}">{p["total"]:,}</span></div>'

    # 绩效明细表
    goals = cfg.get('monthly_goals', {}) if 'cfg' in globals() else {}
    detail_rows = ''
    for p in persons:
        cls = ' class="ld"' if p['role'] == '剪辑组长' else ''
        st = p['status']
        bd = '<span class="b ok">是</span>' if st == '是' else ('<span class="b no">否</span>' if st == '否' else '<span class="b na">N/A</span>')
        ex = f'{p["extra"]:,}' if p['extra'] else '-'
        pn = f'<span class="an">-{p["penalty"]:,}</span>' if p['penalty'] else '-'
        tc = 'ap' if p['total'] > 0 else ('an' if p['total'] < 0 else '')
        qt = f'{p["quota"]}集' if p['quota'] > 0 else '无'
        # 目标进度
        g = goals.get(p['name'], {})
        goal_html = '-'
        if g.get('episodes'):
            geps = g['episodes']
            if geps > 0:
                gpct = min(100, int(p['episodes'] / geps * 100))
                gcolor = '#27ae60' if gpct >= 100 else ('#e67e22' if gpct >= 80 else '#e74c3c')
                goal_html = f'<div style="display:flex;align-items:center;gap:4px"><div style="flex:1;height:12px;background:#eef2f7;border-radius:4px"><div style="height:100%;width:{gpct}%;background:{gcolor};border-radius:4px"></div></div><span style="font-size:11px;color:{gcolor}">{p["episodes"]}/{geps}</span></div>'
        if g.get('income') and p['total'] > 0:
            ginc = g['income']
            prev = goal_html
            if ginc > 0:
                ipct = min(100, int(p['total'] / ginc * 100))
                icolor = '#27ae60' if ipct >= 100 else ('#e67e22' if ipct >= 80 else '#e74c3c')
                goal_html = f'{prev}<div style="display:flex;align-items:center;gap:4px;margin-top:2px"><div style="flex:1;height:12px;background:#eef2f7;border-radius:4px"><div style="height:100%;width:{ipct}%;background:{icolor};border-radius:4px"></div></div><span style="font-size:11px;color:{icolor}">{p["total"]:,}/{ginc:,}</span></div>'
        detail_rows += f'<tr{cls}><td><b>{p["name"]}</b></td><td>{p["role"]}</td><td>{p["episodes"]}</td><td>{qt}</td><td>{bd}</td><td>{ex}</td><td>{pn}</td><td class="{tc}">{p["total"]:,}</td><td class="tl">{goal_html}</td></tr>\n'

    t_ex = sum(p['extra'] for p in persons)
    t_pn = sum(p['penalty'] for p in persons)

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>剪辑一组 · {TEMPLATE_DATE} · 月度统计</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI','Microsoft YaHei',sans-serif;background:#f0f4f8;color:#2c3e50}}
.hd{{background:linear-gradient(135deg,#1a5276,#2e86c1);color:#fff;padding:32px 40px;text-align:center}}
.hd h1{{font-size:28px;font-weight:600;letter-spacing:2px}}
.hd p{{font-size:14px;opacity:.85;margin-top:6px}}
.ct{{max-width:1300px;margin:0 auto;padding:24px 20px}}
.cds{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;margin-bottom:28px}}
.cd{{background:#fff;border-radius:12px;padding:20px 24px;box-shadow:0 2px 12px rgba(0,0,0,.06);text-align:center;transition:transform .2s}}
.cd:hover{{transform:translateY(-2px);box-shadow:0 4px 20px rgba(0,0,0,.1)}}
.cd .v{{font-size:36px;font-weight:700;margin:8px 0}}
.cd .l{{font-size:13px;color:#7f8c8d;text-transform:uppercase;letter-spacing:1px}}
.cd .s{{font-size:12px;color:#95a5a6}}
.cd.g .v{{color:#27ae60}}.cd.r .v{{color:#e74c3c}}.cd.b .v{{color:#2980b9}}.cd.p .v{{color:#8e44ad}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:28px}}
@media(max-width:900px){{.g2{{grid-template-columns:1fr}}}}
.pn{{background:#fff;border-radius:12px;padding:24px;box-shadow:0 2px 12px rgba(0,0,0,.06);margin-bottom:28px}}
.pn h2{{font-size:18px;color:#1a5276;margin-bottom:16px;border-bottom:2px solid #d6e4f0;padding-bottom:10px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#d6e4f0;color:#1a5276;font-weight:600;padding:10px 8px;text-align:center;border-bottom:2px solid #b0c4de}}
td{{padding:8px;text-align:center;border-bottom:1px solid #eef2f7}}
tr:hover td{{background:#f8fafc}}.tl{{text-align:left}}
.b{{display:inline-block;padding:2px 10px;border-radius:10px;font-size:12px;font-weight:600}}
.b.ok{{background:#d5f5e3;color:#27ae60}}.b.no{{background:#fadbd8;color:#e74c3c}}.b.na{{background:#eaf2f8;color:#5d6d7e}}
.ap{{color:#27ae60;font-weight:600}}.an{{color:#e74c3c;font-weight:600}}
.ld td{{background:#fef9e7}}
.ft{{text-align:center;padding:20px;color:#95a5a6;font-size:12px}}
.br{{display:flex;align-items:center;margin:6px 0;gap:10px}}
.bl{{min-width:60px;font-size:13px;font-weight:500;text-align:right}}
.rn{{min-width:28px;text-align:center;font-size:16px}}
.bt{{flex:1;height:22px;background:#eef2f7;border-radius:4px;position:relative;overflow:visible}}
.bf{{height:100%;border-radius:4px;transition:width .8s ease;min-width:2px}}
.ql{{position:absolute;top:-2px;bottom:-2px;width:2px;background:#e67e22;z-index:2}}
.ql::after{{content:'基准';position:absolute;top:-18px;left:-10px;font-size:9px;color:#e67e22;white-space:nowrap}}
.bv{{min-width:80px;font-size:13px;font-weight:600;text-align:left}}
.lg{{display:flex;gap:20px;margin-bottom:12px;font-size:12px;color:#7f8c8d}}
.lg span{{display:flex;align-items:center;gap:4px}}
.lg .d{{width:12px;height:12px;border-radius:3px;display:inline-block}}
</style></head>
<body>
<div class="hd"><h1>🎬 后期剪辑部 · {TEMPLATE_DATE} · 月度统计仪表盘</h1>
<p>剪辑一组 | 生成日期：2026-07-29 | {total_people}人 · {total_projects}个项目 · 总提成 {total_comm_all:,} 元</p></div>
<div class="ct">
<div class="cds">
<div class="cd b"><div class="l">全组总提成</div><div class="v">{total_comm_all:,}</div><div class="s">元</div></div>
<div class="cd"><div class="l">总集数</div><div class="v">{total_ep_all:,}</div><div class="s">集</div></div>
<div class="cd"><div class="l">完成项目</div><div class="v">{total_projects}</div><div class="s">部</div></div>
<div class="cd g"><div class="l">绩效达标</div><div class="v">{done}</div><div class="s">/ {total_people} 人</div></div>
<div class="cd r"><div class="l">未达标</div><div class="v">{fail}</div><div class="s">人</div></div>
<div class="cd p"><div class="l">角色分布</div><div class="v" style="font-size:20px">一卡{card1} · 二卡{card2}</div><div class="s">人</div></div>
</div>
<div class="g2">
<div class="pn"><h2>📊 集数完成情况 (vs 基准)</h2>
<div class="lg"><span><span class="d" style="background:#27ae60"></span>达标</span><span><span class="d" style="background:#e74c3c"></span>未达标</span><span><span class="d" style="background:#e67e22"></span>基准线</span></div>
{ep_bars}</div>
<div class="pn"><h2>🏆 提成排行</h2>{rank_bars}</div>
</div>
<div class="pn"><h2>👥 人员绩效明细</h2><div style="overflow-x:auto"><table>
<thead><tr><th>姓名</th><th>角色</th><th>总集数</th><th>基准</th><th>绩效</th><th>超额/组奖</th><th>缺集扣</th><th>提成合计</th><th>目标进度</th></tr></thead>
<tbody>{detail_rows}
<tr style="font-weight:bold;background:#d6e4f0"><td>合计</td><td></td><td>{total_ep_all}</td><td></td><td></td><td>{t_ex:,}</td><td>{t_pn:,}</td><td>{total_comm_all:,}</td></tr></tbody></table></div></div>
<div class="pn"><h2>📋 本月项目清单（{total_projects}部）</h2><div style="overflow-x:auto"><table>
<thead><tr><th>#</th><th>项目ID</th><th>项目名称</th><th>集数</th><th>参与人数</th><th>交付日期</th></tr></thead>
<tbody>{proj_rows}</tbody></table></div></div>
</div>
<div class="ft">AI后期剪辑提成表生成工具 v4.0 · 自动生成于 2026-07-29 · 纯静态离线可用</div>
</body></html>'''

    html_path = excel_path.replace('.xlsx', '_仪表盘.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"📊 HTML仪表盘已生成: {html_path}")
    return html_path

# ===================== 打印汇总 =====================

def print_summary(records, commission_data):
    print()
    print("=" * 70)
    print("  剪辑一组 - 完整绩效及提成汇总")
    print("=" * 70)
    print(f"{'姓名':　<6s} {'角色':　<6s} {'集数':>4s} {'基准':>4s} {'绩效':>4s} {'超额/组奖':>8s} {'缺集扣':>8s} {'提成合计':>8s}")
    print("-" * 70)

    for nm in NAME_ORDER:
        cd = commission_data.get(nm, {})
        total = cd.get('total_episodes', 0)
        role = cd.get('role', '')
        quota = cd.get('quota', 0)
        is_ok = cd.get('is_complete', '')
        overtime = cd.get('overtime_bonus', 0)
        group_b = cd.get('group_project_bonus', 0)
        shortage = cd.get('shortage_penalty', 0)
        total_c = cd.get('total_commission', 0)
        extra = overtime + group_b if (overtime + group_b) > 0 else 0
        extra_str = str(extra) if extra else ''
        short_str = str(shortage) if shortage else ''
        quota_str = str(quota) if quota > 0 else '无'
        print(f"  {nm:　<4s}  {role:　<6s}  {total:>4d}  {quota_str:>4s}  {is_ok:>4s}  {extra_str:>8s}  {short_str:>8s}  {total_c:>8d}")

    total_all = sum(cd.get('total_commission', 0) for cd in commission_data.values())
    print("-" * 70)
    print(f"  {'总提成合计':　>20s}  {total_all:>8d}  元")
    print("=" * 70)


# ===================== 主程序 =====================

def main():
    print()
    print("=" * 70)
    print("  AI后期剪辑提成表生成工具 v4.0")
    print("=" * 70)

    # 检查文件
    missing = []
    for f, desc in [(PROJECT_FILE, '一组AI项目.xlsx'),
                    (TEMPLATE_FILE, '模板文件'),
                    (CONFIG_FILE, 'config.json')]:
        if not os.path.exists(f):
            missing.append(f"{desc} ({os.path.basename(f)})")
    if missing:
        print(f"\n❌ 找不到: {', '.join(missing)}")
        print(f"   请确保这些文件与 generate_commission.py 在同一目录。")
        if len(sys.argv) < 2:
            input("\n按任意键退出...")
        return

    print(f"\n📖 数据源: {os.path.basename(PROJECT_FILE)}")
    print(f"📋 模板:   {os.path.basename(TEMPLATE_FILE)}")
    print(f"⚙️  配置:   {os.path.basename(CONFIG_FILE)}")
    print(f"   角色: 一卡{cfg['rules']['一卡剪辑']['基准集数']}集 | 二卡/助理{cfg['rules']['二卡剪辑']['基准集数']}集 | 组长无限制")

    # 读取和解析
    try:
        df = pd.read_excel(PROJECT_FILE, header=None)
    except Exception as e:
        print(f"❌ 无法读取项目数据文件: {e}")
        if len(sys.argv) < 2:
            input("\n按任意键退出...")
        return
    print("🔍 解析数据中...")

    # 用数据文件的真实月份覆盖模板月份（输出文件名/报表标题/仪表盘均以数据月份为准）
    global OUTPUT_MONTH, TEMPLATE_DATE, TEMPLATE_YEAR, OUTPUT_FILE
    data_cn, data_month_num = get_month_from_data(df)
    if data_cn and data_month_num:
        # 年份优先从模板取，否则取当前年份
        template_year_match = re.match(r'(\d{4})年', TEMPLATE_DATE)
        cur_year = template_year_match.group(1) if template_year_match else str(datetime.date.today().year)
        OUTPUT_MONTH = data_cn
        TEMPLATE_DATE = f'{cur_year}年{data_month_num:02d}月'
        TEMPLATE_YEAR = int(cur_year)
        OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'AI后期剪辑提成一组{OUTPUT_MONTH}.xlsx')
        print(f"📅 数据文件月份: {TEMPLATE_DATE}（已按数据月份覆盖模板）")
    else:
        print(f"📅 未在数据文件中识别到月份，使用模板月份: {TEMPLATE_DATE}")
    try:
        overtime_map = load_overtime_map(OVERTIME_FILE)
    except (OSError, ValueError, json.JSONDecodeError) as e:
        print(f"❌ 无法读取超时标记: {e}")
        return
    if overtime_map:
        total_ot = sum(len(v) for v in overtime_map.values())
        print(f'⏱️ 已加载本次超时集数: {total_ot}集 ({len(overtime_map)}个项目)')

    records, group_pids = parse_projects(
        df,
        default_year=TEMPLATE_YEAR,
        overtime_map=overtime_map,
    )
    print(f"   共 {len(records)} 条记录")
    if not records:
        print("⚠️  未解析到任何有效记录，请检查项目数据文件的格式。")
        if len(sys.argv) < 2:
            input("\n按任意键退出...")
        return

    # 统计全局项目数
    global_pids = set()
    for pids in group_pids.values():
        global_pids |= pids
    print(f"   全组去重项目数: {len(global_pids)} 部")

    # 自查
    self_check(records)

    # 计算提成
    print("\n🧮 计算绩效和提成...")
    commission_data = compute_commission(records, group_pids)

    # 汇总
    print_summary(records, commission_data)

    # 生成
    final_path, html_path = generate_excel(records, commission_data, TEMPLATE_FILE, OUTPUT_FILE)

    # 输出文件路径供 GUI 读取
    print(f"OUTPUT_EXCEL={final_path}")
    print(f"OUTPUT_HTML={html_path}")
    print(f"\n🎉 完成！")
    print()
    # 如果是从 GUI 调用的（传了命令行参数），不要阻塞等待输入
    if len(sys.argv) < 2:
        input("按任意键退出...")


if __name__ == '__main__':
    main()
