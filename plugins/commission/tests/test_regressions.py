import json
import tempfile
import unittest
import sys, os
from datetime import date
from pathlib import Path
from unittest.mock import patch

# 确保能找到 src/ 下的模块
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'src'))

import pandas as pd
from openpyxl import Workbook, load_workbook

import config_loader
import generate_commission as gc


class RegressionTests(unittest.TestCase):
    def test_parser_does_not_reuse_previous_project_state(self):
        frame = pd.DataFrame([
            ["1111-甲", None, None, "7.1下午18点交"],
            [None, None, "任显翔：1-2", None],
            ["无效项目标题", None, None, None],
            [None, None, "任显翔：3", None],
            ["2222-乙", None, None, "7.2下午18点交"],
            [None, None, "任显翔：4", None],
        ])
        records, _ = gc.parse_projects(frame, default_year=2025)
        editor_records = [r for r in records if r["身份证姓名"] == "任显翔"]
        self.assertEqual([(r["项目ID"], r["AI项目名称"]) for r in editor_records],
                         [("1111", "甲"), ("2222", "乙")])
        self.assertEqual(editor_records[0]["结束日期"], date(2025, 7, 1))

    def test_leader_gets_all_projects_without_episode_commission(self):
        frame = pd.DataFrame([
            ["3333-组内项目", None, None, "7.3下午18点交"],
            [None, None, "任显翔：1-2", None],
        ])
        records, groups = gc.parse_projects(frame, default_year=2025)
        leader = next(
            name for name in gc.NAME_ORDER
            if gc.normalize_role(gc.ROLE_MAP[name]) == "剪辑组长"
        )
        leader_record = next(r for r in records if r["身份证姓名"] == leader)
        self.assertFalse(leader_record["参与剪辑"])
        self.assertEqual(leader_record["项目ID"], "3333")
        self.assertEqual(leader_record["单项目数/集数"], 0)

        with tempfile.TemporaryDirectory() as tmp:
            template = Path(tmp) / "template.xlsx"
            output = Path(tmp) / "output.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.cell(1, 2, "2025年07月提成表")
            for col in range(1, 21):
                sheet.cell(3, col, col)
            workbook.save(template)

            commission = gc.compute_commission(records, groups)
            with patch.object(gc, "generate_html_dashboard", return_value=""), \
                 patch.object(gc.os, "startfile", create=True):
                gc.generate_excel(records, commission, str(template), str(output))

            sheet = load_workbook(output).active
            leader_row = next(
                row for row in range(4, sheet.max_row + 1)
                if sheet.cell(row, 2).value == leader
            )
            self.assertIsNone(sheet.cell(leader_row, 11).value)
            self.assertIsNone(sheet.cell(leader_row, 12).value)

    def test_config_rejects_unknown_role_and_completes_order(self):
        raw = json.loads(Path("config.json").read_text(encoding="utf-8"))
        raw["人员角色"] = {"甲": "一卡剪辑", "乙": "未知角色"}
        raw["人员排序"] = ["甲"]
        raw["小组"] = {"测试组": {"组长": "甲", "成员": ["甲", "乙"]}}
        with self.assertRaises(ValueError):
            config_loader._parse_config_dict(raw)

        raw["人员角色"]["乙"] = "二卡剪辑"
        config = config_loader._parse_config_dict(raw)
        self.assertEqual(config.person_order, ["甲", "乙"])

    def test_single_record_writes_summary_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            template = tmp_path / "template.xlsx"
            output = tmp_path / "output.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.cell(1, 2, "2025年07月提成表")
            for col in range(1, 21):
                sheet.cell(3, col, col)
            workbook.save(template)

            record = {
                "身份证姓名": "任显翔", "角色": "一卡剪辑", "项目ID": "1111",
                "项目类型": "AI海外真人", "AI项目名称": "甲",
                "开始日期": date(2025, 7, 1), "结束日期": date(2025, 7, 2),
                "完成明细": "1-80", "单项目数/集数": 80,
                "原始集数": 80, "超时集数": 0, "超时明细": [],
            }
            commission = gc.compute_commission([record], {})
            with patch.object(gc, "generate_html_dashboard", return_value=""), \
                 patch.object(gc.os, "startfile", create=True):
                gc.generate_excel([record], commission, str(template), str(output))

            sheet = load_workbook(output).active
            self.assertEqual(sheet["M4"].value, 80)
            # R列现在是算式文本（一卡：超额(80-70)×20=200）
            self.assertEqual(sheet["R4"].value, "(80-70)×20=200")

    def test_overtime_is_explicit_input(self):
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", encoding="utf-8", delete=False) as f:
            json.dump({"1111": [1, 2]}, f)
            path = f.name
        try:
            self.assertEqual(gc.load_overtime_map(path), {"1111": {1, 2}})
        finally:
            Path(path).unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
