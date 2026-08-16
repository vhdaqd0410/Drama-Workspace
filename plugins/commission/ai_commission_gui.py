# -*- coding: utf-8 -*-
"""AI后期剪辑提成表生成工具 - GUI v6.0"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import subprocess, threading, os, sys, json, re, tempfile
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_DIR = os.path.join(SCRIPT_DIR, 'src')
sys.path.insert(0, SRC_DIR)  # 让 Python 找到 src/ 下的模块
CLI_SCRIPT = os.path.join(SRC_DIR, 'generate_commission.py')
CONFIG_PATH = os.path.join(SCRIPT_DIR, 'config.json')
BACKUP_DIR = os.path.join(SCRIPT_DIR, 'backup')
CARDS_DIR = os.path.join(SCRIPT_DIR, '个人绩效卡片')
HISTORY_DIR = os.path.join(SCRIPT_DIR, 'history')

# 导入功能模块
try:
    from features import (generate_next_month_template, export_to_pdf,
                          generate_person_cards, compare_months, backup_output,
                          data_preview, generate_ranking_html,
                          generate_project_management_html, smart_episode_assignment,
                          generate_person_trend_html, validate_project_data,
                          list_backups, cleanup_backups,
                          advanced_filter,
                          generate_project_template,
                          create_config_snapshot, list_config_snapshots,
                          restore_config_snapshot, start_web_server,
                          validate_episode_assignments)
    HAS_FEATURES = True
except ImportError as e:
    HAS_FEATURES = False
    print(f'features import error: {e}')

# Python 解释器路径——启动时探测一次，之后缓存
_CACHE_FILE = os.path.join(SCRIPT_DIR, '.python_cache')
def _find_python_with_openpyxl():
    # 读缓存
    if os.path.exists(_CACHE_FILE):
        try:
            with open(_CACHE_FILE, 'r') as f:
                cached = f.read().strip()
            if os.path.exists(cached):
                return cached
        except: pass
    # 探测
    candidates = []
    if os.path.exists(r'C:\Users\Admin\AppData\Local\Programs\Python\Python313\python.exe'):
        candidates.append(r'C:\Users\Admin\AppData\Local\Programs\Python\Python313\python.exe')
    candidates.append(sys.executable)
    import shutil
    for cmd in ['python', 'python3']:
        p = shutil.which(cmd)
        if p and p not in candidates:
            candidates.append(p)
    for exe in candidates:
        try:
            r = subprocess.run([exe, '-c', 'import openpyxl'], capture_output=True, timeout=5)
            if r.returncode == 0:
                with open(_CACHE_FILE, 'w') as f: f.write(exe)
                return exe
        except: continue
    return sys.executable

PYTHON_EXE = _find_python_with_openpyxl()

ROLES = ['一卡剪辑', '二卡剪辑', '剪辑助理', '剪辑组长', '小组长']

# ============ 配色方案 v7.0 - 清晰工作台 ============
C = {
    # 全局背景
    'bg':            '#f4f7f8',
    'card':          '#ffffff',
    'card_hover':    '#f6fbfa',
    'card_border':   '#dce6e5',
    # 主色调
    'accent':        '#007c70',
    'accent_h':      '#00675e',
    'accent_a':      '#00554e',
    'accent_l':      '#e3f4f1',
    # 功能色
    'green':         '#0ea16b',
    'green_l':       '#e6f7f0',
    'blue':          '#2b7fff',
    'blue_l':        '#eaf2ff',
    'purple':        '#6d5bd0',
    'purple_l':      '#f0effb',
    'orange':        '#f0641a',
    'orange_l':      '#fff3eb',
    'red':           '#e53e3e',
    'red_l':         '#fef0f0',
    'teal':          '#0e9388',
    'teal_l':        '#e6f7f5',
    'pink':          '#db2777',
    'amber':         '#d97706',
    'cyan':          '#0e8fa6',
    'indigo':        '#536acb',
    'gray':          '#6b7280',
    'slate':         '#475569',
    # 文字
    'text':          '#111827',
    'text2':         '#4b5563',
    'text3':         '#9ca3af',
    'placeholder':   '#d1d5db',
    # 边框
    'border':        '#e5e7eb',
    'border_l':      '#f3f4f6',
    # 头部 (深色)
    'hdr_bg':        '#102827',
    'hdr_text':      '#f2f8f7',
    'hdr_sub':       '#8ed9d0',
    # 日志终端
    'log_bg':        '#0d1117',
    'log_fg':        '#c9d1d9',
    # 状态栏
    'status_bg':     '#f9fafb',
    'status_fg':     '#9ca3af',
}

# 暗色主题配色
C_DARK = {
    'bg':            '#1e1e2e',
    'card':          '#2a2a3c',
    'card_hover':    '#313245',
    'card_border':   '#45475a',
    'accent':        '#89b4fa',
    'accent_h':      '#74a5f0',
    'accent_a':      '#5f8fd5',
    'accent_l':      '#313244',
    'green':         '#a6e3a1',
    'green_l':       '#2a3a2e',
    'blue':          '#89b4fa',
    'blue_l':        '#3a3a52',
    'purple':        '#cba6f7',
    'purple_l':      '#3a3050',
    'orange':        '#fab387',
    'orange_l':      '#4a3a2a',
    'red':           '#f38ba8',
    'red_l':         '#4a2a33',
    'teal':          '#94e2d5',
    'teal_l':        '#2a3a3a',
    'pink':          '#f5c2e7',
    'amber':         '#f9e2af',
    'cyan':          '#89dceb',
    'indigo':        '#b4befe',
    'gray':          '#a6adc8',
    'slate':         '#7f849c',
    'text':          '#cdd6f4',
    'text2':         '#a6adc8',
    'text3':         '#7f849c',
    'placeholder':   '#585b70',
    'border':        '#45475a',
    'border_l':      '#313244',
    'hdr_bg':        '#11111b',
    'hdr_text':      '#cdd6f4',
    'hdr_sub':       '#89b4fa',
    'log_bg':        '#0d1117',
    'log_fg':        '#a6adc8',
    'status_bg':     '#1e1e2e',
    'status_fg':     '#a6adc8',
}

# 角色配色 (前景色, 背景色)
ROLE_COLORS = {
    '一卡剪辑':  ('#0ea16b', '#e6f7f0'),
    '二卡剪辑':  ('#2b7fff', '#eaf2ff'),
    '剪辑助理':  ('#7c3aed', '#f3f0ff'),
    '剪辑组长':  ('#f0641a', '#fff3eb'),
    '小组长':    ('#d97706', '#fef6ec'),
}

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("提成表生成 · AI后期剪辑组")
        self.root.geometry("1180x800")
        self.root.minsize(960, 680)
        self.root.configure(bg=C['bg'])

        self._setup_ttk_style()

        try: self.root.iconbitmap(os.path.join(SCRIPT_DIR, 'icon.ico'))
        except: pass

        self.cfg = self._load_config()
        self.app_settings = self.cfg.setdefault('app_settings', {})

        # 默认路径（若配置里有上次路径则优先恢复）
        default_project = os.path.join(SCRIPT_DIR, 'data', '一组AI项目-8月.xlsx')
        if not os.path.exists(default_project):
            default_project = os.path.join(SCRIPT_DIR, '一组AI项目.xlsx')
        default_template = os.path.join(SCRIPT_DIR, 'data', 'AI后期剪辑提成一组模板.xlsx')
        if not os.path.exists(default_template):
            default_template = os.path.join(SCRIPT_DIR, 'AI后期剪辑提成一组模板.xlsx')
        default_output = os.path.join(SCRIPT_DIR, 'output')
        if not os.path.exists(default_output):
            os.makedirs(default_output, exist_ok=True)

        self.project_file = self.app_settings.get('project_file') or default_project
        self.template_file = self.app_settings.get('template_file') or default_template
        self.output_dir = self.app_settings.get('output_dir') or default_output

        # 如果上次路径失效，安全回退到默认路径
        if not os.path.exists(self.project_file):
            self.project_file = default_project
        if not os.path.exists(self.template_file):
            self.template_file = default_template
        if not os.path.exists(self.output_dir):
            self.output_dir = default_output
        self.auto_backup = tk.BooleanVar(value=self.app_settings.get('auto_backup', True))
        self._current_overtime_map = {}
        self._generation_year = None
        self._watcher = False
        self._web_server_instance = None
        self._apply_theme()
        self.build_ui()

        self.root.bind('<Control-g>', lambda e: self.run())
        self.root.bind('<Control-o>', lambda e: self._open(self.output_dir))
        self.root.bind('<Control-r>', lambda e: self.open_role_editor())
        self.root.bind('<Control-G>', lambda e: self.run())
        self.root.bind('<Control-O>', lambda e: self._open(self.output_dir))
        self.root.bind('<Control-R>', lambda e: self.open_role_editor())
        self._bind_hotkeys()
        self.root.protocol('WM_DELETE_WINDOW', self._on_close)
        self.check_files()
        # 启动 2 秒后静默检查更新
        if self.cfg.get('update_check', True):
            try:
                self.root.after(2000, lambda: self._check_update(silent=True))
            except Exception:
                pass
        # 启动时检查自动备份是否到期
        try:
            self._check_auto_backup()
        except Exception:
            pass
        # 启动定时生成检查器
        try:
            self._start_schedule_checker()
        except Exception:
            pass

    def _bind_hotkeys(self):
        """全局快捷键体系（扩展常用操作）"""
        hotkeys = {
            '<Control-p>': lambda e: self._preview_data(),
            '<Control-P>': lambda e: self._preview_data(),
            '<Control-d>': lambda e: self._smart_assign(),
            '<Control-D>': lambda e: self._smart_assign(),
            '<Control-s>': lambda e: self._data_correction(),
            '<Control-S>': lambda e: self._data_correction(),
            '<Control-b>': lambda e: self._manage_backups(),
            '<Control-B>': lambda e: self._manage_backups(),
            '<Control-m>': lambda e: self._compare_months(),
            '<Control-M>': lambda e: self._compare_months(),
            '<F5>': lambda e: self._refresh_role_tags(),
            '<Control-f>': lambda e: self._focus_search(),
            '<Control-F>': lambda e: self._focus_search(),
        }
        for seq, fn in hotkeys.items():
            try:
                self.root.bind(seq, fn)
            except Exception:
                pass

    def _focus_search(self):
        """聚焦主界面搜索框（若存在）"""
        if hasattr(self, 'search_entry') and self.search_entry.winfo_exists():
            self.search_entry.focus_set()
            self.search_entry.selection_range(0, 'end')

    def _apply_theme(self):
        """应用主题（light/dark）。启动时根据 config 的 theme 字段更新全局 C 字典。"""
        self.theme = self.cfg.get('theme', 'light')
        if self.theme == 'dark':
            C.clear()
            C.update(C_DARK)

    def _set_theme(self, theme):
        """切换主题并保存配置（重启后完全生效）"""
        self.cfg['theme'] = theme
        try:
            self._save_config()
        except Exception:
            pass
        messagebox.showinfo(
            '主题设置',
            f'已切换为{"暗色" if theme == "dark" else "亮色"}主题。\n重启软件后完全生效。',
            parent=self.root)

    def _do_search(self):
        """全局搜索：读取项目数据，按 项目ID/名称/人员 搜索并弹出结果窗口"""
        kw = self.search_entry.get().strip()
        if not kw:
            messagebox.showinfo('搜索', '请输入搜索关键词', parent=self.root)
            return
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()
            df = pd.read_excel(self.project_file, header=None)
            records, _ = gc.parse_projects(df)
            _sys.path.pop(0)
        except Exception as e:
            messagebox.showerror('搜索失败', f'无法读取数据：{e}', parent=self.root)
            return
        kw_l = kw.lower()
        hits = []
        for r in records:
            pid = str(r.get('项目ID', '') or '')
            name = str(r.get('AI项目名称', '') or '')
            person = str(r.get('身份证姓名', '') or '')
            if kw_l in pid.lower() or kw_l in name.lower() or kw_l in person.lower():
                hits.append(r)
        if not hits:
            messagebox.showinfo('搜索', f'未找到与"{kw}"匹配的记录', parent=self.root)
            return
        dlg = tk.Toplevel(self.root)
        dlg.title(f'🔍 搜索结果（{len(hits)} 条）')
        dlg.geometry('760x420')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text=f'关键词：{kw} · 共 {len(hits)} 条匹配',
                 font=('Microsoft YaHei', 10, 'bold'), bg=C['bg'],
                 fg=C['text']).pack(padx=14, pady=(10, 4), anchor='w')
        tree = ttk.Treeview(dlg, columns=('person', 'role', 'pid', 'name', 'detail'),
                            show='headings', height=12)
        for cid, txt, w in [('person', '人员', 90), ('role', '角色', 80),
                            ('pid', '项目ID', 90), ('name', '项目名称', 260),
                            ('detail', '完成明细', 160)]:
            tree.heading(cid, text=txt)
            tree.column(cid, width=w, anchor='w')
        for r in hits:
            tree.insert('', 'end', values=(
                r.get('身份证姓名', ''), r.get('角色', ''),
                r.get('项目ID', ''), r.get('AI项目名称', '')[:30],
                r.get('完成明细', '')[:40]))
        tree.pack(fill='both', expand=True, padx=14, pady=(0, 10))
        self._btn(dlg, '关闭', C['gray'], dlg.destroy, font_size=9, padx=16, pady=4).pack(pady=(0, 10))

    def _gen_charts(self):
        """统计图表可视化：人员集数柱状图 + 提成柱状图 + 项目类型分布"""
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()
            df = pd.read_excel(self.project_file, header=None)
            records, group_pids = gc.parse_projects(df)
            cd = gc.compute_commission(records, group_pids)
            _sys.path.pop(0)
        except Exception as e:
            messagebox.showerror('图表失败', f'无法读取数据：{e}', parent=self.root)
            return
        try:
            import matplotlib
            matplotlib.use('TkAgg')
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        except Exception as e:
            messagebox.showerror('图表失败', f'缺少 matplotlib：{e}', parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('📈 统计图表')
        dlg.geometry('900x700')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root)

        fig = Figure(figsize=(9, 11), dpi=90, facecolor=C['card'])
        # 图1：人员集数柱状图
        ax1 = fig.add_subplot(3, 1, 1)
        names = gc.NAME_ORDER
        eps = [cd.get(n, {}).get('total_episodes', 0) for n in names]
        colors = [ROLE_COLORS.get(cd.get(n, {}).get('role', ''), ('#888', '#fff'))[0] for n in names]
        ax1.bar(names, eps, color=colors)
        ax1.set_title(f'人员集数（{gc.TEMPLATE_DATE}）', fontsize=11)
        ax1.set_ylabel('集数')
        ax1.tick_params(axis='x', rotation=45, labelsize=7)
        # 图2：人员提成柱状图
        ax2 = fig.add_subplot(3, 1, 2)
        comm = [cd.get(n, {}).get('total_commission', 0) for n in names]
        ax2.bar(names, comm, color=colors)
        ax2.set_title('人员提成', fontsize=11)
        ax2.set_ylabel('元')
        ax2.tick_params(axis='x', rotation=45, labelsize=7)
        # 图3：项目类型分布饼图
        ax3 = fig.add_subplot(3, 1, 3)
        type_counts = {}
        for r in records:
            t = r.get('项目类型', '未知')
            type_counts[t] = type_counts.get(t, 0) + 1
        if type_counts:
            ax3.pie(list(type_counts.values()), labels=list(type_counts.keys()),
                    autopct='%1.1f%%')
            ax3.set_title('项目类型分布', fontsize=11)
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=dlg)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True, padx=8, pady=8)
        self._btn(dlg, '关闭', C['gray'], dlg.destroy, font_size=9, padx=16, pady=4).pack(pady=(0, 8))

    def _compare_groups(self):
        """多组横向对比：选择两个组/团队的提成表，对比集数与提成差异"""
        if not self._require_features(): return
        f1 = filedialog.askopenfilename(
            title='选择第一组提成表',
            filetypes=[('Excel文件', '*.xlsx')],
            initialdir=SCRIPT_DIR)
        if not f1: return
        f2 = filedialog.askopenfilename(
            title='选择第二组提成表',
            filetypes=[('Excel文件', '*.xlsx')],
            initialdir=os.path.dirname(f1))
        if not f2: return
        self._log(f'👥 正在对比两个组的提成数据...')
        try:
            label1, label2, diffs = compare_months(f1, f2)
            if not diffs:
                self._log(f'✅ 两组数据完全一致，无差异。')
                messagebox.showinfo('对比结果', f'{label1} ↔ {label2}\n\n数据完全一致，无差异。')
                return
            self._log(f'{label1} ↔ {label2} 对比: {len(diffs)}人有差异')
            msg = f'{label1}  →  {label2}\n{"="*60}\n'
            msg += f'{"姓名":　<6s} {"组1集数":>6s} {"组2集数":>6s} {"集数差":>6s} {"组1提成":>8s} {"组2提成":>8s} {"提成差":>8s}\n'
            for nm, e1, e2, de, c1, c2, dc in diffs:
                de_str = f'+{de}' if de > 0 else str(de)
                dc_str = f'+{dc}' if dc > 0 else str(dc)
                msg += f'{nm:　<6s} {e1:>6d} {e2:>6d} {de_str:>6s} {c1:>8d} {c2:>8d} {dc_str:>8s}\n'
            self._log(msg)
            dlg = tk.Toplevel(self.root)
            dlg.title('多组对比')
            dlg.geometry('720x520')
            dlg.configure(bg=C['bg'])
            tk.Label(dlg, text=f'{label1}  ↔  {label2}', font=('Microsoft YaHei', 14, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(pady=10)
            txt = tk.Text(dlg, font=('Microsoft YaHei', 10), bg=C['log_bg'], fg=C['log_fg'],
                          relief='flat', padx=12, pady=10)
            txt.insert('1.0', msg)
            txt.configure(state='disabled')
            txt.pack(fill='both', expand=True, padx=20, pady=0)
        except Exception as e:
            self._log(f'❌ 对比失败: {e}')
            messagebox.showerror('失败', f'对比失败:\n{e}')

    def _gen_trend_report(self):
        """#1 跨月趋势：基于历史数据生成跨月收入/集数趋势 HTML"""
        if not self._require_features(): return
        try:
            from features import generate_trend_report
            html_path = generate_trend_report(HISTORY_DIR, self.output_dir)
            if not html_path:
                messagebox.showinfo('跨月趋势', 'history/ 目录暂无历史数据。\n请先点击"一键生成"至少一次以归档数据。', parent=self.root)
                return
            self._log(f'📈 已生成跨月趋势: {os.path.basename(html_path)}')
            self._open(html_path)
        except Exception as e:
            self._log(f'❌ 跨月趋势失败: {e}')
            messagebox.showerror('失败', f'生成跨月趋势失败:\n{e}', parent=self.root)

    def _gen_annual_summary(self):
        """#3 年度汇总：基于历史数据批量生成年度对比 Excel"""
        if not self._require_features(): return
        try:
            from features import generate_annual_summary
            excel_path = generate_annual_summary(HISTORY_DIR, self.output_dir)
            if not excel_path:
                messagebox.showinfo('年度汇总', 'history/ 目录暂无历史数据。\n请先点击"一键生成"至少一次以归档数据。', parent=self.root)
                return
            self._log(f'📅 已生成年度汇总: {os.path.basename(excel_path)}')
            self._open(excel_path)
        except Exception as e:
            self._log(f'❌ 年度汇总失败: {e}')
            messagebox.showerror('失败', f'生成年度汇总失败:\n{e}', parent=self.root)

    def _interactive_ranking(self):
        """#2 交互式绩效排名面板：Treeview 排序/筛选/TopN"""
        if not self._require_features(): return
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()
            df = pd.read_excel(self.project_file, header=None)
            records, group_pids = gc.parse_projects(df)
            cd = gc.compute_commission(records, group_pids)
            _sys.path.pop(0)
        except Exception as e:
            messagebox.showerror('排名失败', f'无法读取数据：{e}', parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('🏅 交互式绩效排名')
        dlg.geometry('820x560')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text='🏅 交互式绩效排名（点击列头排序 · 可筛选）',
                 font=('Microsoft YaHei', 14, 'bold'), bg=C['bg'],
                 fg=C['text']).pack(pady=(12, 4))

        # 筛选栏
        fb = tk.Frame(dlg, bg=C['bg']); fb.pack(fill='x', padx=14, pady=(0, 6))
        tk.Label(fb, text='角色:', font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack(side='left')
        role_var = tk.StringVar(value='全部')
        roles_all = ['全部'] + sorted(set(v for v in gc.ROLE_MAP.values()))
        ttk.Combobox(fb, textvariable=role_var, values=roles_all, state='readonly',
                     font=('Microsoft YaHei', 9), width=10).pack(side='left', padx=(4, 12))
        tk.Label(fb, text='Top N:', font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack(side='left')
        top_var = tk.StringVar(value='全部')
        ttk.Combobox(fb, textvariable=top_var, values=['全部', '5', '10', '15', '20'],
                     state='readonly', font=('Microsoft YaHei', 9), width=6).pack(side='left', padx=(4, 12))
        self._btn(fb, '应用筛选', C['accent'], lambda: _refresh(), font_size=9, padx=12, pady=3).pack(side='left')

        # 排名表
        cols = ('rank', 'name', 'role', 'eps', 'comm', 'status')
        tree = ttk.Treeview(dlg, columns=cols, show='headings')
        for cid, txt, w in [('rank', '排名', 60), ('name', '姓名', 90), ('role', '角色', 90),
                            ('eps', '集数', 80), ('comm', '提成', 100), ('status', '绩效', 60)]:
            tree.heading(cid, text=txt, command=lambda c=cid: _sort_by(c))
            tree.column(cid, width=w, anchor='center')
        tree.pack(fill='both', expand=True, padx=14, pady=(0, 8))

        sort_col = 'comm'
        sort_desc = True

        def _get_rows():
            rows = []
            for nm, c in cd.items():
                if role_var.get() != '全部' and c.get('role') != role_var.get():
                    continue
                rows.append((nm, c.get('role', ''), c.get('total_episodes', 0),
                             c.get('total_commission', 0), c.get('is_complete', '')))
            return rows

        def _sort_by(col):
            nonlocal sort_col, sort_desc
            if sort_col == col:
                sort_desc = not sort_desc
            else:
                sort_col = col
                sort_desc = True
            _refresh()

        def _refresh():
            for i in tree.get_children():
                tree.delete(i)
            rows = _get_rows()
            key_idx = {'rank': None, 'name': 0, 'role': 1, 'eps': 2, 'comm': 3, 'status': 4}[sort_col]
            if key_idx is not None:
                rows.sort(key=lambda r: r[key_idx], reverse=sort_desc)
            top = top_var.get()
            if top != '全部':
                rows = rows[:int(top)]
            for i, r in enumerate(rows, 1):
                tree.insert('', 'end', values=(i, r[0], r[1], r[2], r[3], r[4]))

        _refresh()
        self._btn(dlg, '关闭', C['gray'], dlg.destroy, font_size=9, padx=16, pady=4).pack(pady=(0, 10))

    def _risk_warning(self):
        """#4 异常/风险预警：扫描集数骤变/未达标/目标滞后等异常"""
        if not self._require_features(): return
        try:
            import pandas as pd
            from features import scan_anomalies
            gc, _sys = self._load_gc_module()
            df = pd.read_excel(self.project_file, header=None)
            records, group_pids = gc.parse_projects(df)
            cd = gc.compute_commission(records, group_pids)
            _sys.path.pop(0)
            goals = self.cfg.get('monthly_goals', {})
            warnings = scan_anomalies(records, cd, HISTORY_DIR, goals)
        except Exception as e:
            messagebox.showerror('预警失败', f'扫描异常失败：{e}', parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('⚠️ 风险预警')
        dlg.geometry('760x480')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text=f'⚠️ 风险预警（发现 {len(warnings)} 项异常）',
                 font=('Microsoft YaHei', 14, 'bold'), bg=C['bg'],
                 fg=C['text']).pack(pady=(12, 4))
        if not warnings:
            tk.Label(dlg, text='✅ 未发现异常，一切正常。', font=('Microsoft YaHei', 12),
                     bg=C['bg'], fg=C['green']).pack(pady=30)
            self._btn(dlg, '关闭', C['gray'], dlg.destroy, font_size=9, padx=16, pady=4).pack(pady=(0, 10))
            return
        tree = ttk.Treeview(dlg, columns=('level', 'type', 'person', 'msg'), show='headings')
        for cid, txt, w in [('level', '等级', 50), ('type', '类型', 90), ('person', '人员', 90), ('msg', '说明', 380)]:
            tree.heading(cid, text=txt)
            tree.column(cid, width=w, anchor='w')
        color_map = {'高': '#e74c3c', '中': '#e67e22', '低': '#f1c40f'}
        for w in warnings:
            iid = tree.insert('', 'end', values=(w.get('level', ''), w.get('type', ''),
                                                 w.get('person', ''), w.get('msg', '')))
            tree.tag_configure(iid, foreground=color_map.get(w.get('level', ''), '#2c3e50'))
        tree.pack(fill='both', expand=True, padx=14, pady=(0, 8))
        self._btn(dlg, '关闭', C['gray'], dlg.destroy, font_size=9, padx=16, pady=4).pack(pady=(0, 10))

    def _manage_environments(self):
        """#8 配置多环境：多套配置切换（不同团队各自环境），互不覆盖"""
        from tkinter import simpledialog
        try:
            from config_loader import list_environments, load_environment, save_environment
            from models import AppConfig
        except Exception as e:
            messagebox.showerror('环境切换', f'加载环境模块失败：{e}', parent=self.root)
            return
        envs = list_environments(CONFIG_PATH)

        dlg = tk.Toplevel(self.root)
        dlg.title('🖥 配置环境切换')
        dlg.geometry('460x380')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text='🖥 配置环境切换', font=('Microsoft YaHei', 14, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=(12, 4))
        tk.Label(dlg, text='不同团队/小组可各自保存一套配置，切换后互不覆盖。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack()

        listbox = tk.Listbox(dlg, font=('Microsoft YaHei', 11), height=8)
        listbox.pack(fill='both', expand=True, padx=16, pady=10)
        for e in envs:
            listbox.insert('end', e)
        if envs:
            listbox.selection_set(0)

        def _switch():
            sel = listbox.curselection()
            if not sel: return
            name = envs[sel[0]]
            try:
                new_cfg = load_environment(CONFIG_PATH, name).to_dict()
                self.cfg.clear()
                self.cfg.update(new_cfg)
                # 保存当前环境名
                self.cfg.setdefault('app_settings', {})['active_environment'] = name
                try: self._save_config()
                except Exception: pass
                self._log(f'🖥 已切换到环境: {name}')
                dlg.destroy()
                messagebox.showinfo('环境切换', f'已切换到环境「{name}」。\n请重启软件后完全生效。', parent=self.root)
            except Exception as e:
                messagebox.showerror('切换失败', str(e), parent=self.root)

        def _save_current():
            name = simpledialog.askstring('保存环境', '输入环境名称（如：一组、二组）：', parent=dlg)
            if not name: return
            name = name.strip()
            if not name: return
            try:
                save_environment(CONFIG_PATH, name, AppConfig.from_dict(self.cfg))
                self.cfg.setdefault('app_settings', {})['active_environment'] = name
                try: self._save_config()
                except Exception: pass
                self._log(f'🖥 已保存当前配置为环境: {name}')
                messagebox.showinfo('保存环境', f'当前配置已保存为环境「{name}」。', parent=self.root)
                dlg.destroy()
            except Exception as e:
                messagebox.showerror('保存失败', str(e), parent=self.root)

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=16, pady=(0, 12))
        self._btn(bf, '✅ 切换', C['green'], _switch, font_size=9, padx=14, pady=4).pack(side='left', padx=3)
        self._btn(bf, '💾 保存当前', C['blue'], _save_current, font_size=9, padx=14, pady=4).pack(side='left', padx=3)
        self._btn(bf, '❌ 关闭', C['gray'], dlg.destroy, font_size=9, padx=14, pady=4).pack(side='left', padx=3)

    def _backup_policy(self):
        """#9 数据备份自动策略：设置定时备份配置和历史数据"""
        from tkinter import simpledialog
        policy = self.cfg.setdefault('backup_policy', {'enabled': False, 'interval_days': 7, 'keep': 15})

        dlg = tk.Toplevel(self.root)
        dlg.title('🗄 备份策略')
        dlg.geometry('460x360')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text='🗄 数据备份自动策略', font=('Microsoft YaHei', 14, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=(12, 4))
        tk.Label(dlg, text='定时自动备份 config.json 和历史数据，防止丢失。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack()

        # 启用开关
        enable_var = tk.BooleanVar(value=bool(policy.get('enabled')))
        tk.Checkbutton(dlg, text='启用自动备份', variable=enable_var,
                       font=('Microsoft YaHei', 11), bg=C['bg'], fg=C['text'],
                       selectcolor=C['bg'], activebackground=C['bg']).pack(pady=(14, 4))

        # 间隔天数
        f1 = tk.Frame(dlg, bg=C['bg']); f1.pack(fill='x', padx=30, pady=4)
        tk.Label(f1, text='备份间隔（天）:', font=('Microsoft YaHei', 10), bg=C['bg'],
                 fg=C['text2']).pack(side='left')
        interval_var = tk.StringVar(value=str(policy.get('interval_days', 7)))
        tk.Spinbox(f1, from_=1, to=90, textvariable=interval_var, font=('Microsoft YaHei', 10),
                   width=6, relief='solid', borderwidth=1).pack(side='left', padx=6)

        # 保留份数
        f2 = tk.Frame(dlg, bg=C['bg']); f2.pack(fill='x', padx=30, pady=4)
        tk.Label(f2, text='保留份数:', font=('Microsoft YaHei', 10), bg=C['bg'],
                 fg=C['text2']).pack(side='left')
        keep_var = tk.StringVar(value=str(policy.get('keep', 15)))
        tk.Spinbox(f2, from_=1, to=99, textvariable=keep_var, font=('Microsoft YaHei', 10),
                   width=6, relief='solid', borderwidth=1).pack(side='left', padx=6)

        # 目标目录
        f3 = tk.Frame(dlg, bg=C['bg']); f3.pack(fill='x', padx=30, pady=4)
        tk.Label(f3, text='备份目录:', font=('Microsoft YaHei', 10), bg=C['bg'],
                 fg=C['text2']).pack(side='left')
        target_var = tk.StringVar(value=policy.get('target_dir', os.path.join(SCRIPT_DIR, 'backup')))
        tk.Entry(f3, textvariable=target_var, font=('Microsoft YaHei', 9),
                 relief='solid', borderwidth=1).pack(side='left', fill='x', expand=True, padx=6)
        def _pick_dir():
            d = filedialog.askdirectory(title='选择备份目录', initialdir=target_var.get())
            if d: target_var.set(d)
        self._btn(f3, '浏览', C['blue'], _pick_dir, font_size=8, padx=8, pady=2).pack(side='left')

        def _save():
            policy['enabled'] = enable_var.get()
            try: policy['interval_days'] = int(interval_var.get())
            except ValueError: pass
            try: policy['keep'] = int(keep_var.get())
            except ValueError: pass
            policy['target_dir'] = target_var.get().strip() or os.path.join(SCRIPT_DIR, 'backup')
            self.cfg['backup_policy'] = policy
            try: self._save_config()
            except Exception: pass
            self._log(f'🗄 备份策略已保存: {"启用" if policy["enabled"] else "停用"} 间隔{policy["interval_days"]}天 保留{policy["keep"]}份')
            dlg.destroy()
            messagebox.showinfo('备份策略', f'备份策略已保存。\n启用后每次启动若到期将自动备份。', parent=self.root)

        # 立即备份按钮
        def _backup_now():
            try:
                from features import auto_backup_config
                target = target_var.get().strip() or os.path.join(SCRIPT_DIR, 'backup')
                created = auto_backup_config(CONFIG_PATH, HISTORY_DIR, target, int(keep_var.get() or 15))
                self._log(f'🗄 已立即备份 {len(created)} 个文件到 {target}')
                messagebox.showinfo('备份完成', f'已备份 {len(created)} 个文件到：\n{target}', parent=self.root)
            except Exception as e:
                messagebox.showerror('备份失败', str(e), parent=self.root)

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=20, pady=(16, 12))
        self._btn(bf, '💾 保存策略', C['green'], _save, font_size=10, padx=14, pady=4).pack(side='left', padx=4)
        self._btn(bf, '⏱ 立即备份', C['blue'], _backup_now, font_size=9, padx=12, pady=4).pack(side='left', padx=4)
        self._btn(bf, '❌ 关闭', C['gray'], dlg.destroy, font_size=9, padx=12, pady=4).pack(side='left', padx=4)

    def _check_auto_backup(self):
        """启动时检查自动备份策略是否到期，若到期则自动备份"""
        policy = self.cfg.get('backup_policy', {})
        if not policy.get('enabled'):
            return
        try:
            from features import auto_backup_config, backup_due
            target = policy.get('target_dir') or os.path.join(SCRIPT_DIR, 'backup')
            last_str = policy.get('last_backup_time', '')
            last = None
            if last_str:
                try:
                    from datetime import datetime as _dt
                    last = _dt.strptime(last_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    last = None
            if backup_due(policy, last):
                created = auto_backup_config(CONFIG_PATH, HISTORY_DIR, target,
                                             int(policy.get('keep', 15)))
                if created:
                    from datetime import datetime as _dt
                    policy['last_backup_time'] = _dt.now().strftime('%Y-%m-%d %H:%M:%S')
                    self.cfg['backup_policy'] = policy
                    try: self._save_config()
                    except Exception: pass
                    self._log(f'🗄 自动备份完成: {len(created)} 个文件 -> {target}')
        except Exception as e:
            self._log(f'⚠️ 自动备份检查失败: {e}')

    def _schedule_setup(self):
        """#10 报表定时生成：设置到点自动生成"""
        schedule = self.cfg.setdefault('schedule', {'enabled': False, 'time': '18:00'})

        dlg = tk.Toplevel(self.root)
        dlg.title('⏰ 定时生成')
        dlg.geometry('420x300')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text='⏰ 报表定时生成', font=('Microsoft YaHei', 14, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=(12, 4))
        tk.Label(dlg, text='软件运行期间，到设定时间自动生成提成表。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack()

        enable_var = tk.BooleanVar(value=bool(schedule.get('enabled')))
        tk.Checkbutton(dlg, text='启用定时生成', variable=enable_var,
                       font=('Microsoft YaHei', 11), bg=C['bg'], fg=C['text'],
                       selectcolor=C['bg'], activebackground=C['bg']).pack(pady=(14, 4))

        f1 = tk.Frame(dlg, bg=C['bg']); f1.pack(fill='x', padx=30, pady=4)
        tk.Label(f1, text='生成时间（HH:MM）:', font=('Microsoft YaHei', 10), bg=C['bg'],
                 fg=C['text2']).pack(side='left')
        time_var = tk.StringVar(value=schedule.get('time', '18:00'))
        tk.Entry(f1, textvariable=time_var, font=('Microsoft YaHei', 11), width=8,
                 relief='solid', borderwidth=1).pack(side='left', padx=6)

        def _save():
            schedule['enabled'] = enable_var.get()
            schedule['time'] = time_var.get().strip() or '18:00'
            self.cfg['schedule'] = schedule
            try: self._save_config()
            except Exception: pass
            self._log(f'⏰ 定时生成已保存: {"启用" if schedule["enabled"] else "停用"} 时间{schedule["time"]}')
            dlg.destroy()
            messagebox.showinfo('定时生成', f'定时生成设置已保存。\n{"到点将自动生成。" if schedule["enabled"] else "已停用。"}', parent=self.root)

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=20, pady=(16, 12))
        self._btn(bf, '💾 保存', C['green'], _save, font_size=10, padx=14, pady=4).pack(side='left', padx=4)
        self._btn(bf, '❌ 关闭', C['gray'], dlg.destroy, font_size=9, padx=12, pady=4).pack(side='left', padx=4)

    def _data_encryption(self):
        """K: 数据加密 —— 为配置文件设置密码保护"""
        from tkinter import simpledialog

        dlg = tk.Toplevel(self.root)
        dlg.title('🔐 数据加密')
        dlg.geometry('440x280')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text='🔐 配置文件加密', font=('Microsoft YaHei', 14, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=(12, 4))
        tk.Label(dlg, text='为 config.json 设置密码，防止他人查看真实人员数据。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack()

        # 检测是否已加密
        is_enc = False
        try:
            import json as _j
            with open(CONFIG_PATH, 'r', encoding='utf-8') as _f:
                _raw = _j.load(_f)
            is_enc = isinstance(_raw, dict) and _raw.get('__encrypted__')
        except Exception:
            pass
        tk.Label(dlg, text='当前状态：' + ('🔒 已加密' if is_enc else '🔓 未加密'),
                 font=('Microsoft YaHei', 10, 'bold'), bg=C['bg'],
                 fg=C['green'] if not is_enc else C['amber']).pack(pady=(12, 6))

        # 密码输入
        f1 = tk.Frame(dlg, bg=C['bg']); f1.pack(fill='x', padx=30, pady=4)
        tk.Label(f1, text='密码:', font=('Microsoft YaHei', 10), bg=C['bg'],
                 fg=C['text2']).pack(side='left')
        pwd_var = tk.StringVar()
        tk.Entry(f1, textvariable=pwd_var, font=('Microsoft YaHei', 11), width=22,
                 show='●', relief='solid', borderwidth=1).pack(side='left', padx=6)

        def _encrypt():
            pwd = pwd_var.get()
            if not pwd:
                messagebox.showwarning('提示', '请输入密码', parent=dlg); return
            try:
                from config_loader import save_config_encrypted
                from models import AppConfig
                # 当前 cfg 是明文（启动时已加载）
                save_config_encrypted(AppConfig.from_dict(self.cfg), CONFIG_PATH, pwd)
                self._log('🔐 配置已加密保存')
                messagebox.showinfo('加密成功', '配置已加密。\n重启后需输入密码才能加载。', parent=dlg)
                dlg.destroy()
            except Exception as e:
                messagebox.showerror('加密失败', str(e), parent=dlg)

        def _decrypt():
            pwd = pwd_var.get()
            try:
                from config_loader import load_config_encrypted
                cfg = load_config_encrypted(CONFIG_PATH, pwd)
                # 解密后保存为明文
                from config_loader import save_config
                save_config(cfg, CONFIG_PATH)
                self._log('🔓 配置已解密（明文保存）')
                messagebox.showinfo('解密成功', '配置已解密为明文。', parent=dlg)
                dlg.destroy()
            except Exception as e:
                messagebox.showerror('解密失败', str(e), parent=dlg)

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=20, pady=(16, 12))
        self._btn(bf, '🔒 加密', C['red'], _encrypt, font_size=10, padx=14, pady=4).pack(side='left', padx=4)
        self._btn(bf, '🔓 解密', C['blue'], _decrypt, font_size=10, padx=14, pady=4).pack(side='left', padx=4)
        self._btn(bf, '❌ 关闭', C['gray'], dlg.destroy, font_size=9, padx=12, pady=4).pack(side='left', padx=4)

    def _config_migrate(self):
        """L: 配置迁移向导 —— 从旧版 config 迁移到当前结构"""
        try:
            import json as _j
            from config_loader import migrate_config
            with open(CONFIG_PATH, 'r', encoding='utf-8') as _f:
                raw = _j.load(_f)
        except Exception as e:
            messagebox.showerror('迁移失败', f'无法读取配置：{e}', parent=self.root)
            return

        migrated, changed = migrate_config(raw)
        if not changed:
            messagebox.showinfo('配置迁移', '✅ 当前配置已是新结构，无需迁移。', parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('🔄 配置迁移')
        dlg.geometry('480x320')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text='🔄 配置迁移向导', font=('Microsoft YaHei', 14, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=(12, 4))
        tk.Label(dlg, text='检测到旧版配置字段，迁移后将兼容新结构。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack()
        tk.Label(dlg, text='\n'.join(f'• {k} → 新字段' for k in changed),
                 font=('Microsoft YaHei', 10), bg=C['bg'], fg=C['amber'],
                 justify='left').pack(pady=12)

        def _do_migrate():
            try:
                # 备份旧配置
                import shutil as _sh
                bak = CONFIG_PATH + '.bak_premigrate'
                if not os.path.exists(bak):
                    _sh.copy2(CONFIG_PATH, bak)
                # 保存迁移结果
                from config_loader import save_config
                from models import AppConfig
                save_config(AppConfig.from_dict(migrated), CONFIG_PATH)
                self._log(f'🔄 配置已迁移: {", ".join(changed)} -> 新字段')
                messagebox.showinfo('迁移完成',
                    f'✅ 配置已迁移到新结构！\n\n旧配置已备份为:\n{os.path.basename(bak)}',
                    parent=self.root)
                dlg.destroy()
            except Exception as e:
                messagebox.showerror('迁移失败', str(e), parent=self.root)

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=20, pady=(16, 12))
        self._btn(bf, '✅ 执行迁移', C['green'], _do_migrate, font_size=10, padx=14, pady=4).pack(side='left', padx=4)
        self._btn(bf, '❌ 取消', C['gray'], dlg.destroy, font_size=9, padx=12, pady=4).pack(side='left', padx=4)

    def _start_schedule_checker(self):
        """启动定时生成检查器：每秒检查一次是否到点"""
        schedule = self.cfg.get('schedule', {})
        if not schedule.get('enabled'):
            return
        target = schedule.get('time', '18:00')
        fired_today = [False]  # 同一天不重复触发

        def _tick():
            if not self.cfg.get('schedule', {}).get('enabled'):
                return
            try:
                now = datetime.now()
                cur_time = now.strftime('%H:%M')
                if cur_time == target and not fired_today[0]:
                    fired_today[0] = True
                    self._log(f'⏰ 定时触发：到点 {target}，开始自动生成...')
                    self._do_generate(self.project_file, self.output_dir, self.template_file, {})
                elif cur_time != target:
                    fired_today[0] = False  # 过了这个时间点，明天可再次触发
            except Exception as e:
                self._log(f'⚠️ 定时检查失败: {e}')
            try:
                self.root.after(30000, _tick)  # 每 30 秒检查一次
            except Exception:
                pass
        self.root.after(30000, _tick)

    def _export_multi(self):
        """多格式导出：导出人员提成/项目明细为 CSV、JSON 和增强版"""
        try:
            from features import export_to_csv, export_to_json, export_enhanced
            gc, _sys = self._load_gc_module()
            df = pd.read_excel(self.project_file, header=None)
            records, group_pids = gc.parse_projects(df)
            cd = gc.compute_commission(records, group_pids)
            _sys.path.pop(0)
        except Exception as e:
            messagebox.showerror('导出失败', f'无法读取数据：{e}', parent=self.root)
            return
        out_dir = filedialog.askdirectory(title='选择导出目录', initialdir=self.output_dir)
        if not out_dir:
            return
        # 选择导出选项
        opt = tk.Toplevel(self.root)
        opt.title('导出选项')
        opt.geometry('400x240')
        opt.configure(bg=C['bg'])
        opt.transient(self.root); opt.grab_set()
        tk.Label(opt, text='💾 选择导出方式', font=('Microsoft YaHei', 13, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=(16, 6))
        split_var = tk.BooleanVar(value=False)
        tk.Checkbutton(opt, text='按角色拆分 CSV（每个角色一个文件）', variable=split_var,
                       font=('Microsoft YaHei', 10), bg=C['bg'], fg=C['text'],
                       selectcolor=C['bg'], activebackground=C['bg']).pack(pady=4)
        def _do_export():
            try:
                created_csv = export_to_csv(records, cd, out_dir)
                created_json = export_to_json(records, cd, out_dir)
                created_enh = export_enhanced(records, cd, out_dir, split_by_role=split_var.get())
                total = len(created_csv) + len(created_enh)
                opt.destroy()
                self._log(f'💾 已导出 {total} 个文件到 {out_dir}')
                messagebox.showinfo('导出完成',
                    f'已导出到：\n{out_dir}\n\n'
                    f'基础 CSV: {len(created_csv)} 个\n'
                    f'增强 CSV/Excel: {len(created_enh)} 个\n'
                    f'JSON: 1 个', parent=self.root)
                self._open(out_dir)
            except Exception as e:
                self._log(f'❌ 导出失败: {e}')
                messagebox.showerror('导出失败', str(e), parent=self.root)
        bf = tk.Frame(opt, bg=C['bg']); bf.pack(fill='x', padx=20, pady=(16, 12))
        self._btn(bf, '✅ 导出', C['green'], _do_export, font_size=10, padx=18, pady=5).pack(side='left', padx=6)
        self._btn(bf, '❌ 取消', C['gray'], opt.destroy, font_size=9, padx=14, pady=4).pack(side='left', padx=6)

    def _import_data(self):
        """数据导入向导：选择 CSV/Excel，预览并追加到项目数据"""
        path = filedialog.askopenfilename(
            title='选择要导入的数据文件',
            filetypes=[('数据文件', '*.xlsx *.csv'), ('Excel', '*.xlsx'), ('CSV', '*.csv')],
            initialdir=SCRIPT_DIR)
        if not path:
            return
        try:
            import pandas as pd
            from features import clean_import_data
            if path.lower().endswith('.csv'):
                df_in = pd.read_csv(path, encoding='utf-8-sig')
            else:
                df_in = pd.read_excel(path, header=None)
            df_clean, clean_report = clean_import_data(df_in)
        except Exception as e:
            messagebox.showerror('导入失败', f'无法读取文件：{e}', parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('📥 数据导入预览')
        dlg.geometry('760x520')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text=f'文件：{os.path.basename(path)} · 共 {len(df_in)} 行',
                 font=('Microsoft YaHei', 10, 'bold'), bg=C['bg'],
                 fg=C['text']).pack(padx=14, pady=(10, 4), anchor='w')
        # 清洗报告
        tk.Label(dlg, text='🧹 清洗结果：' + '；'.join(clean_report),
                 font=('Microsoft YaHei', 8), bg=C['bg'],
                 fg=C['green'] if '移除重复' not in '；'.join(clean_report) else C['amber'],
                 wraplength=700, justify='left').pack(padx=14, anchor='w')
        tk.Label(dlg, text='预览前 10 行（确认格式后导入）：',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack(padx=14, anchor='w')
        preview = tk.Text(dlg, font=('Microsoft YaHei', 9), bg=C['log_bg'], fg=C['log_fg'],
                          relief='flat', padx=10, pady=8, height=12)
        preview.pack(fill='both', expand=True, padx=14, pady=(4, 8))
        lines = []
        for i in range(min(10, len(df_clean))):
            row = [str(x) for x in df_clean.iloc[i].tolist() if pd.notna(x)]
            lines.append(' | '.join(row))
        preview.insert('1.0', '\n'.join(lines) if lines else '（空文件）')
        preview.configure(state='disabled')

        def _do_import():
            try:
                from openpyxl import load_workbook
                wb = load_workbook(self.project_file)
                ws = wb.active
                last = ws.max_row
                for r in range(last, 0, -1):
                    if ws.cell(r, 1).value or ws.cell(r, 3).value:
                        last = r; break
                next_row = last + 2
                added = 0
                for i in range(len(df_clean)):
                    vals = [df_clean.iloc[i, j] for j in range(min(10, df_clean.shape[1]))]
                    if all(pd.isna(v) for v in vals):
                        continue
                    for j, v in enumerate(vals):
                        if pd.notna(v):
                            ws.cell(next_row, j + 1, v)
                    next_row += 1
                    added += 1
                wb.save(self.project_file)
                wb.close()
                self._log(f'📥 已导入 {added} 行到项目数据')
                dlg.destroy()
                messagebox.showinfo('导入完成', f'成功导入 {added} 行数据。\n请用「数据预览」或重新生成核对。', parent=self.root)
            except Exception as e:
                messagebox.showerror('导入失败', f'写入失败：{e}', parent=self.root)

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=14, pady=(0, 10))
        self._btn(bf, '✅ 确认导入', C['green'], _do_import, font_size=9, padx=16, pady=4).pack(side='left', padx=3)
        self._btn(bf, '❌ 取消', C['gray'], dlg.destroy, font_size=9, padx=16, pady=4).pack(side='left', padx=3)

    def _data_entry(self):
        """#2 GUI数据录入：图形化录入项目数据，无需手动填 Excel"""
        from datetime import datetime as _dt, timedelta as _td
        names = sorted(self.cfg.get('人员角色', {}).keys())

        dlg = tk.Toplevel(self.root)
        dlg.title('📝 数据录入')
        dlg.geometry('860x620')
        dlg.minsize(700, 500)
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        # ---- 标题 ----
        tk.Label(dlg, text='📝 图形化数据录入', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=(12, 2))
        tk.Label(dlg, text='填写项目信息与人员分配，确认后自动写入项目数据（无需编辑Excel）。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack(pady=(0, 8))

        # ---- 项目信息卡 ----
        c1 = tk.Frame(dlg, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        c1.pack(fill='x', padx=14, pady=(0, 8))
        tk.Label(c1, text='📋 项目信息', font=('Microsoft YaHei', 10, 'bold'),
                 bg=C['card'], fg=C['text']).pack(anchor='w', padx=12, pady=(8, 4))
        g1 = tk.Frame(c1, bg=C['card']); g1.pack(fill='x', padx=12)
        tk.Label(g1, text='项目名称:', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).grid(row=0, column=0, sticky='w', pady=2)
        proj_name_var = tk.StringVar()
        tk.Entry(g1, textvariable=proj_name_var, font=('Microsoft YaHei', 10), width=32,
                 relief='solid', borderwidth=1).grid(row=0, column=1, padx=6, pady=2)
        tk.Label(g1, text='项目ID:', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).grid(row=0, column=2, sticky='w', padx=(12, 0), pady=2)
        proj_id_var = tk.StringVar()
        tk.Entry(g1, textvariable=proj_id_var, font=('Microsoft YaHei', 10), width=12,
                 relief='solid', borderwidth=1).grid(row=0, column=3, padx=6, pady=2)
        tk.Label(g1, text='交付日期:', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).grid(row=1, column=0, sticky='w', pady=2)
        date_var = tk.StringVar(value='8.1下午18点交')
        tk.Entry(g1, textvariable=date_var, font=('Microsoft YaHei', 10), width=16,
                 relief='solid', borderwidth=1).grid(row=1, column=1, sticky='w', padx=6, pady=2)

        # ---- 人员分配卡 ----
        c2 = tk.Frame(dlg, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        c2.pack(fill='both', expand=True, padx=14, pady=(0, 8))
        hdr2 = tk.Frame(c2, bg=C['card']); hdr2.pack(fill='x', padx=12, pady=(8, 4))
        tk.Label(hdr2, text='👥 人员分配', font=('Microsoft YaHei', 10, 'bold'),
                 bg=C['card'], fg=C['text']).pack(side='left')
        self._btn(hdr2, '➕ 添加分配行', C['blue'], lambda: _add_assign_row(), font_size=8, padx=10, pady=2).pack(side='right')

        # 选人模版行（复用分集工具的 personnel_templates）
        tpl_row = tk.Frame(c2, bg=C['card']); tpl_row.pack(fill='x', padx=12, pady=(0, 4))
        tk.Label(tpl_row, text='📋 选人模版:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text2']).pack(side='left', padx=(0, 4))
        p_templates = self.cfg.get('personnel_templates', {})
        tpl_var = tk.StringVar()
        tpl_names = list(p_templates.keys())
        tpl_combo = ttk.Combobox(tpl_row, textvariable=tpl_var, values=['（暂无模版）'] + tpl_names,
                                 state='readonly', font=('Microsoft YaHei', 9), width=14)
        tpl_combo.pack(side='left', padx=(0, 6))
        if tpl_names:
            tpl_combo.current(0)

        def _load_tpl():
            name = tpl_var.get()
            if not name or name.startswith('（'):
                messagebox.showwarning('提示', '请先在智能分集中保存选人模版', parent=dlg); return
            if name not in p_templates:
                messagebox.showwarning('提示', f'模版"{name}"不存在', parent=dlg); return
            people = p_templates[name]
            if not people:
                messagebox.showwarning('提示', '该模版为空', parent=dlg); return
            # 清除现有分配行（assign_rows 为 6 元组：name, f1, t1, f2, t2, frame）
            for _, _, _, _, _, f in list(assign_rows):
                f.destroy()
            assign_rows.clear()
            # 按模版创建分配行（集数留空待填）
            for nm in people:
                _add_assign_row(nm=nm)
            self._log(f'📋 已按模版"{name}"创建 {len(people)} 行分配')
        self._btn(tpl_row, '📥 加载', C['green'], _load_tpl, font_size=8, padx=10, pady=2).pack(side='left')

        # 分配行容器（可滚动）
        ac = tk.Frame(c2, bg=C['card'])
        ac.pack(fill='both', expand=True, padx=12, pady=(0, 8))
        acanvas = tk.Canvas(ac, bg=C['card'], highlightthickness=0, height=180)
        asb = tk.Scrollbar(ac, orient='vertical', command=acanvas.yview)
        ainner = tk.Frame(acanvas, bg=C['card'])
        ainner.bind('<Configure>', lambda e: acanvas.configure(scrollregion=acanvas.bbox('all')))
        awin = acanvas.create_window((0, 0), window=ainner, anchor='nw')
        acanvas.configure(yscrollcommand=asb.set)
        acanvas.pack(side='left', fill='both', expand=True)
        asb.pack(side='right', fill='y')
        def _aresize(e): acanvas.itemconfig(awin, width=e.width)
        acanvas.bind('<Configure>', _aresize)

        assign_rows = []  # [(name_var, from_var, to_var, from2_var, to2_var, frame)]

        def _add_assign_row(nm='', frm='', to='', frm2='', to2=''):
            row = tk.Frame(ainner, bg=C['card'])
            row.pack(fill='x', pady=2)
            nv = tk.StringVar(value=nm)
            cb = ttk.Combobox(row, textvariable=nv, values=names, state='readonly',
                              font=('Microsoft YaHei', 9), width=10)
            cb.pack(side='left', padx=(0, 6))
            # 第一段
            tk.Label(row, text='从', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).pack(side='left')
            fv = tk.StringVar(value=frm)
            tk.Entry(row, textvariable=fv, font=('Microsoft YaHei', 9), width=5,
                     relief='solid', borderwidth=1).pack(side='left', padx=3)
            tk.Label(row, text='到', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).pack(side='left')
            tv = tk.StringVar(value=to)
            tk.Entry(row, textvariable=tv, font=('Microsoft YaHei', 9), width=5,
                     relief='solid', borderwidth=1).pack(side='left', padx=3)
            # 第二段（可不填）
            tk.Label(row, text='｜', font=('Microsoft YaHei', 9, 'bold'), bg=C['card'],
                     fg=C['text3']).pack(side='left', padx=(6, 0))
            tk.Label(row, text='第二段', font=('Microsoft YaHei', 7), bg=C['card'],
                     fg=C['text3']).pack(side='left', padx=(4, 2))
            fv2 = tk.StringVar(value=frm2)
            tk.Entry(row, textvariable=fv2, font=('Microsoft YaHei', 9), width=5,
                     relief='solid', borderwidth=1).pack(side='left', padx=3)
            tv2 = tk.StringVar(value=to2)
            tk.Entry(row, textvariable=tv2, font=('Microsoft YaHei', 9), width=5,
                     relief='solid', borderwidth=1).pack(side='left', padx=3)
            tk.Label(row, text='集', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).pack(side='left', padx=(0, 8))
            def _del_row(r=row):
                r.destroy()
                assign_rows[:] = [x for x in assign_rows if x[5] is not r]
            tk.Button(row, text='✕', font=('Microsoft YaHei', 8, 'bold'), bg=C['red'], fg='white',
                      relief='flat', cursor='hand2', padx=6, pady=0, command=_del_row).pack(side='left')
            assign_rows.append((nv, fv, tv, fv2, tv2, row))

        # 默认加 3 行
        _add_assign_row()
        _add_assign_row()
        _add_assign_row()

        # ---- 写入逻辑 ----
        def _save():
            proj_name = proj_name_var.get().strip()
            proj_id = proj_id_var.get().strip()
            date_str = date_var.get().strip()
            if not proj_name or not proj_id:
                messagebox.showwarning('提示', '请填写项目名称和项目ID', parent=dlg); return
            # 收集分配行（支持两段不连续区间）
            assigns = []
            for nv, fv, tv, fv2, tv2, _ in assign_rows:
                nm = nv.get().strip()
                if not nm: continue
                try:
                    f = int(fv.get()); t = int(tv.get())
                except (ValueError, TypeError):
                    messagebox.showwarning('提示', f'"{nm}" 的第一段集数范围格式不正确', parent=dlg); return
                if f < 1 or t < f:
                    messagebox.showwarning('提示', f'"{nm}" 的第一段集数范围无效（{f}-{t}）', parent=dlg); return
                # 第二段（可选）
                seg2 = None
                f2s, t2s = fv2.get().strip(), tv2.get().strip()
                if f2s or t2s:
                    try:
                        f2 = int(f2s); t2 = int(t2s)
                    except (ValueError, TypeError):
                        messagebox.showwarning('提示', f'"{nm}" 的第二段集数范围格式不正确', parent=dlg); return
                    if f2 < 1 or t2 < f2:
                        messagebox.showwarning('提示', f'"{nm}" 的第二段集数范围无效（{f2}-{t2}）', parent=dlg); return
                    seg2 = (f2, t2)
                assigns.append((nm, f, t, seg2))
            if not assigns:
                messagebox.showwarning('提示', '请至少添加一条人员分配', parent=dlg); return

            # #7 前置校验：人员重复 / 区间重叠 / 单行两段重叠
            seen_person = set()
            for nm, f, t, seg2 in assigns:
                if nm in seen_person:
                    messagebox.showwarning('校验提示', f'人员"{nm}"出现了多次，请合并为一行', parent=dlg); return
                seen_person.add(nm)
                if seg2:
                    f2, t2 = seg2
                    if not (t < f2 or t2 < f):
                        messagebox.showwarning('校验提示',
                            f'"{nm}" 的两段区间重叠（{f}-{t} 与 {f2}-{t2}）', parent=dlg); return
            # 不同人员之间的区间重叠
            for i in range(len(assigns)):
                nm1, f1, t1, seg2_1 = assigns[i]
                segs1 = [(f1, t1)] + ([seg2_1] if seg2_1 else [])
                for j in range(i + 1, len(assigns)):
                    nm2, f2, t2, seg2_2 = assigns[j]
                    segs2 = [(f2, t2)] + ([seg2_2] if seg2_2 else [])
                    for (a1, b1) in segs1:
                        for (a2, b2) in segs2:
                            if not (b1 < a2 or b2 < a1):
                                messagebox.showwarning('校验提示',
                                    f'集数区间重叠："{nm1}"（{a1}-{b1}）与 "{nm2}"（{a2}-{b2}）', parent=dlg); return
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, Alignment
                wb = load_workbook(self.project_file)
                ws = wb.active
                # 找到末尾
                last = ws.max_row
                for r in range(last, 0, -1):
                    if ws.cell(r, 1).value or ws.cell(r, 3).value:
                        last = r; break
                next_row = last + 2  # 空一行
                # 项目标题行
                dir_label = f'O:\\AI漫剧剪辑一组\\{proj_name}'
                font_title = Font(name='宋体', size=14, bold=True)
                font_normal = Font(name='宋体', size=14)
                align_center = Alignment(horizontal='center', vertical='center')
                ws.cell(next_row, 1, proj_name).font = font_title
                ws.cell(next_row, 2, dir_label).font = font_title
                ws.cell(next_row, 4, date_str).font = font_title
                ws.cell(next_row, 5, '已分集').font = font_title
                # 人员分配行（支持两段：1-30，45-60）
                for item in assigns:
                    nm, f, t, seg2 = item
                    if f != t:
                        eps_text = f'{f}-{t}'
                    else:
                        eps_text = str(f)
                    if seg2:
                        f2, t2 = seg2
                        eps_text += '，' + (f'{f2}-{t2}' if f2 != t2 else str(f2))
                    ws.cell(next_row, 3, f'{nm}：{eps_text}').font = font_normal
                    next_row += 1
                wb.save(self.project_file)
                wb.close()
                self._log(f'📝 已录入项目"{proj_name}"（{len(assigns)}人分配）')
                dlg.destroy()
                messagebox.showinfo('录入完成',
                    f'✅ 项目"{proj_name}"已写入项目数据。\n'
                    f'共 {len(assigns)} 条人员分配。\n\n'
                    f'可用「数据预览」或"一键生成"核对。', parent=self.root)
            except Exception as e:
                messagebox.showerror('录入失败', f'写入失败：{e}', parent=self.root)

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=14, pady=(0, 12))
        self._btn(bf, '✅ 写入项目数据', C['green'], _save, font_size=11, padx=18, pady=5).pack(side='left', padx=3)
        self._btn(bf, '❌ 取消', C['gray'], dlg.destroy, font_size=9, padx=14, pady=4).pack(side='left', padx=3)

    def _all_tools(self):
        """返回所有可收藏功能列表：[(名称, 图标, 方法)]"""
        return [
            ('数据预览', '📋', self._preview_data),
            ('数据校验', '✅', self._validate_data),
            ('智能分集', '📐', self._smart_assign),
            ('数据录入', '📝', self._data_entry),
            ('批量处理', '📦', self._batch_process),
            ('交互排名', '🏅', self._interactive_ranking),
            ('统计图表', '📈', self._gen_charts),
            ('跨月趋势', '📈', self._gen_trend_report),
            ('年度汇总', '📅', self._gen_annual_summary),
            ('多格式导出', '💾', self._export_multi),
            ('数据导入', '📥', self._import_data),
            ('月度目标', '🎯', self._monthly_goals),
            ('备份策略', '🗄', self._backup_policy),
            ('定时生成', '⏰', self._schedule_setup),
            ('数据加密', '🔐', self._data_encryption),
            ('配置迁移', '🔄', self._config_migrate),
            ('风险预警', '⚠️', self._risk_warning),
            ('环境切换', '🖥', self._manage_environments),
            ('月份对比', '📊', self._compare_months),
            ('组内排名', '🏆', self._gen_ranking),
            ('项目管理', '🗂', self._gen_project_mgmt),
            ('绩效卡片', '🃏', self._gen_cards),
            ('数据修正', '✏', self._data_correction),
            ('Web服务', '🌐', self._web_server),
            ('高级筛选', '🔎', self._advanced_filter),
            ('下月模板', '📅', self._gen_next_template),
        ]

    def _refresh_favorites_bar(self):
        """刷新顶部收藏快捷按钮栏"""
        for w in self.fav_bar.winfo_children():
            w.destroy()
        favs = self.cfg.get('favorites', [])
        if not favs:
            tk.Label(self.fav_bar, text='⭐ 点击右上角收藏常用功能，快速直达',
                     font=('Microsoft YaHei', 8), bg=C['bg'], fg=C['text3']).pack(side='left', padx=4)
            return
        tk.Label(self.fav_bar, text='⭐ 快捷:', font=('Microsoft YaHei', 8, 'bold'),
                 bg=C['bg'], fg=C['text2']).pack(side='left', padx=(0, 4))
        tools_map = {name: (icon, fn) for name, icon, fn in self._all_tools()}
        for name in favs:
            if name not in tools_map:
                continue
            icon, fn = tools_map[name]
            b = self._btn(self.fav_bar, f'{icon} {name}', C['accent'], fn, font_size=8, padx=10, pady=3)
            b.pack(side='left', padx=2)

    def _manage_favorites(self):
        """⭐ 常用操作收藏管理：勾选收藏，保存到 config"""
        dlg = tk.Toplevel(self.root)
        dlg.title('⭐ 常用操作收藏')
        dlg.geometry('420x480')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text='⭐ 常用操作收藏', font=('Microsoft YaHei', 14, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=(12, 4))
        tk.Label(dlg, text='勾选常用功能，将显示在顶部快捷栏。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack()

        current = set(self.cfg.get('favorites', []))
        tools = self._all_tools()

        sf = tk.Frame(dlg, bg=C['bg']); sf.pack(fill='both', expand=True, padx=16, pady=8)
        canvas = tk.Canvas(sf, bg=C['bg'], highlightthickness=0)
        sbar = tk.Scrollbar(sf, orient='vertical', command=canvas.yview)
        inner = tk.Frame(canvas, bg=C['bg'])
        inner.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        wid = canvas.create_window((0, 0), window=inner, anchor='nw')
        canvas.configure(yscrollcommand=sbar.set)
        canvas.pack(side='left', fill='both', expand=True)
        sbar.pack(side='right', fill='y')
        def _resize(e): canvas.itemconfig(wid, width=e.width)
        canvas.bind('<Configure>', _resize)

        vars_ = {}
        for name, icon, _ in tools:
            v = tk.BooleanVar(value=name in current)
            vars_[name] = v
            tk.Checkbutton(inner, text=f'{icon} {name}', variable=v,
                           font=('Microsoft YaHei', 10), bg=C['bg'], fg=C['text'],
                           selectcolor=C['bg'], activebackground=C['bg'],
                           anchor='w').pack(fill='x', padx=8, pady=1)

        def _save():
            selected = [n for n, v in vars_.items() if v.get()]
            self.cfg['favorites'] = selected
            try: self._save_config()
            except Exception: pass
            self._refresh_favorites_bar()
            self._log(f'⭐ 已保存 {len(selected)} 个收藏')
            dlg.destroy()

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=16, pady=(0, 12))
        self._btn(bf, '💾 保存收藏', C['green'], _save, font_size=10, padx=14, pady=4).pack(side='left', padx=4)
        self._btn(bf, '❌ 关闭', C['gray'], dlg.destroy, font_size=9, padx=12, pady=4).pack(side='left', padx=4)

    def _monthly_goals(self):
        """月度目标设定与跟踪：为每人设定集数/收入目标，保存到 config.json"""
        goals = self.cfg.setdefault('monthly_goals', {})
        dlg = tk.Toplevel(self.root)
        dlg.title('🎯 月度目标设定')
        dlg.geometry('720x520')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()
        tk.Label(dlg, text='🎯 月度目标设定（集数 / 收入）',
                 font=('Microsoft YaHei', 14, 'bold'), bg=C['bg'],
                 fg=C['text']).pack(pady=(12, 4))
        tk.Label(dlg, text='为每人设定目标，仪表盘将显示完成进度。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack()

        sf = tk.Frame(dlg, bg=C['bg']); sf.pack(fill='both', expand=True, padx=16, pady=8)
        canvas = tk.Canvas(sf, bg=C['bg'], highlightthickness=0)
        sbar = tk.Scrollbar(sf, orient='vertical', command=canvas.yview)
        inner = tk.Frame(canvas, bg=C['bg'])
        inner.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        wid = canvas.create_window((0, 0), window=inner, anchor='nw')
        canvas.configure(yscrollcommand=sbar.set)
        canvas.pack(side='left', fill='both', expand=True)
        sbar.pack(side='right', fill='y')
        def _resize(e): canvas.itemconfig(wid, width=e.width)
        canvas.bind('<Configure>', _resize)

        hdr = tk.Frame(inner, bg=C['bg'])
        hdr.pack(fill='x', pady=(0, 4))
        for t, w in [('人员', 120), ('角色', 80), ('目标集数', 100), ('目标收入', 100)]:
            tk.Label(hdr, text=t, font=('Microsoft YaHei', 9, 'bold'), bg=C['bg'],
                     fg=C['text2'], width=w // 10).pack(side='left', padx=4)

        entries = {}
        roles = self.cfg.get('人员角色', {})
        for nm in roles:
            row = tk.Frame(inner, bg=C['bg']); row.pack(fill='x', pady=2)
            tk.Label(row, text=nm, font=('Microsoft YaHei', 9), bg=C['bg'],
                     fg=C['text'], width=12, anchor='w').pack(side='left', padx=4)
            tk.Label(row, text=roles.get(nm, ''), font=('Microsoft YaHei', 8), bg=C['bg'],
                     fg=C['text3'], width=8, anchor='w').pack(side='left', padx=4)
            g = goals.get(nm, {})
            ev = tk.StringVar(value=str(g.get('episodes', '')))
            iv = tk.StringVar(value=str(g.get('income', '')))
            ep = tk.Entry(row, textvariable=ev, font=('Microsoft YaHei', 9), width=10)
            ep.pack(side='left', padx=4)
            inc = tk.Entry(row, textvariable=iv, font=('Microsoft YaHei', 9), width=10)
            inc.pack(side='left', padx=4)
            entries[nm] = (ev, iv)

        def _save():
            for nm, (ev, iv) in entries.items():
                ep_val = ev.get().strip()
                inc_val = iv.get().strip()
                g = {}
                if ep_val:
                    try: g['episodes'] = int(ep_val)
                    except ValueError: pass
                if inc_val:
                    try: g['income'] = int(inc_val)
                    except ValueError: pass
                if g:
                    goals[nm] = g
                else:
                    goals.pop(nm, None)
            self.cfg['monthly_goals'] = goals
            try: self._save_config()
            except Exception: pass
            self._log(f'🎯 已保存 {len(goals)} 人月度目标')
            dlg.destroy()
            messagebox.showinfo('完成', f'已保存 {len(goals)} 人月度目标。', parent=self.root)

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=16, pady=(0, 12))
        self._btn(bf, '💾 保存目标', C['green'], _save, font_size=10, padx=18, pady=5).pack(side='left', padx=3)
        self._btn(bf, '❌ 取消', C['gray'], dlg.destroy, font_size=9, padx=16, pady=4).pack(side='left', padx=3)

    def _manage_project_templates(self, parent, proj_templates, combo, names):
        """项目管理模板：新增/编辑/删除项目模板（总集数、一卡区间）"""
        from tkinter import simpledialog
        tdlg = tk.Toplevel(parent)
        tdlg.title('🗂 项目模板管理')
        tdlg.geometry('520x420')
        tdlg.configure(bg=C['bg'])
        tdlg.transient(parent); tdlg.grab_set()
        tk.Label(tdlg, text='🗂 项目模板管理', font=('Microsoft YaHei', 14, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=(12, 4))
        tk.Label(tdlg, text='模板可预设总集数、一卡区间，一键套用。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack()

        listbox = tk.Listbox(tdlg, font=('Microsoft YaHei', 10), height=10)
        listbox.pack(fill='both', expand=True, padx=16, pady=8)
        def _refresh_list():
            listbox.delete(0, 'end')
            for nm in names:
                t = proj_templates.get(nm, {})
                roles_str = ''
                if t.get('roles'):
                    roles_str = ' 配比:' + ','.join(f'{k}{v}' for k, v in t['roles'].items())
                listbox.insert('end', f'{nm}  (总集数:{t.get("total","")} 一卡区间:{t.get("range","")}{roles_str})')
        _refresh_list()

        def _add():
            name = simpledialog.askstring('新增模板', '模板名称：', parent=tdlg)
            if not name: return
            name = name.strip()
            if not name: return
            total = simpledialog.askinteger('新增模板', '总集数：', parent=tdlg, minvalue=1)
            if total is None: return
            rng = simpledialog.askinteger('新增模板', '一卡区间（可留空）：', parent=tdlg, minvalue=1)
            t = {'total': total}
            if rng: t['range'] = rng
            # 角色配比：格式如 "小组长:1, 一卡剪辑:3"
            roles_str = simpledialog.askstring('新增模板', '角色配比（可留空，格式：角色:人数, 角色:人数）\n例：小组长:1, 一卡剪辑:3', parent=tdlg)
            if roles_str:
                roles = {}
                for part in roles_str.replace('，', ',').split(','):
                    part = part.strip()
                    if ':' in part:
                        rname, cnt = part.split(':', 1)
                        rname = rname.strip()
                        try: roles[rname] = int(cnt.strip())
                        except ValueError: pass
                if roles:
                    t['roles'] = roles
            proj_templates[name] = t
            names.append(name)
            self.cfg['project_templates'] = proj_templates
            try: self._save_config()
            except Exception: pass
            combo['values'] = ['（自定义）'] + names
            combo.set(name)
            _refresh_list()
        def _delete():
            sel = listbox.curselection()
            if not sel: return
            nm = names[sel[0]]
            if messagebox.askyesno('确认删除', f'删除模板"{nm}"？', parent=tdlg):
                proj_templates.pop(nm, None)
                names.remove(nm)
                self.cfg['project_templates'] = proj_templates
                try: self._save_config()
                except Exception: pass
                combo['values'] = ['（自定义）'] + names
                if combo.current() >= len(names):
                    combo.current(0)
                _refresh_list()

        bf = tk.Frame(tdlg, bg=C['bg']); bf.pack(fill='x', padx=16, pady=(0, 12))
        self._btn(bf, '➕ 新增', C['green'], _add, font_size=9, padx=14, pady=4).pack(side='left', padx=3)
        self._btn(bf, '🗑 删除', C['red'], _delete, font_size=9, padx=14, pady=4).pack(side='left', padx=3)
        self._btn(bf, '❌ 关闭', C['gray'], tdlg.destroy, font_size=9, padx=14, pady=4).pack(side='left', padx=3)

    def _check_update(self, silent=True):
        """检查版本更新：优先读取本地 version.txt，否则从 GitHub releases 获取。"""
        cur = self.cfg.get('version', '1.0.0')
        try:
            latest = None
            # 1) 本地 version.txt（内网/共享文件夹场景）
            local_v = os.path.join(SCRIPT_DIR, 'version.txt')
            if os.path.exists(local_v):
                with open(local_v, 'r', encoding='utf-8') as f:
                    latest = f.read().strip()
            # 2) GitHub releases API（可选，超时静默失败）
            if not latest:
                try:
                    import urllib.request
                    req = urllib.request.Request(
                        'https://api.github.com/repos/OWNER/REPO/releases/latest',
                        headers={'User-Agent': 'Mozilla/5.0'})
                    with urllib.request.urlopen(req, timeout=5) as resp:
                        import json as _j
                        latest = _j.loads(resp.read().decode('utf-8')).get('tag_name', '').lstrip('v')
                except Exception:
                    latest = None
            if latest and latest != cur:
                self._log(f'🆕 发现新版本 v{latest}（当前 v{cur}）')
                # J: 若配置了 update_url 则支持直接下载更新
                update_url = self.cfg.get('update_url', '')
                if update_url:
                    do_up = messagebox.askyesno(
                        '发现新版本',
                        f'发现新版本 v{latest}（当前 v{cur}）\n\n是否立即下载并更新？',
                        parent=self.root)
                    if do_up:
                        self._download_update(update_url)
                        return
                messagebox.showinfo('发现新版本',
                                     f'发现新版本 v{latest}（当前 v{cur}）\n\n请前往下载或联系管理员更新。',
                                     parent=self.root)
            else:
                if not silent:
                    messagebox.showinfo('检查更新', f'当前已是最新版本 v{cur}。', parent=self.root)
                self._log(f'✅ 已是最新版本 v{cur}')
        except Exception as e:
            if not silent:
                messagebox.showerror('检查失败', f'无法检查更新：{e}', parent=self.root)
            self._log(f'⚠️ 更新检查失败: {e}')

    def _download_update(self, url):
        """J: 下载新版本文件并替换当前程序文件"""
        import urllib.request
        try:
            self._log(f'⬇️ 正在下载更新: {url}')
            tmp = os.path.join(SCRIPT_DIR, '_update_tmp.py')
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=30) as resp, open(tmp, 'wb') as f:
                f.write(resp.read())
            self._log(f'✅ 下载完成，正在替换...')
            # 备份当前文件后替换
            cur_file = os.path.abspath(__file__)
            bak = cur_file + '.bak_update'
            try:
                if os.path.exists(bak):
                    os.remove(bak)
                os.rename(cur_file, bak)
            except OSError:
                pass
            os.replace(tmp, cur_file)
            messagebox.showinfo('更新完成',
                f'✅ 已更新到新版本！\n\n请重启软件生效。\n（旧版本已备份为 .bak_update）',
                parent=self.root)
            self._log(f'✅ 更新完成，请重启软件')
        except Exception as e:
            self._log(f'❌ 更新失败: {e}')
            messagebox.showerror('更新失败', f'下载或替换失败：\n{e}', parent=self.root)
            try:
                if os.path.exists(tmp := os.path.join(SCRIPT_DIR, '_update_tmp.py')):
                    os.remove(tmp)
            except OSError:
                pass

    def _batch_process(self):
        """批量处理：选择多个数据文件，逐个生成提成表（带进度条/停止/失败重试）"""
        files = filedialog.askopenfilenames(
            title='选择多个数据文件（可多选）',
            filetypes=[('Excel文件', '*.xlsx')],
            initialdir=SCRIPT_DIR)
        if not files:
            return
        self._log(f'📦 批量处理：选中 {len(files)} 个数据文件')

        # 进度对话框
        dlg = tk.Toplevel(self.root)
        dlg.title('📦 批量处理进度')
        dlg.geometry('420x200')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root)
        tk.Label(dlg, text='📦 批量处理中...', font=('Microsoft YaHei', 12, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=(14, 6))
        self._batch_progress = ttk.Progressbar(dlg, mode='determinate', maximum=len(files))
        self._batch_progress.pack(fill='x', padx=24, pady=6)
        self._batch_status = tk.Label(dlg, text='准备开始...', font=('Microsoft YaHei', 9),
                                      bg=C['bg'], fg=C['text2'])
        self._batch_status.pack(pady=4)
        stop_flag = [False]
        tk.Button(dlg, text='⏹ 停止', font=('Microsoft YaHei', 10), bg=C['red'], fg='white',
                  relief='flat', cursor='hand2', padx=16, pady=4,
                  command=lambda: stop_flag.__setitem__(0, True)).pack(pady=(10, 8))

        def worker():
            ok, failed = 0, []
            total = len(files)
            for i, pf in enumerate(files):
                if stop_flag[0]:
                    break
                try:
                    gc, _sys = self._load_gc_module()
                    import pandas as pd
                    df = pd.read_excel(pf, header=None)
                    data_cn, data_mnum = gc.get_month_from_data(df)
                    if data_cn and data_mnum:
                        import re as _re
                        _tm = _re.match(r'(\d{4})年', gc.TEMPLATE_DATE)
                        _yr = _tm.group(1) if _tm else str(datetime.now().year)
                        gc.OUTPUT_MONTH = data_cn
                        gc.TEMPLATE_DATE = f'{_yr}年{data_mnum:02d}月'
                    records, group_pids = gc.parse_projects(df)
                    if not records:
                        self.root.after(0, lambda p=pf, i=i, n=total: self._log(f'  ⚠️ [{i+1}/{n}] {os.path.basename(p)} 无有效记录，跳过'))
                    else:
                        cd = gc.compute_commission(records, group_pids)
                        out = os.path.join(self.output_dir, f'AI后期剪辑提成一组{gc.OUTPUT_MONTH}.xlsx')
                        gc.generate_excel(records, cd, self.template_file, out)
                        _sys.path.pop(0)
                        ok += 1
                        self.root.after(0, lambda p=pf, i=i, n=total: self._log(f'  ✅ [{i+1}/{n}] {os.path.basename(p)} 已生成'))
                except Exception as e:
                    failed.append((pf, e))
                    self.root.after(0, lambda p=pf, e=e, i=i, n=total: self._log(f'  ❌ [{i+1}/{n}] {os.path.basename(p)} 失败: {e}'))
                self.root.after(0, lambda i=i: (self._batch_progress.configure(value=i + 1),
                                                self._batch_status.configure(text=f'已完成 {i+1}/{total}')))

            # H: 失败自动重试一次
            if failed and not stop_flag[0]:
                self.root.after(0, lambda: self._log(f'🔁 对 {len(failed)} 个失败文件重试一次...'))
                retry_ok = 0
                for pf, _ in failed:
                    try:
                        gc, _sys = self._load_gc_module()
                        import pandas as pd
                        df = pd.read_excel(pf, header=None)
                        data_cn, data_mnum = gc.get_month_from_data(df)
                        if data_cn and data_mnum:
                            import re as _re
                            _tm = _re.match(r'(\d{4})年', gc.TEMPLATE_DATE)
                            _yr = _tm.group(1) if _tm else str(datetime.now().year)
                            gc.OUTPUT_MONTH = data_cn
                            gc.TEMPLATE_DATE = f'{_yr}年{data_mnum:02d}月'
                        records, group_pids = gc.parse_projects(df)
                        if records:
                            cd = gc.compute_commission(records, group_pids)
                            out = os.path.join(self.output_dir, f'AI后期剪辑提成一组{gc.OUTPUT_MONTH}.xlsx')
                            gc.generate_excel(records, cd, self.template_file, out)
                            _sys.path.pop(0)
                            retry_ok += 1
                            self.root.after(0, lambda p=pf: self._log(f'  ✅ 重试成功: {os.path.basename(p)}'))
                    except Exception as e2:
                        self.root.after(0, lambda p=pf, e2=e2: self._log(f'  ❌ 重试仍失败: {os.path.basename(p)} - {e2}'))
                ok += retry_ok

            def _done():
                self._batch_progress.configure(value=total)
                self._batch_status.configure(text=f'完成：成功 {ok} 个，失败 {total-ok} 个' + ('（已停止）' if stop_flag[0] else ''))
                dlg.after(1200, dlg.destroy)
                self._log(f'📦 批量完成：成功 {ok} 个，失败 {total-ok} 个' + ('（已停止）' if stop_flag[0] else ''))
                self._open(self.output_dir)
            self.root.after(0, _done)
        threading.Thread(target=worker, daemon=True).start()

    def _setup_ttk_style(self):
        s = ttk.Style()
        s.theme_use('clam')

        s.configure('TNotebook', background=C['bg'], borderwidth=0, tabmargins=(0, 0, 0, 8))
        s.configure('TNotebook.Tab',
            font=('Microsoft YaHei', 10),
            padding=(20, 9),
            background=C['bg'],
            foreground=C['text2'],
            borderwidth=0)
        s.map('TNotebook.Tab',
            background=[('selected', C['accent_l'])],
            foreground=[('selected', C['accent'])],
            expand=[('selected', [0, 0, 0, 0])])

        s.configure('TProgressbar', thickness=3,
            background=C['accent'], troughcolor=C['border_l'])

        s.configure('Treeview', font=('Microsoft YaHei', 9),
            rowheight=28, background=C['card'],
            fieldbackground=C['card'], foreground=C['text'])
        s.configure('Treeview.Heading', font=('Microsoft YaHei', 9, 'bold'),
            background='#edf3f2', foreground=C['text'], relief='flat', padding=(6, 4))
        s.map('Treeview',
            background=[('selected', C['accent_l'])],
            foreground=[('selected', C['text'])])

        s.configure('TCombobox',
            fieldbackground=C['card'], background=C['card'],
            arrowcolor=C['text'])

    # ============ 通用组件 ============

    @staticmethod
    def _card(parent, **pack):
        """白色卡片，统一圆角风格（通过 highlightthickness 模拟边框）"""
        f = tk.Frame(parent, bg=C['card'], highlightthickness=1,
                     highlightbackground=C['card_border'])
        if pack: f.pack(**pack)
        return f

    @staticmethod
    def _hdr(parent, icon, title, right=None):
        """区块标题：左侧色条 + 图标 + 文字，可选右侧操作区"""
        bar = tk.Frame(parent, bg=C['card'])
        bar.pack(fill='x', padx=14, pady=(12, 6))
        # 左侧色条
        tk.Frame(bar, bg=C['accent'], width=3).pack(side='left', fill='y', padx=(0, 8))
        tk.Label(bar, text=icon, font=('Microsoft YaHei', 13), bg=C['card'],
                 fg=C['text']).pack(side='left', padx=(0, 6))
        tk.Label(bar, text=title, font=('Microsoft YaHei', 12, 'bold'),
                 bg=C['card'], fg=C['text']).pack(side='left')
        if right:
            right(bar)
        sep = tk.Frame(parent, bg=C['border_l'], height=1)
        sep.pack(fill='x', padx=14, pady=(0, 6))

    @staticmethod
    def _btn(parent, text, color, cmd, font_size=10, padx=16, pady=8):
        """统一命令按钮，带悬停加深与按下反馈"""
        hc = C['accent_a'] if color == C['accent'] else color
        btn = tk.Label(parent, text=text, font=('Microsoft YaHei', font_size, 'bold'),
                       bg=color, fg='white', padx=padx, pady=pady,
                       cursor='hand2', relief='flat')
        original = color
        # 注意：<Button-1> 与 <ButtonPress-1> 是同一事件，用 bind(..., add='+') 追加回调，
        # 避免覆盖 cmd()。按下触发命令并加深，松开恢复悬停色。
        btn.bind('<ButtonPress-1>', lambda e: (cmd(), btn.configure(bg=C['accent_a'] if color == C['accent'] else C['slate'])))
        btn.bind('<ButtonRelease-1>', lambda e: btn.configure(bg=C['accent_h'] if color == C['accent'] else hc), add='+')
        btn.bind('<Enter>', lambda e: btn.configure(bg=C['accent_h'] if color == C['accent'] else hc))
        btn.bind('<Leave>', lambda e: btn.configure(bg=original))
        return btn

    @staticmethod
    def _tooltip(widget, text):
        """为图标按钮提供鼠标悬停说明。"""
        tip = [None]

        def show(_event):
            if tip[0] is not None:
                return
            popup = tk.Toplevel(widget)
            popup.wm_overrideredirect(True)
            popup.configure(bg='#172120')
            x = widget.winfo_rootx()
            y = widget.winfo_rooty() + widget.winfo_height() + 6
            popup.geometry(f'+{x}+{y}')
            tk.Label(popup, text=text, bg='#172120', fg='white',
                     font=('Microsoft YaHei', 8), padx=8, pady=4).pack()
            tip[0] = popup

        def hide(_event):
            if tip[0] is not None:
                tip[0].destroy()
                tip[0] = None

        widget.bind('<Enter>', show, add='+')
        widget.bind('<Leave>', hide, add='+')

    def _load_config(self):
        from config_loader import load_config
        return load_config(CONFIG_PATH).to_dict()

    def _save_config(self):
        from config_loader import save_config
        from models import AppConfig
        save_config(AppConfig.from_dict(self.cfg), CONFIG_PATH)

    def _save_app_settings(self):
        """保存 GUI 上次使用设置（文件路径、输出目录、界面选项）"""
        self.cfg.setdefault('app_settings', {})
        self.cfg['app_settings'].update({
            'project_file': self.project_file,
            'template_file': self.template_file,
            'output_dir': self.output_dir,
            'auto_backup': bool(self.auto_backup.get()) if hasattr(self, 'auto_backup') else True,
        })
        self._save_config()

    def _on_close(self):
        """关闭程序前保存界面设置"""
        try:
            self._save_app_settings()
        except Exception as e:
            self._log(f'⚠️ 保存界面设置失败: {e}')
        self.root.destroy()

    def _load_gc_module(self):
        """加载 generate_commission 模块并注入当前配置"""
        import sys as _sys
        _sys.path.insert(0, SRC_DIR)
        import generate_commission as gc
        gc.set_config(self.cfg)
        return gc, _sys

    def _require_features(self):
        """检查功能模块是否可用，不可用则弹窗并返回 False"""
        if not HAS_FEATURES:
            messagebox.showerror('错误', '功能模块(features.py)未找到')
            return False
        return True

    @staticmethod
    def _open(path):
        """在系统默认应用中打开文件/目录"""
        try:
            os.startfile(path)
        except Exception:
            pass

    # ============ UI 构建 ============

    def build_ui(self):
        # ---- 应用栏 ----
        hdr = tk.Frame(self.root, bg=C['hdr_bg'], height=72)
        hdr.pack(fill='x'); hdr.pack_propagate(False)
        hl = tk.Frame(hdr, bg=C['hdr_bg'])
        hl.pack(side='left', padx=24, pady=12)
        tk.Label(hl, text='AI 后期剪辑', font=('Microsoft YaHei', 15, 'bold'),
                 fg=C['hdr_text'], bg=C['hdr_bg']).pack(anchor='w')
        tk.Label(hl, text='提成工作台  ·  月度核算与报表生成',
                 font=('Microsoft YaHei', 8), fg=C['hdr_sub'],
                 bg=C['hdr_bg']).pack(anchor='w', pady=(1, 0))

        actions = tk.Frame(hdr, bg=C['hdr_bg'])
        actions.pack(side='right', padx=20)
        open_btn = self._btn(actions, '📂', '#234846',
                             lambda: self._open(self.output_dir), padx=12, pady=7)
        open_btn.pack(side='left', padx=3)
        self._tooltip(open_btn, '打开输出目录')
        role_btn = self._btn(actions, '⚙', '#234846', self.open_role_editor, padx=12, pady=7)
        role_btn.pack(side='left', padx=3)
        self._tooltip(role_btn, '管理人员角色')
        fav_btn = self._btn(actions, '⭐', '#234846', self._manage_favorites, padx=12, pady=7)
        fav_btn.pack(side='left', padx=3)
        self._tooltip(fav_btn, '常用操作收藏')
        # 日期显示
        from datetime import datetime as _dt
        _today = _dt.now()
        tk.Label(actions, text=f'{_today.month}月{_today.day}日', font=('Microsoft YaHei', 9, 'bold'),
                 fg=C['hdr_text'], bg=C['hdr_bg']).pack(side='left', padx=(8, 0))
        tk.Label(actions, text='v8.0', font=('Consolas', 9), fg='#89aaa7',
                 bg=C['hdr_bg']).pack(side='left', padx=(12, 0))

        # ---- 主导航：左侧边栏 ----
        main = tk.Frame(self.root, bg=C['bg'])
        main.pack(fill='both', expand=True, padx=0, pady=0)

        # 左侧边栏
        sidebar = tk.Frame(main, bg=C['hdr_bg'], width=176)
        sidebar.pack(side='left', fill='y')
        sidebar.pack_propagate(False)

        tk.Label(sidebar, text='导 航', font=('Microsoft YaHei', 9, 'bold'),
                 bg=C['hdr_bg'], fg='#89aaa7', anchor='w').pack(
                     fill='x', padx=18, pady=(18, 8))

        nav_items = [
            ('🏠', '工作台', self._build_tab_main),
            ('🧰', '工具箱', self._build_tab_tools),
            ('📊', '分析管理', self._build_tab_advanced),
        ]
        self._nav_buttons = []  # [(label_widget, content_frame)]

        # 内容区容器
        content = tk.Frame(main, bg=C['bg'])
        content.pack(side='left', fill='both', expand=True)
        self._content_frames = {}  # name -> frame

        # 收藏快捷栏（顶部）
        self.fav_bar = tk.Frame(content, bg=C['bg'])
        self.fav_bar.pack(fill='x', padx=10, pady=(8, 0))
        self._refresh_favorites_bar()

        def _make_nav(i, icon, name, builder):
            # 内容页
            page = tk.Frame(content, bg=C['bg'])
            self._content_frames[name] = page
            builder(page)

            # 侧边栏按钮
            btn = tk.Frame(sidebar, bg=C['hdr_bg'], cursor='hand2')
            btn.pack(fill='x', pady=1)
            lbl = tk.Label(btn, text=f'{icon}  {name}', font=('Microsoft YaHei', 11, 'bold'),
                           bg=C['hdr_bg'], fg=C['hdr_text'], anchor='w')
            lbl.pack(fill='x', padx=16, pady=11)
            self._nav_buttons.append((lbl, page))

            def _select(e=None):
                # 高亮当前，隐藏其他
                for bl, p in self._nav_buttons:
                    p.pack_forget()
                    bl.configure(bg=C['hdr_bg'], fg=C['hdr_text'])
                page.pack(fill='both', expand=True)
                lbl.configure(bg=C['accent'], fg='#ffffff')
            btn.bind('<Button-1>', _select)
            lbl.bind('<Button-1>', _select)

            # 悬停高亮
            def _nav_enter(e, l=lbl):
                if l.cget('bg') != C['accent']:
                    l.configure(bg='#1b3a38', fg=C['hdr_text'])
            def _nav_leave(e, l=lbl):
                if l.cget('bg') != C['accent']:
                    l.configure(bg=C['hdr_bg'], fg=C['hdr_text'])
            btn.bind('<Enter>', _nav_enter)
            lbl.bind('<Enter>', _nav_enter)
            btn.bind('<Leave>', _nav_leave)
            lbl.bind('<Leave>', _nav_leave)

        for i, (ic, nm, bd) in enumerate(nav_items):
            _make_nav(i, ic, nm, bd)

        # 侧边栏底部版本信息
        tk.Frame(sidebar, bg=C['hdr_bg']).pack(fill='both', expand=True)
        tk.Label(sidebar, text='v8.0 · 提成工作台', font=('Consolas', 8),
                 bg=C['hdr_bg'], fg='#5f807d', anchor='w').pack(
                     fill='x', side='bottom', padx=18, pady=(0, 14))

        # 默认显示第一个 tab（工作台）
        if self._nav_buttons:
            self._nav_buttons[0][1].pack(fill='both', expand=True)
            self._nav_buttons[0][0].configure(bg=C['accent'], fg='#ffffff')

        # ---- 状态栏 ----
        bar = tk.Frame(self.root, bg=C['status_bg'], height=28)
        bar.pack(fill='x', side='bottom'); bar.pack_propagate(False)
        self.st = tk.Label(bar, text='  ●  就绪 · 配置已加载', font=('Microsoft YaHei', 9),
                           bg=C['status_bg'], fg=C['status_fg'],
                           anchor='w', padx=14)
        self.st.pack(side='left', fill='x')
        # 右侧信息
        tk.Frame(bar, bg=C['border'], width=1).pack(side='right', fill='y', padx=8)
        tk.Label(bar, text='Ctrl+G 生成 · Ctrl+P 预览 · Ctrl+D 分集 · F5 刷新',
                 font=('Microsoft YaHei', 7), bg=C['status_bg'], fg=C['text3']).pack(side='right', padx=2)
        tk.Label(bar, text=f'{len(self.cfg.get("人员角色", {}))} 人', font=('Microsoft YaHei', 8),
                 bg=C['status_bg'], fg=C['text3']).pack(side='right', padx=2)

    def _build_tab_main(self, p):
        """主工作台：准备输入、发起生成、查看执行结果。"""
        intro = tk.Frame(p, bg=C['bg'])
        intro.pack(fill='x', pady=(2, 10))
        tk.Label(intro, text='月度提成生成', font=('Microsoft YaHei', 18, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(anchor='w')
        tk.Label(intro, text='核对文件与角色后，生成 Excel、统计简报、仪表盘和个人绩效卡片。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text2']).pack(anchor='w', pady=(3, 0))

        # 全局搜索栏
        search_card = self._card(p, fill='x', pady=(0, 10))
        search_row = tk.Frame(search_card, bg=C['card']); search_row.pack(fill='x', padx=14, pady=8)
        tk.Label(search_row, text='🔍', font=('Microsoft YaHei', 12), bg=C['card'],
                 fg=C['text']).pack(side='left', padx=(0, 6))
        self.search_entry = tk.Entry(search_row, font=('Microsoft YaHei', 11),
                                     relief='solid', borderwidth=1)
        self.search_entry.pack(side='left', fill='x', expand=True, padx=(0, 8))
        self.search_entry.bind('<Return>', lambda e: self._do_search())
        self._btn(search_row, '搜索', C['accent'], self._do_search, font_size=9, padx=14, pady=4).pack(side='left')
        tk.Label(search_row, text='按项目ID / 名称 / 人员搜索', font=('Microsoft YaHei', 7),
                 bg=C['card'], fg=C['text3']).pack(side='left', padx=8)

        c1 = self._card(p, fill='x', pady=(0, 12))
        header = tk.Frame(c1, bg=C['card']); header.pack(fill='x', padx=16, pady=(12, 5))
        tk.Label(header, text='输入与输出', font=('Microsoft YaHei', 11, 'bold'),
                 bg=C['card'], fg=C['text']).pack(side='left')
        tk.Label(header, text='生成前可直接更换文件', font=('Microsoft YaHei', 8),
                 bg=C['card'], fg=C['text3']).pack(side='right')
        file_grid = tk.Frame(c1, bg=C['card']); file_grid.pack(fill='x', padx=12, pady=(0, 12))
        for col, (label, attr, cmd) in enumerate([
            ('项目数据', 'pf_label', self._select_project),
            ('提成模板', 'tf_label', self._select_template),
            ('输出目录', 'od_label', self._select_output_dir)]):
            cell = tk.Frame(file_grid, bg='#f8fbfa', highlightthickness=1,
                            highlightbackground=C['card_border'])
            cell.grid(row=0, column=col, sticky='nsew', padx=4)
            # 悬停高亮
            def _cell_enter(e, c=cell):
                c.configure(highlightbackground=C['accent'], highlightthickness=2)
            def _cell_leave(e, c=cell):
                c.configure(highlightbackground=C['card_border'], highlightthickness=1)
            cell.bind('<Enter>', _cell_enter)
            cell.bind('<Leave>', _cell_leave)
            cell.bind('<Button-1>', lambda e, c=cmd: c())
            tk.Label(cell, text=label, font=('Microsoft YaHei', 8, 'bold'),
                     bg='#f8fbfa', fg=C['text2']).pack(anchor='w', padx=10, pady=(8, 1))
            lbl = tk.Label(cell, text='', font=('Microsoft YaHei', 8), bg='#f8fbfa',
                           fg=C['text3'], anchor='w')
            lbl.pack(fill='x', padx=10, pady=(0, 7))
            setattr(self, attr, lbl)
            change = tk.Label(cell, text='更换', font=('Microsoft YaHei', 8, 'bold'),
                              bg=C['accent_l'], fg=C['accent'], cursor='hand2', padx=9, pady=2)
            change.pack(anchor='e', padx=8, pady=(0, 8))
            change.bind('<Button-1>', lambda e, c=cmd: c())
            file_grid.grid_columnconfigure(col, weight=1, uniform='file')
        self._refresh_file_labels()

        body = tk.Frame(p, bg=C['bg']); body.pack(fill='both', expand=True)
        left = tk.Frame(body, bg=C['bg'], width=385)
        left.pack(side='left', fill='both', padx=(0, 12)); left.pack_propagate(False)
        right = tk.Frame(body, bg=C['bg'])
        right.pack(side='left', fill='both', expand=True)

        cc = self._card(left, fill='both', expand=True, pady=(0, 8))
        self._hdr(cc, '👥', '本组角色配置')
        self.role_tags = tk.Frame(cc, bg=C['card'])
        self.role_tags.pack(fill='both', expand=True, padx=10, pady=(0, 8))
        self._refresh_role_tags()

        run_card = self._card(left, fill='x')
        tk.Label(run_card, text='准备完成后开始生成', font=('Microsoft YaHei', 9, 'bold'),
                 bg=C['card'], fg=C['text']).pack(anchor='w', padx=14, pady=(12, 2))
        tk.Label(run_card, text='将先标记超时集数，再生成全部报表。', font=('Microsoft YaHei', 8),
                 bg=C['card'], fg=C['text3']).pack(anchor='w', padx=14, pady=(0, 10))
        btn_row = tk.Frame(run_card, bg=C['card']); btn_row.pack(fill='x')
        self.run_btn = self._btn(btn_row, '▶  开始生成', C['accent'],
                                  self.run, font_size=12, padx=28, pady=10)
        self.run_btn.pack(side='left', fill='x', expand=True, padx=(14, 4), pady=(0, 12))
        role_small = self._btn(btn_row, '⚙', C['orange'], self.open_role_editor, padx=13, pady=10)
        role_small.pack(side='left', padx=4, pady=(0, 12))
        self._tooltip(role_small, '管理角色')
        self.progress = ttk.Progressbar(run_card, mode='indeterminate')
        self.progress.pack(fill='x', padx=14, pady=(0, 8))
        aux = tk.Frame(run_card, bg=C['card']); aux.pack(fill='x', padx=14, pady=(0, 10))
        tk.Checkbutton(aux, text='自动备份', variable=self.auto_backup,
                       font=('Microsoft YaHei', 8), bg=C['card'], fg=C['text3'],
                       selectcolor=C['card'], activebackground=C['card'],
                       command=self._save_app_settings).pack(side='left')
        tk.Label(aux, text='Ctrl+G 生成  ·  Ctrl+R 角色', font=('Microsoft YaHei', 7),
                 bg=C['card'], fg=C['text3']).pack(side='right')

        c3 = self._card(right, fill='both', expand=True)
        def _clear_log(bar):
            clear = tk.Label(bar, text='清空', font=('Microsoft YaHei', 8, 'bold'),
                             bg=C['accent_l'], fg=C['accent'], cursor='hand2', padx=8, pady=1)
            clear.pack(side='right')
            def _do():
                self.log_txt.configure(state='normal')
                self.log_txt.delete('1.0', 'end')
                self.log_txt.configure(state='disabled')
            clear.bind('<Button-1>', lambda e: _do())
        self._hdr(c3, '📝', '运行记录', right=_clear_log)
        tw = tk.Frame(c3, bg=C['log_bg'])
        tw.pack(fill='both', expand=True, padx=10, pady=(0, 10))
        self.log_txt = tk.Text(tw, font=('Microsoft YaHei', 9),
                               bg=C['log_bg'], fg=C['log_fg'],
                               insertbackground='#58a6ff',
                               relief='flat', padx=10, pady=8,
                               wrap='word', state='disabled',
                               selectbackground='#1f6feb')
        self.log_txt.pack(side='left', fill='both', expand=True)
        sb = tk.Scrollbar(tw, bg=C['log_bg'], troughcolor=C['log_bg'])
        sb.pack(side='right', fill='y')
        self.log_txt.configure(yscrollcommand=sb.set)
        sb.configure(command=self.log_txt.yview)

    def _build_tab_tools(self, p):
        canvas = tk.Canvas(p, bg=C['bg'], highlightthickness=0)
        sb = ttk.Scrollbar(p, orient='vertical', command=canvas.yview)
        sf = tk.Frame(canvas, bg=C['bg'])
        sf.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        wid = canvas.create_window((0,0), window=sf, anchor='nw')
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side='left', fill='both', expand=True, padx=(0, 6))
        sb.pack(side='right', fill='y')
        # 局部滚轮：进入工具箱区域才激活，离开即解除，避免影响其他页面
        def _wheel(e):
            canvas.yview_scroll(int(-e.delta/120), 'units')
        canvas.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>', _wheel))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))
        def _fw(e): canvas.itemconfig(wid, width=e.width)
        canvas.bind('<Configure>', _fw)

        # Use Unicode text directly instead of escape sequences
        groups = [
            ('📊 数据工具', [
                ('📋','数据预览', C['teal'], self._preview_data, '预览解析后的人员集数、绩效和提成汇总'),
                ('✅','数据校验', C['green'], self._validate_data, '校验项目数据日期、人名、重复分配等问题'),
                ('🔍','项目去重', C['red'], self._check_duplicates, '检查相同项目ID是否对应了不同项目名称'),
                ('📐','智能分集', C['pink'], self._smart_assign, '按角色权重自动分配各人负责集数'),
                ('📝','数据录入', C['blue'], self._data_entry, '图形化录入项目数据，无需手动填Excel'),
                ('📦','批量处理', C['indigo'], self._batch_process, '选择多个数据文件批量生成提成表'),
                ('🏅','交互排名', C['amber'], self._interactive_ranking, '交互式排序/筛选的绩效排名面板'),
            ]),
            ('📈 报表生成', [
                ('📊','月份对比', C['orange'], self._compare_months, '选择两个月提成表对比集数和提成变化'),
                ('📈','统计图表', C['cyan'], self._gen_charts, '可视化收入趋势、集数柱状与项目分布'),
                ('👥','多组对比', C['teal'], self._compare_groups, '对比两个组/团队的集数与提成差异'),
                ('📈','跨月趋势', C['indigo'], self._gen_trend_report, '基于历史数据生成跨月收入/集数趋势'),
                ('📅','年度汇总', C['purple'], self._gen_annual_summary, '批量汇总各月数据生成年度对比表'),
                ('🏆','组内排名', C['amber'], self._gen_ranking, '生成月度集数排行和提成排行HTML'),
                ('🗂','项目管理', C['purple'], self._gen_project_mgmt, '生成项目清单视图按交付日期排序'),
                ('🃏','绩效卡片', C['blue'], self._gen_cards, '每人独立HTML绩效卡片含集数达标提成'),
            ]),
            ('🔧 辅助工具', [
                ('📅','下月模板', C['indigo'], self._gen_next_template, '基于当前模板自动创建下月空白模板'),
                ('📤','导出PDF', C['cyan'], self._export_pdf, '将Excel提成表导出为A3横版PDF'),
                ('💾','多格式导出', C['teal'], self._export_multi, '导出人员提成/项目明细为CSV和JSON'),
                ('📥','数据导入', C['blue'], self._import_data, '从CSV/Excel导入数据并映射列名'),
                ('🗄','备份策略', C['slate'], self._backup_policy, '设置定时自动备份配置和历史数据'),
                ('🏷','提成规则', C['slate'], self._edit_rules, '可视化编辑各角色的基准集数单价'),
                ('📥','模板下载', C['teal'], self._download_template, '下载标准化项目数据录入模板Excel'),
                ('🔄','配置迁移', C['indigo'], self._config_migrate, '从旧版配置迁移到当前结构'),
            ]),
            ('⚡ 高级工具', [
                ('🎯','月度目标', C['amber'], self._monthly_goals, '为每人设定集数/收入目标并跟踪进度'),
                ('🔎','高级筛选', C['blue'], self._advanced_filter, '多条件组合筛选支持导出CSV'),
                ('✏','数据修正', '#f0641a', self._data_correction, '双击修正集数或删除重新生成提成表'),
                ('🌐','Web服务', C['accent'], self._web_server, '启动本地HTTP团队浏览器查看报表'),
                ('📸','配置快照', C['purple'], self._config_snapshot, '保存配置历史支持一键回滚'),
                ('🔄','文件监控', C['green'], self._toggle_watch, '检测数据变化自动提示重新生成'),
                ('🆕','更新检查', C['cyan'], self._check_update, '检查版本更新并提示下载'),
                ('⚠️','风险预警', C['red'], self._risk_warning, '扫描集数骤变/未达标/目标滞后等异常'),
                ('🖥','环境切换', C['teal'], self._manage_environments, '多套配置切换（不同团队各自环境）'),
                ('⏰','定时生成', C['cyan'], self._schedule_setup, '到点自动生成报表'),
                ('🔐','数据加密', C['slate'], self._data_encryption, '为配置文件设置密码保护'),
            ]),
        ]

        for gtitle, tools in groups:
            gf = tk.Frame(sf, bg=C['bg'])
            gf.pack(fill='x', padx=2, pady=(14, 7))
            tk.Label(gf, text=gtitle, font=('Microsoft YaHei', 11, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(side='left')
            tk.Frame(gf, bg=C['border'], height=1).pack(
                side='left', fill='x', expand=True, padx=(10,0))

            grid = tk.Frame(sf, bg=C['bg'])
            grid.pack(fill='x', padx=2)
            for i, (icon, name, color, cmd, desc) in enumerate(tools):
                card = tk.Frame(grid, bg=C['card'], highlightthickness=1,
                                highlightbackground=C['card_border'],
                                cursor='hand2')
                card.grid(row=i//3, column=i%3, padx=5, pady=5, sticky='nsew')
                # 图标圆底
                icon_cell = tk.Frame(card, bg=C['card'])
                icon_cell.pack(pady=(14, 6))
                icon_bg = tk.Frame(icon_cell, bg=color, width=44, height=44)
                icon_bg.pack_propagate(False)
                icon_bg.pack()
                tk.Label(icon_bg, text=icon, font=('Microsoft YaHei', 19),
                         bg=color, fg='white').pack(expand=True)
                tk.Label(card, text=name, font=('Microsoft YaHei', 10, 'bold'),
                         bg=C['card'], fg=C['text']).pack(pady=(2, 0))
                tk.Label(card, text=desc, font=('Microsoft YaHei', 7),
                         bg=C['card'], fg=C['text3'], wraplength=200,
                         justify='center').pack(padx=10, pady=(2, 12))
                # bottom accent bar
                tk.Frame(card, bg=color, height=3).pack(fill='x', side='bottom')
                for ch in list(card.children.values()):
                    ch.bind('<Button-1>', lambda e, c=cmd: c())
                card.bind('<Button-1>', lambda e, c=cmd: c())
                def _enter(e, f=card, c=color):
                    f.configure(bg=C['card_hover'], highlightbackground=c)
                    f.lift()
                def _leave(e, f=card):
                    f.configure(bg=C['card'], highlightbackground=C['card_border'])
                card.bind('<Enter>', _enter)
                card.bind('<Leave>', _leave)
            for c in range(3): grid.grid_columnconfigure(c, weight=1, uniform='tool')

    def _build_tab_advanced(self, p):
        # 个人趋势
        c1 = self._card(p, fill='x', padx=10, pady=(10,6))
        self._hdr(c1, '📈', '个人月度趋势')
        sf = tk.Frame(c1, bg=C['card']); sf.pack(fill='x', padx=14, pady=(0,10))
        self._trend_var = tk.StringVar()
        names = sorted(self.cfg.get('人员角色', {}).keys())
        if names:
            self._trend_var.set(names[0])
            ttk.Combobox(sf, textvariable=self._trend_var, values=names,
                         state='readonly', font=('Microsoft YaHei', 10),
                         width=10).pack(side='left', padx=(0,8))
        self._btn(sf, '生成趋势图', C['accent'], self._gen_trend).pack(side='left')

        # 备份管理
        c2 = self._card(p, fill='both', expand=True, padx=10, pady=(4,8))
        self._hdr(c2, '🗄', '备份管理')
        hf = tk.Frame(c2, bg=C['card']); hf.pack(fill='x', padx=14, pady=(0,4))
        self._btn(hf, '清理旧备份', C['red'], self._manage_backups,
                  font_size=9, padx=10, pady=4).pack(side='right')
        self._backup_list = tk.Text(c2, font=('Microsoft YaHei', 9),
                                    bg=C['log_bg'], fg=C['log_fg'],
                                    relief='flat', padx=10, pady=8,
                                    height=10, wrap='word',
                                    selectbackground='#1f6feb')
        self._backup_list.pack(fill='both', expand=True, padx=10, pady=(0,10))
        self._refresh_backups()

        # 主题设置
        c3 = self._card(p, fill='x', padx=10, pady=(4, 8))
        self._hdr(c3, '🎨', '界面主题')
        tf = tk.Frame(c3, bg=C['card']); tf.pack(fill='x', padx=14, pady=(0, 10))
        cur = '暗色' if self.theme == 'dark' else '亮色'
        tk.Label(tf, text=f'当前主题：{cur}', font=('Microsoft YaHei', 9),
                 bg=C['card'], fg=C['text2']).pack(side='left', padx=(0, 12))
        self._btn(tf, '☀️ 亮色', C['blue'], lambda: self._set_theme('light'),
                  font_size=9, padx=12, pady=4).pack(side='left', padx=3)
        self._btn(tf, '🌙 暗色', C['purple'], lambda: self._set_theme('dark'),
                  font_size=9, padx=12, pady=4).pack(side='left', padx=3)
        tk.Label(tf, text='切换后重启完全生效', font=('Microsoft YaHei', 7),
                 bg=C['card'], fg=C['text3']).pack(side='left', padx=8)

    def _refresh_role_tags(self):
        for w in self.role_tags.winfo_children():
            w.destroy()
        if not self.cfg: return
        rm = self.cfg.get('人员角色', {})
        if not rm:
            tk.Label(self.role_tags, text='暂无人员，请点击「角色配置」添加',
                     font=('Microsoft YaHei', 9), bg=C['card'],
                     fg=C['text3']).pack(anchor='w', pady=10)
            return

        total = len(rm)
        tk.Label(self.role_tags, text=f'共 {total} 人',
                 font=('Microsoft YaHei', 8), bg=C['card'],
                 fg=C['text3']).pack(anchor='w', pady=2)

        for role in ROLES:
            names = [n for n, r in rm.items() if r == role]
            if not names:
                continue
            fg, bg_c = ROLE_COLORS.get(role, ('#666', '#f0f0f0'))
            tag = tk.Frame(self.role_tags, bg=bg_c, highlightthickness=1,
                           highlightbackground=fg)
            tag.pack(fill='x', pady=3, padx=2)
            tk.Label(tag, text=f'  {role}  ', font=('Microsoft YaHei', 9, 'bold'),
                     bg=bg_c, fg=fg, padx=6, pady=3).pack(side='left')
            sep = tk.Frame(tag, bg=fg, width=1)
            sep.pack(side='left', fill='y', padx=5, pady=4)
            tk.Label(tag, text='、'.join(names),
                     font=('Microsoft YaHei', 9),
                     bg=bg_c, fg=C['text2'], padx=2, pady=3).pack(side='left')

    # ============ 角色编辑器 ============

    def open_role_editor(self):
        if not self.cfg:
            messagebox.showwarning('错误', '无法加载 config.json'); return
        dlg = tk.Toplevel(self.root)
        dlg.title('角色配置编辑器'); dlg.geometry('620x620')
        dlg.minsize(500, 500); dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        tk.Label(dlg, text='👥 人员角色配置', font=('Microsoft YaHei', 16, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=16)
        tk.Label(dlg, text='修改角色 / 添加人员 / 删除人员 · 保存后即时生效',
                 font=('Microsoft YaHei', 9), bg=C['bg'],
                 fg=C['text3']).pack(pady=0)

        # ---- 新增人员栏 ----
        add_bar = tk.Frame(dlg, bg=C['bg'])
        add_bar.pack(fill='x', padx=20, pady=0)
        tk.Label(add_bar, text='新增:', font=('Microsoft YaHei', 10),
                 bg=C['bg'], fg=C['text']).pack(side='left', padx=0)
        add_name = tk.Entry(add_bar, font=('Microsoft YaHei', 10), width=12)
        add_name.pack(side='left', padx=0)
        add_role_var = tk.StringVar(value='一卡剪辑')
        ttk.Combobox(add_bar, textvariable=add_role_var, values=ROLES,
                     state='readonly', font=('Microsoft YaHei', 10), width=10).pack(side='left', padx=0)
        rm_ref = self.cfg['人员角色']
        order_ref = self.cfg.get('人员排序', list(rm_ref.keys()))

        # 刷新函数——重建整个列表
        def rebuild_person_list():
            for w in sf.winfo_children():
                w.destroy()
            for i, name in enumerate(order_ref):
                if name not in rm_ref:
                    continue
                row = tk.Frame(sf, bg='white', highlightthickness=1,
                               highlightbackground=C['border'])
                row.grid(row=i, column=0, padx=5, pady=2, sticky='ew')
                # 删除按钮
                btn_del = tk.Button(row, text='✕', font=('Microsoft YaHei', 9, 'bold'),
                                    bg='#fee2e2', fg='#dc2626', relief='flat',
                                    cursor='hand2', padx=6, pady=3,
                                    activebackground='#fecaca',
                                    command=lambda n=name: _delete_person(n))
                btn_del.pack(side='left', padx=6, pady=5)
                # 姓名
                tk.Label(row, text=name, font=('Microsoft YaHei', 10, 'bold'),
                         bg='white', fg=C['text'], width=8, anchor='w').pack(side='left', padx=2, pady=5)
                # 角色下拉
                current = rm_ref.get(name, '一卡剪辑')
                var = tk.StringVar(value=current)
                role_vars[name] = var
                cb = ttk.Combobox(row, textvariable=var, values=ROLES,
                                  state='readonly', font=('Microsoft YaHei', 10), width=10)
                cb.pack(side='left', padx=0, pady=5)
                # 颜色点
                color_map = {'一卡剪辑': '#27ae60', '二卡剪辑': '#3b82f6',
                             '剪辑助理': '#8b5cf6', '剪辑组长': '#f59e0b'}
                dot = tk.Label(row, text='●', font=('Microsoft YaHei', 14),
                               fg=color_map.get(current, '#333'), bg='white')
                dot.pack(side='left', pady=5)
                def _on_change(*_, lbl=dot, v=var):
                    lbl.configure(fg=color_map.get(v.get(), '#333'))
                var.trace_add('write', _on_change)
            sf.grid_columnconfigure(0, weight=1)

        def _delete_person(name):
            if not messagebox.askyesno('确认删除', f'确定要删除人员 "{name}" 吗？\n\n此操作不可恢复。'):
                return
            rm_ref.pop(name, None)
            if name in order_ref:
                order_ref.remove(name)
            role_vars.pop(name, None)
            self.cfg['人员排序'] = order_ref
            rebuild_person_list()

        def _add_person():
            name = add_name.get().strip()
            if not name:
                messagebox.showwarning('提示', '请输入人员姓名'); return
            if name in rm_ref:
                messagebox.showwarning('提示', f'人员 "{name}" 已存在'); return
            rm_ref[name] = add_role_var.get()
            order_ref.append(name)
            self.cfg['人员排序'] = order_ref
            add_name.delete(0, 'end')
            rebuild_person_list()
            # 滚到底部
            canvas.yview_moveto(1.0)

        tk.Button(add_bar, text='＋ 添加', font=('Microsoft YaHei', 10, 'bold'),
                  bg=C['blue'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=3, command=_add_person).pack(side='left', padx=4)
        # 绑定回车
        add_name.bind('<Return>', lambda e: _add_person())

        # ---- 可滚动人员列表 ----
        canvas = tk.Canvas(dlg, bg=C['bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(dlg, orient='vertical', command=canvas.yview)
        sf = tk.Frame(canvas, bg=C['bg'])
        sf.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((10, 0), window=sf, anchor='nw', width=570)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side='left', fill='both', expand=True, padx=20, pady=0)
        scrollbar.pack(side='right', fill='y', pady=0)

        role_vars = {}
        rebuild_person_list()

        # ---- 底部按钮 ----
        bf = tk.Frame(dlg, bg=C['bg'])
        bf.pack(fill='x', padx=20, pady=0)

        def save():
            for n, v in role_vars.items():
                if n in rm_ref:
                    rm_ref[n] = v.get()
            self.cfg['人员角色'] = rm_ref
            self.cfg['人员排序'] = order_ref

            # 同步更新小组中的成员列表
            all_names = set(rm_ref.keys())
            groups = self.cfg.get('小组', {})
            for gname, ginfo in groups.items():
                members = ginfo.get('成员', [])
                new_members = [m for m in members if m in all_names]
                ginfo['成员'] = new_members
                leader = ginfo.get('组长', '')
                if leader not in all_names:
                    ginfo['组长'] = new_members[0] if new_members else ''

            self._save_config()
            self._refresh_role_tags()
            self._log('✅ 角色配置已更新并保存')
            dlg.destroy()
            messagebox.showinfo('保存成功', '角色配置已保存！\n\n下次生成时生效。')

        tk.Button(bf, text='💾 保存配置', font=('Microsoft YaHei', 12, 'bold'),
                  bg=C['green'], fg='white', relief='flat', cursor='hand2',
                  padx=30, pady=8, command=save).pack(side='right', padx=10)
        tk.Button(bf, text='取消', font=('Microsoft YaHei', 11),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=24, pady=8, command=dlg.destroy).pack(side='right')
        canvas.bind_all('<MouseWheel>', lambda e: canvas.yview_scroll(int(-e.delta/120), 'units'))
        dlg.protocol('WM_DELETE_WINDOW', lambda: [canvas.unbind_all('<MouseWheel>'), dlg.destroy()])

    # ============ 文件选择 ============

    def _select_project(self):
        path = filedialog.askopenfilename(
            title='选择项目数据文件',
            filetypes=[('Excel文件', '*.xlsx'), ('所有文件', '*.*')],
            initialdir=SCRIPT_DIR)
        if path:
            self.project_file = path
            self._save_app_settings()
            self._refresh_file_labels()
            self.check_files()

    def _select_template(self):
        path = filedialog.askopenfilename(
            title='选择模板文件',
            filetypes=[('Excel文件', '*.xlsx'), ('所有文件', '*.*')],
            initialdir=SCRIPT_DIR)
        if path:
            self.template_file = path
            self._save_app_settings()
            self._refresh_file_labels()
            self.check_files()

    def _select_output_dir(self):
        path = filedialog.askdirectory(
            title='选择输出目录',
            initialdir=self.output_dir)
        if path:
            self.output_dir = path
            self._save_app_settings()
            self._refresh_file_labels()
            self.check_files()

    def _refresh_file_labels(self):
        def _short(fpath):
            if not fpath: return '（未选择）'
            name = os.path.basename(fpath)
            exists = os.path.exists(fpath)
            icon = '✅' if exists else '❌'
            return f'{icon} {name}'
        self.pf_label.configure(text=_short(self.project_file))
        self.tf_label.configure(text=_short(self.template_file))
        self.od_label.configure(text=f'📁 {self.output_dir}')

    # ============ 功能：生成下月模板 ============
    def _gen_next_template(self):
        if not self._require_features(): return
        try:
            self._log('📅 正在生成下月模板...')
            self._log(f'   源模板: {os.path.basename(self.template_file)}')
            path, msg = generate_next_month_template(self.template_file, self.output_dir)
            if path:
                self._log(f'✅ {msg}')
                self._log(f'   文件: {os.path.basename(path)}')
                self._open(path)
                messagebox.showinfo('完成', f'{msg}\n\n文件已自动打开。')
            else:
                self._log(f'❌ {msg}')
                messagebox.showerror('失败', msg)
        except Exception as e:
            self._log(f'❌ 生成下月模板失败: {e}')
            messagebox.showerror('失败', f'生成失败:\n{e}')

    # ============ 功能：导出PDF ============
    def _export_pdf(self):
        if not self._require_features(): return
        path = filedialog.askopenfilename(
            title='选择要导出的Excel提成表',
            filetypes=[('Excel文件', '*.xlsx')],
            initialdir=SCRIPT_DIR)
        if not path:
            return
        self._log(f'📤 正在导出PDF: {os.path.basename(path)}')
        try:
            pdf_path = export_to_pdf(path)
            self._log(f'✅ PDF已生成: {os.path.basename(pdf_path)}')
            self._open(pdf_path)
            messagebox.showinfo('完成', f'PDF导出成功！\n\n📄 {os.path.basename(pdf_path)}')
        except Exception as e:
            self._log(f'❌ PDF导出失败: {e}')
            messagebox.showerror('失败', f'PDF导出失败:\n{e}')

    # ============ 功能：个人绩效卡片 ============
    def _gen_cards(self):
        if not self._require_features(): return
        self._log('🃏 正在生成个人绩效卡片...')
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()

            df = pd.read_excel(self.project_file, header=None)
            records, group_pids = gc.parse_projects(df)
            if not records:
                self._log('⚠️ 无数据，请检查项目文件')
                _sys.path.pop(0)
                return
            cd = gc.compute_commission(records, group_pids)
            card_paths = generate_person_cards(records, cd, CARDS_DIR)
            _sys.path.pop(0)

            if card_paths:
                self._log(f'✅ 已生成 {len(card_paths)-1} 张个人绩效卡片 -> 个人绩效卡片/')
                self._open(card_paths[0])
            else:
                self._log('⚠️ 未能生成卡片')
        except Exception as e:
            self._log(f'❌ 卡片生成失败: {e}')

    # ============ 功能：多月份对比 ============
    def _compare_months(self):
        if not self._require_features(): return
        f1 = filedialog.askopenfilename(
            title='选择第一个月份的提成表',
            filetypes=[('Excel文件', '*.xlsx')],
            initialdir=SCRIPT_DIR)
        if not f1: return
        f2 = filedialog.askopenfilename(
            title='选择第二个月份的提成表',
            filetypes=[('Excel文件', '*.xlsx')],
            initialdir=os.path.dirname(f1))
        if not f2: return
        self._log(f'📊 正在对比两个月份的提成表...')
        try:
            label1, label2, diffs = compare_months(f1, f2)
            if not diffs:
                self._log(f'✅ 两个月份数据完全一致，无差异。')
                messagebox.showinfo('对比结果', f'{label1} ↔ {label2}\n\n数据完全一致，无差异。')
                return
            self._log(f'{label1} ↔ {label2} 对比: {len(diffs)}人有变化')
            msg = f'{label1}  →  {label2}\n{"="*60}\n'
            msg += f'{"姓名":　<6s} {"上月集数":>6s} {"本月集数":>6s} {"集数变化":>6s} {"上月提成":>8s} {"本月提成":>8s} {"提成变化":>8s}\n'
            for nm, e1, e2, de, c1, c2, dc in diffs:
                de_str = f'+{de}' if de > 0 else str(de)
                dc_str = f'+{dc}' if dc > 0 else str(dc)
                msg += f'{nm:　<6s} {e1:>6d} {e2:>6d} {de_str:>6s} {c1:>8d} {c2:>8d} {dc_str:>8s}\n'
            self._log(msg)
            # 弹窗显示
            dlg = tk.Toplevel(self.root)
            dlg.title('多月份对比')
            dlg.geometry('700x500')
            dlg.configure(bg=C['bg'])
            tk.Label(dlg, text=f'{label1}  ↔  {label2}', font=('Microsoft YaHei', 14, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(pady=10)
            txt = tk.Text(dlg, font=('Microsoft YaHei', 10), bg=C['log_bg'], fg=C['log_fg'],
                          relief='flat', padx=12, pady=10)
            txt.insert('1.0', msg)
            txt.configure(state='disabled')
            txt.pack(fill='both', expand=True, padx=20, pady=0)
        except Exception as e:
            self._log(f'❌ 对比失败: {e}')
            messagebox.showerror('失败', f'对比失败:\n{e}')

    # ============ 功能：提成规则面板 ============
    def _edit_rules(self):
        dlg = tk.Toplevel(self.root)
        dlg.title('提成规则设置')
        dlg.geometry('500x420')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        tk.Label(dlg, text='🏷️ 提成规则配置', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=20)
        tk.Label(dlg, text='修改后点击保存，下次生成生效', font=('Microsoft YaHei', 9),
                 bg=C['bg'], fg=C['text3']).pack()

        rules = self.cfg.get('rules', {})
        vars_map = {}
        for role_name in ['一卡剪辑', '二卡剪辑', '剪辑助理']:
            frm = tk.Frame(dlg, bg=C['card'], highlightthickness=1,
                           highlightbackground=C['border'])
            frm.pack(fill='x', padx=20, pady=4)
            tk.Label(frm, text=f'{role_name}', font=('Microsoft YaHei', 11, 'bold'),
                     bg=C['card'], fg=C['text'], width=12, anchor='w').pack(side='left', padx=10, pady=8)

            r = rules.get(role_name, {})
            for key, label in [('基准集数', '基准'), ('超额每集', '超额/集'), ('缺集每集扣', '缺扣/集')]:
                v = tk.IntVar(value=r.get(key, 70 if role_name == '一卡剪辑' else 120))
                vars_map[f'{role_name}|{key}'] = v
                sub = tk.Frame(frm, bg=C['card'])
                sub.pack(side='left', padx=4, pady=4)
                tk.Label(sub, text=label, font=('Microsoft YaHei', 8), bg=C['card'],
                         fg=C['text3']).pack()
                s = tk.Scale(sub, from_=0, to=300 if '集数' in key else 100,
                             orient='horizontal', variable=v, length=70,
                             bg=C['card'], fg=C['text'], highlightthickness=0)
                s.pack()

        # 组长规则
        frm_l = tk.Frame(dlg, bg=C['card'], highlightthickness=1,
                         highlightbackground=C['border'])
        frm_l.pack(fill='x', padx=20, pady=4)
        tk.Label(frm_l, text='剪辑组长', font=('Microsoft YaHei', 11, 'bold'),
                 bg=C['card'], fg=C['text'], width=12, anchor='w').pack(side='left', padx=10, pady=8)
        r_l = rules.get('剪辑组长', {})
        for key, label in [('每集单价', '单价/集'), ('组内每部提成', '每部奖')]:
            v = tk.IntVar(value=r_l.get(key, 20 if '单价' in key else 100))
            vars_map[f'剪辑组长|{key}'] = v
            sub = tk.Frame(frm_l, bg=C['card'])
            sub.pack(side='left', padx=4, pady=4)
            tk.Label(sub, text=label, font=('Microsoft YaHei', 8), bg=C['card'],
                     fg=C['text3']).pack()
            s = tk.Scale(sub, from_=0, to=300, orient='horizontal', variable=v,
                         length=70, bg=C['card'], fg=C['text'], highlightthickness=0)
            s.pack()

        def save_rules():
            for key, v in vars_map.items():
                role, field = key.split('|', 1)
                rules.setdefault(role, {})[field] = v.get()
            self._save_config()
            self._log('✅ 提成规则已更新并保存')
            dlg.destroy()
            messagebox.showinfo('保存成功', '提成规则已保存！\n\n下次生成时生效。')

        tk.Button(dlg, text='💾 保存规则', font=('Microsoft YaHei', 12, 'bold'),
                  bg=C['green'], fg='white', relief='flat', cursor='hand2',
                  padx=30, pady=8, command=save_rules).pack(pady=16)

    # ============ 功能：数据预览 ============
    def _preview_data(self):
        self._log('📋 正在加载数据预览...')
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()

            self._log(f'   项目文件: {os.path.basename(self.project_file)}')
            self._log(f'   人员数量: {len(gc.ALL_NAMES)}')
            df = pd.read_excel(self.project_file, header=None)
            self._log(f'   Excel行数: {len(df)}')
            records, group_pids = gc.parse_projects(df)
            self._log(f'   解析记录: {len(records)}')
            cd = gc.compute_commission(records, group_pids)
            preview = data_preview(records, cd)
            self._log(f'   预览行数: {len(preview)}')
            _sys.path.pop(0)

            if not preview:
                self._log('⚠️ 无数据可预览（可能项目文件名不在config人员名单中）')
                messagebox.showwarning('无数据', '解析到0条记录。\n\n请确认项目数据文件中的人员姓名与角色配置中的一致。')
                return

            # 弹窗展示
            dlg = tk.Toplevel(self.root)
            dlg.title('数据预览')
            dlg.geometry('750x520')
            dlg.configure(bg=C['bg'])
            tk.Label(dlg, text='📋 数据预览', font=('Microsoft YaHei', 15, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(pady=12)

            # Treeview
            cols = ('姓名', '角色', '集数', '项目数', '基准', '绩效', '提成')
            tree = ttk.Treeview(dlg, columns=cols, show='headings', height=18)
            widths = [80, 90, 60, 60, 60, 60, 80]
            for c, w in zip(cols, widths):
                tree.heading(c, text=c)
                tree.column(c, width=w, anchor='center')
            tree.pack(fill='both', expand=True, padx=20, pady=0)

            # 颜色标签
            tree.tag_configure('ok', foreground='#16a34a')
            tree.tag_configure('fail', foreground='#dc2626')
            for p in preview:
                tag = 'ok' if p['status'] == '是' else 'fail'
                tree.insert('', 'end', values=(
                    p['name'], p['role'], p['episodes'], p['projects'],
                    f'{p["quota"]}集' if p['quota'] > 0 else '无', p['status'],
                    f'{p["commission"]:,}'), tags=(tag,))
        except Exception as e:
            self._log(f'❌ 预览失败: {e}')
            import traceback; self._log(traceback.format_exc())

    # ============ 功能：组内排名 ============
    def _gen_ranking(self):
        self._log('🏆 正在生成组内排名...')
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()

            df = pd.read_excel(self.project_file, header=None)
            records, group_pids = gc.parse_projects(df)
            cd = gc.compute_commission(records, group_pids)
            _sys.path.pop(0)

            path = os.path.join(self.output_dir, '组内排名.html')
            generate_ranking_html(cd, path)
            self._log(f'✅ 组内排名已生成: {os.path.basename(path)}')
            self._open(path)
        except Exception as e:
            self._log(f'❌ 排名生成失败: {e}')

    # ============ 功能：项目管理视图 ============
    def _gen_project_mgmt(self):
        self._log('🗂️ 正在生成项目管理视图...')
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()

            df = pd.read_excel(self.project_file, header=None)
            records, _ = gc.parse_projects(df)
            _sys.path.pop(0)

            path = os.path.join(self.output_dir, '项目管理_项目清单.html')
            generate_project_management_html(records, path)
            self._log(f'✅ 项目管理视图已生成: {os.path.basename(path)}')
            self._open(path)
        except Exception as e:
            self._log(f'❌ 项目管理视图生成失败: {e}')

    # ============ 功能：文件监控 ============
    def _toggle_watch(self):
        if hasattr(self, '_watcher') and self._watcher:
            self._stop_watch()
            return
        self._watcher = True
        self._watch_mtime = os.path.getmtime(self.project_file)
        btn = getattr(self, 'btn_watch', None)
        if btn:
            btn.configure(text='🔄 监控中...', bg='#059669')
        self._log('🔄 文件监控已开启，检测到项目文件变化将自动提示...')
        self._watch_loop()

    def _watch_loop(self):
        if not getattr(self, '_watcher', False):
            return
        try:
            current_mtime = os.path.getmtime(self.project_file)
            if current_mtime != self._watch_mtime:
                self._watch_mtime = current_mtime
                self._log('🔔 检测到项目数据文件已更新！')
                if messagebox.askyesno('文件更新', '项目数据文件已更新，是否立即重新生成？'):
                    self.run()
        except Exception:
            pass
        if getattr(self, '_watcher', False):
            self.root.after(5000, self._watch_loop)

    def _stop_watch(self):
        self._watcher = False
        btn = getattr(self, 'btn_watch', None)
        if btn:
            btn.configure(text='🔄 开启监控', bg='#059669')
        self._log('🔴 文件监控已关闭')

    # ============ 功能：智能分集 ============
    def _smart_assign(self):
        try:
            self._smart_assign_impl()
        except Exception as e:
            import traceback
            msg = traceback.format_exc()
            # 同时写文件，方便完整查看
            with open(os.path.join(SCRIPT_DIR, '_smart_assign_error.txt'), 'w', encoding='utf-8') as f:
                f.write(msg)
            # 只显示前 500 字符
            short = msg[:600] + ('...' if len(msg) > 600 else '')
            messagebox.showerror('智能分集出错', f'完整错误已写入 _smart_assign_error.txt\n\n{short}')

    def _smart_assign_impl(self):
        if not self._require_features(): return
        dlg = tk.Toplevel(self.root)
        dlg.title('📐 智能分集')
        dlg.geometry('1180x760')
        dlg.minsize(960, 600)
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        # ===== 顶部标题栏 =====
        title_bar = tk.Frame(dlg, bg=C['hdr_bg'])
        title_bar.pack(fill='x')
        tk.Label(title_bar, text='📐 智能分集工具', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['hdr_bg'], fg=C['hdr_text'], padx=18, pady=8).pack(side='left')

        # ---- 顶部操作按钮（一直可见）----
        btn_bar = tk.Frame(dlg, bg=C['bg'])
        btn_bar.pack(fill='x', padx=12, pady=(8, 4))

        def _mk_btn(parent, text, color, hover, cmd, size=13, bold=True):
            return tk.Button(parent, text=text, font=('Microsoft YaHei', size, 'bold' if bold else 'normal'),
                             bg=color, fg='white', relief='flat', cursor='hand2',
                             padx=18, pady=8, activebackground=hover, command=cmd)

        # 主按钮先占位，稍后绑定
        btn_assign = _mk_btn(btn_bar, '🎲 随机分集', '#e11d48', '#be123c', lambda: None)
        btn_assign.pack(side='left', padx=3, fill='x', expand=True)
        btn_reroll = _mk_btn(btn_bar, '🔄 再次随机', '#d97706', '#b45309', lambda: None, 12, False)
        btn_reroll.pack(side='left', padx=3, fill='x', expand=True)
        btn_restore = _mk_btn(btn_bar, '↩️ 恢复上次', '#7c3aed', '#6d28d9', lambda: None, 12, False)
        btn_restore.pack(side='left', padx=3, fill='x', expand=True)
        btn_confirm = _mk_btn(btn_bar, '✅ 确认入库', '#16a34a', '#15803d', lambda: None, 12, False)
        btn_confirm.pack(side='left', padx=3, fill='x', expand=True)
        btn_confirm.configure(state='disabled')

        # ===== 左右分栏 =====
        paned = tk.PanedWindow(dlg, orient='horizontal', bg=C['border'], sashwidth=3)
        paned.pack(fill='both', expand=True, padx=0, pady=0)

        # ---- 左栏：设置区（可滚动）----
        left = tk.Frame(paned, bg=C['bg'], width=430)
        paned.add(left, minsize=380, width=430)

        scroll_container = tk.Frame(left, bg=C['bg'])
        scroll_container.pack(fill='both', expand=True)
        l_canvas = tk.Canvas(scroll_container, bg=C['bg'], highlightthickness=0)
        l_scroll = tk.Scrollbar(scroll_container, orient='vertical', command=l_canvas.yview)
        l_canvas.configure(yscrollcommand=l_scroll.set)
        l_scroll.pack(side='right', fill='y')
        l_canvas.pack(side='left', fill='both', expand=True)
        l_inner = tk.Frame(l_canvas, bg=C['bg'])
        l_inner.bind('<Configure>', lambda e: l_canvas.configure(scrollregion=l_canvas.bbox('all')))
        l_win = l_canvas.create_window((0, 0), window=l_inner, anchor='nw')
        def _l_resize(event): l_canvas.itemconfig(l_win, width=event.width)
        l_canvas.bind('<Configure>', _l_resize)
        def _l_wheel(event): l_canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')
        l_canvas.bind('<Enter>', lambda e: l_canvas.bind_all('<MouseWheel>', _l_wheel))
        l_canvas.bind('<Leave>', lambda e: l_canvas.unbind_all('<MouseWheel>'))

        # ===== 记住上次设置 =====
        sa = self.cfg.setdefault('app_settings', {}).setdefault('smart_assign', {})
        last_name = sa.get('name', '')
        last_eps = sa.get('eps', 70)
        last_range = sa.get('range', 15)
        last_selected = set(sa.get('selected', []))
        self._last_sa = sa

        # ① 项目信息区（紧凑）
        c1 = tk.Frame(l_inner, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        c1.pack(fill='x', padx=10, pady=(8, 4))
        in1 = tk.Frame(c1, bg=C['card']); in1.pack(fill='x', padx=10, pady=8)
        tk.Label(in1, text='📋 项目信息', font=('Microsoft YaHei', 10, 'bold'),
                 bg=C['card'], fg=C['text']).pack(anchor='w', pady=(0, 4))

        # 项目名称行
        nm_row = tk.Frame(in1, bg=C['card']); nm_row.pack(fill='x')
        tk.Label(nm_row, text='名称', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).pack(side='left')
        name_var = tk.StringVar(value=last_name)
        tk.Entry(nm_row, textvariable=name_var, font=('Microsoft YaHei', 10),
                 relief='solid', borderwidth=1).pack(side='left', fill='x', expand=True, padx=6)
        def _browse_project_name():
            selected = filedialog.askdirectory(title='选择项目文件夹')
            if not selected: return
            name_var.set(os.path.basename(selected.rstrip('/\\')))
        tk.Button(nm_row, text='📁 浏览', font=('Microsoft YaHei', 9), bg=C['blue'], fg='white',
                  relief='flat', cursor='hand2', padx=8, pady=1, command=_browse_project_name).pack(side='left')

        # 集数 + 区间行
        param_row = tk.Frame(in1, bg=C['card']); param_row.pack(fill='x', pady=(6, 0))
        tk.Label(param_row, text='总集数', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).pack(side='left')
        eps_var = tk.IntVar(value=last_eps)
        tk.Spinbox(param_row, from_=1, to=2000, textvariable=eps_var, font=('Microsoft YaHei', 10),
                   width=6, relief='solid', borderwidth=1).pack(side='left', padx=(4, 14))
        tk.Label(param_row, text='一卡区间', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).pack(side='left')
        range_var = tk.IntVar(value=last_range)
        tk.Spinbox(param_row, from_=1, to=500, textvariable=range_var, font=('Microsoft YaHei', 10),
                   width=6, relief='solid', borderwidth=1).pack(side='left', padx=4)

        # 项目模板行
        proj_templates = self.cfg.get('project_templates', {})
        tpl_row = tk.Frame(in1, bg=C['card']); tpl_row.pack(fill='x', pady=(6, 0))
        tk.Label(tpl_row, text='模板', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).pack(side='left')
        tpl_var = tk.StringVar()
        tpl_names = list(proj_templates.keys())
        tpl_combo = ttk.Combobox(tpl_row, textvariable=tpl_var, values=['（自定义）'] + tpl_names,
                                 state='readonly', font=('Microsoft YaHei', 9), width=12)
        tpl_combo.pack(side='left', padx=(4, 6))
        tpl_combo.current(0)
        def _apply_tpl(_e=None):
            name = tpl_var.get()
            if name not in proj_templates:
                return
            t = proj_templates[name]
            if t.get('total'):
                try: eps_var.set(int(t['total']))
                except ValueError: pass
            if t.get('range'):
                try: range_var.set(int(t['range']))
                except ValueError: pass
        tpl_combo.bind('<<ComboboxSelected>>', _apply_tpl)
        def _manage_tpl():
            self._manage_project_templates(dlg, proj_templates, tpl_combo, tpl_names)
        tk.Button(tpl_row, text='🗂 管理', font=('Microsoft YaHei', 8), bg=C['purple'], fg='white',
                  relief='flat', cursor='hand2', padx=8, pady=1, command=_manage_tpl).pack(side='left')

        # ② 人员选择区
        c2 = tk.Frame(l_inner, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        c2.pack(fill='x', padx=10, pady=4)
        in2 = tk.Frame(c2, bg=C['card']); in2.pack(fill='x', padx=10, pady=8)
        hdr_row = tk.Frame(in2, bg=C['card']); hdr_row.pack(fill='x', pady=(0, 4))
        tk.Label(hdr_row, text='👥 选择剪辑人员', font=('Microsoft YaHei', 10, 'bold'),
                 bg=C['card'], fg=C['text']).pack(side='left')

        # ---- 选人模版 ----
        templates = self.cfg.get('personnel_templates', {})
        template_var = tk.StringVar(value='')
        tmpl_row = tk.Frame(in2, bg=C['card']); tmpl_row.pack(fill='x', pady=(0, 6))
        def _refresh_template_combo():
            names = list(templates.keys())
            if not names: names = ['（暂无模版）']
            tpl_combo['values'] = names
            if template_var.get() not in names: template_var.set(names[0])
        tpl_combo = ttk.Combobox(tmpl_row, textvariable=template_var, values=[],
                                  state='readonly', font=('Microsoft YaHei', 9), width=14)
        tpl_combo.pack(side='left', padx=(0, 4))
        _refresh_template_combo()

        def _load_template():
            name = template_var.get()
            if not name or name.startswith('（'): return
            if name not in templates:
                messagebox.showwarning('提示', f'模版"{name}"不存在', parent=dlg); return
            for v in check_vars.values(): v.set(False)
            for nm in templates[name]:
                if nm in check_vars: check_vars[nm].set(True)
        def _save_template():
            selected = [n for n, v in check_vars.items() if v.get()]
            if not selected:
                messagebox.showwarning('提示', '请先勾选至少一位剪辑人员', parent=dlg); return
            from tkinter import simpledialog
            name = simpledialog.askstring('保存模版', '请输入模版名称：', parent=dlg)
            if not name: return
            name = name.strip()
            if not name: return
            if name in templates:
                if not messagebox.askyesno('确认覆盖', f'模版"{name}"已存在，是否覆盖？', parent=dlg): return
            templates[name] = selected
            self.cfg['personnel_templates'] = templates
            self._save_config()
            _refresh_template_combo(); template_var.set(name)
            self._log(f'💾 选人模版"{name}"已保存 ({len(selected)}人)')
        def _delete_template():
            name = template_var.get()
            if not name or name.startswith('（') or name not in templates: return
            if messagebox.askyesno('确认删除', f'确定删除模版"{name}"？', parent=dlg):
                del templates[name]
                self.cfg['personnel_templates'] = templates
                self._save_config()
                _refresh_template_combo()
                self._log(f'🗑 选人模版"{name}"已删除')

        tk.Button(tmpl_row, text='📥 加载', font=('Microsoft YaHei', 8), bg=C['blue'], fg='white', relief='flat',
                  cursor='hand2', padx=8, pady=1, command=_load_template).pack(side='left', padx=2)
        tk.Button(tmpl_row, text='💾 保存', font=('Microsoft YaHei', 8), bg=C['green'], fg='white', relief='flat',
                  cursor='hand2', padx=8, pady=1, command=_save_template).pack(side='left', padx=2)
        tk.Button(tmpl_row, text='🗑 删除', font=('Microsoft YaHei', 8), bg=C['red'], fg='white', relief='flat',
                  cursor='hand2', padx=8, pady=1, command=_delete_template).pack(side='left', padx=2)

        # ---- 快捷选择 ----
        sel_frame = tk.Frame(in2, bg=C['card']); sel_frame.pack(fill='x', pady=(0, 4))
        def _mk_quick(text, color, fn):
            b = tk.Button(sel_frame, text=text, font=('Microsoft YaHei', 9), bg=color, fg='white',
                          relief='flat', cursor='hand2', padx=10, pady=2, command=fn)
            b.pack(side='left', padx=3)
        def _select_all():
            for v in check_vars.values(): v.set(True)
        def _select_none():
            for v in check_vars.values(): v.set(False)
        def _select_card1():
            for n, v in check_vars.items():
                v.set('一卡' in roles_map.get(n, '') or '小组长' in roles_map.get(n, ''))
        def _select_card2():
            for n, v in check_vars.items():
                v.set('二卡' in roles_map.get(n, '') or '助理' in roles_map.get(n, '') or ('组长' in roles_map.get(n, '') and '小组长' not in roles_map.get(n, '')))
        _mk_quick('全选', C['blue'], _select_all)
        _mk_quick('全不选', C['gray'], _select_none)
        _mk_quick('小组长+一卡', '#eab308', _select_card1)
        _mk_quick('二卡/助理/组长', '#d97706', _select_card2)

        check_vars = {}
        roles_map = self.cfg.get('人员角色', {})
        role_order = {'小组长': 0, '一卡剪辑': 1, '二卡剪辑': 2, '剪辑助理': 3, '剪辑组长': 4}
        sorted_people = sorted(roles_map.keys(), key=lambda n: role_order.get(roles_map[n], 99))
        role_labels = [
            ('小组长', '🟡', lambda r: '小组长' in r),
            ('一卡剪辑', '🟢', lambda r: '一卡' in r and '小组长' not in r),
            ('二卡剪辑', '🔵', lambda r: '二卡' in r),
            ('剪辑助理', '🟣', lambda r: '助理' in r),
            ('剪辑组长', '🟠', lambda r: '组长' in r and '小组长' not in r),
        ]
        cb_frame = tk.Frame(in2, bg=C['card']); cb_frame.pack(fill='x')
        for role_name, icon, matcher in role_labels:
            people_in_role = [n for n in sorted_people if matcher(roles_map[n])]
            if not people_in_role: continue
            tk.Label(cb_frame, text=f'{icon} {role_name}', font=('Microsoft YaHei', 9, 'bold'),
                     bg=C['card'], fg=C['text']).pack(anchor='w', pady=(6, 2))
            grid = tk.Frame(cb_frame, bg=C['card']); grid.pack(fill='x')
            for col in range(2): grid.columnconfigure(col, weight=1)
            row, col = 0, 0
            for nm in people_in_role:
                v = tk.BooleanVar(value=nm in last_selected)
                check_vars[nm] = v
                tk.Checkbutton(grid, text=nm, variable=v, font=('Microsoft YaHei', 9),
                               bg=C['card'], fg=C['text'], selectcolor=C['card'],
                               activebackground=C['card'], anchor='w', padx=4, pady=1).grid(row=row, column=col, sticky='w')
                col += 1
                if col >= 2: col = 0; row += 1

        # 增强模板套用：按角色配比自动勾选人员
        def _apply_tpl_roles(_e=None):
            name = tpl_var.get()
            if name not in proj_templates:
                return
            t = proj_templates[name]
            roles_cfg = t.get('roles')
            if not roles_cfg:
                return
            # 先清空，再按配比勾选（每个角色勾选前 N 人）
            for v in check_vars.values():
                v.set(False)
            for rname, cnt in roles_cfg.items():
                matched = [nm for nm in check_vars if roles_map.get(nm, '') == rname
                           or (rname == '小组长' and '小组长' in roles_map.get(nm, ''))
                           or (rname == '一卡剪辑' and '一卡' in roles_map.get(nm, '') and '小组长' not in roles_map.get(nm, ''))]
                for nm in matched[:cnt]:
                    check_vars[nm].set(True)
        tpl_combo.bind('<<ComboboxSelected>>', lambda e: (_apply_tpl(e), _apply_tpl_roles(e)))

        # ---- 右栏：结果预览 + 编辑 ----
        right = tk.Frame(paned, bg=C['bg']); paned.add(right, minsize=360)

        c3 = tk.Frame(right, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        c3.pack(fill='both', expand=True, padx=(0, 10), pady=8)
        c3_header = tk.Frame(c3, bg=C['accent_l']); c3_header.pack(fill='x')
        tk.Label(c3_header, text='📋 分集结果（编辑区）', font=('Microsoft YaHei', 10, 'bold'),
                 bg=C['accent_l'], fg=C['text'], padx=12, pady=5).pack(side='left')
        tk.Button(c3_header, text='📋 复制结果', font=('Microsoft YaHei', 8), bg=C['blue'], fg='white',
                  relief='flat', cursor='hand2', padx=10, pady=2,
                  command=lambda: (dlg.clipboard_clear(), dlg.clipboard_append(result_text.get('1.0', 'end-1c')), None)
                  ).pack(side='right', padx=8, pady=3)

        # 结果预览文本（顶部）
        result_text = tk.Text(c3, font=('Microsoft YaHei', 10), bg=C['log_bg'], fg=C['log_fg'],
                              relief='flat', padx=10, pady=6, wrap='word', height=8)
        result_text.pack(fill='x', padx=8, pady=(8, 4))

        # 编辑区（可滚动）
        edit_container = tk.Frame(c3, bg=C['card'])
        edit_container.pack(fill='both', expand=True, padx=8, pady=(0, 8))
        ec = tk.Canvas(edit_container, bg=C['card'], highlightthickness=0)
        es = tk.Scrollbar(edit_container, orient='vertical', command=ec.yview)
        ec.configure(yscrollcommand=es.set)
        es.pack(side='right', fill='y'); ec.pack(side='left', fill='both', expand=True)
        eframe = tk.Frame(ec, bg=C['card'])
        eframe.bind('<Configure>', lambda e: ec.configure(scrollregion=ec.bbox('all')))
        ewin = ec.create_window((0, 0), window=eframe, anchor='nw')
        def _e_resize(event): ec.itemconfig(ewin, width=event.width)
        ec.bind('<Configure>', _e_resize)
        def _e_wheel(event): ec.yview_scroll(int(-1 * (event.delta / 120)), 'units')
        ec.bind('<Enter>', lambda e: ec.bind_all('<MouseWheel>', _e_wheel))
        ec.bind('<Leave>', lambda e: ec.unbind_all('<MouseWheel>'))

        hint = tk.Label(eframe, text='点击「🎲 随机分集」生成分配结果，\n然后在此逐人调整区间。',
                        font=('Microsoft YaHei', 10), bg=C['card'], fg=C['text3'], justify='center')
        hint.pack(pady=30)

        # 状态
        last_result = [None]
        last_selected_list = []

        def _display_result(result, project_name):
            result_text.delete('1.0', 'end')
            result_text.insert('end', f'项目: {project_name}  ({eps_var.get()}集)\n')
            result_text.insert('end', '=' * 55 + '\n')
            for nm in last_selected_list:
                fmt = result['formatted'].get(nm, '')
                if fmt: result_text.insert('end', f'{nm}: {fmt}\n')

        def _save_txt(proj_name, result):
            try:
                safe_name = re.sub(r'[\\/:*?"<>|]', '_', proj_name)
                txt_path = os.path.join(self.output_dir, f'{safe_name}_分集.txt')
                lines = [f'项目: {proj_name}  ({eps_var.get()}集)', '=' * 55, '']
                for nm in result['formatted']:
                    f = result['formatted'].get(nm, '')
                    if f: lines.append(f'{nm}: {f}')
                with open(txt_path, 'w', encoding='utf-8') as f: f.write('\n'.join(lines))
                self._log(f'📄 分集TXT已保存: {os.path.basename(txt_path)}')
            except Exception as e: self._log(f'⚠️ 保存TXT失败: {e}')

        def _append_to_project(proj_name, result, selected_people, total_eps_val):
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, Alignment
                from datetime import datetime, timedelta
                now = datetime.now(); end_date = now + timedelta(days=1)
                date_str = f'{end_date.month}.{end_date.day}下午18点交'
                proj_label = proj_name
                dir_label = f'O:\\AI漫剧剪辑一组\\{proj_name}'
                font_title = Font(name='宋体', size=14, bold=True)
                font_normal = Font(name='宋体', size=14)
                align_center = Alignment(horizontal='center', vertical='center')
                wb = load_workbook(self.project_file); ws = wb.active
                last = ws.max_row
                for r in range(last, 0, -1):
                    if ws.cell(r, 1).value or ws.cell(r, 3).value: last = r; break
                next_row = last + 2
                title_row = next_row
                c = ws.cell(next_row, 1, proj_label); c.font = font_title; c.alignment = align_center
                c = ws.cell(next_row, 2, dir_label); c.font = font_title; c.alignment = align_center
                c = ws.cell(next_row, 4, date_str); c.font = font_title; c.alignment = align_center
                c = ws.cell(next_row, 5, '已分集'); c.font = font_title; c.alignment = align_center
                next_row += 1
                data_start = next_row
                for nm in selected_people:
                    fmt = result['formatted'].get(nm, '')
                    if not fmt: continue
                    c = ws.cell(next_row, 3, f'{nm}：{fmt}'); c.font = font_normal; c.alignment = align_center
                    next_row += 1
                data_end = next_row - 1
                if data_end >= title_row:
                    for col in ['A', 'B', 'D', 'E']:
                        ws.merge_cells(f'{col}{title_row}:{col}{data_end}')
                wb.save(self.project_file); wb.close()
                self._log(f'📎 已追加到项目数据: {os.path.basename(self.project_file)}')
                self._open(self.project_file)
            except Exception as e:
                self._log(f'⚠️ 追加项目数据失败: {e}'); raise

        # ---- 编辑行管理 ----
        rows = []  # 每行 ri

        def _refresh_tags(ri):
            for w in ri['tag_container'].winfo_children(): w.destroy()
            ri['ranges'].sort()
            if not ri['ranges']:
                tk.Label(ri['tag_container'], text='（暂无区间）', font=('Microsoft YaHei', 9, 'italic'),
                         bg=C['card'], fg=C['text3']).pack(side='left'); return
            for s, e in ri['ranges']:
                label_text = f'{s}-{e}' if s != e else str(s)
                tag = tk.Frame(ri['tag_container'], bg=C['accent_l'], padx=0, pady=0)
                tag.pack(side='left', padx=2)
                tk.Label(tag, text=label_text, font=('Microsoft YaHei', 10), bg=C['accent_l'],
                         fg=C['text'], padx=6, pady=1).pack(side='left')
                x_btn = tk.Label(tag, text=' ×', font=('Microsoft YaHei', 9, 'bold'), bg=C['accent_l'],
                                 fg=C['red'], cursor='hand2', padx=1, pady=1)
                x_btn.pack(side='left')
                ri_cap, se_cap = ri, (s, e)
                x_btn.bind('<Button-1>', lambda e, r=ri_cap, se=se_cap: (
                    r['ranges'].remove(se) if se in r['ranges'] else None, _refresh_tags(r)))

        def _do_add(ri, total):
            try:
                s = int(ri['spin_from'].get()); e = int(ri['spin_to'].get())
            except (ValueError, TypeError): return
            if not (1 <= s <= total and 1 <= e <= total): return
            if s > e: s, e = e, s
            new_set = set(range(s, e + 1))
            surv = []
            for (a, b) in ri['ranges']:
                old_set = set(range(a, b + 1))
                if not (old_set & new_set): surv.append((a, b))
            ri['ranges'] = surv
            for other in rows:
                if other is ri: continue
                surv_other = []
                for (a, b) in other['ranges']:
                    old_set = set(range(a, b + 1))
                    if not (old_set & new_set): surv_other.append((a, b))
                    else:
                        kept = []
                        if a < s: kept.append((a, s - 1))
                        if b > e: kept.append((e + 1, b))
                        for (ka, kb) in kept:
                            kset = set(range(ka, kb + 1))
                            if not (kset & new_set): surv_other.append((ka, kb))
                other['ranges'] = surv_other; _refresh_tags(other)
            ri['ranges'].append((s, e)); _refresh_tags(ri)

        def _swap(idx_a, idx_b):
            a, b = rows[idx_a], rows[idx_b]
            a['name_var'].set(b['name_var'].get()); b['name_var'].set(a['name_var'].get())
            a['ranges'], b['ranges'] = b['ranges'], a['ranges']
            _refresh_tags(a); _refresh_tags(b)

        def _build_edit_rows(total):
            """基于 last_result 重建右侧编辑卡片"""
            for w in eframe.winfo_children(): w.destroy()
            rows.clear()
            result = last_result[0]
            for i, nm in enumerate(last_selected_list):
                card = tk.Frame(eframe, bg=C['card'], highlightthickness=1,
                                highlightbackground=C['border'], padx=8, pady=6)
                card.pack(fill='x', pady=3)
                top = tk.Frame(card, bg=C['card']); top.pack(fill='x')
                bcol = tk.Frame(top, bg=C['card']); bcol.pack(side='left', padx=(0, 4))
                nv = tk.StringVar(value=nm)
                cb = ttk.Combobox(top, textvariable=nv, values=last_selected_list,
                                  state='readonly', font=('Microsoft YaHei', 10), width=12)
                cb.pack(side='left', padx=4)
                role = roles_map.get(nm, '')
                role_colors = {'小组长': '#fef3c7', '一卡': '#d1fae5', '二卡': '#dbeafe', '助理': '#ede9fe'}
                rc = next((v for k, v in role_colors.items() if k in role), C['accent_l'])
                role_short = role.replace('剪辑', '') if role else ''
                if role_short:
                    tk.Label(top, text=role_short, font=('Microsoft YaHei', 8), bg=rc, fg=C['text2'],
                             padx=5, pady=1).pack(side='left', padx=4)
                tag_row = tk.Frame(card, bg=C['card']); tag_row.pack(fill='x', pady=(4, 0))
                tag_container = tk.Frame(tag_row, bg=C['card']); tag_container.pack(side='left')
                add_row = tk.Frame(card, bg=C['card']); add_row.pack(fill='x', pady=(4, 0))
                raw_ranges = result['assignments'].get(nm, [])
                ri = {'name_var': nv, 'ranges': list(raw_ranges), 'frame': card,
                      'tag_container': tag_container}
                _refresh_tags(ri)
                f = tk.Frame(add_row, bg=C['card']); f.pack(side='left')
                ri['spin_from'] = tk.IntVar(value=1); ri['spin_to'] = tk.IntVar(value=1)
                tk.Label(f, text='从', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).pack(side='left')
                sp_from = tk.Spinbox(f, from_=1, to=total, textvariable=ri['spin_from'], font=('Microsoft YaHei', 10),
                           width=5, relief='solid', borderwidth=1, justify='center')
                sp_from.pack(side='left', padx=3)
                ri['sp_from'] = sp_from
                tk.Label(f, text='到', font=('Microsoft YaHei', 9), bg=C['card'], fg=C['text2']).pack(side='left')
                sp_to = tk.Spinbox(f, from_=1, to=total, textvariable=ri['spin_to'], font=('Microsoft YaHei', 10),
                           width=5, relief='solid', borderwidth=1, justify='center')
                sp_to.pack(side='left', padx=3)
                ri['sp_to'] = sp_to

                # #4 区间冲突实时提示：输入变化时检查与其他行的重叠
                def _mk_check(r):
                    def _check(*_a):
                        try:
                            s = int(r['spin_from'].get()); e = int(r['spin_to'].get())
                        except (ValueError, TypeError):
                            return
                        if not (1 <= s <= total and 1 <= e <= total) or s > e:
                            for sp in (r.get('sp_from'), r.get('sp_to')):
                                if sp: sp.configure(highlightbackground='#e53e3e', highlightthickness=2)
                            return
                        # 检查是否与其他行的区间冲突
                        conflict = False
                        new_set = set(range(s, e + 1))
                        for other in rows:
                            if other is r: continue
                            for (a, b) in other['ranges']:
                                if set(range(a, b + 1)) & new_set:
                                    conflict = True; break
                            if conflict: break
                        color = '#e53e3e' if conflict else '#dce6e5'
                        for sp in (r.get('sp_from'), r.get('sp_to')):
                            if sp:
                                sp.configure(highlightbackground=color, highlightthickness=2 if conflict else 1)
                    return _check
                ri['spin_from'].trace_add('write', _mk_check(ri))
                ri['spin_to'].trace_add('write', _mk_check(ri))

                tk.Button(f, text='添加', font=('Microsoft YaHei', 9), bg=C['green'], fg='white', relief='flat',
                          cursor='hand2', padx=10, pady=1, command=lambda r=ri: _do_add(r, total)).pack(side='left', padx=4)
                rows.append(ri)
                idx = len(rows) - 1
                def _mk_up(i):
                    return lambda: _swap(i, i - 1) if i > 0 else None
                def _mk_down(i):
                    return lambda: _swap(i, i + 1) if i < len(rows) - 1 else None
                tk.Button(bcol, text='▲', font=('Microsoft YaHei', 9, 'bold'), bg=C['blue_l'], fg=C['blue'],
                          relief='flat', cursor='hand2', padx=5, pady=0, command=_mk_up(idx)).pack(side='top')
                tk.Button(bcol, text='▼', font=('Microsoft YaHei', 9, 'bold'), bg=C['blue_l'], fg=C['blue'],
                          relief='flat', cursor='hand2', padx=5, pady=0, command=_mk_down(idx)).pack(side='top')

        def _do_assign_inner():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning('提示', '请输入项目名称', parent=dlg); return
            selected = [n for n, v in check_vars.items() if v.get()]
            if not selected:
                messagebox.showwarning('提示', '请至少选择一位剪辑人员', parent=dlg); return
            total = eps_var.get(); rng = range_var.get()
            result = smart_episode_assignment(total, selected, roles_map, rng)
            last_result[0] = result
            last_selected_list[:] = selected
            _display_result(result, name)
            _build_edit_rows(total)
            btn_confirm.configure(state='normal')
            # 记住设置（含分集结果区间，供"恢复上次"使用）
            sa['name'] = name; sa['eps'] = total; sa['range'] = rng; sa['selected'] = selected
            sa['assignments'] = {k: [list(x) for x in v] for k, v in result['assignments'].items()}
            self.cfg.setdefault('app_settings', {})['smart_assign'] = sa
            self._save_config()
            self._log(f'🎲 已生成分集结果（{total}集 / {len(selected)}人）')

        def _do_assign_common():
            try: _do_assign_inner()
            except Exception as e:
                import traceback; msg = traceback.format_exc()
                messagebox.showerror('分集出错', msg, parent=dlg)

        def _restore_last():
            """#3 恢复上次分集结果（人员、总集数、区间）"""
            try:
                if not sa.get('assignments'):
                    messagebox.showinfo('恢复上次', '暂无上次分集记录。\n请先进行一次随机分集。', parent=dlg)
                    return
                # 恢复项目名/总集数/区间
                if sa.get('name'):
                    name_var.set(sa['name'])
                if sa.get('eps'):
                    eps_var.set(int(sa['eps']))
                if sa.get('range'):
                    range_var.set(int(sa['range']))
                # 恢复勾选人员
                saved_sel = sa.get('selected', [])
                for v in check_vars.values():
                    v.set(False)
                for nm in saved_sel:
                    if nm in check_vars:
                        check_vars[nm].set(True)
                # 恢复分集结果
                last_result[0] = {'assignments': {k: [tuple(x) for x in v] for k, v in sa['assignments'].items()},
                                   'formatted': {}, 'summary': {}}
                last_selected_list[:] = [n for n in saved_sel if n in check_vars]
                # 构造 formatted 用于显示
                fmt = {}
                for k, v in sa['assignments'].items():
                    parts = []
                    for s, e in v:
                        parts.append(f'{s}-{e}' if s != e else str(s))
                    fmt[k] = '，'.join(parts)
                last_result[0]['formatted'] = fmt
                _display_result(last_result[0], sa.get('name', ''))
                _build_edit_rows(eps_var.get())
                btn_confirm.configure(state='normal')
                self._log(f'↩️ 已恢复上次分集结果（{len(last_selected_list)}人 / {sa.get("eps","")}集）')
            except Exception as e:
                import traceback
                messagebox.showerror('恢复失败', f'{e}\n\n{traceback.format_exc()}', parent=dlg)

        def _confirm():
            try:
                lines = []; seen = {}
                for ri in rows:
                    nm = ri['name_var'].get().strip()
                    if not nm: continue
                    if nm in seen: raise ValueError(f'"{nm}" 出现多次，请合并为一行')
                    seen[nm] = True
                    parts = [f'{s}-{e}' if s != e else str(s) for s, e in sorted(ri['ranges'])]
                    if not parts: raise ValueError(f'"{nm}" 没有分配任何集数区间')
                    lines.append(f'{nm}：' + ', '.join(parts))
                if not lines:
                    messagebox.showwarning('提示', '没有有效的分配行', parent=dlg); return
                validated = validate_episode_assignments('\n'.join(lines), last_selected_list, eps_var.get())
                proj_name = name_var.get().strip()
                _display_result(validated, proj_name)
                _save_txt(proj_name, validated)
                _append_to_project(proj_name, validated, last_selected_list, eps_var.get())
                last_result[0] = validated
                btn_confirm.configure(state='disabled')
                self._log(f'✅ 分集结果已确认入库')
            except ValueError as e:
                messagebox.showerror('校验失败', str(e), parent=dlg)
            except Exception as e:
                import traceback; msg = f'{type(e).__name__}: {e}\n\n{traceback.format_exc()}'
                messagebox.showerror('确认失败', msg, parent=dlg)

        # 绑定主按钮
        btn_assign.configure(command=_do_assign_common)
        btn_reroll.configure(command=_do_assign_common)
        btn_restore.configure(command=_restore_last)
        btn_confirm.configure(command=_confirm)

    def _check_duplicates(self):
        self._log('🔍 正在检查项目去重...')
        try:
            import pandas as pd
            gc, _sys = self._load_gc_module()

            df = pd.read_excel(self.project_file, header=None)
            records, _ = gc.parse_projects(df)
            if not records:
                self._log('⚠️ 未解析到有效记录')
                _sys.path.pop(0)
                return

            ok = gc.self_check(records)
            id_names = {}
            for r in records:
                pid = r['项目ID']
                if pid:
                    if pid not in id_names:
                        id_names[pid] = set()
                    id_names[pid].add(r['AI项目名称'][:30])

            issues = {pid: names for pid, names in id_names.items() if len(names) > 1}
            if issues:
                msg = f'⚠️ 发现 {len(issues)} 个项目ID存在名称不一致:\n'
                for pid, names in list(issues.items())[:10]:
                    msg += f'  ID={pid}: {names}\n'
                self._log(msg)
                messagebox.showwarning('去重检查', msg)
            else:
                self._log(f'✅ 全部 {len(id_names)} 个项目ID名称一致，无重复。')
                messagebox.showinfo('去重检查', f'✅ 检查通过！\n\n{len(id_names)} 个项目ID，名称均一致，无重复。')
            _sys.path.pop(0)
        except Exception as e:
            self._log(f'❌ 检查失败: {e}')
            messagebox.showerror('失败', str(e))

    # ============ 新增功能：数据校验 ============
    def _validate_data(self):
        self._log('🔍 正在校验项目数据...')
        try:
            issues = validate_project_data(self.project_file, self.cfg.get('人员角色', {}))
            if not issues:
                self._log('✅ 数据校验通过，无问题！')
                messagebox.showinfo('校验通过', '✅ 项目数据校验通过！\n\n未发现日期异常、人名不匹配、或集数重复分配等问题。')
                return

            self._log(f'⚠️ 发现 {len(issues)} 个问题:')
            for loc, item, detail in issues[:20]:
                self._log(f'  [{loc}] {item}: {detail}')

            dlg = tk.Toplevel(self.root)
            dlg.title('数据校验结果')
            dlg.geometry('650x400')
            dlg.configure(bg=C['bg'])
            tk.Label(dlg, text=f'⚠️ 发现 {len(issues)} 个问题', font=('Microsoft YaHei', 14, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(pady=12)
            txt = tk.Text(dlg, font=('Microsoft YaHei', 10), bg=C['log_bg'], fg=C['log_fg'],
                          relief='flat', padx=10, pady=8)
            for loc, item, detail in issues[:50]:
                txt.insert('end', f'[{loc}] {item}: {detail}\n')
            txt.configure(state='disabled')
            txt.pack(fill='both', expand=True, padx=20, pady=0)
        except Exception as e:
            self._log(f'❌ 校验失败: {e}')

    # ============ 新增功能：个人趋势 ============
    def _gen_trend(self):
        name = self._trend_var.get()
        if not name:
            messagebox.showwarning('提示', '请选择人员'); return
        self._log(f'📊 正在生成 {name} 的月度趋势...')
        try:
            roles_map = self.cfg.get('人员角色', {})
            path = generate_person_trend_html(name, roles_map, self.output_dir)
            if path:
                self._log(f'✅ 趋势图已生成: {os.path.basename(path)}')
                self._open(path)
            else:
                self._log('⚠️ 未找到该人员的历史数据（可能需要先生成几个月的提成表）')
                messagebox.showinfo('提示', '未找到历史数据。\n\n请确保在输出目录中有多个以 "AI后期剪辑提成一组" 开头的 Excel 文件。')
        except Exception as e:
            self._log(f'❌ 趋势生成失败: {e}')

    # ============ 功能4：高级查询与筛选 ============
    def _advanced_filter(self):
        if not self._require_features(): return
        dlg = tk.Toplevel(self.root)
        dlg.title('🔎 高级查询与筛选')
        dlg.geometry('850x650')
        dlg.minsize(700, 500)
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        tk.Label(dlg, text='🔎 高级查询与筛选', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=14)
        tk.Label(dlg, text='按条件组合筛选，支持导出筛选结果', font=('Microsoft YaHei', 9),
                 bg=C['bg'], fg=C['text3']).pack(pady=0)

        # 筛选条件区
        cf = tk.Frame(dlg, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        cf.pack(fill='x', padx=16, pady=0)
        cfi = tk.Frame(cf, bg=C['card']); cfi.pack(fill='x', padx=12, pady=8)

        # 第一行：姓名 + 角色
        r1 = tk.Frame(cfi, bg=C['card']); r1.pack(fill='x', pady=2)
        tk.Label(r1, text='姓名:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=8, anchor='w').pack(side='left')
        name_entry = tk.Entry(r1, font=('Microsoft YaHei', 10), width=14)
        name_entry.pack(side='left', padx=0)

        tk.Label(r1, text='角色:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=8, anchor='w').pack(side='left')
        role_vars = {}
        for role in ROLES:
            v = tk.BooleanVar(value=True)
            role_vars[role] = v
            tk.Checkbutton(r1, text=role, variable=v, font=('Microsoft YaHei', 8),
                           bg=C['card'], fg=C['text'], selectcolor=C['card'],
                           activebackground=C['card']).pack(side='left', padx=0)

        # 第二行：集数范围 + 项目ID
        r2 = tk.Frame(cfi, bg=C['card']); r2.pack(fill='x', pady=2)
        tk.Label(r2, text='集数范围:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=8, anchor='w').pack(side='left')
        eps_min = tk.Entry(r2, font=('Microsoft YaHei', 10), width=6)
        eps_min.pack(side='left')
        tk.Label(r2, text='~', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text']).pack(side='left', padx=2)
        eps_max = tk.Entry(r2, font=('Microsoft YaHei', 10), width=6)
        eps_max.pack(side='left', padx=0)

        tk.Label(r2, text='项目ID:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=8, anchor='w').pack(side='left')
        pid_entry = tk.Entry(r2, font=('Microsoft YaHei', 10), width=10)
        pid_entry.pack(side='left', padx=0)

        tk.Label(r2, text='绩效:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=6, anchor='w').pack(side='left')
        status_var = tk.StringVar(value='全部')
        ttk.Combobox(r2, textvariable=status_var, values=['全部', '是', '否'],
                     state='readonly', font=('Microsoft YaHei', 10), width=6).pack(side='left', padx=0)

        tk.Label(r2, text='提成范围:', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text'], width=8, anchor='w').pack(side='left')
        comm_min = tk.Entry(r2, font=('Microsoft YaHei', 10), width=7)
        comm_min.pack(side='left')
        tk.Label(r2, text='~', font=('Microsoft YaHei', 9), bg=C['card'],
                 fg=C['text']).pack(side='left', padx=2)
        comm_max = tk.Entry(r2, font=('Microsoft YaHei', 10), width=7)
        comm_max.pack(side='left')

        # 按钮
        btn_f = tk.Frame(cfi, bg=C['card']); btn_f.pack(fill='x', pady=8)

        # 结果区
        result_frame = tk.Frame(dlg, bg=C['bg'])
        result_frame.pack(fill='both', expand=True, padx=16, pady=0)

        cols = ('姓名', '角色', '集数', '项目数', '基准', '绩效', '提成')
        tree = ttk.Treeview(result_frame, columns=cols, show='headings', height=14)
        col_widths = [90, 95, 65, 60, 60, 55, 85]
        for c, w in zip(cols, col_widths):
            tree.heading(c, text=c)
            tree.column(c, width=w, anchor='center')
        tree.pack(side='left', fill='both', expand=True)
        tree.tag_configure('ok', foreground='#16a34a')
        tree.tag_configure('fail', foreground='#dc2626')

        sb = ttk.Scrollbar(result_frame, orient='vertical', command=tree.yview)
        sb.pack(side='right', fill='y')
        tree.configure(yscrollcommand=sb.set)

        def _int_or_none(s):
            try: return int(s.strip())
            except: return None

        def _do_filter():
            for item in tree.get_children():
                tree.delete(item)
            try:
                import pandas as pd
                gc, _sys = self._load_gc_module()
                df = pd.read_excel(self.project_file, header=None)
                records, group_pids = gc.parse_projects(df)
                cd = gc.compute_commission(records, group_pids)
                _sys.path.pop(0)

                filters = {
                    'name_keyword': name_entry.get().strip() or None,
                    'roles': [r for r, v in role_vars.items() if v.get()],
                    'min_eps': _int_or_none(eps_min.get()),
                    'max_eps': _int_or_none(eps_max.get()),
                    'project_id': pid_entry.get().strip() or None,
                    'status': status_var.get(),
                    'min_commission': _int_or_none(comm_min.get()),
                    'max_commission': _int_or_none(comm_max.get()),
                }
                results = advanced_filter(records, cd, filters)

                if not results:
                    self._log('🔎 筛选结果：0 条匹配')
                    return

                total_eps = sum(r['episodes'] for r in results)
                total_comm = sum(r['commission'] for r in results)
                self._log(f'🔎 筛选结果：{len(results)} 人，总集数 {total_eps}，总提成 {total_comm:,}')

                for r in results:
                    tag = 'ok' if r['status'] == '是' else 'fail'
                    tree.insert('', 'end', values=(
                        r['name'], r['role'], r['episodes'], r['projects'],
                        f'{r["quota"]}集' if r['quota'] > 0 else '无', r['status'],
                        f'{r["commission"]:,}'), tags=(tag,))
            except Exception as e:
                self._log(f'❌ 筛选失败: {e}')

        def _export_csv():
            items = tree.get_children()
            if not items:
                messagebox.showwarning('提示', '无数据可导出，请先执行筛选')
                return
            import csv
            path = filedialog.asksaveasfilename(
                title='导出筛选结果',
                defaultextension='.csv',
                filetypes=[('CSV文件', '*.csv')],
                initialdir=self.output_dir,
                initialfile='筛选结果.csv')
            if not path: return
            with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(cols)
                for item in items:
                    writer.writerow(tree.item(item)['values'])
            self._log(f'📥 已导出: {os.path.basename(path)}')
            self._open(os.path.dirname(path))

        tk.Button(btn_f, text='🔍 执行筛选', font=('Microsoft YaHei', 11, 'bold'),
                  bg=C['accent'], fg='white', relief='flat', cursor='hand2',
                  padx=20, pady=6, command=_do_filter).pack(side='left', padx=0)
        tk.Button(btn_f, text='📥 导出CSV', font=('Microsoft YaHei', 10),
                  bg=C['green'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=6, command=_export_csv).pack(side='left', padx=0)
        tk.Button(btn_f, text='关闭', font=('Microsoft YaHei', 10),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=6, command=dlg.destroy).pack(side='right')

        name_entry.bind('<Return>', lambda e: _do_filter())

    # ============ 功能5：数据修正工具 ============
    def _data_correction(self):
        if not self._require_features(): return
        dlg = tk.Toplevel(self.root)
        dlg.title('✏️ 数据修正工具')
        dlg.geometry('900x650')
        dlg.minsize(750, 500)
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        tk.Label(dlg, text='✏️ 数据修正工具', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=14)
        tk.Label(dlg, text='双击某行可修正集数或删除。修改后自动重算提成并更新Excel。',
                 font=('Microsoft YaHei', 9), bg=C['bg'], fg=C['text3']).pack(pady=0)

        # 数据加载
        cols = ('#', '姓名', '角色', '项目ID', '项目名', '集数', '明细')
        tree = ttk.Treeview(dlg, columns=cols, show='headings', height=16)
        col_widths = [35, 80, 85, 70, 200, 55, 140]
        for c, w in zip(cols, col_widths):
            tree.heading(c, text=c)
            tree.column(c, width=w, anchor='center')
        tree.pack(fill='both', expand=True, padx=16, pady=0)

        self._correction_records = []
        self._correction_gc = None
        self._correction_sys = None
        # 撤销/重做栈
        self._undo_stack = []
        self._redo_stack = []
        def _push_undo():
            import copy
            self._undo_stack.append(copy.deepcopy(self._correction_records))
            if len(self._undo_stack) > 50:
                self._undo_stack.pop(0)
            self._redo_stack.clear()
            self._log(f'↩️ 已记录撤销点（剩余 {len(self._undo_stack)} 步）')
        def _undo():
            if not self._undo_stack:
                messagebox.showinfo('撤销', '没有可撤销的操作', parent=dlg); return
            import copy
            self._redo_stack.append(copy.deepcopy(self._correction_records))
            self._correction_records = self._undo_stack.pop()
            _load_data()
            self._log(f'↩️ 已撤销（剩余 {len(self._undo_stack)} 步）')
        def _redo():
            if not self._redo_stack:
                messagebox.showinfo('重做', '没有可重做的操作', parent=dlg); return
            import copy
            self._undo_stack.append(copy.deepcopy(self._correction_records))
            self._correction_records = self._redo_stack.pop()
            _load_data()
            self._log(f'↪️ 已重做')
        dlg.bind('<Control-z>', lambda e: _undo())
        dlg.bind('<Control-Z>', lambda e: _undo())
        dlg.bind('<Control-y>', lambda e: _redo())
        dlg.bind('<Control-Y>', lambda e: _redo())

        def _load_data():
            for item in tree.get_children():
                tree.delete(item)
            try:
                import pandas as pd
                gc, _sys = self._load_gc_module()
                df = pd.read_excel(self.project_file, header=None)
                # 用数据文件真实月份更新输出文件名/标题
                data_cn, data_mnum = gc.get_month_from_data(df)
                if data_cn and data_mnum:
                    import re as _re
                    _tm = _re.match(r'(\d{4})年', gc.TEMPLATE_DATE)
                    _yr = _tm.group(1) if _tm else str(datetime.now().year)
                    gc.OUTPUT_MONTH = data_cn
                    gc.TEMPLATE_DATE = f'{_yr}年{data_mnum:02d}月'
                records, group_pids = gc.parse_projects(df)
                self._correction_records = records
                self._correction_gc = gc
                self._correction_sys = _sys

                for i, r in enumerate(records):
                    pname = r['AI项目名称'][:35] if r['AI项目名称'] else ''
                    detail = r['完成明细'][:25] if r['完成明细'] else ''
                    tree.insert('', 'end', iid=str(i), values=(
                        i + 1, r['身份证姓名'], r['角色'],
                        r['项目ID'], pname, r['单项目数/集数'], detail))
                self._log(f'✏️ 已加载 {len(records)} 条记录，双击可编辑')
            except Exception as e:
                self._log(f'❌ 加载失败: {e}')

        _load_data()

        def _edit_record(event):
            sel = tree.selection()
            if not sel: return
            idx = int(sel[0])
            rec = self._correction_records[idx]

            ed = tk.Toplevel(dlg)
            ed.title('修正记录')
            ed.geometry('420x310')
            ed.configure(bg=C['bg'])
            ed.transient(dlg); ed.grab_set()

            tk.Label(ed, text=f'修正记录 #{idx+1}', font=('Microsoft YaHei', 13, 'bold'),
                     bg=C['bg'], fg=C['text']).pack(pady=14)

            cf = tk.Frame(ed, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
            cf.pack(fill='x', padx=20, pady=0)
            ci = tk.Frame(cf, bg=C['card']); ci.pack(fill='x', padx=12, pady=10)

            fields = [
                ('姓名:', 'name', rec['身份证姓名']),
                ('项目ID:', 'pid', rec['项目ID']),
                ('项目名:', 'pname', rec['AI项目名称']),
                ('集数:', 'eps', str(rec['单项目数/集数'])),
            ]
            entries = {}
            for j, (label, key, val) in enumerate(fields):
                tk.Label(ci, text=label, font=('Microsoft YaHei', 9), bg=C['card'],
                         fg=C['text'], width=8, anchor='w').grid(row=j, column=0, sticky='w', pady=3)
                e = tk.Entry(ci, font=('Microsoft YaHei', 10), width=35)
                e.insert(0, val)
                e.grid(row=j, column=1, padx=6, pady=3)
                entries[key] = e

            def _save_edit():
                _push_undo()
                rec['身份证姓名'] = entries['name'].get().strip()
                rec['项目ID'] = entries['pid'].get().strip()
                rec['AI项目名称'] = entries['pname'].get().strip()
                try:
                    new_eps = int(entries['eps'].get())
                    rec['单项目数/集数'] = new_eps
                    rec['完成明细'] = ','.join(str(i) for i in range(1, new_eps + 1))
                except:
                    pass
                tree.set(str(idx), column='姓名', value=rec['身份证姓名'])
                tree.set(str(idx), column='项目ID', value=rec['项目ID'])
                tree.set(str(idx), column='集数', value=rec['单项目数/集数'])
                tree.set(str(idx), column='明细', value=rec['完成明细'][:25])
                self._log(f'✏️ 已修正记录 #{idx+1}')
                ed.destroy()

            def _delete_rec():
                if messagebox.askyesno('确认删除', f'确定要删除记录 #{idx+1} 吗？\n\n此操作不可恢复。'):
                    _push_undo()
                    self._correction_records.pop(idx)
                    tree.delete(str(idx))
                    self._log(f'🗑️ 已删除记录 #{idx+1}')
                    ed.destroy()
                    _load_data()

            bf = tk.Frame(ed, bg=C['bg']); bf.pack(fill='x', padx=20, pady=0)
            tk.Button(bf, text='💾 保存修改', font=('Microsoft YaHei', 11, 'bold'),
                      bg=C['green'], fg='white', relief='flat', cursor='hand2',
                      padx=20, pady=6, command=_save_edit).pack(side='left', padx=0)
            tk.Button(bf, text='🗑️ 删除', font=('Microsoft YaHei', 11),
                      bg=C['red'], fg='white', relief='flat', cursor='hand2',
                      padx=16, pady=6, command=_delete_rec).pack(side='left')
            tk.Button(bf, text='取消', font=('Microsoft YaHei', 11),
                      bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                      padx=16, pady=6, command=ed.destroy).pack(side='right')

        tree.bind('<Double-1>', _edit_record)

        def _regenerate():
            if not self._correction_records:
                messagebox.showwarning('提示', '无数据可生成'); return
            if not messagebox.askyesno('确认重新生成', '将使用修正后的数据重新生成提成表，确定吗？'):
                return
            try:
                gc = self._correction_gc
                # 重新计算
                group_pids = {g: set() for g in gc.GROUPS}
                for r in self._correction_records:
                    pid = r['项目ID']
                    for gname, ginfo in gc.GROUPS.items():
                        if r['身份证姓名'] in ginfo['成员'] and pid:
                            group_pids[gname].add(pid)

                cd = gc.compute_commission(self._correction_records, group_pids)
                excel_path = os.path.join(self.output_dir,
                    f'AI后期剪辑提成一组{self._correction_gc.OUTPUT_MONTH}.xlsx')
                path, html_path = gc.generate_excel(
                    self._correction_records, cd,
                    self._correction_gc.TEMPLATE_FILE, excel_path)
                self._log(f'✅ 已用修正数据重新生成')
                if self._correction_sys:
                    self._correction_sys.path.pop(0)
                dlg.destroy()
                self._open(path)
                self._open(html_path)
            except Exception as e:
                self._log(f'❌ 重新生成失败: {e}')

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=16, pady=0)
        tk.Button(bf, text='🔄 重新加载', font=('Microsoft YaHei', 10),
                  bg=C['blue'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=5, command=_load_data).pack(side='left', padx=0)
        tk.Button(bf, text='▶ 用修正数据重新生成', font=('Microsoft YaHei', 11, 'bold'),
                  bg=C['green'], fg='white', relief='flat', cursor='hand2',
                  padx=18, pady=5, command=_regenerate).pack(side='left', padx=0)
        tk.Button(bf, text='关闭', font=('Microsoft YaHei', 10),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=5, command=dlg.destroy).pack(side='right')

    # ============ 功能7：项目数据模板下载 ============
    def _download_template(self):
        if not self._require_features(): return
        try:
            path = filedialog.asksaveasfilename(
                title='保存模板到',
                defaultextension='.xlsx',
                filetypes=[('Excel文件', '*.xlsx')],
                initialdir=self.output_dir,
                initialfile='AI项目数据录入模板.xlsx')
            if not path: return
            generate_project_template(path)
            self._log(f'📥 模板已生成: {os.path.basename(path)}')
            self._open(path)
            messagebox.showinfo('完成', f'✅ 项目数据录入模板已生成！\n\n📄 {os.path.basename(path)}\n\n包含：\n• 示例数据行\n• 录入规范说明Sheet')
        except Exception as e:
            self._log(f'❌ 模板生成失败: {e}')
            messagebox.showerror('失败', str(e))

    # ============ 功能8：配置快照与回滚 ============
    def _config_snapshot(self):
        if not self._require_features(): return
        snap_dir = os.path.join(SCRIPT_DIR, 'config_snapshots')
        snapshots = list_config_snapshots(snap_dir)

        dlg = tk.Toplevel(self.root)
        dlg.title('📸 配置快照管理')
        dlg.geometry('650x500')
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root); dlg.grab_set()

        tk.Label(dlg, text='📸 配置快照管理', font=('Microsoft YaHei', 15, 'bold'),
                 bg=C['bg'], fg=C['text']).pack(pady=14)
        tk.Label(dlg, text='保存配置历史版本，支持一键回滚', font=('Microsoft YaHei', 9),
                 bg=C['bg'], fg=C['text3']).pack(pady=0)

        # 列表
        list_frame = tk.Frame(dlg, bg=C['card'], highlightthickness=1, highlightbackground=C['border'])
        list_frame.pack(fill='both', expand=True, padx=20, pady=0)

        cols = ('快照文件', '时间')
        tree = ttk.Treeview(list_frame, columns=cols, show='headings', height=12)
        tree.heading('快照文件', text='快照文件')
        tree.heading('时间', text='时间')
        tree.column('快照文件', width=350, anchor='w')
        tree.column('时间', width=180, anchor='center')
        tree.pack(side='left', fill='both', expand=True, padx=4, pady=4)

        sb = ttk.Scrollbar(list_frame, orient='vertical', command=tree.yview)
        sb.pack(side='right', fill='y')
        tree.configure(yscrollcommand=sb.set)

        def _refresh_list():
            for item in tree.get_children():
                tree.delete(item)
            snaps = list_config_snapshots(snap_dir)
            for s in snaps:
                tree.insert('', 'end', values=(s['name'], s['mtime'].strftime('%Y-%m-%d %H:%M:%S')))

        _refresh_list()

        def _create_snap():
            path = create_config_snapshot(CONFIG_PATH, snap_dir)
            self._log(f'📸 配置快照已保存: {os.path.basename(path)}')
            _refresh_list()
            messagebox.showinfo('完成', f'✅ 快照已保存！\n\n{os.path.basename(path)}')

        def _restore_snap():
            sel = tree.selection()
            if not sel: return
            val = tree.item(sel[0])['values']
            fname = val[0]
            if not messagebox.askyesno('确认回滚', f'确定要恢复到快照 "{fname}" 吗？\n\n当前配置将被覆盖。'):
                return
            restore_config_snapshot(os.path.join(snap_dir, fname), CONFIG_PATH)
            self.cfg = self._load_config()
            self._refresh_role_tags()
            self._log(f'📸 已回滚到快照: {fname}')
            messagebox.showinfo('完成', f'✅ 已回滚到快照！\n\n{fname}\n\n配置已更新，下次生成时生效。')
            _refresh_list()

        def _delete_snap():
            sel = tree.selection()
            if not sel: return
            val = tree.item(sel[0])['values']
            if messagebox.askyesno('确认删除', f'确定要删除快照 "{val[0]}" 吗？'):
                try:
                    os.remove(os.path.join(snap_dir, val[0]))
                    _refresh_list()
                except Exception as e:
                    messagebox.showerror('失败', str(e))

        bf = tk.Frame(dlg, bg=C['bg']); bf.pack(fill='x', padx=20, pady=0)
        tk.Button(bf, text='📸 创建快照', font=('Microsoft YaHei', 11, 'bold'),
                  bg=C['accent'], fg='white', relief='flat', cursor='hand2',
                  padx=18, pady=6, command=_create_snap).pack(side='left', padx=0)
        tk.Button(bf, text='↩ 回滚到此', font=('Microsoft YaHei', 11),
                  bg=C['orange'], fg='white', relief='flat', cursor='hand2',
                  padx=16, pady=6, command=_restore_snap).pack(side='left', padx=0)
        tk.Button(bf, text='🗑️ 删除', font=('Microsoft YaHei', 11),
                  bg=C['red'], fg='white', relief='flat', cursor='hand2',
                  padx=14, pady=6, command=_delete_snap).pack(side='left', padx=0)
        tk.Button(bf, text='关闭', font=('Microsoft YaHei', 11),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=16, pady=6, command=dlg.destroy).pack(side='right')

    # ============ 功能9：Web仪表盘服务 ============
    def _web_server(self):
        if not self._require_features(): return
        if hasattr(self, '_web_server_instance') and self._web_server_instance:
            self._web_server_instance.shutdown()
            self._web_server_instance.server_close()
            self._web_server_instance = None
            self._log('🛑 Web服务已停止')
            return

        try:
            # 生成只读看板页（供主管在线查看）
            try:
                from features import generate_readonly_dashboard
                ro = generate_readonly_dashboard(HISTORY_DIR, self.output_dir)
                if ro:
                    self._log(f'📊 只读看板已生成: {os.path.basename(ro)}')
            except Exception as roe:
                self._log(f'⚠️ 只读看板生成跳过: {roe}')
            self._web_server_instance = start_web_server(self.output_dir, 8080)
            port = self._web_server_instance.server_address[1]
            self._log(f'🌐 Web仪表盘已启动: http://localhost:{port}')
            self._log(f'   服务目录: {self.output_dir}')
            self._log(f'   浏览器访问查看全部报表和卡片')

            import webbrowser
            webbrowser.open(f'http://localhost:{port}')

            # 后台运行
            def serve():
                self._web_server_instance.serve_forever()

            threading.Thread(target=serve, daemon=True).start()
            messagebox.showinfo('Web服务已启动',
                f'✅ Web仪表盘已启动！\n\n'
                f'📍 地址: http://localhost:{port}\n'
                f'📂 目录: {os.path.basename(self.output_dir)}\n\n'
                f'浏览器已自动打开。\n'
                f'请勿关闭此窗口。')
        except Exception as e:
            self._log(f'❌ Web服务启动失败: {e}')
            messagebox.showerror('失败', str(e))

    # 功能10：移动端适配绩效卡片 —— 在 features.py 中改进 generate_person_cards
    def _refresh_backups(self):
        if not self._require_features(): return
        backups = list_backups(BACKUP_DIR)
        self._backup_list.configure(state='normal')
        self._backup_list.delete('1.0', 'end')
        if not backups:
            self._backup_list.insert('1.0', '暂无备份文件')
        else:
            total_size = sum(b['size'] for b in backups)
            self._backup_list.insert('1.0', f'共 {len(backups)} 个备份文件，占用 {total_size/1024:.1f} KB\n')
            self._backup_list.insert('end', '-' * 50 + '\n')
            for b in backups[:30]:
                self._backup_list.insert('end',
                    f'{b["mtime"].strftime("%m/%d %H:%M")}  {b["size"]/1024:6.1f}KB  {b["name"]}\n')
        self._backup_list.configure(state='disabled')

    def _manage_backups(self):
        if not self._require_features(): return
        removed = cleanup_backups(BACKUP_DIR, keep=30)
        self._log(f'🗄️ 已清理 {removed} 个旧备份，保留最近30个')
        self._refresh_backups()
        messagebox.showinfo('完成', f'已清理 {removed} 个旧备份文件。\n当前保留最近30个。')

    def check_files(self):
        items = [
            (self.project_file, '项目数据'),
            (CONFIG_PATH, '角色配置'),
            (self.template_file, '模板文件'),
        ]
        ok = True
        for fpath, desc in items:
            if not os.path.exists(fpath):
                ok = False
        self.run_btn.configure(state='normal' if ok else 'disabled',
                                bg=C['green'] if ok else C['gray'])
        if ok:
            self._log('✅ 文件就绪，点击"一键生成"开始。')
        else:
            self._log('❌ 文件不完整，请检查。')

    # ============ 日志 ============

    def _log(self, msg):
        self.log_txt.configure(state='normal')
        self.log_txt.insert('end', f'[{datetime.now().strftime("%H:%M:%S")}] {msg}\n')
        self.log_txt.see('end'); self.log_txt.configure(state='disabled')

    def _filter_line(self, line):
        line = line.strip()
        if not line or line.startswith('PS '):
            return
        skip = ['所在位置', 'CategoryInfo', 'FullyQualifiedErrorId', 'NativeCommandError',
                'RemoteException', '~~~~~~~~~~', 'EOFError']
        if any(s in line for s in skip): return
        self._log(line)

    # ============ 生成 ============

    def run(self):
        """第一步：解析项目 → 第二步：弹出超时对话框 → 第三步：生成"""
        self.run_btn.configure(state='disabled', bg=C['gray'], text='⏳ 解析中...')
        self.st.configure(text='⏳ 正在解析项目数据...')
        self.progress.start(10)
        self._log('🔍 第一步：解析项目数据...')
        self._log('—' * 50)

        project_file = self.project_file
        output_dir = self.output_dir
        template_file = self.template_file

        def parse_worker():
            try:
                import pandas as pd
                gc, _sys = self._load_gc_module()

                df = pd.read_excel(project_file, header=None)
                self._log(f'   项目文件: {os.path.basename(project_file)} ({len(df)}行)')

                # 先用不含超时的解析收集项目信息
                records, group_pids = gc.parse_projects(df)
                _sys.path.pop(0)

                # 收集每个项目的集数范围
                proj_info = {}  # pid -> {name, eps_set, people}
                for r in records:
                    pid = r['项目ID']
                    if not pid:
                        continue
                    if pid not in proj_info:
                        proj_info[pid] = {
                            'name': r['AI项目名称'][:40],
                            'eps_set': set(),
                            'people': {}
                        }
                    detail = r['完成明细']
                    for part in detail.split(','):
                        part = part.strip()
                        if part.isdigit():
                            proj_info[pid]['eps_set'].add(int(part))
                    nm = r['身份证姓名']
                    if nm not in proj_info[pid]['people']:
                        proj_info[pid]['people'][nm] = set()
                    for part in detail.split(','):
                        part = part.strip()
                        if part.isdigit():
                            proj_info[pid]['people'][nm].add(int(part))

                self.root.after(0, lambda: self._show_overtime_dialog(
                    proj_info, project_file, output_dir, template_file))
            except Exception as e:
                self.root.after(0, lambda: self._log(f'❌ 解析失败: {e}'))
                import traceback
                self.root.after(0, lambda: self._log(traceback.format_exc()))
                self.root.after(0, self._done)

        threading.Thread(target=parse_worker, daemon=True).start()

    def _show_overtime_dialog(self, proj_info, project_file, output_dir, template_file):
        """弹出超时集数标记对话框：搜索 → 选中 → 输入集数 → 确认"""
        self.progress.stop()
        self.run_btn.configure(state='normal', bg=C['accent'], text='▶  一键生成提成表')
        self._log(f'📋 已解析 {len(proj_info)} 个项目，请标记超4分钟的集数')

        dlg = tk.Toplevel(self.root)
        dlg.title('⏱️ 标记超4分钟集数')
        dlg.geometry('860x620')
        dlg.minsize(720, 500)
        dlg.configure(bg=C['bg'])
        dlg.transient(self.root)
        dlg.grab_set()

        # ---- 标题 ----
        hdr_f = tk.Frame(dlg, bg=C['hdr_bg'])
        hdr_f.pack(fill='x')
        tk.Label(hdr_f, text='⏱️  标记超4分钟集数（超时算2集）',
                 font=('Microsoft YaHei', 14, 'bold'), fg=C['hdr_text'],
                 bg=C['hdr_bg']).pack(pady=10)

        # ---- 搜索栏 + 已标记摘要 ----
        top_bar = tk.Frame(dlg, bg=C['bg'])
        top_bar.pack(fill='x', padx=12, pady=(8, 4))

        tk.Label(top_bar, text='🔍 搜索:', font=('Microsoft YaHei', 10),
                 bg=C['bg'], fg=C['text']).pack(side='left', padx=(0, 6))
        search_var = tk.StringVar()
        search_entry = tk.Entry(top_bar, textvariable=search_var,
                                font=('Microsoft YaHei', 11), width=30,
                                relief='solid', borderwidth=1)
        search_entry.pack(side='left', padx=(0, 8))

        summary_frame = tk.Frame(top_bar, bg=C['bg'])
        summary_frame.pack(side='left', fill='x', expand=True)
        summary_label = tk.Label(summary_frame, text='',
                                 font=('Microsoft YaHei', 8), bg=C['bg'],
                                 fg=C['text2'], anchor='w')
        summary_label.pack(side='left')

        # ---- 主内容区：左列表 + 右输入 ----
        main_frame = tk.Frame(dlg, bg=C['bg'])
        main_frame.pack(fill='both', expand=True, padx=12, pady=(4, 0))

        # 左：项目列表
        left_panel = tk.Frame(main_frame, bg=C['card'], highlightthickness=1,
                              highlightbackground=C['border'], width=300)
        left_panel.pack(side='left', fill='both', padx=(0, 6))
        left_panel.pack_propagate(False)

        tk.Label(left_panel, text='项目列表', font=('Microsoft YaHei', 10, 'bold'),
                 bg=C['card'], fg=C['text']).pack(padx=10, pady=(8, 4), anchor='w')

        list_frame = tk.Frame(left_panel, bg=C['card'])
        list_frame.pack(fill='both', expand=True, padx=6, pady=(0, 6))

        listbox = tk.Listbox(list_frame, font=('Microsoft YaHei', 9),
                             bg='white', fg=C['text'], relief='flat',
                             highlightthickness=1, highlightbackground=C['border'],
                             selectbackground=C['accent'], selectforeground='white',
                             activestyle='none')
        list_scroll = ttk.Scrollbar(list_frame, orient='vertical', command=listbox.yview)
        listbox.configure(yscrollcommand=list_scroll.set)
        listbox.pack(side='left', fill='both', expand=True)
        list_scroll.pack(side='right', fill='y')

        # 右：输入面板
        right_panel = tk.Frame(main_frame, bg=C['card'], highlightthickness=1,
                               highlightbackground=C['border'])
        right_panel.pack(side='left', fill='both', expand=True)

        tk.Label(right_panel, text='超时集数输入', font=('Microsoft YaHei', 10, 'bold'),
                 bg=C['card'], fg=C['text']).pack(padx=12, pady=(8, 4), anchor='w')

        selected_info = tk.Label(right_panel, text='← 请先在左侧搜索并点击项目',
                                 font=('Microsoft YaHei', 9), bg=C['card'],
                                 fg=C['text3'], justify='left')
        selected_info.pack(padx=12, pady=(0, 6), anchor='w')

        input_frame = tk.Frame(right_panel, bg=C['card'])
        input_frame.pack(fill='x', padx=12, pady=(0, 4))

        tk.Label(input_frame, text='超4分钟的集数:',
                 font=('Microsoft YaHei', 9, 'bold'), bg=C['card'],
                 fg=C['text']).pack(anchor='w', pady=(0, 4))

        ot_var = tk.StringVar()
        ot_entry = tk.Entry(input_frame, textvariable=ot_var,
                            font=('Consolas', 11), width=35,
                            relief='solid', borderwidth=1)
        ot_entry.pack(fill='x', pady=(0, 4))

        tk.Label(input_frame, text='支持: 1,3,5-8,10  或  1 3 5-8 10（空格/逗号/分号均可）',
                 font=('Microsoft YaHei', 7), bg=C['card'],
                 fg=C['text3']).pack(anchor='w')

        preview_label = tk.Label(input_frame, text='',
                                 font=('Microsoft YaHei', 8), bg=C['card'],
                                 fg=C['orange'], anchor='w')
        preview_label.pack(fill='x', pady=(4, 0))

        save_btn = tk.Button(right_panel, text='💾 保存此项目标记',
                             font=('Microsoft YaHei', 10, 'bold'),
                             bg=C['green'], fg='white', relief='flat',
                             cursor='hand2', padx=16, pady=6,
                             activebackground=C['green_l'])
        save_btn.pack(padx=12, pady=(8, 12), anchor='w')

        # ---- 数据模型 ----
        overtime_data = {}   # {pid: set of int}
        current_pid = [None]

        all_projects = []
        for pid in sorted(proj_info.keys(), key=lambda p: int(p) if p.isdigit() else 0):
            info = proj_info[pid]
            all_projects.append((f'{pid}  {info["name"]}', pid, info))

        def _populate_list(filter_text=''):
            listbox.delete(0, 'end')
            ft = filter_text.strip().lower()
            for disp, pid, info in all_projects:
                if ft and ft not in disp.lower():
                    continue
                marker = ' ●' if pid in overtime_data and overtime_data[pid] else ''
                listbox.insert('end', disp + marker)
                if pid in overtime_data and overtime_data[pid]:
                    listbox.itemconfig('end', bg='#fef3c7')

        def _update_summary():
            parts = []
            for pid in sorted(overtime_data.keys(), key=lambda p: int(p) if p.isdigit() else 0):
                eps = sorted(overtime_data[pid])
                parts.append(f'{pid}({",".join(str(e) for e in eps)})')
            summary_label.configure(
                text='已标记: ' + ' | '.join(parts) if parts else '暂未标记任何超时集数')

        def _on_list_select(event):
            sel = listbox.curselection()
            if not sel: return
            target_idx = sel[0]
            ft = search_var.get().strip().lower()
            count = 0
            for disp, pid, info in all_projects:
                if ft and ft not in disp.lower():
                    continue
                if count == target_idx:
                    current_pid[0] = pid
                    selected_info.configure(
                        text=f'📁 ID: {pid}\n📝 {info["name"]}\n📊 总集数: {len(info["eps_set"])}集',
                        fg=C['text'])
                    if pid in overtime_data:
                        eps = sorted(overtime_data[pid])
                        ot_var.set(','.join(str(e) for e in eps))
                        _show_preview(eps)
                    else:
                        ot_var.set('')
                        preview_label.configure(text='')
                    return
                count += 1

        def _show_preview(eps_list=None):
            if eps_list is not None:
                preview_label.configure(text=f'当前: {len(eps_list)}集 → 实际算 {len(eps_list)*2}集', fg=C['orange'])
                return
            text = ot_var.get().strip()
            if not text:
                preview_label.configure(text='')
                return
            try:
                from features import parse_overtime_episodes
                eps = parse_overtime_episodes(text)
                if eps:
                    preview_label.configure(text=f'解析: {len(eps)}集 → {sorted(eps)[:15]} 实际算{len(eps)*2}集', fg=C['orange'])
                else:
                    preview_label.configure(text='未识别到有效集数', fg=C['red'])
            except:
                preview_label.configure(text='格式错误', fg=C['red'])

        def _on_input_change(*args):
            _show_preview()

        def _save_current():
            pid = current_pid[0]
            if not pid:
                messagebox.showwarning('提示', '请先在左侧列表中选择一个项目')
                return
            text = ot_var.get().strip()
            if not text:
                overtime_data.pop(pid, None)
            else:
                try:
                    from features import parse_overtime_episodes
                    eps = parse_overtime_episodes(text)
                    if not eps:
                        messagebox.showwarning('提示', '未能识别到有效的集数，请检查格式')
                        return
                    overtime_data[pid] = eps
                except Exception as e:
                    messagebox.showwarning('错误', f'解析失败: {e}')
                    return
            _update_summary()
            _populate_list(search_var.get())
            self._log(f'⏱️ 项目 {pid}: {text if text else "(已清除)"}')

        def _on_search_change(*args):
            _populate_list(search_var.get())

        search_var.trace_add('write', _on_search_change)
        ot_var.trace_add('write', _on_input_change)
        listbox.bind('<<ListboxSelect>>', _on_list_select)
        save_btn.configure(command=_save_current)
        search_entry.bind('<Return>', lambda e: listbox.focus_set())

        _populate_list()

        # ---- 底部按钮 ----
        btn_frame = tk.Frame(dlg, bg=C['bg'])
        btn_frame.pack(fill='x', padx=12, pady=12)

        def _confirm():
            save_data = {pid: sorted(eps) for pid, eps in overtime_data.items() if eps}
            total_ot = sum(len(v) for v in save_data.values())
            self._log(f'⏱️ 已标记 {total_ot} 个超时集数 ({len(save_data)}个项目)')
            dlg.destroy()
            self._do_generate(project_file, output_dir, template_file, save_data)

        def _skip():
            self._log('⏭️ 跳过超时标记')
            dlg.destroy()
            self._do_generate(project_file, output_dir, template_file, {})

        tk.Button(btn_frame, text='✅ 确认并生成', font=('Microsoft YaHei', 12, 'bold'),
                  bg=C['accent'], fg='white', relief='flat', cursor='hand2',
                  padx=28, pady=8, activebackground=C['accent_a'],
                  command=_confirm).pack(side='left', padx=(0, 10))
        tk.Button(btn_frame, text='⏭️ 跳过（无超时）', font=('Microsoft YaHei', 10),
                  bg=C['orange'], fg='white', relief='flat', cursor='hand2',
                  padx=18, pady=8, command=_skip).pack(side='left', padx=(0, 10))
        tk.Button(btn_frame, text='取消', font=('Microsoft YaHei', 10),
                  bg=C['gray'], fg='white', relief='flat', cursor='hand2',
                  padx=18, pady=8,
                  command=lambda: [dlg.destroy(), self._done()]).pack(side='right')

        dlg.protocol('WM_DELETE_WINDOW', lambda: [dlg.destroy(), self._done()])

    def _do_generate(self, project_file, output_dir, template_file, overtime_data):
        """实际执行生成：用 CLI 脚本生成提成表"""
        self.run_btn.configure(state='disabled', bg=C['gray'], text='⏳ 生成中...')
        self.st.configure(text='⏳ 正在计算绩效和生成表格...')
        self.progress.start(10)
        self._log('🔍 第二步：计算提成并生成表格...')
        self._log('—' * 50)

        self._output_files = {}
        self._current_overtime_map = {
            str(pid): set(episodes) for pid, episodes in overtime_data.items()
        }

        with tempfile.NamedTemporaryFile(
                mode='w', encoding='utf-8', suffix='.json',
                prefix='overtime_', dir=SCRIPT_DIR, delete=False) as temp_file:
            json.dump(overtime_data, temp_file, ensure_ascii=False, indent=2)
            overtime_file = temp_file.name

        gc, gc_sys = self._load_gc_module()
        _, template_date = gc.get_month_from_template(template_file)
        year_match = re.match(r'(\d{4})年', template_date)
        self._generation_year = int(year_match.group(1)) if year_match else None
        gc_sys.path.pop(0)

        def worker():
            try:
                env = os.environ.copy()
                env['PYTHONIOENCODING'] = 'utf-8'
                if 'PYTHONPATH' in env:
                    env['PYTHONPATH'] = SRC_DIR + os.pathsep + env['PYTHONPATH']
                else:
                    env['PYTHONPATH'] = SRC_DIR
                self._log(f'  子进程 Python: {PYTHON_EXE}')
                p = subprocess.Popen([PYTHON_EXE, CLI_SCRIPT,
                                      project_file, template_file,
                                      output_dir, '--overtime-file', overtime_file],
                                     cwd=SCRIPT_DIR,
                                     stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                                     stderr=subprocess.STDOUT, text=True,
                                     encoding='utf-8', errors='replace',
                                     env=env)
                try: p.stdin.write('\n'); p.stdin.flush()
                except: pass
                for ln in iter(p.stdout.readline, ''):
                    line = ln.strip()
                    if line.startswith('OUTPUT_EXCEL='):
                        self._output_files['excel'] = line.split('=', 1)[1]
                    elif line.startswith('OUTPUT_HTML='):
                        self._output_files['html'] = line.split('=', 1)[1]
                    self.root.after(0, lambda l=ln: self._filter_line(l))
                p.wait()
                self.root.after(0, self._done)
            except Exception as e:
                self.root.after(0, lambda: self._log(f'❌ 错误: {e}'))
                self.root.after(0, self._done)
            finally:
                try:
                    os.unlink(overtime_file)
                except OSError:
                    pass

        threading.Thread(target=worker, daemon=True).start()

    def _done(self):
        self.progress.stop()
        self.run_btn.configure(state='normal', bg=C['green'], text='▶  一键生成提成表')
        self._log('—' * 50)

        excel_path = self._output_files.get('excel', '')
        html_path = self._output_files.get('html', '')

        excel_ok = excel_path and os.path.exists(excel_path)
        html_ok = html_path and os.path.exists(html_path)

        if excel_ok or html_ok:
            self._log('🎉 全部完成！')
            self.st.configure(text='● 就绪 · 生成完成')

            # 自动备份
            if self.auto_backup.get() and HAS_FEATURES:
                try:
                    backed = backup_output(excel_path, html_path, BACKUP_DIR)
                    if backed:
                        self._log(f'💾 已备份 {len(backed)} 个文件到 backup/')
                except Exception as e:
                    self._log(f'⚠️ 备份失败: {e}')

            # 自动生成个人绩效卡片
            if excel_ok and HAS_FEATURES:
                try:
                    import pandas as pd
                    gc, _sys = self._load_gc_module()

                    df = pd.read_excel(self.project_file, header=None)
                    records, group_pids = gc.parse_projects(
                        df,
                        default_year=self._generation_year,
                        overtime_map=self._current_overtime_map,
                    )
                    cd = gc.compute_commission(records, group_pids)
                    card_paths = generate_person_cards(records, cd, CARDS_DIR)
                    if card_paths:
                        self._log(f'🃏 已生成 {len(card_paths)-1} 张个人绩效卡片 -> 个人绩效卡片/')
                        self._open(card_paths[0])
                    # 自动归档到历史数据库（跨月趋势用）
                    try:
                        from features import archive_history
                        month_label = gc.TEMPLATE_DATE
                        archive_history(records, cd, HISTORY_DIR, month_label)
                        self._log(f'🗂 已归档历史数据: {month_label} -> history/')
                    except Exception as ae:
                        self._log(f'⚠️ 历史归档跳过: {ae}')
                    _sys.path.pop(0)
                except Exception as e:
                    self._log(f'⚠️ 卡片生成跳过: {e}')

            # 打开文件
            if excel_ok:
                self._open(excel_path)
            if html_ok:
                self._open(html_path)
            messagebox.showinfo('完成', f'✅ 提成表、统计简报和仪表盘已生成完毕！\n\n📊 {os.path.basename(excel_path)}\n📈 {os.path.basename(html_path)}')
        else:
            self._log('❌ 生成失败：未找到输出文件')
            self.st.configure(text='❌ 生成失败 · 请查看日志')
            messagebox.showerror('失败', '❌ 生成失败！\n\n请检查日志了解详细错误信息。\n常见原因：\n  1. 项目数据文件格式不正确\n  2. 模板文件表头不匹配\n  3. config.json 人员配置有误')


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()

if __name__ == '__main__':
    main()
