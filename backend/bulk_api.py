# -*- coding: utf-8 -*-
"""批量操作 + 任务中心 + 月度报告 API"""
from datetime import datetime
from flask import jsonify, request, send_file


def backfill_months(now, count=6):
    """返回含 now 在内的最近 count 个月的 'YYYY-MM' 列表（降序，最新在前）。
    安全处理跨年回溯：now.month - i 在 1~5 月会 <=0，datetime 会抛 ValueError。
    """
    out = []
    for i in range(count - 1, -1, -1):
        _y, _m = now.year, now.month - i
        while _m <= 0:
            _m += 12
            _y -= 1
        out.append(datetime(_y, _m, 1).strftime("%Y-%m"))
    return out


def register_routes(app, db):
    # ================================================================
    # 批量操作
    # ================================================================

    @app.route("/api/bulk/update_month", methods=["POST"])
    def bulk_update_month():
        body = request.get_json(force=True) or {}
        names = body.get("names") or []
        month = body.get("month", "")
        if not names:
            return jsonify({"ok": False, "msg": "未选择项目"}), 400
        try:
            with db.get_conn() as conn:
                placeholders = ",".join(["?"] * len(names))
                if month:
                    conn.execute(
                        f"UPDATE projects SET project_month=? WHERE name IN ({placeholders})",
                        [month] + list(names),
                    )
                else:
                    conn.execute(
                        f"UPDATE projects SET project_month=NULL WHERE name IN ({placeholders})",
                        list(names),
                    )
            return jsonify({"ok": True, "updated": len(names)})
        except Exception as e:
            return jsonify({"ok": False, "msg": str(e)}), 500

    @app.route("/api/bulk/update_status", methods=["POST"])
    def bulk_update_status():
        body = request.get_json(force=True) or {}
        names = body.get("names") or []
        status = body.get("custom_status") or ""
        if not names or not status:
            return jsonify({"ok": False, "msg": "参数缺失"}), 400
        try:
            with db.get_conn() as conn:
                placeholders = ",".join(["?"] * len(names))
                conn.execute(
                    f"UPDATE projects SET custom_status=? WHERE name IN ({placeholders})",
                    [status] + list(names),
                )
            # 审计：批量改状态
            try:
                for n in names:
                    db.add_audit_log(n, "批量改状态", "状态 -> " + status)
            except Exception:
                pass
            return jsonify({"ok": True, "updated": len(names)})
        except Exception as e:
            return jsonify({"ok": False, "msg": str(e)}), 500

    @app.route("/api/bulk/update_delivered_date", methods=["POST"])
    def bulk_update_delivered_date():
        """批量设置交付日期。body: {names:[], date:'YYYY-MM-DD' or '' 清除}"""
        body = request.get_json(force=True) or {}
        names = body.get("names") or []
        date = (body.get("date") or "").strip()
        if not names:
            return jsonify({"ok": False, "msg": "未选择项目"}), 400
        try:
            with db.get_conn() as conn:
                placeholders = ",".join(["?"] * len(names))
                conn.execute(
                    f"UPDATE projects SET delivered_date=? WHERE name IN ({placeholders})",
                    [date] + list(names),
                )
            return jsonify({"ok": True, "updated": len(names)})
        except Exception as e:
            return jsonify({"ok": False, "msg": str(e)}), 500

    @app.route("/api/bulk/export", methods=["POST"])
    def bulk_export():
        """批量导出所选项目的档案 CSV（按 name 列表过滤）。"""
        import io as _io
        import csv as _csv
        body = request.get_json(force=True) or {}
        names = body.get("names") or []
        if not names:
            return jsonify({"ok": False, "msg": "未选择项目"}), 400
        projs = db.get_all_projects() or []
        name_set = set(names)
        projs = [p for p in projs if p.get("name") in name_set]
        buf = _io.StringIO()
        writer = _csv.writer(buf)
        writer.writerow(["项目名", "部门", "当前状态", "交付状态", "总集数", "已生成集数", "交付日期", "组内路径", "制作路径", "创建时间"])
        for p in projs:
            writer.writerow([
                p.get("name", ""), p.get("department", ""), p.get("custom_status", ""),
                p.get("delivery_status", ""), p.get("total_episodes", 0),
                p.get("current_episodes", 0), p.get("delivered_date", ""),
                p.get("group_path", ""), p.get("production_path", ""), p.get("created_at", ""),
            ])
        data = "\ufeff" + buf.getvalue()
        from flask import Response
        return Response(data, mimetype="text/csv; charset=utf-8",
                        headers={"Content-Disposition": "attachment; filename=projects-selected.csv"})

    # ================================================================
    # 任务中心
    # ================================================================

    @app.route("/api/activity_log", methods=["GET"])
    def api_activity_log():
        limit = int(request.args.get("limit", 200))
        project = request.args.get("project", "").strip()
        type_filter = request.args.get("type", "").strip()
        status_filter = request.args.get("status", "").strip()

        try:
            all_logs = []

            if type_filter != "deliver":
                sync_logs = db.get_sync_logs(project_name=project or None, limit=limit)
                for l in sync_logs:
                    all_logs.append({
                        "type": "sync",
                        "project_name": l.get("project_name", ""),
                        "action": l.get("action", ""),
                        "file_name": l.get("file_path", ""),
                        "file_size": l.get("file_size", 0) or 0,
                        "status": l.get("status", ""),
                        "message": l.get("message", ""),
                        "created_at": l.get("created_at", ""),
                    })

            if type_filter != "sync":
                del_logs = db.get_delivery_logs(project_name=project or None, limit=limit)
                for l in del_logs:
                    all_logs.append({
                        "type": "deliver",
                        "project_name": l.get("project_name", ""),
                        "action": "",
                        "file_name": l.get("file_name", ""),
                        "file_size": l.get("file_size", 0) or 0,
                        "status": l.get("status", ""),
                        "message": l.get("message", ""),
                        "created_at": l.get("created_at", ""),
                    })

            if status_filter:
                all_logs = [l for l in all_logs if l["status"] == status_filter]

            all_logs.sort(key=lambda x: x.get("created_at", ""), reverse=True)
            all_logs = all_logs[:limit]

            active_runs = db.get_deliver_runs(limit=20)
            active_list = []
            for r in active_runs:
                if r.get("status") in ("pending", "running", "syncing"):
                    active_list.append(dict(r))

            return jsonify({"ok": True, "logs": all_logs, "active_runs": active_list})
        except Exception as e:
            import traceback; traceback.print_exc()
            return jsonify({"ok": False, "msg": str(e)}), 500

    # ================================================================
    # 月度报告
    # ================================================================

    def _query_report_data(month):
        with db.get_conn() as conn:
            cur = conn.cursor()

            cur.execute(
                """
                SELECT name, department, custom_status, delivery_status,
                       total_episodes, current_episodes, project_month
                FROM projects
                WHERE project_month = ?
                ORDER BY department, name
                """,
                (month,),
            )
            projects_rows = [dict(r) for r in cur.fetchall()]

            cur.execute(
                """
                SELECT department,
                       COUNT(*) as total,
                       SUM(CASE WHEN custom_status='已完成' THEN 1 ELSE 0 END) as completed,
                       SUM(CASE WHEN custom_status IN ('剪辑中','审核中','修改中') THEN 1 ELSE 0 END) as editing,
                       SUM(CASE WHEN delivery_status='delivered' THEN 1 ELSE 0 END) as delivered,
                       SUM(IFNULL(total_episodes,0)) as total_eps
                FROM projects
                WHERE project_month = ?
                GROUP BY department
                ORDER BY total DESC
                """,
                (month,),
            )
            summary_by_dept = [dict(r) for r in cur.fetchall()]

            cur.execute(
                """
                SELECT COALESCE(SUM(file_size),0) as total_bytes,
                       COUNT(*) as file_count,
                       COUNT(DISTINCT project_name) as project_count
                FROM delivery_logs
                WHERE created_at LIKE ? || '%'
                  AND status = 'success'
                """,
                (month,),
            )
            row = cur.fetchone()
            delivery_stats = dict(row) if row else {"total_bytes": 0, "file_count": 0, "project_count": 0}

        return {
            "ok": True,
            "month": month,
            "projects": projects_rows,
            "summary_by_department": summary_by_dept,
            "delivery_stats": delivery_stats,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    @app.route("/api/report/monthly", methods=["GET"])
    def api_report_monthly():
        month = request.args.get("month", "").strip() or datetime.now().strftime("%Y-%m")
        try:
            data = _query_report_data(month)
            return jsonify(data)
        except Exception as e:
            import traceback; traceback.print_exc()
            return jsonify({"ok": False, "msg": str(e)}), 500

    @app.route("/api/report/monthly/export", methods=["GET"])
    def api_report_export():
        month = request.args.get("month", "").strip() or datetime.now().strftime("%Y-%m")
        try:
            data = _query_report_data(month)
        except Exception as e:
            import traceback; traceback.print_exc()
            return jsonify({"ok": False, "msg": str(e)}), 500

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            import io as _io

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "汇总"

            header_font = Font(bold=True, color="FFFFFF", name="微软雅黑")
            header_fill = PatternFill("solid", fgColor="4472C4")
            title_font = Font(bold=True, size=14, name="微软雅黑")

            ws1["A1"] = f"视频工作台 月度报告 — {month}"
            ws1["A1"].font = title_font
            ws1.merge_cells("A1:F1")
            ws1["A3"] = "生成时间：" + data.get("generated_at", "")

            ds = data.get("delivery_stats", {})
            ws1["A5"] = "回传概览"
            ws1["A5"].font = Font(bold=True, name="微软雅黑")
            ws1["A6"] = "总回传数据量"
            ws1["B6"] = (ds.get("total_bytes") or 0)
            ws1["C6"] = "总回传文件数"
            ws1["D6"] = (ds.get("file_count") or 0)
            ws1["E6"] = "涉及项目数"
            ws1["F6"] = (ds.get("project_count") or 0)

            ws1["A8"] = "按部门汇总"
            ws1["A8"].font = Font(bold=True, name="微软雅黑")

            headers = ["部门", "项目数", "已完成", "制作中", "已交付", "总集数"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws1.cell(row=9, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill

            row = 10
            for dept in data.get("summary_by_department", []):
                vals = [
                    dept.get("department") or "(未分类)",
                    dept.get("total", 0),
                    dept.get("completed", 0),
                    dept.get("editing", 0),
                    dept.get("delivered", 0),
                    dept.get("total_eps", 0),
                ]
                for ci, v in enumerate(vals):
                    ws1.cell(row=row, column=ci + 1, value=v)
                row += 1

            for col in ["A", "B", "C", "D", "E", "F"]:
                ws1.column_dimensions[col].width = 16

            ws2 = wb.create_sheet("项目清单")
            headers2 = ["项目名称", "部门", "状态", "交付状态", "总集数", "当前集数", "月份"]
            for ci, h in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill

            for ri, p in enumerate(data.get("projects", []), 2):
                ws2.cell(row=ri, column=1, value=p.get("name", ""))
                ws2.cell(row=ri, column=2, value=p.get("department", "") or "")
                ws2.cell(row=ri, column=3, value=p.get("custom_status", "") or "")
                ws2.cell(row=ri, column=4, value=p.get("delivery_status", "") or "")
                ws2.cell(row=ri, column=5, value=p.get("total_episodes", 0) or 0)
                ws2.cell(row=ri, column=6, value=p.get("current_episodes", 0) or 0)
                ws2.cell(row=ri, column=7, value=p.get("project_month", "") or "")

            for col in ["A", "B", "C", "D", "E", "F", "G"]:
                ws2.column_dimensions[col].width = 22

            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            filename = f"video_workbench_report_{month}.xlsx"
            return send_file(
                buf,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            import traceback; traceback.print_exc()
            return jsonify({"ok": False, "msg": str(e)}), 500

    # ================================================================
    # 数据看板统计 API
    # ================================================================

    @app.route("/api/stats/dashboard", methods=["GET"])
    def api_stats_dashboard():
        """数据看板：剪辑师工作量 + 部门统计 + 产能趋势。"""
        import json as _json
        from datetime import datetime as _dt

        now = _dt.now()
        cur_month = request.args.get("month", "") or now.strftime("%Y-%m")

        # ---------- 1. 剪辑师工作量（统一口径：从 episode_plan 聚合） ----------
        try:
            from features import aggregate_editor_workload
        except Exception:
            aggregate_editor_workload = None
        editor_list = []
        if aggregate_editor_workload:
            # 统一口径：仅统计当月(project_month=cur_month)项目的分集工作量
            editor_list = aggregate_editor_workload(db, month=cur_month)
        # 若 helper 不可用，回退到内联实现
        if not editor_list:
            import json as _json
            editor_workload = {}
            with db.get_conn() as conn:
                rows = conn.execute(
                    "SELECT name, episode_plan FROM projects"
                ).fetchall()
            for row in rows:
                r = dict(row)
                ep = r.get("episode_plan") or ""
                if not ep or ep == "{}":
                    continue
                try:
                    plan = _json.loads(ep) if isinstance(ep, str) else (ep or {})
                except Exception:
                    continue
                if not isinstance(plan, dict):
                    continue
                for ep_num, editor in plan.items():
                    if not editor:
                        continue
                    ed = editor.strip()
                    if not ed:
                        continue
                    item = editor_workload.setdefault(ed, {"assigned": 0, "projects": set()})
                    item["assigned"] += 1
                    item["projects"].add(r.get("name"))
            editor_list = [
                {"name": name, "assigned": item["assigned"], "projects": len(item["projects"])}
                for name, item in sorted(editor_workload.items(), key=lambda x: -x[1]["assigned"])
            ]

        # 为剪辑师附上角色/提成卡点(基准集数)，供工作量看板标注（功能：70集卡点标记）
        try:
            from commission_service import editor_quota_map
            _quota = editor_quota_map()
            for e in editor_list:
                q = _quota.get(e.get("name"))
                e["quota"] = q["quota"] if q else None
                e["editor_role"] = q["role"] if q else ""
        except Exception:
            pass

        # ---------- 2. 部门项目统计 ----------
        with db.get_conn() as conn:
            rows = conn.execute(
                """SELECT department,
                          COUNT(*) as total,
                          SUM(CASE WHEN custom_status='已完成' THEN 1 ELSE 0 END) as completed,
                          SUM(CASE WHEN custom_status IN ('剪辑中','审核中','修改中') THEN 1 ELSE 0 END) as editing,
                          SUM(CASE WHEN delivery_status='delivered' THEN 1 ELSE 0 END) as delivered
                   FROM projects GROUP BY department ORDER BY total DESC"""
            ).fetchall()
        dept_stats = [dict(r) for r in rows]

        # ---------- 3. 产能趋势（近6个月） ----------
        months = backfill_months(now, 6)
        trend = []
        for m in months:
            with db.get_conn() as conn:
                r = conn.execute(
                    """SELECT COUNT(*) as total,
                              SUM(CASE WHEN custom_status='已完成' THEN 1 ELSE 0 END) as done,
                              SUM(CASE WHEN delivery_status='delivered' THEN 1 ELSE 0 END) as delivered
                       FROM projects WHERE project_month=?""",
                    (m,),
                ).fetchone()
            trend.append({
                "month": m,
                "total": r["total"] or 0,
                "done": r["done"] or 0,
                "delivered": r["delivered"] or 0,
            })

        return jsonify({
            "ok": True,
            "month": cur_month,
            "editors": editor_list,
            "dept_stats": dept_stats,
            "trend": trend,
            "summary": {
                "total_editors": len(editor_list),
                "total_assigned": sum(e["assigned"] for e in editor_list),
            },
        })

