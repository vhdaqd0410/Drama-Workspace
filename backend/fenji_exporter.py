# -*- coding: utf-8 -*-
"""
分集导出到 Excel 模板 —— 从独立分集程序搬过来的 openpyxl 核心逻辑。
被 enhanced_routes.py 的 /api/fenji/export_excel 路由调用。
"""
import os as _os
import io as _io
import base64 as _b64
import datetime as _dt
import logging as _logging

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

_logger = _logging.getLogger(__name__)

# ======== 美化样式配置（与分集程序保持一致） ========
HEADER_FILL = "DDEBF7"
HEADER_FONT = "1F4E79"
TITLE_FONT_SIZE = 16
HEADER_FONT_SIZE = 12
BODY_FONT_SIZE = 11
DONE_FILL = "E2EFDA"
DONE_FONT = "006100"

THIN_GRAY = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

COL_WIDTHS = {'A': 32, 'B': 52, 'C': 26, 'D': 18, 'E': 12, 'F': 12}


def _clean_text(value):
    if not isinstance(value, str):
        return value
    return ''.join(c for c in value if ord(c) >= 32 or c in '\n\t')


def _find_last_project_end(ws):
    last_end = 0
    for rng in ws.merged_cells.ranges:
        if rng.min_col == 1 and rng.max_col == 1:
            top_val = ws.cell(row=rng.min_row, column=1).value
            if top_val:
                last_end = max(last_end, rng.max_row)
    if last_end == 0:
        last_end = ws.max_row
    return last_end


def _append_project(ws, start_row, project_name, path, assign_list, time_text, status_text):
    n = len(assign_list)
    end_row = start_row + n - 1
    for col in (1, 2, 4, 5):
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=col, end_column=col)
    ws.cell(row=start_row, column=1, value=_clean_text(project_name))
    ws.cell(row=start_row, column=2, value=_clean_text(path))
    ws.cell(row=start_row, column=4, value=_clean_text(time_text) if time_text else None)
    ws.cell(row=start_row, column=5, value=_clean_text(status_text) if status_text else None)
    for i, d in enumerate(assign_list):
        ws.cell(row=start_row + i, column=3,
                value=_clean_text(f"{d['person']}：{d['range']}"))
    return end_row


def _beautify(ws):
    last_end = _find_last_project_end(ws)
    blocks = []
    for rng in ws.merged_cells.ranges:
        if rng.min_col == 1 and rng.max_col == 1:
            top_val = ws.cell(row=rng.min_row, column=1).value
            if top_val:
                blocks.append((rng.min_row, rng.max_row))
    blocks.sort()

    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(name='微软雅黑', size=TITLE_FONT_SIZE, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 7):
        for r in range(1, 3):
            ws.cell(row=r, column=col).border = BORDER_ALL

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    for (blk_start, blk_end) in blocks:
        for col in (1, 2, 4, 5):
            cell = ws.cell(row=blk_start, column=col)
            cell.font = Font(name='微软雅黑', size=HEADER_FONT_SIZE, bold=True, color=HEADER_FONT)
            cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
            cell.alignment = CENTER
            cell.border = BORDER_ALL
        c_cell = ws.cell(row=blk_start, column=3)
        c_cell.font = Font(name='微软雅黑', size=BODY_FONT_SIZE)
        c_cell.alignment = CENTER
        c_cell.border = BORDER_ALL
        for r in range(blk_start, blk_end + 1):
            ws.row_dimensions[r].height = 22
            cell_c = ws.cell(row=r, column=3)
            cell_c.font = Font(name='微软雅黑', size=BODY_FONT_SIZE)
            cell_c.alignment = CENTER
            cell_c.border = BORDER_ALL
            for col in (1, 2, 4, 5):
                cell = ws.cell(row=r, column=col)
                cell.border = BORDER_ALL
                cell.alignment = CENTER
                if r != blk_start:
                    cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
        e_val = ws.cell(row=blk_start, column=5).value
        if e_val and '已分集' in str(e_val):
            for r in range(blk_start, blk_end + 1):
                cell = ws.cell(row=r, column=5)
                cell.font = Font(name='微软雅黑', size=HEADER_FONT_SIZE, bold=True, color=DONE_FONT)
                cell.fill = PatternFill(start_color=DONE_FILL, end_color=DONE_FILL, fill_type='solid')
    return blocks


def export_from_template(tpl_bytes, project_name, path, assign_list,
                         time_text='', status_text='已分集'):
    """
    核心函数：给定模板 bytes + 分集参数，返回新文件 bytes。
    assign_list: [{"person": "张三", "range": "1-25"}, ...]
    """
    wb = openpyxl.load_workbook(_io.BytesIO(tpl_bytes))
    ws = wb[wb.sheetnames[0]]

    last_end = _find_last_project_end(ws)
    start_row = last_end + 1

    _append_project(ws, start_row, project_name, path, assign_list, time_text, status_text)
    _beautify(ws)

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def backup_template(tpl_bytes, original_name, backup_dir):
    """把原始模板备份到指定目录，返回备份文件名。"""
    _os.makedirs(backup_dir, exist_ok=True)
    base = _os.path.splitext(original_name)[0] if original_name else '模板'
    ts = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    bak_name = f'{base}_备份_{ts}.xlsx'
    bak_path = _os.path.join(backup_dir, bak_name)
    with open(bak_path, 'wb') as f:
        f.write(tpl_bytes)
    return bak_name, bak_path


def list_templates(tpl_dir):
    """列出模板目录下所有 .xlsx 文件。"""
    if not _os.path.isdir(tpl_dir):
        return []
    return [f for f in _os.listdir(tpl_dir)
            if f.lower().endswith(('.xlsx', '.xlsm')) and not f.startswith('~$')]
