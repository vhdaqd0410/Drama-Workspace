# -*- coding: utf-8 -*-
"""融合功能路由（自「项目档案管理器」）：待办事项 / 时间轴 / 数据备份 / 视频缩略图。

统一在 register_routes(app, db) 内用 @app.route 注册，随 app.py 启动加载。
所有路由默认受 _auth_gate 保护（缩略图端点由 app 侧加入 /api/thumbnail/ 公共前缀）。
"""
import os
import json
import logging
import subprocess
import tempfile
from flask import request, jsonify, send_file, abort

logger = logging.getLogger("features")

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_THUMB_DIR = os.path.join(_BASE, "data", "thumbs")
_DATA_DIR = os.path.join(_BASE, "data")


def _yaml_cfg():
    try:
        import yaml
        with open(os.path.join(_BASE, "backend", "config.yaml"), "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception as e:
        logger.warning("读取 config.yaml 失败: %s", e)
        return {}


def _yaml_get(dotted, default=None):
    cur = _yaml_cfg()
    for part in dotted.split("."):
        if isinstance(cur, dict) and part in cur:
            cur = cur[part]
        else:
            return default
    return cur


def _now():
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def compute_audit_alerts(db, stale_days=3):
    """审核流超时提醒（功能5）：找出卡在审核/质检/修改等状态超过 stale_days 天的项目。
    返回 [{'name','owner','status','status_changed_at','days_stuck','department'}]。
    责任到人：owner 为空时回退到分集里第一个剪辑师。
    """
    import json as _json
    from datetime import datetime
    now = datetime.now()
    review_states = ("待质检", "质检中", "质检中", "审核中", "修改中")
    try:
        projs = db.get_all_projects() or []
    except Exception:
        projs = []
    alerts = []
    for p in projs:
        st = (p.get("custom_status") or "").strip()
        if st not in review_states:
            continue
        changed = (p.get("status_changed_at") or "").strip()
        days_stuck = None
        if changed:
            try:
                dt = datetime.strptime(changed[:19], "%Y-%m-%d %H:%M:%S")
                days_stuck = (now - dt).total_seconds() / 86400.0
            except Exception:
                days_stuck = None
        if days_stuck is None or days_stuck < stale_days:
            continue
        # 责任到人：owner，否则从分集取第一个剪辑师
        owner = (p.get("owner") or "").strip()
        if not owner:
            plan = p.get("episode_plan") or "{}"
            try:
                plan = _json.loads(plan) if isinstance(plan, str) else plan
            except Exception:
                plan = {}
            editors = [e for e in (plan.values() if isinstance(plan, dict) else []) if e]
            owner = editors[0] if editors else ""
        alerts.append({
            "name": p.get("name") or "",
            "owner": owner,
            "status": st,
            "status_changed_at": changed,
            "days_stuck": round(days_stuck, 1),
            "department": p.get("department") or "",
        })
    alerts.sort(key=lambda a: -a["days_stuck"])
    return {"stale_days": stale_days, "alerts": alerts}


def _parse_time_text(text):
    """解析分集目标表格的胶片日期文本（如 '8.15上午10点交'、'8.15'、'2026-8-15'）
    为 'YYYY-MM-DD'。解析失败返回空串。"""
    import re as _re
    from datetime import datetime
    if not text:
        return ""
    s = str(text).strip()
    # 完整日期：2026-8-15 / 2026/8/15
    m = _re.search(r"(20\d{2})[-/年.](\d{1,2})[-/月.](\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        # 短格式：M.D 或 M.D上午/下午…点交
        m = _re.search(r"(\d{1,2})[.月](\d{1,2})", s)
        if not m:
            # 纯数字日期如 815 / 8.15
            m = _re.search(r"(\d{1,2})\s*[-./]\s*(\d{1,2})", s)
            if not m:
                return ""
        y, mo, d = datetime.now().year, int(m.group(1)), int(m.group(2))
    try:
        dt = datetime(y, mo, d)
    except Exception:
        return ""
    return dt.strftime("%Y-%m-%d")


def _log_audit(db, project_name, action, detail="", username=""):
    try:
        db.add_audit_log(project_name, action, detail, username)
    except Exception as e:
        logger.warning("写入审计日志失败: %s", e)


def aggregate_editor_workload(db, month=None):
    """统一剪辑师工作量口径：从项目 episode_plan（分集详情）聚合每人集数。
    以当前 DB 分集详情为最终数据（用户在分集管理手动校准过的 {集号:剪辑师}）。
    month: 若指定 'YYYY-MM'，仅统计 project_month==month 的项目；否则统计全部。
    返回 [{'name','assigned','projects'}]（按 assigned 降序），剔除空/空白剪辑师。
    """
    import json as _json
    from collections import Counter as _Counter
    try:
        projs = db.get_all_projects() or []
    except Exception:
        projs = []
    editor_workload = {}
    for p in projs:
        if month and (p.get("project_month") or "") != month:
            continue
        # 从 episode_plan（分集详情，最终数据）聚合
        ep = p.get("episode_plan") or ""
        if not ep or ep == "{}":
            continue
        try:
            plan = _json.loads(ep) if isinstance(ep, str) else (ep or {})
        except Exception:
            continue
        if not isinstance(plan, dict) or not plan:
            continue
        tmp = _Counter()
        for _epn, editor in plan.items():
            if editor and str(editor).strip():
                tmp[str(editor).strip()] += 1
        for editor, cnt in tmp.items():
            ed = str(editor).strip()
            if not ed or not cnt:
                continue
            item = editor_workload.setdefault(ed, {"assigned": 0, "projects": set()})
            item["assigned"] += cnt
            item["projects"].add(p.get("name") or "")
    return [
        {"name": name, "assigned": item["assigned"], "projects": len(item["projects"])}
        for name, item in sorted(editor_workload.items(), key=lambda x: -x[1]["assigned"])
    ]


def sync_workload_from_episode_plan(db=None):
    """以当前分集数据(episode_plan)为准，重建各项目的 editor_workload。
    统计口径始终跟随分集数据：本项目循环读取每个有 episode_plan 的项目，
    按 {集号:剪辑师} 聚合每人集数，写入 editor_workload（供向下兼容）。
    不读取任何目标 Excel，不覆盖 episode_plan。返回统计结果。
    """
    if db is None:
        from db import db as _db
        db = _db
    import json as _json
    from collections import Counter as _Counter
    try:
        projs = db.get_all_projects() or []
    except Exception:
        projs = []
    synced = []
    total = 0
    for p in projs:
        ep = p.get("episode_plan") or ""
        if not ep or ep == "{}":
            continue
        try:
            plan = _json.loads(ep) if isinstance(ep, str) else (ep or {})
        except Exception:
            continue
        if not isinstance(plan, dict) or not plan:
            continue
        cnt = _Counter()
        for _epn, editor in plan.items():
            if editor and str(editor).strip():
                cnt[str(editor).strip()] += 1
        per_editor = dict(cnt)
        if per_editor:
            db.set_editor_workload(p.get("name") or "", per_editor)
            synced.append({"name": p.get("name") or "", "episodes": len(plan),
                           "editor_count": len(per_editor)})
            total += len(plan)
    return {"synced": synced, "total_projects": len(synced),
            "total_episodes": total, "skipped": []}


def compute_insights_summary(db, month=None):
    """计算数据洞察 KPI 汇总（统一口径，与 compute_overview_stats 一致）：
    - 本月项目：project_month==month 且有制作痕迹
    - 本月已完成：本月项目里 custom_status=='已完成'
    - 制作中：本月项目里处于进行中状态
    纯函数，供 /api/insights/summary 调用，便于单测。
    """
    import json as _json
    from datetime import datetime
    if not month:
        month = datetime.now().strftime("%Y-%m")
    try:
        from scan import _is_active_project, _is_producing
    except Exception:
        _is_active_project = lambda p: bool((p.get("custom_status") or "").strip()
                                            or (p.get("total_episodes") or 0) > 0
                                            or (p.get("delivery_status") or "") not in ("", "pending"))
        _is_producing = lambda p: str(p.get("custom_status") or "").strip() not in ("", "已完成")
    projs = db.get_all_projects() or []
    active_projs = [p for p in projs if _is_active_project(p)]
    month_projs = [p for p in active_projs if (p.get("project_month") or "") == month]
    month_completed = [p for p in month_projs
                       if str(p.get("custom_status") or "").strip() == "已完成"]
    in_progress = [p for p in month_projs if _is_producing(p)]
    status_map = {}
    for p in month_projs:
        st = (p.get("custom_status") or p.get("delivery_status") or "未设置").strip() or "未设置"
        status_map[st] = status_map.get(st, 0) + 1
    editor_eps = {}
    for ed in aggregate_editor_workload(db, month=month):
        editor_eps[ed["name"]] = ed["assigned"]
    members = 0
    try:
        members = len(db.list_members() or []) if hasattr(db, "list_members") else 0
    except Exception:
        pass
    return {
        "month": month,
        "projectCount": len(active_projs),
        "monthProjectCount": len(month_projs),
        "monthCompleted": len(month_completed),
        "inProgress": len(in_progress),
        "memberCount": members,
        "statusMap": status_map,
        "editorEpisodes": editor_eps,
    }


def _parse_assign_range_str(plan, line):
    """解析一行 '剪辑师：1-3，44-45' 加入 plan 字典（{集号str: 剪辑师}）。

    统一委托给 fenji_parser.parse_assign_line（唯一实现，避免多份漂移）。
    """
    from fenji_parser import parse_assign_line
    parse_assign_line(plan, line)


def sync_episode_plan_from_target(target_path, db=None, clear_stale=True):
    """从分集目标表格（权威来源）重建各项目的 episode_plan 与 editor_workload。
    读取每个项目块第 C 列的 '剪辑师：集数范围'：
      - episode_plan  = {集号: 剪辑师}（供分集 UI 用，重叠会被后写覆盖）
      - editor_workload = {剪辑师: 集数}（独立计数，不被重叠覆盖，作为统计口径唯一来源）
    返回 {synced, skipped, cleared, total_episodes, total_projects}。
    若 clear_stale=True：把 DB 中已有数据但不在目标文件里的项目清空。
    """
    if db is None:
        from db import db as _db
        db = _db
    import openpyxl
    import re as _re
    wb = openpyxl.load_workbook(target_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    months = ['一月份','二月份','三月份','四月份','五月份','六月份',
              '七月份','八月份','九月份','十月份','十一月份','十二月份']
    # 项目名 -> (episode_plan dict, editor_workload dict{剪辑师:独立集数})
    raw_plans = {}
    raw_workloads = {}
    cur_name = None
    for row in ws.iter_rows(values_only=True):
        if not row or not any(row):
            continue
        a = str(row[0] or '').strip() if row[0] else ''
        c = str(row[2] or '').strip() if len(row) > 2 and row[2] else ''
        if a and a not in months:
            cur_name = a
            raw_plans.setdefault(cur_name, {})
            raw_workloads.setdefault(cur_name, {})
        elif c and cur_name:
            _parse_assign_range_str(raw_plans[cur_name], c)
            # 独立计数（逐行算每人出现的集数，不与其他行去重）
            if '：' in c:
                parts = c.split('：', 1)
            elif ':' in c:
                parts = c.split(':', 1)
            else:
                parts = []
            if len(parts) == 2:
                editor = parts[0].strip()
                if editor:
                    from fenji_parser import expand_ranges
                    _eps = expand_ranges(parts[1])
                    if _eps:
                        raw_workloads[cur_name][editor] = raw_workloads[cur_name].get(editor, 0) + len(_eps)
    # 空项目块剔除
    raw_plans = {k: v for k, v in raw_plans.items() if v}
    # 匹配 DB 项目并覆盖写入
    all_proj = db.get_all_projects() or []
    name_to_db = {}
    for p in all_proj:
        n = p.get("name") or ""
        name_to_db[n] = p
    synced = []
    skipped = []
    matched_db_names = set()
    for tname, plan in raw_plans.items():
        proj = name_to_db.get(tname)
        if proj is None:
            # 兜底：按项目名模糊匹配
            t_clean = tname.split('（')[0].split('(')[0].strip()
            for n, p in name_to_db.items():
                if n == tname or (t_clean and t_clean in n) or (tname and n in tname):
                    proj = p
                    break
        if proj is None:
            skipped.append({'name': tname, 'episodes': len(plan), 'reason': '未找到对应项目'})
            continue
        db.set_episode_plan(proj['name'], plan)
        db.set_editor_workload(proj['name'], raw_workloads.get(tname, {}))
        matched_db_names.add(proj['name'])
        total = int(proj.get('total_episodes') or 0)
        if total < len(plan):
            try:
                db.set_episodes(proj['name'], len(plan), len(plan))
            except Exception:
                pass
        synced.append({'name': proj['name'], 'episodes': len(plan),
                       'editor_count': len(set(plan.values()))})
    # 清空 DB 中已有 plan 但不在目标文件的项目（避免过时数据污染）
    cleared = []
    if clear_stale:
        for p in all_proj:
            n = p.get("name") or ""
            ep = p.get("episode_plan") or ""
            if n in matched_db_names:
                continue
            if ep and ep != "{}":
                try:
                    plan = json.loads(ep)
                except Exception:
                    plan = {}
                if plan:
                    db.set_episode_plan(n, {})
                    db.set_editor_workload(n, {})
                    cleared.append(n)
    return {'synced': synced, 'skipped': skipped, 'cleared': cleared,
            'total_episodes': sum(s['episodes'] for s in synced),
            'total_projects': len(synced)}


def sync_delivery_dates_from_target(target_path, db=None):
    """从分集目标表格读取胶片日期（col4），解析为 YYYY-MM-DD，写入项目 delivered_date。
    返回已更新项目列表 [{project, date}]。失败抛异常。供路由与分集导出自动调用。
    """
    if db is None:
        from db import db as _db
        db = _db
    import openpyxl
    wb = openpyxl.load_workbook(target_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    updated = []
    seen_names = set()
    cur_name = None
    cur_date = ""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        name = row[0].value if len(row) > 0 else None
        if name and str(name).strip():
            # 新项目块开始
            if cur_name and cur_date and cur_name not in seen_names:
                seen_names.add(cur_name)
                parsed = _parse_time_text(cur_date)
                if parsed:
                    db.update_project_status(cur_name, delivered_date=parsed)
                    # 功能8：目标文件里的胶片日期即"预计交付日期"，同步为 due_date
                    db.update_project_status(cur_name, due_date=parsed)
                    updated.append({"project": cur_name, "date": parsed})
            cur_name = str(name).strip()
            cur_date = row[3].value if len(row) > 3 else None
    # 处理最后一个块
    if cur_name and cur_date and cur_name not in seen_names:
        seen_names.add(cur_name)
        parsed = _parse_time_text(cur_date)
        if parsed:
            db.update_project_status(cur_name, delivered_date=parsed)
            db.update_project_status(cur_name, due_date=parsed)
            updated.append({"project": cur_name, "date": parsed})
    return updated


def register_routes(app, db):
    # ==================== 项目待办事项 ====================
    @app.route("/api/project/<name>/todos", methods=["GET"])
    def features_todos_get(name):
        return jsonify({"ok": True, "todos": db.get_project_todos(name)})

    @app.route("/api/project/<name>/todos", methods=["POST"])
    def features_todos_add(name):
        data = request.get_json(silent=True) or {}
        text = (data.get("text") or "").strip()
        if not text:
            return jsonify({"ok": False, "message": "请输入待办内容"}), 400
        todo_id = db.add_project_todo(name, text, data.get("priority", 0))
        _log_audit(db, name, "添加待办", text)
        return jsonify({"ok": True, "id": todo_id})

    @app.route("/api/project/<name>/todos/<int:todo_id>", methods=["PUT"])
    def features_todos_update(name, todo_id):
        data = request.get_json(silent=True) or {}
        if "done" in data:
            db.update_project_todo(todo_id, done=bool(data.get("done")))
            _log_audit(db, name, "完成待办" if data.get("done") else "取消待办")
        if "text" in data:
            db.update_project_todo(todo_id, text=data.get("text"))
        if "priority" in data:
            db.update_project_todo(todo_id, priority=data.get("priority"))
        return jsonify({"ok": True})

    @app.route("/api/project/<name>/todos/<int:todo_id>", methods=["DELETE"])
    def features_todos_delete(name, todo_id):
        db.delete_project_todo(todo_id)
        _log_audit(db, name, "删除待办")
        return jsonify({"ok": True})

    @app.route("/api/todos/global", methods=["GET"])
    def features_todos_global():
        """跨项目全局待办：?done=1 含已完成；?q=关键词 过滤；?project=指定项目。"""
        include_done = (request.args.get("done") or "") == "1"
        kw = (request.args.get("q") or "").strip()
        proj = (request.args.get("project") or "").strip()
        rows = db.get_all_todos(include_done=include_done, keyword=kw or proj)
        # 若只按项目过滤，需保留该项目所有待办（含已完成按需）
        return jsonify({"ok": True, "todos": rows, "count": len(rows)})

    # ==================== 项目时间轴 ====================
    @app.route("/api/project/<name>/timeline")
    def features_timeline(name):
        events = []
        # 1. 项目创建
        proj = db.get_project(name)
        if proj and proj.get("created_at"):
            events.append({
                "time": proj["created_at"], "type": "create",
                "title": "📁 项目创建",
                "detail": "项目「%s」创建" % name,
                "icon": "📁", "color": "#0071e3",
            })
        # 2. 交付日志
        try:
            logs = db.get_delivery_logs(name, limit=500)
            for l in logs:
                st = (l.get("status") or "").lower()
                if st in ("fail", "error"):
                    icon, color = "❌", "#ff3b30"
                elif st in ("warn", "warning"):
                    icon, color = "⚠️", "#ff9500"
                else:
                    icon, color = "📋", "#34c759"
                events.append({
                    "time": l.get("created_at", ""), "type": "delivery",
                    "title": "📋 交付 " + (l.get("file_name") or ""),
                    "detail": (l.get("message") or l.get("dest_path") or ""),
                    "icon": icon, "color": color,
                })
        except Exception as e:
            logger.warning("时间轴交付日志失败: %s", e)
        # 3. 同步日志
        try:
            logs = db.get_sync_logs(name, limit=200)
            for l in logs:
                icon, color = "🔃", "#5856d6"
                events.append({
                    "time": l.get("created_at", ""), "type": "sync",
                    "title": "🔃 同步 " + (l.get("action") or ""),
                    "detail": (l.get("file_path") or l.get("message") or ""),
                    "icon": icon, "color": color,
                })
        except Exception:
            pass
        # 4. QA 质检
        try:
            qas = db.list_qa_runs_for_project(name, limit=50)
            for q in qas:
                st = (q.get("status") or "").lower()
                if st == "pass":
                    icon, color = "✅", "#34c759"
                elif st == "fail":
                    icon, color = "❌", "#ff3b30"
                else:
                    icon, color = "🔍", "#ff9500"
                events.append({
                    "time": q.get("started_at") or q.get("created_at", ""), "type": "qa",
                    "title": "🔍 质检",
                    "detail": "通过 %s / 失败 %s" % (q.get("passed", 0), q.get("failed", 0)),
                    "icon": icon, "color": color,
                })
        except Exception:
            pass
        # 5. 审计日志
        try:
            for l in db.get_audit_logs(name, limit=100):
                events.append({
                    "time": l.get("created_at", ""), "type": "audit",
                    "title": "🔧 " + (l.get("action") or "操作"),
                    "detail": l.get("detail") or "",
                    "icon": "🔧", "color": "#8e8e93",
                })
        except Exception:
            pass

        # 按时间排序
        events.sort(key=lambda x: x.get("time") or "")
        # 摘要
        summary = {
            "totalEvents": len(events),
            "firstEvent": events[0]["time"] if events else None,
            "lastEvent": events[-1]["time"] if events else None,
            "totalDeliveries": sum(1 for e in events if e["type"] == "delivery"),
        }
        return jsonify({"ok": True, "project": name, "events": events, "summary": summary})

    # ==================== 数据备份 ====================
    @app.route("/api/backup/list")
    def features_backup_list():
        import backup_service
        return jsonify({"ok": True, "backups": backup_service.list_backups()})

    @app.route("/api/backup/create", methods=["POST"])
    def features_backup_create():
        import backup_service
        try:
            b = backup_service.backup_now()
            return jsonify({"ok": True, "backup": b})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/backup/restore", methods=["POST"])
    def features_backup_restore():
        import backup_service
        data = request.get_json(silent=True) or {}
        name = data.get("backupName") or data.get("name") or ""
        if not name:
            return jsonify({"ok": False, "message": "请选择备份文件"}), 400
        try:
            backup_service.restore(name)
            return jsonify({"ok": True, "message": "已恢复，请重启软件生效"})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    # ==================== 数据洞察（可选增强：大屏 / 日历 / 导出）====================
    @app.route("/api/insights/summary")
    def features_insights_summary():
        """KPI 汇总（以当月为基准），口径与 compute_overview_stats 一致。"""
        try:
            from datetime import datetime
            month = datetime.now().strftime("%Y-%m")
            data = compute_insights_summary(db, month=month)
            return jsonify({"ok": True, **data})
        except Exception as e:
            import traceback; traceback.print_exc()
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/insights/calendar")
    def features_insights_calendar():
        """交付日历：按项目 delivered_date 分组，返回每天交付的项目。
        ?month=YYYY-MM；?dept=部门名 可选按部门过滤。
        days: {YYYY-MM-DD: [{name, department}]}
        """
        month = request.args.get("month", "")
        if not month or len(month) != 7:
            from datetime import datetime
            month = datetime.now().strftime("%Y-%m")
        dept = (request.args.get("dept") or "").strip()
        prefix = month + "-"
        days = {}
        try:
            projs = db.get_all_projects() or []
            for p in projs:
                if dept and (p.get("department") or "") != dept:
                    continue
                dd = (p.get("delivered_date") or "").strip()
                if dd.startswith(prefix):
                    d = dd[:10]
                    days.setdefault(d, []).append({
                        "name": p.get("name") or "",
                        "department": p.get("department") or "",
                    })
        except Exception:
            pass
        for k in days:
            days[k] = sorted(days[k], key=lambda x: x.get("name") or "")
        return jsonify({"ok": True, "month": month, "dept": dept, "days": days})


def compute_delivery_stats(db, month=None):
    """交付日历增强统计（功能8）：
    - 当月项目交付情况：已交付/未交付
    - 按时交付率：有 due_date 且 delivered_date <= due_date 视为按时
    - 延迟预警：due_date 已过但未交付，或 delivered_date > due_date
    纯函数，供 /api/insights/delivery_stats 调用，便于单测。
    """
    from datetime import datetime
    if not month:
        month = datetime.now().strftime("%Y-%m")
    try:
        projs = db.get_all_projects() or []
    except Exception:
        projs = []
    # 筛选当月项目（project_month 或 delivered_date 落在本月）
    month_projs = [p for p in projs
                   if (p.get("project_month") or "").startswith(month)
                   or (p.get("delivered_date") or "").startswith(month)]
    delivered = []
    undelivered = []
    overdue = []       # due_date 已过且未交付
    late = []          # 已交付但 delivered_date > due_date
    ontime = 0
    today = datetime.now().date()

    def _d(s):
        s = (s or "").strip()
        if len(s) >= 10 and s[4] == "-":
            try:
                return datetime.strptime(s[:10], "%Y-%m-%d").date()
            except Exception:
                return None
        return None

    for p in month_projs:
        name = p.get("name") or ""
        dd = _d(p.get("delivered_date"))
        due = _d(p.get("due_date"))
        status = (p.get("custom_status") or "").strip()
        rec = {
            "name": name, "department": p.get("department") or "",
            "delivered_date": (p.get("delivered_date") or "").strip()[:10],
            "due_date": (p.get("due_date") or "").strip()[:10],
            "status": status,
        }
        if dd is not None:
            delivered.append(rec)
            if due is not None and dd <= due:
                ontime += 1
            elif due is not None and dd > due:
                rec["late"] = True
                late.append(rec)
        else:
            undelivered.append(rec)
            if due is not None and due < today:
                rec["overdue"] = True
                overdue.append(rec)
    total_delivered = len(delivered)
    with_due = [r for r in delivered if _d(r.get("due_date"))]
    on_time_rate = (ontime / len(with_due) * 100) if with_due else None
    return {
        "month": month,
        "delivered_count": total_delivered,
        "undelivered_count": len(undelivered),
        "on_time_rate": round(on_time_rate, 1) if on_time_rate is not None else None,
        "ontime_count": ontime,
        "late_count": len(late),
        "overdue_count": len(overdue),
        "delivered": sorted(delivered, key=lambda x: x["delivered_date"]),
        "undelivered": undelivered,
        "overdue": sorted(overdue, key=lambda x: x["due_date"]),
    }

    @app.route("/api/insights/delivery_stats")
    def features_insights_delivery_stats():
        """交付日历增强统计：按时交付率 + 延迟预警。?month=YYYY-MM"""
        try:
            from datetime import datetime
            month = request.args.get("month", "") or datetime.now().strftime("%Y-%m")
            data = compute_delivery_stats(db, month=month)
            return jsonify({"ok": True, **data})
        except Exception as e:
            import traceback; traceback.print_exc()
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/project/<name>/due_date", methods=["POST"])
    def features_set_due_date(name):
        """设置项目预计交付日期（due_date）。body: {date:'YYYY-MM-DD' 或 ''清空}"""
        data = request.get_json(silent=True) or {}
        date = (data.get("date") or "").strip()
        if date and len(date) == 10 and date[4] == "-" and date[7] == "-":
            db.update_project_status(name, due_date=date)
            _log_audit(db, name, "设置预计交付日期", date)
            return jsonify({"ok": True, "date": date})
        db.update_project_status(name, due_date="")
        _log_audit(db, name, "清除预计交付日期")
        return jsonify({"ok": True, "date": ""})

    @app.route("/api/project/<name>/owner", methods=["POST"])
    def features_set_owner(name):
        """设置项目负责人（审核流责任到人）。body: {owner: '姓名' 或 ''清空}"""
        data = request.get_json(silent=True) or {}
        owner = (data.get("owner") or "").strip()
        db.update_project_status(name, owner=owner)
        _log_audit(db, name, "设置负责人" if owner else "清除负责人", owner)
        return jsonify({"ok": True, "owner": owner})

    @app.route("/api/audit/alerts", methods=["GET"])
    def features_audit_alerts():
        """审核流超时提醒。?stale_days=3"""
        try:
            stale = int(request.args.get("stale_days", "3") or "3")
        except Exception:
            stale = 3
        data = compute_audit_alerts(db, stale_days=stale)
        return jsonify({"ok": True, **data})

    @app.route("/api/insights/calendar/export")
    def features_insights_calendar_export():
        """导出某月交付清单 CSV。?month=YYYY-MM；?dept=可选按部门过滤。"""
        import io as _io
        import csv as _csv
        month = request.args.get("month", "")
        if not month or len(month) != 7:
            from datetime import datetime
            month = datetime.now().strftime("%Y-%m")
        dept = (request.args.get("dept") or "").strip()
        prefix = month + "-"
        rows = []
        try:
            projs = db.get_all_projects() or []
            for p in projs:
                if dept and (p.get("department") or "") != dept:
                    continue
                dd = (p.get("delivered_date") or "").strip()
                if dd.startswith(prefix):
                    rows.append([dd[:10], p.get("name") or "", p.get("department") or ""])
        except Exception:
            pass
        rows.sort(key=lambda r: r[0])
        buf = _io.StringIO()
        writer = _csv.writer(buf)
        writer.writerow(["交付日期", "项目名", "部门"])
        writer.writerows(rows)
        data = "\ufeff" + buf.getvalue()
        from flask import Response
        return Response(data, mimetype="text/csv; charset=utf-8",
                        headers={"Content-Disposition": f"attachment; filename=delivery-{month}.csv"})

    @app.route("/api/project/<name>/delivered_date", methods=["POST"])
    def features_set_delivered_date(name):
        """设置项目交付/归档日期（交付日历用）。body: {date:'YYYY-MM-DD' 或 ''清空}"""
        data = request.get_json(silent=True) or {}
        date = (data.get("date") or "").strip()
        if date and len(date) == 10 and date[4] == "-" and date[7] == "-":
            db.update_project_status(name, delivered_date=date)
            _log_audit(db, name, "设置交付日期", date)
            return jsonify({"ok": True, "date": date})
        db.update_project_status(name, delivered_date="")
        _log_audit(db, name, "清除交付日期")
        return jsonify({"ok": True, "date": ""})

    @app.route("/api/insights/sync_delivery_dates", methods=["POST"])
    def features_sync_delivery_dates():
        """从分集目标文件表格读取胶片日期（col4），解析为 YYYY-MM-DD，
        更新到项目 delivered_date 并刷新日历。body: {target_path?: 路径}"""
        # 目标文件路径：优先 body，其次设置，再扫描 data/fenji_targets
        data = request.get_json(silent=True) or {}
        target = (data.get("target_path") or "").strip()
        if not target:
            try:
                target = db.get_all_settings().get("fj_target_path", "") or ""
            except Exception:
                target = ""
        if not target or not os.path.isfile(target):
            # 尝试自动查找最近的目标文件
            tgt_dir = os.path.join(_DATA_DIR, "fenji_targets")
            if os.path.isdir(tgt_dir):
                xlsx = [f for f in os.listdir(tgt_dir) if f.lower().endswith((".xlsx", ".xlsm", ".xls"))]
                if xlsx:
                    target = os.path.join(tgt_dir, sorted(xlsx)[-1])
        if not target or not os.path.isfile(target):
            return jsonify({"ok": False, "message": "未找到目标文件，请先在分集管理设置目标文件"}), 400
        try:
            updated = sync_delivery_dates_from_target(target, db)
            return jsonify({"ok": True, "updated": len(updated), "target": target, "items": updated})
        except Exception as e:
            return jsonify({"ok": False, "message": "解析失败: " + str(e)}), 500

    @app.route("/api/insights/export")
    def features_insights_export():
        """导出项目档案 CSV（项目名/状态/总集数/已生成/部门/路径/交付日期/创建时间）。
        ?save=1 时保存到 data/exports/ 并返回路径（桌面端用），否则返回流式下载。"""
        import io as _io
        import csv as _csv
        projs = db.get_all_projects() or []
        buf = _io.StringIO()
        writer = _csv.writer(buf)
        writer.writerow(["项目名", "部门", "当前状态", "交付状态", "总集数", "已生成集数", "交付日期", "组内路径", "制作路径", "创建时间"])
        for p in projs:
            writer.writerow([
                p.get("name", ""),
                p.get("department", ""),
                p.get("custom_status", ""),
                p.get("delivery_status", ""),
                p.get("total_episodes", 0),
                p.get("current_episodes", 0),
                p.get("delivered_date", ""),
                p.get("group_path", ""),
                p.get("production_path", ""),
                p.get("created_at", ""),
            ])
        data = "\ufeff" + buf.getvalue()  # BOM 便于 Excel 打开中文
        save_mode = (request.args.get("save") or "") == "1"
        if save_mode:
            from datetime import datetime as _dt
            try:
                exp_dir = os.path.join(_DATA_DIR, "exports")
                os.makedirs(exp_dir, exist_ok=True)
                fname = "项目档案-" + _dt.now().strftime("%Y%m%d-%H%M%S") + ".csv"
                fpath = os.path.join(exp_dir, fname)
                with open(fpath, "w", encoding="utf-8-sig", newline="") as f:
                    f.write(data)
                return jsonify({
                    "ok": True,
                    "path": fpath,
                    "filename": fname,
                    "count": len(projs),
                    "dir": exp_dir,
                })
            except Exception as e:
                return jsonify({"ok": False, "message": "保存失败: " + str(e)}), 500
        from flask import Response
        return Response(
            data,
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=projects.csv"})

    @app.route("/api/insights/export/open_folder", methods=["POST"])
    def features_insights_export_open_folder():
        """打开导出文件夹（桌面端用 explorer 打开）。"""
        exp_dir = os.path.join(_DATA_DIR, "exports")
        os.makedirs(exp_dir, exist_ok=True)
        try:
            if os.name == "nt":
                subprocess.Popen(["explorer", exp_dir])
            else:
                subprocess.Popen(["xdg-open", exp_dir])
            return jsonify({"ok": True, "dir": exp_dir})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/search", methods=["GET"])
    def features_global_search():
        """全局搜索：?q=关键词，跨 项目名/月份/部门/状态/剪辑师/交付日期/待办内容 匹配。
        返回分组结果：projects / editors / todos / delivered_dates。
        """
        import json as _json
        q = (request.args.get("q") or "").strip().lower()
        if not q:
            return jsonify({"ok": True, "projects": [], "editors": [], "todos": [], "delivered_dates": []})
        try:
            projs = db.get_all_projects() or []
        except Exception:
            projs = []
        projects = []
        editors = {}
        delivered_dates = []
        for p in projs:
            name = str(p.get("name") or "")
            month = str(p.get("project_month") or "")
            dept = str(p.get("department") or "")
            status = str(p.get("custom_status") or "")
            dd = str(p.get("delivered_date") or "")
            hay = " ".join([name.lower(), month.lower(), dept.lower(), status.lower(), dd.lower()])
            # 剪辑师匹配（episode_plan）
            plan = p.get("episode_plan") or "{}"
            try:
                plan = _json.loads(plan) if isinstance(plan, str) else (plan or {})
            except Exception:
                plan = {}
            if isinstance(plan, dict):
                for ep, ed in plan.items():
                    if ed and q in str(ed).lower():
                        editors.setdefault(str(ed).strip(), {"count": 0, "projects": set()})
                        editors[str(ed).strip()]["count"] += 1
                        editors[str(ed).strip()]["projects"].add(name)
            if q in hay:
                projects.append({
                    "name": name, "department": dept, "month": month, "status": status,
                    "delivered_date": dd,
                })
            if dd and q in dd.lower():
                delivered_dates.append({"name": name, "date": dd, "department": dept})
        # 待办匹配
        try:
            todos = db.get_all_todos(include_done=True, keyword=q) or []
        except Exception:
            todos = []
        todo_results = [{"project": t.get("project_name"), "text": t.get("text"), "id": t.get("id")} for t in todos]
        return jsonify({
            "ok": True,
            "q": q,
            "projects": projects[:50],
            "editors": [{"name": k, "count": v["count"], "projects": len(v["projects"])} for k, v in editors.items()],
            "todos": todo_results[:50],
            "delivered_dates": delivered_dates[:50],
        })

    @app.route("/api/notifications", methods=["GET"])
    def features_notifications():
        """通知中心：交付提醒 + 待办提醒。
        交付：逾期(已到交付日未完成) / 今日交付 / 未来3天内交付。
        待办：未完成待办按项目分组。
        """
        from datetime import datetime as _dt, timedelta as _td
        today = _dt.now().strftime("%Y-%m-%d")
        day3 = (_dt.now() + _td(days=3)).strftime("%Y-%m-%d")
        try:
            projs = db.get_all_projects() or []
        except Exception:
            projs = []
        overdue = []
        today_deliver = []
        upcoming = []
        for p in projs:
            dd = (p.get("delivered_date") or "").strip()
            st = str(p.get("custom_status") or "").strip()
            if not dd or len(dd) != 10:
                continue
            done = (st == "已完成")
            if dd == today:
                today_deliver.append({"name": p.get("name") or "", "department": p.get("department") or ""})
            elif dd < today and not done:
                overdue.append({"name": p.get("name") or "", "department": p.get("department") or "", "date": dd})
            elif today < dd <= day3 and not done:
                upcoming.append({"name": p.get("name") or "", "department": p.get("department") or "", "date": dd})
        # 待办提醒：未完成待办按项目聚合
        try:
            todos = db.get_all_todos(include_done=False) or []
        except Exception:
            todos = []
        todo_by_project = {}
        for t in todos:
            todo_by_project.setdefault(t.get("project_name") or "", []).append(t)
        todo_reminders = [
            {"project": pname, "count": len(items),
             "items": [{"id": t.get("id"), "text": t.get("text")} for t in items]}
            for pname, items in todo_by_project.items()
        ]
        overdue.sort(key=lambda x: x.get("date", ""))
        upcoming.sort(key=lambda x: x.get("date", ""))
        return jsonify({
            "ok": True,
            "today": today,
            "overdue": overdue,
            "today_deliver": today_deliver,
            "upcoming": upcoming,
            "todos": todo_reminders,
            "count": len(overdue) + len(today_deliver) + len(upcoming) + len(todo_reminders),
        })

    # ==================== 视频缩略图（ffmpeg）====================
    @app.route("/api/thumbnail")
    def features_thumbnail():
        """生成视频缩略图：?path=绝对路径。生成后返回图片，带缓存。"""
        file_path = request.args.get("path", "")
        if not file_path or not os.path.isfile(file_path):
            return jsonify({"ok": False, "message": "文件不存在"}), 404
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in (".mp4", ".mov", ".mxf", ".avi", ".mkv", ".webm", ".wmv", ".m4v"):
            return jsonify({"ok": False, "message": "非视频文件"}), 400
        # 生成缩略图（缓存到 data/thumbs）
        try:
            os.makedirs(_THUMB_DIR, exist_ok=True)
            import hashlib
            digest = hashlib.md5(file_path.encode("utf-8", "ignore")).hexdigest()
            thumb = os.path.join(_THUMB_DIR, digest + ".jpg")
            if not os.path.exists(thumb):
                _gen_thumbnail(file_path, thumb)
            if os.path.exists(thumb):
                return send_file(thumb, mimetype="image/jpeg", conditional=True)
            return jsonify({"ok": False, "message": "无法生成缩略图"}), 422
        except Exception as e:
            logger.warning("缩略图失败: %s", e)
            return jsonify({"ok": False, "message": str(e)}), 500


def _gen_thumbnail(file_path, out_path):
    """用 ffmpeg 抽取视频中段一帧生成为 JPG 缩略图。"""
    ffmpeg = _resolve_ffmpeg_bin()
    if not ffmpeg:
        raise RuntimeError("未找到 ffmpeg")
    cmd = [ffmpeg, "-y", "-i", file_path,
           "-vf", "scale=320:-1", "-frames:v", "1", "-q:v", "5", out_path]
    try:
        subprocess.run(cmd, capture_output=True, timeout=20, check=False)
    except Exception:
        pass


def _resolve_ffmpeg_bin():
    """查找 ffmpeg：config.yaml 的 qa.ffmpeg_path → PATH → 常见目录。"""
    p = _yaml_get("qa.ffmpeg_path", "")
    if p and os.path.isfile(p):
        return p
    import shutil
    p = shutil.which("ffmpeg")
    if p:
        return p
    candidates = [
        r"C:\Users\Admin\Desktop\视频检查工具\视频质检工具\dist\视频质检工具\ffmpeg.exe",
    ]
    for c in candidates:
        if os.path.isfile(c):
            return c
    return None
