"""Additional routes for workbench enhancements."""
import os as _os
import re as _re
import io as _io
import base64 as _b64
import json as _j
import yaml as _y
import logging as _logging
from flask import request, jsonify, send_file, send_from_directory, abort
import config as _cfg
from utils import scan_dir
from fenji_exporter import export_from_template, backup_template, list_templates as _list_templates

_logger = _logging.getLogger(__name__)

# =====================================================================
# 配置收敛：统一以 backend/config.yaml 为唯一配置源。
# 旧实现通过 config.py 读写 data/config.json（与 app.py 使用的 config.yaml
# 不一致，导致设置页面改动对主流程无效）。这里新增 yaml 读写封装并替换
# NAS 路径相关读取，使 enhanced_routes 与 app.py / sync_engine 同源。
# =====================================================================
_CONFIG_YAML = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "config.yaml")


def _yaml_cfg():
    """读取 config.yaml 全量 dict（失败返回空 dict）。"""
    try:
        with open(_CONFIG_YAML, "r", encoding="utf-8") as f:
            return _y.safe_load(f) or {}
    except Exception:
        return {}


def _yaml_get(dotted, default=None):
    """按 a.b.c 点路径读取 yaml 配置。"""
    cur = _yaml_cfg()
    for part in dotted.split("."):
        if isinstance(cur, dict) and part in cur:
            cur = cur[part]
        else:
            return default
    return cur


def _yaml_save(new_cfg):
    """整体写回 config.yaml（保留未知键，仅更新给定结构）。"""
    merged = _yaml_cfg()
    merged.update(new_cfg)
    tmp = _CONFIG_YAML + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        _y.dump(merged, f, allow_unicode=True, default_flow_style=False,
                sort_keys=False)
    _os.replace(tmp, _CONFIG_YAML)


# 模板/备份存放目录（相对于 backend/ 父目录的 data/fenji_templates 和 data/fenji_backups）
_DATA_DIR = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), 'data')
_TEMPLATE_DIR = _os.path.join(_DATA_DIR, 'fenji_templates')
_BACKUP_DIR = _os.path.join(_DATA_DIR, 'fenji_backups')


def _nas_search_roots():
    """Build [(root, path_key), ...] from config.yaml NAS section."""
    group_root = _yaml_get("nas.group_root", "")
    production_roots = _yaml_get("nas.production_roots", []) or []
    roots = []
    if group_root:
        roots.append((group_root, "group_path"))
    for p in production_roots:
        roots.append((p, "production_path"))
    if group_root:
        roots.append((_os.path.join(group_root, "00已完成"), "group_path"))
    return roots


def _production_labels():
    """Return path -> label map from config."""
    return _yaml_get("nas.production_labels", {}) or {}


# episode_summary 扫描参数：key = 部门标签（production_labels 的 value 或 "group_root"）
_EPISODE_SCAN_OVERRIDES = {
    "group_root": {"skip": ["00已完成", "0000新人"], "depth": 0},
    "AI漫剧二部": {"skip": ["00序列_ID_剧名"], "depth": 0},
    "AI漫剧一部海外": {"skip": ["AAA《xxxxxxxx》文件夹模板", "00序列_ID_剧名"], "depth": 0},
    "AI漫剧九部海外": {"skip": [], "depth": 1},
    "AI漫剧六部中转": {"skip": [], "depth": 0},
}


def _episode_summary_scan_tasks():
    """Build [(path, skip_names, depth), ...] from config NAS + scan overrides."""
    tasks = []
    group_root = _yaml_get("nas.group_root", "")
    production_roots = _yaml_get("nas.production_roots", []) or []
    labels = _production_labels()

    if group_root:
        o = _EPISODE_SCAN_OVERRIDES["group_root"]
        tasks.append((group_root, o["skip"], o["depth"]))

    for root in production_roots:
        label = labels.get(root, "")
        o = _EPISODE_SCAN_OVERRIDES.get(label, {"skip": [], "depth": 0})
        tasks.append((root, o["skip"], o["depth"]))
    return tasks

def _parse_assign_line(plan, line):
    """解析一行 '剪辑师：1-3，44-45' 加入 plan（模块级辅助函数）。"""
    line = (line or "").strip()
    if '：' in line:
        parts = line.split('：', 1)
    elif ':' in line:
        parts = line.split(':', 1)
    else:
        return
    if len(parts) < 2:
        return
    editor = parts[0].strip()
    ranges = parts[1].strip()
    if not editor or not ranges:
        return
    for r in _re.split(r'[,，、;\s]+', ranges):
        r = r.strip()
        if not r:
            continue
        m = _re.match(r'^(\d+)\s*[-—~]\s*(\d+)$', r)
        if m:
            s, e = int(m.group(1)), int(m.group(2))
            for ep in range(min(s, e), max(s, e) + 1):
                plan[str(ep)] = editor
        else:
            m2 = _re.match(r'^(\d+)$', r)
            if m2:
                plan[m2.group(1)] = editor

def _auto_sync_delivery_dates(target_path):
    """分集导出写入目标文件后，自动从表格读取胶片日期更新项目交付日期。"""
    try:
        import features as _feat
        from db import db as _db
        updated = _feat.sync_delivery_dates_from_target(target_path, _db)
        _logger.info("自动同步交付日期完成: %d 个项目", len(updated))
    except Exception as e:
        _logger.warning("自动同步交付日期失败: %s", e)


def _register_enhanced_routes(app, db, qa_engine=None, sync_engine=None):
    @app.route("/api/project/<path:project_name>", methods=["GET"])
    def api_project_detail(project_name):
        p = None
        try:
            p = db.get_project(project_name)
        except Exception:
            pass
        if p:
            return jsonify({"ok": True, "project": p})
        search_roots = _nas_search_roots()
        for root, path_key in search_roots:
            candidate = _os.path.join(root, project_name)
            if _os.path.isdir(candidate):
                proj = {
                    "name": project_name,
                    "production_path": candidate if path_key == "production_path" else "",
                    "group_path": candidate if path_key == "group_path" else "",
                    "custom_status": "",
                    "delivery_status": "",
                    "total_episodes": 0,
                    "current_episodes": 0,
                }
                return jsonify({"ok": True, "project": proj})
        return jsonify({"ok": False, "message": "项目不存在"}), 404

    @app.route("/api/project", methods=["POST"])
    def api_project_create():
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"ok": False, "message": "项目名不能为空"}), 400
        try:
            db.upsert_project(
                name,
                production_path=data.get("production_path", ""),
                group_path=data.get("group_path", ""),
                source_root=data.get("source_root", ""),
            )
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500
        return jsonify({"ok": True, "project": name})

    @app.route("/api/project/<path:project_name>", methods=["DELETE"])
    def api_project_delete(project_name):
        try:
            db.delete_project(project_name)
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500
        return jsonify({"ok": True})

    # ============================================================
    # 新版 QA 路由（完全照搬独立视频质检工具的功能）
    # ============================================================

    @app.route("/api/qa/project_dir", methods=["GET"])
    def api_qa_project_dir():
        """根据项目名返回质检目录（优先 000交付 文件夹路径）。

        Query: name=项目名
        返回: {ok, project_path, project_name, source}
            - project_path: 质检目录绝对路径（000交付 的父目录或项目根目录）
            - source: "delivery" | "group" | "production" | "not_found"
        """
        name = request.args.get("name", "").strip()
        if not name:
            return jsonify({"ok": False, "message": "项目名不能为空"}), 400

        # 1. 先从 DB 查项目记录
        proj = None
        try:
            proj = db.get_project(name)
        except Exception:
            pass

        group_path = ""
        production_path = ""
        if proj:
            group_path = proj.get("group_path", "") or ""
            production_path = proj.get("production_path", "") or ""

        # 2. 如果 DB 没有，从 NAS 搜索根目录查找
        if not group_path and not production_path:
            for root, path_key in _nas_search_roots():
                candidate = _os.path.join(root, name)
                if _os.path.isdir(candidate):
                    if path_key == "group_path":
                        group_path = candidate
                    else:
                        production_path = candidate
                    break

        # 3. 优先查找 000交付 文件夹
        for base in [group_path, production_path]:
            if not base:
                continue
            delivery = _os.path.join(base, "000交付")
            if _os.path.isdir(delivery):
                return jsonify({
                    "ok": True,
                    "project_path": delivery,
                    "project_name": name,
                    "source": "delivery",
                    "group_path": group_path,
                    "production_path": production_path,
                })

        # 4. 没有 000交付，回退到项目根目录
        for base in [group_path, production_path]:
            if base and _os.path.isdir(base):
                return jsonify({
                    "ok": True,
                    "project_path": base,
                    "project_name": name,
                    "source": "group" if base == group_path else "production",
                    "group_path": group_path,
                    "production_path": production_path,
                })

        return jsonify({"ok": False, "message": "未找到项目目录"}), 404

    @app.route("/api/qa/browse_dirs", methods=["GET"])
    def api_qa_browse_dirs():
        """返回 NAS 根目录下的项目文件夹列表，供前端浏览选择。

        Query: root=group|production (默认 group)
        返回: {ok, dirs: [{name, path, has_delivery}]}
        """
        root_type = request.args.get("root", "group")
        dirs = []
        if root_type == "group":
            group_root = _yaml_get("nas.group_root", "")
            if group_root and _os.path.isdir(group_root):
                try:
                    for name in sorted(_os.listdir(group_root)):
                        full = _os.path.join(group_root, name)
                        if not _os.path.isdir(full):
                            continue
                        has_delivery = _os.path.isdir(_os.path.join(full, "000交付"))
                        dirs.append({"name": name, "path": full, "has_delivery": has_delivery})
                except OSError:
                    pass
        else:
            for prod_root in (_yaml_get("nas.production_roots", []) or []):
                if not prod_root or not _os.path.isdir(prod_root):
                    continue
                try:
                    for name in sorted(_os.listdir(prod_root)):
                        full = _os.path.join(prod_root, name)
                        if not _os.path.isdir(full):
                            continue
                        has_delivery = _os.path.isdir(_os.path.join(full, "000交付"))
                        dirs.append({"name": name, "path": full, "has_delivery": has_delivery})
                except OSError:
                    pass
        return jsonify({"ok": True, "dirs": dirs})

    @app.route("/api/qa/scan_dir", methods=["POST"])
    def api_qa_scan_dir():
        """POST JSON {project_path: "..."} → 返回文件夹布局（完全对应视频质检工具 auto_detect_folders）。

        body 字段:
            project_path (必填): 项目目录绝对路径
        """
        import qa_toolkits
        data = request.get_json(silent=True) or {}
        p = data.get("project_path") or ""
        if not p or not _os.path.isdir(p):
            return jsonify({"ok": False, "message": "路径不存在"}), 400
        layout = qa_toolkits.auto_detect_folders(p)
        suggested_name = qa_toolkits.auto_fill_project_name(p)
        return jsonify({"ok": True, "layout": layout, "project_name_suggest": suggested_name})

    @app.route("/api/qa/preview_frame", methods=["POST"])
    def api_qa_preview_frame():
        """POST JSON {video_path: "...", max_width: 720} → 返回 JPEG base64 + 尺寸。

        用于 Web 端字幕区域框选预览图（独立工具 SubRegionPicker 的等价物）。
        """
        import qa_toolkits
        data = request.get_json(silent=True) or {}
        vp = data.get("video_path") or ""
        if not vp or not _os.path.isfile(vp):
            return jsonify({"ok": False, "message": "视频文件不存在"}), 400
        frame, w, h = qa_toolkits.extract_preview_frame(vp, timestamp_fraction=0.25)
        if frame is None:
            return jsonify({"ok": False, "message": "无法提取帧，请检查 ffmpeg"}), 500
        maxw = int(data.get("max_width") or 960)
        jpeg_bytes = qa_toolkits.encode_frame_to_jpeg(frame, quality=85, max_width=maxw)
        if jpeg_bytes is None:
            return jsonify({"ok": False, "message": "JPEG 编码失败"}), 500
        import base64
        return jsonify({
            "ok": True,
            "video_width": int(w),
            "video_height": int(h),
            "jpeg": base64.b64encode(jpeg_bytes).decode("ascii"),
            "mime": "image/jpeg",
        })

    @app.route("/api/project/<path:project_name>/qa_start", methods=["POST"])
    def api_qa_start(project_name):
        """新版质检启动接口（对应独立工具 开始检测 按钮）。

        body 字段:
            project_path (必填): 项目目录绝对路径
            cp_folder, hardsub_folders[], srt_folder: 版本选择（通常来自 /api/qa/scan_dir 返回）
            opt_blackframes / opt_hardsubs / opt_duration: 检测项开关
            workers: 线程数
            sub_region: [x, y, w, h] 或 null（字幕区域框选）
            folder_layout: 可选，预扫描的 layout 数据
        """
        data = request.get_json(silent=True) or {}
        project_path = data.get("project_path") or ""
        if not project_path or not _os.path.isdir(project_path):
            for root, _pk in _nas_search_roots():
                candidate = _os.path.join(root, project_name)
                if _os.path.isdir(candidate):
                    project_path = candidate
                    break
        if not project_path or not _os.path.isdir(project_path):
            return jsonify({"ok": False, "message": "找不到项目路径"}), 404
        if qa_engine is None:
            return jsonify({"ok": False, "message": "QA引擎未初始化"}), 500

        opts = {
            'cp_folder': data.get('cp_folder'),
            'hardsub_folders': data.get('hardsub_folders'),
            'srt_folder': data.get('srt_folder'),
            'opt_blackframes': bool(data.get('opt_blackframes', True)),
            'opt_hardsubs': bool(data.get('opt_hardsubs', True)),
            'opt_duration': bool(data.get('opt_duration', True)),
            'workers': int(data.get('workers', 4)),
            'sub_region': data.get('sub_region'),
        }
        folder_layout = data.get('folder_layout') or None

        ok = qa_engine.run(project_path, project_name,
                           workers=opts['workers'],
                           opts=opts, folder_layout=folder_layout)
        # 质检启动成功 → 工作流状态自动流转为"质检中"
        if ok:
            try:
                cur = (db.get_project(project_name) or {}).get("custom_status") or ""
                if str(cur).strip() != "质检中":
                    db.update_project_status(
                        project_name, custom_status="质检中",
                        sync_progress="质检进行中...")
            except Exception as _e:
                _logger.warning("质检启动后设置质检中失败: %s", _e)
        return jsonify({"ok": ok})

    @app.route("/api/project/<path:project_name>/qa_cancel", methods=["POST"])
    def api_qa_cancel(project_name):
        if qa_engine:
            qa_engine.cancel(project_name)
        return jsonify({"ok": True})

    @app.route("/api/project/<path:project_name>/qa_checkpoint_clear", methods=["POST"])
    def api_qa_checkpoint_clear(project_name):
        """清除项目目录下的断点续检 checkpoint（对应独立工具 清除缓存 按钮）。"""
        import qa_toolkits
        data = request.get_json(silent=True) or {}
        p = data.get("project_path") or ""
        if not p or not _os.path.isdir(p):
            for root, _pk in _nas_search_roots():
                candidate = _os.path.join(root, project_name)
                if _os.path.isdir(candidate):
                    p = candidate
                    break
        if not p or not _os.path.isdir(p):
            return jsonify({"ok": False, "message": "找不到项目路径"}), 404
        qa_toolkits.clear_checkpoint(p)
        return jsonify({"ok": True})

    @app.route("/api/project/<path:project_name>/qa_status", methods=["GET"])
    def api_qa_status(project_name):
        if qa_engine is None:
            return jsonify({"ok": False, "message": "QA引擎未初始化"}), 500
        # get_status 返回 {is_running, progress, current_video, total, done,
        # passed, warnings, failed, results, log, report_html, qa_result_file}
        status = qa_engine.get_status(project_name)
        return jsonify({"ok": True, **status})

    @app.route("/api/qa/summary", methods=["GET"])
    def api_qa_summary():
        """质检结果统计：各部门/项目质检通过率、最近质检记录、运行中任务。"""
        try:
            runs = db.list_all_qa_runs(limit=200) or []
            # 只统计已完成的质检
            done_runs = [r for r in runs if r.get("status") in ("done", "completed")]
            total_run = len(done_runs)
            total_videos = sum(int(r.get("total") or 0) for r in done_runs)
            total_pass = sum(int(r.get("passed") or 0) for r in done_runs)
            total_warn = sum(int(r.get("warnings") or 0) for r in done_runs)
            total_fail = sum(int(r.get("failed") or 0) for r in done_runs)
            pass_rate = round(total_pass * 100.0 / total_videos, 1) if total_videos else 0

            # 按项目聚合最近结果
            by_project = {}
            for r in runs:
                name = r.get("project_name") or ""
                if not name:
                    continue
                item = by_project.setdefault(name, {
                    "project_name": name, "last_total": 0, "last_pass": 0,
                    "last_warn": 0, "last_fail": 0, "last_time": "", "status": "done",
                })
                # 保留最近一次（按 started_at 排序）
                last = item["last_time"]
                if r.get("status") in ("done", "completed") and (not last or str(r.get("started_at") or "") > last):
                    item.update({
                        "last_total": int(r.get("total") or 0),
                        "last_pass": int(r.get("passed") or 0),
                        "last_warn": int(r.get("warnings") or 0),
                        "last_fail": int(r.get("failed") or 0),
                        "last_time": str(r.get("started_at") or ""),
                    })

            # 运行中任务
            running = [r for r in runs if r.get("status") in ("running", "pending")]
            active_names = [r.get("project_name") for r in running]

            # 按状态归类项目
            by_status = {"pass": [], "warn": [], "fail": [], "running": []}
            for name, item in by_project.items():
                if name in active_names:
                    by_status["running"].append(item)
                elif item["last_fail"] > 0:
                    by_status["fail"].append(item)
                elif item["last_warn"] > 0:
                    by_status["warn"].append(item)
                else:
                    by_status["pass"].append(item)

            return jsonify({
                "ok": True,
                "summary": {
                    "total_run": total_run,
                    "total_videos": total_videos,
                    "total_pass": total_pass,
                    "total_warn": total_warn,
                    "total_fail": total_fail,
                    "pass_rate": pass_rate,
                },
                "by_status": {
                    "pass": by_status["pass"],
                    "warn": by_status["warn"],
                    "fail": by_status["fail"],
                    "running": by_status["running"],
                },
                "running_count": len(active_names),
            })
        except Exception as e:
            _logger.exception("质检统计失败")
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/qa/batch_report", methods=["GET"])
    def api_qa_batch_report():
        """生成跨项目的质检批量汇总报告，并下载 HTML。

        Query:
            dl: 1 时强制下载，否则内联显示。
        """
        import qa_toolkits
        force_dl = request.args.get("dl") == "1"
        try:
            runs = db.list_all_qa_runs(limit=200) or []
            html_path = qa_toolkits.generate_batch_report_html(runs)
            if not _os.path.isfile(html_path):
                return jsonify({"ok": False, "message": "批量报告生成失败"}), 500
            directory = _os.path.dirname(html_path)
            filename = _os.path.basename(html_path)
            return send_from_directory(directory, filename, as_attachment=force_dl)
        except Exception as e:
            _logger.exception("生成批量质检报告失败")
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/qa/batch_start", methods=["POST"])
    def api_qa_batch_start():
        """批量启动质检：接收项目名列表，对每个项目定位目录并启动质检。

        body: { projects: ["项目A", "项目B", ...], workers: 4 }
        返回: { ok, started: [已启动项目], skipped: [跳过项目及原因] }
        """
        data = request.get_json(silent=True) or {}
        projects = data.get("projects") or []
        workers = int(data.get("workers") or 4)
        if not projects:
            return jsonify({"ok": False, "message": "未选择项目"}), 400
        if qa_engine is None:
            return jsonify({"ok": False, "message": "QA引擎未初始化"}), 500

        started, skipped = [], []
        for name in projects:
            name = str(name or "").strip()
            if not name:
                continue
            # 定位项目目录（组内优先，其次制作部）
            path = ""
            proj = db.get_project(name)
            if proj:
                gp = proj.get("group_path") or ""
                if gp and _os.path.isdir(gp):
                    path = gp
                else:
                    pp = proj.get("production_path") or ""
                    if pp and _os.path.isdir(pp):
                        path = pp
            if not path:
                for root, _pk in _nas_search_roots():
                    candidate = _os.path.join(root, name)
                    if _os.path.isdir(candidate):
                        path = candidate
                        break
            if not path or not _os.path.isdir(path):
                skipped.append({"name": name, "reason": "找不到项目目录"})
                continue
            try:
                ok = qa_engine.run(path, name, workers=workers)
                if ok:
                    started.append(name)
                    try:
                        cur = (db.get_project(name) or {}).get("custom_status") or ""
                        if str(cur).strip() != "质检中":
                            db.update_project_status(name, custom_status="质检中")
                    except Exception:
                        pass
                else:
                    skipped.append({"name": name, "reason": "质检已在运行或启动失败"})
            except Exception as e:
                skipped.append({"name": name, "reason": str(e)})

        return jsonify({
            "ok": True,
            "started": started,
            "skipped": skipped,
            "started_count": len(started),
            "skipped_count": len(skipped),
        })

    @app.route("/api/project/<path:project_name>/qa_report", methods=["GET"])
    def api_qa_report_download(project_name):
        """下载 / 预览上一次质检生成的 HTML 报告或 JSON 数据。

        Query:
            fmt: 'html'（默认，返回 HTML 并内联浏览器显示） | 'json'（下载 JSON）
            dl:  1 时强制下载（Content-Disposition: attachment）
        """
        import qa_toolkits
        fmt = request.args.get('fmt', 'html').lower()
        force_dl = request.args.get('dl') == '1'
        if qa_engine is None:
            return jsonify({"ok": False, "message": "QA引擎未初始化"}), 500

        cached = qa_engine.get_last_result(project_name)
        if not cached:
            return jsonify({"ok": False, "message": "暂无质检结果，请先运行质检"}), 404

        if fmt == 'json':
            payload = qa_toolkits.generate_report_json(
                cached['project_path'], cached['project_name'],
                cached['results'], cached.get('extra_data'),
            )
            # 用 BytesIO 包装，文件名自动加时间戳
            from datetime import datetime as _dt
            fname = f"质检数据_{project_name}_{_dt.now().strftime('%Y%m%d_%H%M%S')}.json"
            buf = _io.BytesIO(_j.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"))
            buf.seek(0)
            return send_file(buf, as_attachment=force_dl or True,
                             download_name=fname, mimetype="application/json")

        # 默认 html
        html_path = cached.get('html_path')
        if not html_path or not _os.path.isfile(html_path):
            return jsonify({"ok": False, "message": "HTML 报告不存在或未生成"}), 404
        directory = _os.path.dirname(html_path)
        filename = _os.path.basename(html_path)
        return send_from_directory(directory, filename, as_attachment=force_dl)

    @app.route("/api/project/<path:project_name>/qa_lastresult", methods=["GET"])
    def api_qa_lastresult(project_name):
        """取上一次质检结果缓存（含 raw results / extra_data / 报告路径）。"""
        if qa_engine is None:
            return jsonify({"ok": False, "message": "QA引擎未初始化"}), 500
        cached = qa_engine.get_last_result(project_name)
        if not cached:
            return jsonify({"ok": False, "message": "暂无缓存结果"}), 404
        # results 可能很大：只保留摘要 + 报告路径 + 统计
        summary = {
            'total': len(cached['results']),
            'passed': sum(1 for r in cached['results'] if r.get('status') == 'pass'),
            'warnings': sum(1 for r in cached['results'] if r.get('status') == 'warn'),
            'failed': sum(1 for r in cached['results'] if r.get('status') == 'fail'),
        }
        return jsonify({
            "ok": True,
            "summary": summary,
            "html_path": cached.get('html_path'),
            "qa_result_file": cached.get('qa_result_file'),
            "extra_data": cached.get('extra_data'),
            "timestamp": cached.get('timestamp'),
            "results": cached['results'],
        })

    @app.route("/api/project/<path:project_name>/qa_history", methods=["GET"])
    def api_qa_history(project_name):
        try:
            runs = db.list_qa_runs_for_project(project_name, limit=20)
            return jsonify({"ok": True, "runs": runs})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/projects/episode_summary", methods=["GET"])
    def api_episode_summary():
        result = {"ok": True, "projects": []}
        all_names = set()
        for path, skip, depth in _episode_summary_scan_tasks():
            for it in scan_dir(path, skip_names=skip, recursive_depth=depth):
                all_names.add(it["name"])
        for name in all_names:
            try:
                p = db.get_project(name)
                if not p:
                    continue
                total = p.get("total_episodes") or 0
                current = p.get("current_episodes") or 0
                if total > 0 and current < total:
                    result["projects"].append({
                        "name": name, "total": total, "current": current,
                        "missing": total - current,
                    })
            except Exception:
                pass
        result["projects"].sort(key=lambda x: -x["missing"])
        return jsonify(result)

    @app.route("/api/team/members", methods=["GET"])
    def api_team_members_get():
        try:
            if hasattr(db, 'list_team_members'):
                members = db.list_team_members()
            elif hasattr(db, 'list_members'):
                members = db.list_members()
            else:
                members = []
            return jsonify({"ok": True, "members": members})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/team", methods=["GET"])
    def api_team_alias():
        """简洁别名，返回团队成员列表"""
        return api_team_members_get()

    @app.route("/api/team/members", methods=["POST"])
    def api_team_members_add():
        """添加成员"""
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        role = data.get("role") or "editor"
        title = data.get("title") or ""
        department = data.get("department") or ""
        if not name:
            return jsonify({"ok": False, "message": "姓名不能为空"}), 400
        try:
            db.add_member(name=name, role=role, title=title, department=department)
            return jsonify({"ok": True, "message": f"已添加 {name}"})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/team/members/<int:mid>", methods=["PUT"])
    def api_team_members_update(mid):
        """编辑成员：改名、改角色、改称号、改部门"""
        data = request.get_json(silent=True) or {}
        try:
            with db.get_conn() as conn:
                row = conn.execute("SELECT * FROM team_members WHERE id=?", (mid,)).fetchone()
                if not row:
                    return jsonify({"ok": False, "message": "成员不存在"}), 404
                old_name = row["name"]
            if "name" in data and data["name"].strip():
                new_name = data["name"].strip()
                with db.get_conn() as conn:
                    conn.execute("UPDATE team_members SET name=? WHERE id=?", (new_name, mid))
                old_name = new_name
            kwargs = {}
            if "role" in data and data["role"] in ("editor", "reviewer", "pm"):
                kwargs["role"] = data["role"]
            if "title" in data:
                kwargs["title"] = data["title"] or ""
            if "department" in data:
                kwargs["department"] = data["department"] or ""
            if kwargs:
                db.update_member(old_name, **kwargs)
            return jsonify({"ok": True, "message": "已更新"})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/team/members/<int:mid>", methods=["DELETE"])
    def api_team_members_delete(mid):
        """删除成员（按 id）"""
        try:
            with db.get_conn() as conn:
                row = conn.execute("SELECT name FROM team_members WHERE id=?", (mid,)).fetchone()
                if not row:
                    return jsonify({"ok": False, "message": "成员不存在"}), 404
                conn.execute("DELETE FROM team_members WHERE id=?", (mid,))
            return jsonify({"ok": True, "message": "已删除"})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/config", methods=["GET"])
    def api_config_get():
        """返回扁平化配置（yaml 视角）。前端设置页按此读取。"""
        try:
            cfg = _yaml_cfg()
            flat = {
                "group_root": (cfg.get("nas", {}) or {}).get("group_root", ""),
                "qa_workers": ((cfg.get("qa", {}) or {}).get("workers", 4)),
                "ffmpeg_path": ((cfg.get("qa", {}) or {}).get("ffmpeg_path", "")
                                or (cfg.get("paths", {}) or {}).get("ffmpeg", "")),
                "ffprobe_path": ((cfg.get("qa", {}) or {}).get("ffprobe_path", "")
                                 or (cfg.get("paths", {}) or {}).get("ffprobe", "")),
                "suggested_names": (cfg.get("fenji", {}) or {}).get("suggested_names", []),
            }
            return jsonify({"ok": True, "config": flat})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/config", methods=["POST"])
    def api_config_save():
        """把扁平设置写回 config.yaml 的嵌套结构，与 app.py 读取同源。"""
        data = request.get_json(silent=True) or {}
        try:
            cfg = _yaml_cfg()
            cfg.setdefault("nas", {})
            cfg.setdefault("qa", {})
            cfg.setdefault("paths", {})
            cfg.setdefault("fenji", {})

            if "group_root" in data:
                cfg["nas"]["group_root"] = (data.get("group_root") or "").strip()
            if "qa_workers" in data:
                try:
                    cfg["qa"]["workers"] = int(data.get("qa_workers") or 4)
                except (TypeError, ValueError):
                    cfg["qa"]["workers"] = 4
            if "ffmpeg_path" in data:
                cfg["qa"]["ffmpeg_path"] = (data.get("ffmpeg_path") or "").strip()
            if "ffprobe_path" in data:
                cfg["qa"]["ffprobe_path"] = (data.get("ffprobe_path") or "").strip()
            if "suggested_names" in data:
                names = data.get("suggested_names")
                if isinstance(names, list):
                    cfg["fenji"]["suggested_names"] = [str(n).strip() for n in names if str(n).strip()]
                elif isinstance(names, str):
                    cfg["fenji"]["suggested_names"] = [n.strip() for n in names.split("\n") if n.strip()]

            _yaml_save(cfg)
            # 同步回主流程配置（若 app 侧持有同一 dict 引用）
            try:
                import app as _app_mod
                if hasattr(_app_mod, "config"):
                    _app_mod.config.update(cfg)
                if hasattr(_app_mod, "reload_sync_engine"):
                    _app_mod.reload_sync_engine()
            except Exception:
                pass
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/project/<path:project_name>/episodes_status", methods=["GET"])
    def api_episodes_status(project_name):
        try:
            if sync_engine is None:
                return jsonify({"ok": False, "message": "sync_engine 不可用"}), 500
            result = sync_engine.get_episode_status(project_name)
            return jsonify(result)
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500


    @app.route("/api/bulk/import_episodes", methods=["POST"])
    def api_bulk_import_episodes():
        """接收前端分集结果，保存到 DB。"""
        body = request.get_json(silent=True) or {}
        project_name = body.get("project_name") or ""
        total = int(body.get("total_episodes") or 0)
        assign = body.get("assign") or {}
        if not project_name:
            return jsonify({"ok": False, "message": "缺少 project_name"}), 400
        if not isinstance(assign, dict) or len(assign) == 0:
            return jsonify({"ok": False, "message": "assign 为空"}), 400
        try:
            db.set_episode_plan(project_name, assign)
            if total > 0:
                p = db.get_project(project_name)
                cur = int((p or {}).get("current_episodes") or 0)
                db.set_episodes(project_name, total, cur)
            return jsonify({"ok": True, "message": "已保存 " + str(len(assign)) + " 集分集", "count": len(assign)})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/project/<path:project_name>/episodes_plan", methods=["GET"])
    def api_episodes_plan(project_name):
        """读取项目的分集 plan。"""
        try:
            plan = db.get_episode_plan(project_name)
            p = db.get_project(project_name) or {}
            total = int(p.get("total_episodes") or 0)
            return jsonify({"ok": True, "plan": plan, "total_episodes": total, "summary": {}})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    @app.route("/api/fenji/suggest", methods=["GET"])
    def api_fenji_suggest():
        """自动建议均分方案。"""
        total = int(request.args.get("total") or 0)
        count = int(request.args.get("count") or 0)
        if total <= 0 or count <= 0:
            return jsonify({"ok": False, "message": "参数错误"}), 400
        per = total // count
        rem = total % count
        ranges = []
        cur = 1
        for i in range(count):
            sz = per + (1 if i < rem else 0)
            ranges.append({"start": cur, "end": cur + sz - 1})
            cur += sz

    # ===========================
    # 分秒帧上传相关路由
    # ===========================
    _FM_LINKS_FILE = _os.path.join(
        _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
        'data', 'fenmiaozhen_links.json'
    )

    def _fm_load_links():
        try:
            with open(_FM_LINKS_FILE, 'r', encoding='utf-8') as _f:
                return _j.load(_f)
        except (FileNotFoundError, Exception):
            return {}

    def _fm_save_links(data):
        _os.makedirs(_os.path.dirname(_FM_LINKS_FILE), exist_ok=True)
        with open(_FM_LINKS_FILE, 'w', encoding='utf-8') as _f:
            _j.dump(data, _f, ensure_ascii=False, indent=2)

    @app.route('/api/fenmiaozhen/config', methods=['GET'])
    def fenmiaozhen_config():
        _cp = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'config.yaml')
        with open(_cp, 'r', encoding='utf-8') as _f:
            _cfg = _y.safe_load(_f) or {}
        fm = _cfg.get('fenmiaozhen', {}) or {}
        return jsonify(
            ok=True,
            enabled_departments=fm.get('enabled_departments', []),
            web_url=fm.get('web_url', 'https://www.mediatrack.cn/project/new'),
        )

    @app.route('/api/fenmiaozhen/link/<path:name>', methods=['GET'])
    def fenmiaozhen_get_link(name):
        links = _fm_load_links()
        url = links.get(name, '')
        return jsonify(ok=True, url=url, has_link=bool(url))

    @app.route('/api/fenmiaozhen/link/<path:name>', methods=['POST'])
    def fenmiaozhen_save_link(name):
        try:
            body = request.get_json(force=True, silent=True) or {}
            url = (body.get('url') or '').strip()
        except Exception:
            url = ''
        if not url:
            return jsonify(ok=False, msg='链接不能为空'), 400
        links = _fm_load_links()
        links[name] = url
        _fm_save_links(links)
        return jsonify(ok=True, url=url)

    # ============================================================
    # 分集导出到 Excel 模板（从独立分集程序搬过来）
    # ============================================================
    @app.route('/api/fenji/templates', methods=['GET'])
    def fenji_list_templates():
        names = _list_templates(_TEMPLATE_DIR)
        return jsonify(ok=True, templates=names, dir=_TEMPLATE_DIR)

    @app.route('/api/fenji/upload_template', methods=['POST'])
    def fenji_upload_template():
        _os.makedirs(_TEMPLATE_DIR, exist_ok=True)
        if 'file' not in request.files:
            return jsonify(ok=False, msg='无文件'), 400
        f = request.files['file']
        if not f.filename:
            return jsonify(ok=False, msg='文件名为空'), 400
        safe_name = _os.path.basename(f.filename)  # 防路径穿越
        save_path = _os.path.join(_TEMPLATE_DIR, safe_name)
        f.save(save_path)
        size = _os.path.getsize(save_path)
        return jsonify(ok=True, name=safe_name, size=size, dir=_TEMPLATE_DIR)

    # ============ Excel 分集表同步到项目 ============
    @app.route('/api/fenji/sync_from_excel', methods=['POST'])
    def fenji_sync_from_excel():
        """解析用户上传的分集 Excel（项目名 + 每行'剪辑师：集数范围'），
        把每个项目的分集 plan 同步到数据库 episode_plan，供工作量统计。"""
        import openpyxl
        if 'file' not in request.files:
            return jsonify(ok=False, msg='无文件'), 400
        f = request.files['file']
        if not f.filename:
            return jsonify(ok=False, msg='文件名为空'), 400

        projects_plan = {}   # 项目名 -> {集号: 剪辑师}
        try:
            wb = openpyxl.load_workbook(_io.BytesIO(f.read()), data_only=True)
            ws = wb.active if wb.sheetnames else None
            if ws is None:
                return jsonify(ok=False, msg='Excel 无工作表'), 400

            cur_project = None
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or not any(row):
                    continue
                proj_name = str(row[0] or '').strip() if row[0] else ''
                assign_str = str(row[2] or '').strip() if len(row) > 2 else ''
                if proj_name:
                    cur_project = proj_name
                    projects_plan.setdefault(cur_project, {})
                elif not cur_project:
                    continue
                if assign_str and cur_project:
                    _parse_assign_line(projects_plan[cur_project], assign_str)
        except Exception as e:
            return jsonify(ok=False, msg='Excel 解析失败: ' + str(e)), 400

        if not projects_plan:
            return jsonify(ok=False, msg='未解析到任何分集数据'), 400

        synced, skipped = [], []
        for name, plan in projects_plan.items():
            if not plan:
                skipped.append({'name': name, 'episodes': 0, 'reason': '无分集数据'})
                continue
            proj = db.get_project(name)
            if not proj:
                all_proj = db.get_all_projects()
                match = None
                for ap in all_proj:
                    an = ap.get('name') or ''
                    if an == name or an.startswith(name[:15]) or name.startswith(an[:15]):
                        match = ap
                        break
                proj = match
            if not proj:
                skipped.append({'name': name, 'episodes': len(plan), 'reason': '未找到对应项目'})
                continue
            existing = db.get_episode_plan(proj['name'])
            existing.update(plan)
            db.set_episode_plan(proj['name'], existing)
            total = int(proj.get('total_episodes') or 0)
            if total < len(existing):
                db.set_episodes(proj['name'], len(existing), len(existing))
            synced.append({'name': proj['name'], 'episodes': len(plan)})

        return jsonify({
            'ok': True,
            'message': '同步完成: {} 个项目，{} 集'.format(len(synced), sum(s['episodes'] for s in synced)),
            'synced': synced,
            'skipped': skipped,
        })

    # ============ 从权威目标文件重建各项目分集工作量 ============
    @app.route('/api/fenji/sync_episode_plan', methods=['POST'])
    def fenji_sync_episode_plan():
        """以当前分集数据(episode_plan)为准，重建/统一各项目工作量统计。
        统计口径始终以分集数据为准：从每个项目的 episode_plan 聚合每人集数，
        写入 editor_workload（供统计使用），不读取目标 Excel、不覆盖分集数据。
        返回 {ok, synced, total_projects, total_episodes}
        """
        try:
            from features import sync_workload_from_episode_plan
            res = sync_workload_from_episode_plan(db)
            return jsonify(ok=True, mode='分集数据', **res)
        except Exception as e:
            import traceback; traceback.print_exc()
            return jsonify(ok=False, msg='同步失败: ' + str(e)), 500

    # ============ 用户设置持久化（模板选择/目标文件路径等） ============
    @app.route('/api/settings', methods=['GET'])
    def api_get_settings():
        """读取全部用户设置（模板选择、目标文件路径等），重开软件自动恢复。"""
        return jsonify(ok=True, settings=db.get_all_settings())

    @app.route('/api/settings', methods=['PUT'])
    def api_set_settings():
        """批量保存用户设置。body: { key: value, ... }"""
        data = request.get_json(silent=True) or {}
        saved = {}
        for k, v in data.items():
            db.set_setting(k, v)
            saved[k] = v
        return jsonify(ok=True, saved=saved)

    # ============ 分集人员模板（勾选人员保存为模板） ============
    @app.route('/api/fenji/person_templates', methods=['GET'])
    def fenji_list_person_templates():
        """列出已保存的人员模板。存储为 JSON: {"模板名": ["人员1", "人员2", ...]}"""
        raw = db.get_setting('fenji_person_templates', '{}')
        try:
            tpls = _j.loads(raw) if isinstance(raw, str) and raw.strip() else {}
        except Exception:
            tpls = {}
        return jsonify(ok=True, templates=tpls)

    @app.route('/api/fenji/person_templates', methods=['POST'])
    def fenji_save_person_template():
        """保存人员模板。body: { name: "模板名", persons: ["A","B",...] }"""
        data = request.get_json(silent=True) or {}
        name = (data.get('name') or '').strip()
        persons = data.get('persons') or []
        if not name or not isinstance(persons, list) or not persons:
            return jsonify(ok=False, msg='模板名或人员列表不能为空'), 400
        raw = db.get_setting('fenji_person_templates', '{}')
        try:
            tpls = _j.loads(raw) if isinstance(raw, str) and raw.strip() else {}
        except Exception:
            tpls = {}
        tpls[name] = [str(p).strip() for p in persons if str(p).strip()]
        db.set_setting('fenji_person_templates', _j.dumps(tpls, ensure_ascii=False))
        return jsonify(ok=True, templates=tpls)

    @app.route('/api/fenji/person_templates', methods=['DELETE'])
    def fenji_delete_person_template():
        """删除人员模板。body: { name: "模板名" }"""
        data = request.get_json(silent=True) or {}
        name = (data.get('name') or '').strip()
        raw = db.get_setting('fenji_person_templates', '{}')
        try:
            tpls = _j.loads(raw) if isinstance(raw, str) and raw.strip() else {}
        except Exception:
            tpls = {}
        if name in tpls:
            del tpls[name]
            db.set_setting('fenji_person_templates', _j.dumps(tpls, ensure_ascii=False))
        return jsonify(ok=True, templates=tpls)

    # ============ 目标文件上传（选择本地 Excel 作为累积目标） ============
    _TARGET_DIR = _os.path.join(_DATA_DIR, 'fenji_targets')

    @app.route('/api/fenji/targets', methods=['GET'])
    def fenji_list_targets():
        """列出已上传的目标文件（累积汇总 Excel）。"""
        try:
            names = []
            if _os.path.isdir(_TARGET_DIR):
                for fn in sorted(_os.listdir(_TARGET_DIR)):
                    if fn.endswith(('.xlsx', '.xls', '.xlsm')):
                        names.append(fn)
            return jsonify(ok=True, targets=names, dir=_TARGET_DIR)
        except Exception as e:
            return jsonify(ok=False, msg=str(e)), 500

    @app.route('/api/fenji/upload_target', methods=['POST'])
    def fenji_upload_target():
        """上传一个本地 Excel 作为目标汇总文件，返回其服务端路径（供导出追加）。"""
        _os.makedirs(_TARGET_DIR, exist_ok=True)
        if 'file' not in request.files:
            return jsonify(ok=False, msg='无文件'), 400
        f = request.files['file']
        if not f.filename:
            return jsonify(ok=False, msg='文件名为空'), 400
        safe_name = _os.path.basename(f.filename)
        save_path = _os.path.join(_TARGET_DIR, safe_name)
        f.save(save_path)
        return jsonify(ok=True, name=safe_name, path=save_path)

    @app.route('/api/fenji/targets', methods=['DELETE'])
    def fenji_delete_target():
        """删除已上传的目标文件。body: { name: "文件名" }"""
        data = request.get_json(silent=True) or {}
        name = (data.get('name') or '').strip()
        safe = _os.path.basename(name)
        p = _os.path.join(_TARGET_DIR, safe)
        if _os.path.isfile(p):
            try:
                _os.remove(p)
            except Exception:
                pass
        return jsonify(ok=True)

    @app.route('/api/fenji/export_excel', methods=['POST'])
    def fenji_export_excel():
        """
        前端传 JSON:
        {
          template_b64: "...",        // 二选一：base64 编码的模板内容
          template_name: "xxx.xlsx",  // 二选一：已上传到服务端的模板文件名
          originalTemplateName: "xxx.xlsx",  // 备份时用的原始文件名
          projectName: "溯回时光见旧人",
          path: "O:\\...\\溯回时光见旧人",
          timeText: "8.15下午18点交",
          statusText: "已分集",
          assign: [{"person": "张三", "range": "1-25"}, ...]
        }
        返回: {ok, file_b64, fileName, backup: {name, saved}}
        """
        body = request.get_json(silent=True) or {}
        assign = body.get('assign') or []
        if not assign:
            return jsonify(ok=False, msg='assign 为空'), 400

        # 1. 获取基准 bytes
        target_path = (body.get('target_path') or '').strip()
        use_target = bool(target_path)
        tpl_bytes = None
        original_name = (body.get('originalTemplateName')
                         or body.get('template_name')
                         or '模板.xlsx')

        tpl_b64 = body.get('template_b64') or body.get('template')
        if tpl_b64:
            try:
                tpl_bytes = _b64.b64decode(tpl_b64)
            except Exception as e:
                return jsonify(ok=False, msg=f'template_b64 解码失败: {e}'), 400
        else:
            tpl_name = body.get('template_name') or ''
            if not tpl_name:
                return jsonify(ok=False, msg='必须提供 template_b64 或 template_name'), 400
            safe = _os.path.basename(tpl_name)
            tpl_path = _os.path.join(_TEMPLATE_DIR, safe)
            if not _os.path.isfile(tpl_path):
                return jsonify(ok=False, msg=f'模板文件不存在: {safe}'), 404
            with open(tpl_path, 'rb') as f:
                tpl_bytes = f.read()

        # 1.5 关键改进: 如果指定了 target_path 且目标文件已存在, 从目标文件读取 (它是累积源)
        #     只有目标文件不存在时才用干净模板启动
        actually_read_from = 'template'
        if use_target:
            if _os.path.isfile(target_path):
                try:
                    with open(target_path, 'rb') as f:
                        tpl_bytes = f.read()
                    actually_read_from = 'target'
                    _logger.info('从目标文件读取 (已有数据): %s', target_path)
                except Exception as e:
                    _logger.warning('读取目标文件失败, 回退到模板: %s', e)
            else:
                _logger.info('目标文件不存在, 从模板启动: %s', target_path)

        # 2. 备份 (备份的是实际读取的那个)
        backup_name, backup_path = '', ''
        try:
            backup_name, backup_path = backup_template(tpl_bytes, original_name, _BACKUP_DIR)
        except Exception as e:
            _logger.warning('备份模板失败: %s', e)

        # 3. 追加 + 美化
        try:
            new_bytes = export_from_template(
                tpl_bytes,
                project_name=body.get('projectName', ''),
                path=body.get('path', ''),
                assign_list=assign,
                time_text=body.get('timeText', ''),
                status_text=body.get('statusText', '已分集'),
            )
        except Exception as e:
            _logger.exception('export_excel 失败')
            return jsonify(ok=False, msg=f'导出失败: {e}'), 500

        # 4. 写回策略:
        #    - 有 target_path: 只写目标文件, 不动模板 (模板保持干净, 目标才是累积源)
        #    - 没 target_path: 写回模板 (累积效果, 兼容"导出预览"里也保留旧数据)
        actually_saved = ''
        if use_target:
            try:
                safe_tgt = _os.path.abspath(target_path)
                _os.makedirs(_os.path.dirname(safe_tgt), exist_ok=True)
                with open(safe_tgt, 'wb') as f:
                    f.write(new_bytes)
                actually_saved = safe_tgt
                _logger.info('✅ 已追加到目标: %s (读取源=%s)', safe_tgt, actually_read_from)
            except Exception as e:
                _logger.exception('写目标文件失败')
                return jsonify(ok=False, msg=f'保存目标文件失败: {e}'), 500
            # 自动同步：写入目标后，自动从表格读取胶片日期更新项目交付日期（不阻塞响应）
            try:
                import threading as _th
                import features as _feat
                _th.Thread(
                    target=_auto_sync_delivery_dates,
                    args=(safe_tgt,),
                    daemon=True,
                ).start()
            except Exception as e:
                _logger.warning('自动同步交付日期失败: %s', e)
        else:
            # 只在没有 target_path 时才写回模板 (保持旧行为兼容)
            if tpl_name and not tpl_b64:
                try:
                    with open(tpl_path, 'wb') as f:
                        f.write(new_bytes)
                    _logger.info('已写回模板 %s (大小 %d 字节)', tpl_name, len(new_bytes))
                except Exception as e:
                    _logger.warning('写回模板失败: %s', e)

        # 5. 自动打开 Excel
        open_excel = bool(body.get('open_excel'))
        if open_excel and actually_saved:
            try:
                import subprocess
                subprocess.Popen(['start', '', actually_saved], shell=True)
                _logger.info('已用 Excel 打开 %s', actually_saved)
            except Exception as e:
                _logger.warning('打开 Excel 失败: %s', e)

        out_name = _re.sub(r'\.[^.]+$', '', original_name) + '_已分集.xlsx'
        resp = {
            'ok': True,
            'file_b64': _b64.b64encode(new_bytes).decode('ascii'),
            'fileName': out_name,
            'backup': {'name': backup_name, 'saved': backup_path},
        }
        if actually_saved:
            resp['target_saved'] = actually_saved
        return jsonify(resp)

    @app.route('/api/fenji/save_to_folder', methods=['POST'])
    def fenji_save_to_folder():
        """把已生成的 Excel 文件保存到指定文件夹。"""
        body = request.get_json(silent=True) or {}
        folder = (body.get('folder') or '').strip()
        file_b64 = body.get('fileB64') or body.get('file_b64') or ''
        fname = body.get('fileName') or '分集结果.xlsx'
        open_excel = bool(body.get('open'))

        if not folder:
            return jsonify(ok=False, msg='未指定文件夹'), 400
        if not file_b64:
            return jsonify(ok=False, msg='无文件内容'), 400
        if not _os.path.isdir(folder):
            return jsonify(ok=False, msg=f'文件夹不存在: {folder}'), 400

        try:
            file_bytes = _b64.b64decode(file_b64)
        except Exception as e:
            return jsonify(ok=False, msg=f'base64 解码失败: {e}'), 400

        safe_fname = _os.path.basename(fname)
        save_path = _os.path.join(folder, safe_fname)
        with open(save_path, 'wb') as f:
            f.write(file_bytes)

        if open_excel:
            try:
                import subprocess
                subprocess.Popen(['start', save_path], shell=True)
            except Exception:
                pass
        return jsonify(ok=True, saved=save_path)

